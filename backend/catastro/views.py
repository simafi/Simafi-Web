from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.hashers import check_password
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.utils import timezone
from django.db import IntegrityError
from django.db.models import Sum, Count, Q, IntegerField, Value
from django.db.models.functions import Cast, Coalesce
from django.urls import reverse
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
import io

from usuarios.models import Usuario
from usuarios.permisos import permisos_catastro_menu_context
from catastro.decorators import catastro_require_permiso
from catastro.permisos_codigos import (
    CATASTRO_PERM_BIENES_EDITAR,
    CATASTRO_PERM_BIENES_VER,
    CATASTRO_PERM_CONFIG_VER,
    CATASTRO_PERM_MAPA_VER,
    CATASTRO_PERM_MENU_VER,
)
from core.models import Municipio
from core.module_access import codigos_empresa_equivalentes
from .models import *
from .forms import CatastroLoginForm, EdificacionForm, EdificacionEspecialForm, CostosForm, DetalleAdicionalesForm, BarriosForm, TopografiaForm, UsosPredioForm, ConfiTipologiaForm, EspecificacionesForm, TipoMaterialForm, DetEspecificacionForm, ComentariosCatastroForm
from .models import TipoDetalle, Barrios, UsoEdifica

# Inicializar logger antes de usarlo
logger = logging.getLogger(__name__)

User = get_user_model()


def _costos_por_municipio_qs(request):
    """
    Costos básicos unitarios (tabla `costos`) del municipio de la sesión.
    Usa variantes de código (p. ej. 0502 / 502) como en el login modular.
    """
    cod = (request.session.get('catastro_empresa') or request.session.get('empresa') or '').strip()
    if not cod:
        return Costos.objects.none()
    empresas = codigos_empresa_equivalentes(cod)
    if not empresas:
        return Costos.objects.none()
    return Costos.objects.filter(empresa__in=empresas)


# Importar utilidades de conversión de coordenadas
# Inicializar variables globales
COORDENADAS_UTILS_AVAILABLE = False
utm_to_latlng = None
latlng_to_utm = None

def _importar_utilidades_coordenadas():
    """Función helper para importar utilidades de coordenadas"""
    global COORDENADAS_UTILS_AVAILABLE, utm_to_latlng, latlng_to_utm
    
    if COORDENADAS_UTILS_AVAILABLE:
        return True  # Ya están importadas
    
    try:
        import sys
        import os
        # Agregar el directorio de tributario_app al path para importar utils_coordenadas
        current_file = os.path.abspath(__file__)
        current_dir = os.path.dirname(current_file)
        # Desde catastro/views.py: subir un nivel a Scripts, luego a tributario/tributario_app
        scripts_dir = os.path.dirname(current_dir)
        tributario_app_path = os.path.join(scripts_dir, 'tributario', 'tributario_app')
        tributario_app_path = os.path.normpath(tributario_app_path)
        if tributario_app_path not in sys.path:
            sys.path.insert(0, tributario_app_path)
        from utils_coordenadas import utm_to_latlng as _utm_to_latlng, latlng_to_utm as _latlng_to_utm
        utm_to_latlng = _utm_to_latlng
        latlng_to_utm = _latlng_to_utm
        COORDENADAS_UTILS_AVAILABLE = True
        logger.info(f"✓ Utilidades de coordenadas importadas correctamente desde: {tributario_app_path}")
        return True
    except ImportError as e:
        logger.warning(f"No se pudo importar utils_coordenadas desde path relativo: {e}")
        # Intentar path alternativo
        try:
            import sys
            import os
            current_file = os.path.abspath(__file__)
            scripts_dir_alt = os.path.dirname(os.path.dirname(current_file))
            tributario_app_path_alt = os.path.join(scripts_dir_alt, 'tributario', 'tributario_app')
            tributario_app_path_alt = os.path.normpath(tributario_app_path_alt)
            if tributario_app_path_alt not in sys.path:
                sys.path.insert(0, tributario_app_path_alt)
            from utils_coordenadas import utm_to_latlng as _utm_to_latlng, latlng_to_utm as _latlng_to_utm
            utm_to_latlng = _utm_to_latlng
            latlng_to_utm = _latlng_to_utm
            COORDENADAS_UTILS_AVAILABLE = True
            logger.info(f"✓ Utilidades de coordenadas importadas correctamente desde path alternativo: {tributario_app_path_alt}")
            return True
        except ImportError as e2:
            logger.error(f"Error al importar utils_coordenadas (intento alternativo): {e2}")
            return False

# Intentar importar al cargar el módulo
_importar_utilidades_coordenadas()

def calcular_impuesto_bdcata1(bdcata1, empresa_codigo):
    """
    Calcula el impuesto para un registro de bdcata1 usando la misma lógica del frontend.
    
    Args:
        bdcata1: Instancia de BDCata1
        empresa_codigo: Código de la empresa (municipio)
    
    Returns:
        Decimal: Valor del impuesto calculado
    """
    try:
        # Obtener datos del municipio
        municipio = Municipio.objects.filter(codigo=empresa_codigo).first()
        if not municipio:
            logger.warning(f'No se encontró municipio con código {empresa_codigo}')
            return Decimal('0.00')
        
        por_concer = municipio.por_concer or Decimal('0.00')
        vl_exento = municipio.vl_exento or Decimal('0.00')
        
        # Obtener estatus tributario (campo st)
        estatus_tributario_str = bdcata1.st or ''
        # Extraer solo el número del estatus tributario
        import re
        match = re.match(r'^(\d+)', str(estatus_tributario_str))
        estatus_tributario = int(match.group(1)) if match else 0
        
        # Validar estatus tributario
        if estatus_tributario < 1 or estatus_tributario > 4:
            logger.warning(f'Estatus tributario inválido: {estatus_tributario}')
            return Decimal('0.00')
        
        # Obtener valores de los campos para calcular Avalúo total
        valor_tierra = bdcata1.bvl2tie or Decimal('0.00')
        edificaciones = bdcata1.mejoras or Decimal('0.00')
        detalles_adicionales = bdcata1.detalle or Decimal('0.00')
        cultivo_permanente = bdcata1.cultivo or Decimal('0.00')
        valor_declarado = bdcata1.declarado or Decimal('0.00')
        porcentaje_exencion = bdcata1.bexenc or Decimal('0.00')
        
        # 1. Calcular Avalúo total (suma de todos los valores)
        avaluo_total = valor_tierra + edificaciones + detalles_adicionales + cultivo_permanente + valor_declarado
        
        num_viviendas = bdcata1.vivienda or Decimal('0')
        num_cuartos = bdcata1.cuartos or Decimal('0')
        num_apartamentos = bdcata1.apartamentos or Decimal('0')
        
        # 2. Calcular Exención según Estatus tributario
        exencion = Decimal('0.00')
        
        if estatus_tributario == 1:
            # Estatus tributario = 1: Exención = min(Avalúo total, vl_exento)
            if vl_exento > avaluo_total:
                exencion = avaluo_total
            else:
                exencion = vl_exento
        elif estatus_tributario == 2:
            # Estatus tributario = 2: Exención = vl_exento * Porcentaje Exención (%)
            exencion_calculada = vl_exento * (porcentaje_exencion / Decimal('100'))
            if exencion_calculada < avaluo_total:
                exencion = exencion_calculada
            else:
                exencion = avaluo_total  # No puede exceder el avalúo total
        elif estatus_tributario == 3:
            # Estatus tributario = 3
            exencion = Decimal('0.00')
        elif estatus_tributario == 4:
            # Estatus tributario = 4
            exencion = avaluo_total
        
        # 3. Calcular Valor grabable según Estatus tributario y porcentaje de concertación
        valor_grabable = Decimal('0.00')
        
        if estatus_tributario == 1:
            # Para Estatus tributario = 1: Valor grabable = Avalúo total - Exención
            valor_grabable = avaluo_total - exencion
        elif estatus_tributario == 4:
            # Para Estatus tributario = 4: Exención = Avalúo total, por lo tanto Valor grabable = 0
            valor_grabable = avaluo_total - exencion  # Esto será 0 porque exencion = avaluoTotal
        else:
            # Para otros estatus (2 y 3): calcular según porcentaje de concertación
            if por_concer > 0:
                # Si porcentaje de concertación > 0, multiplicar Avalúo total por el porcentaje
                valor_grabable = avaluo_total * (por_concer / Decimal('100'))
            else:
                # Si porcentaje de concertación = 0, Valor grabable = Avalúo total
                valor_grabable = avaluo_total
            
            # Recalcular Avalúo total después de aplicar exención (solo para estatus 2 y 3)
            if por_concer > 0:
                # Si porcentaje de concertación > 0, Avalúo total = Valor grabable - Exención
                avaluo_total = valor_grabable - exencion
            else:
                # Si porcentaje de concertación = 0, Valor grabable = Avalúo total - Exención
                valor_grabable = avaluo_total - exencion
        
        # Asegurar que los valores no sean negativos
        if avaluo_total < 0:
            avaluo_total = Decimal('0.00')
        if valor_grabable < 0:
            valor_grabable = Decimal('0.00')
        if exencion < 0:
            exencion = Decimal('0.00')
        
        # 4. Calcular Impuesto
        # El impuesto es: Valor grabable × Tasa impositiva (por millar, se divide entre 1000)
        # Si Valor grabable = 0, entonces Impuesto = 0
        tasa = bdcata1.tasaimpositiva or Decimal('0.00')
        impuesto_calculado = Decimal('0.00')
        
        if valor_grabable > 0 and tasa > 0:
            # Calcular impuesto y redondear a 2 decimales
            # Fórmula: (Valor grabable × Tasa impositiva) / 1000
            impuesto_calculado = (valor_grabable * tasa) / Decimal('1000')
            # Redondear a 2 decimales
            impuesto_calculado = impuesto_calculado.quantize(Decimal('0.01'), rounding='ROUND_HALF_UP')
        else:
            # Si valorGrabable = 0 o tasa = 0, el impuesto es 0
            impuesto_calculado = Decimal('0.00')
        
        return impuesto_calculado
        
    except Exception as e:
        logger.error(f'Error al calcular impuesto para bdcata1: {str(e)}', exc_info=True)
        return Decimal('0.00')

def actualizar_impuesto_tasas_municipales(bdcata1, impuesto_valor, empresa_codigo):
    """
    Actualiza el valor del impuesto en la tabla tasasmunicipales según el perímetro (urbano/rural).
    
    Args:
        bdcata1: Instancia de BDCata1
        impuesto_valor: Valor del impuesto a actualizar (Decimal)
        empresa_codigo: Código de la empresa (municipio)
    
    Returns:
        bool: True si se actualizó exitosamente, False en caso contrario
    """
    try:
        # Importar modelos necesarios
        try:
            from .models import TasasMunicipales
        except ImportError:
            try:
                from modules.catastro.models import TasasMunicipales
            except ImportError:
                TasasMunicipales = globals().get('TasasMunicipales')
                if not TasasMunicipales:
                    logger.warning("No se pudo importar TasasMunicipales")
                    return False
        
        # Obtener ficha para determinar si es urbano (1) o rural (2)
        ficha_valor = int(bdcata1.ficha) if bdcata1.ficha else 1
        
        # Determinar el rubro según el perímetro
        if ficha_valor == 1:  # Urbano
            codigo_rubro = 'B0001'
        else:  # Rural
            codigo_rubro = 'B0002'
        
        # Obtener clave catastral y empresa
        cocata1 = bdcata1.cocata1
        empresa = bdcata1.empresa or empresa_codigo
        
        if not cocata1 or not empresa:
            logger.warning(f"No se puede actualizar tasasmunicipales: cocata1={cocata1}, empresa={empresa}")
            return False
        
        # Buscar y actualizar el registro en tasasmunicipales
        try:
            tasa_existente = TasasMunicipales.objects.get(
                empresa=empresa,
                clave=cocata1,
                rubro=codigo_rubro
            )
            # Actualizar solo el valor del impuesto
            tasa_existente.valor = impuesto_valor
            tasa_existente.save(update_fields=['valor'])
            logger.info(f"✓ Impuesto actualizado en TasasMunicipales: empresa={empresa}, clave={cocata1}, rubro={codigo_rubro}, valor={impuesto_valor}")
            return True
        except TasasMunicipales.DoesNotExist:
            # Si no existe el registro, intentar crearlo con datos básicos
            try:
                from tributario.models import Rubro
            except ImportError:
                try:
                    from tributario.models import Rubro
                except ImportError:
                    Rubro = None
            
            # Obtener cuenta y cuentarez del rubro si está disponible
            cuenta_rubro = ''
            cuentarez_rubro = ''
            if Rubro:
                try:
                    rubro_obj = Rubro.objects.get(empresa=empresa, codigo=codigo_rubro)
                    cuenta_rubro = rubro_obj.cuenta or ''
                    cuentarez_rubro = rubro_obj.cuentarez or ''
                except Rubro.DoesNotExist:
                    logger.warning(f"Rubro {codigo_rubro} no existe en tabla Rubros para empresa {empresa}")
            
            # Crear nuevo registro en tasasmunicipales
            nueva_tasa = TasasMunicipales.objects.create(
                empresa=empresa,
                clave=cocata1,
                rubro=codigo_rubro,
                cod_tarifa='01',
                valor=impuesto_valor,
                cuenta=cuenta_rubro,
                cuentarez=cuentarez_rubro
            )
            logger.info(f"✓ Nuevo registro creado en TasasMunicipales: ID={nueva_tasa.id}, empresa={empresa}, clave={cocata1}, rubro={codigo_rubro}, valor={impuesto_valor}")
            return True
        except Exception as e:
            logger.error(f"Error al actualizar/crear registro en TasasMunicipales: {str(e)}", exc_info=True)
            return False
            
    except Exception as e:
        logger.error(f"Error en actualizar_impuesto_tasas_municipales: {str(e)}", exc_info=True)
        return False

def calcular_tasas_municipales_automatico(bdcata1, empresa_codigo):
    """
    Calcula automáticamente las tasas municipales (rubros que empiezan con T) 
    después de actualizar el impuesto o avalúo del bien inmueble.
    
    Args:
        bdcata1: Instancia de BDCata1
        empresa_codigo: Código de la empresa (municipio)
    
    Returns:
        dict: {'calculadas': int, 'errores': list}
    """
    try:
        # Importar modelos necesarios
        try:
            from .models import TasasMunicipales
        except ImportError:
            try:
                from modules.catastro.models import TasasMunicipales
            except ImportError:
                TasasMunicipales = globals().get('TasasMunicipales')
                if not TasasMunicipales:
                    logger.warning("No se pudo importar TasasMunicipales")
                    return {'calculadas': 0, 'errores': ['No se pudo importar TasasMunicipales']}
        
        try:
            from tributario.models import Tarifas, PlanArbitrio
        except ImportError:
            try:
                from tributario.models import Tarifas, PlanArbitrio
            except ImportError:
                logger.warning("No se pudieron importar Tarifas y PlanArbitrio")
                return {'calculadas': 0, 'errores': ['No se pudieron importar Tarifas y PlanArbitrio']}
        
        from datetime import datetime
        from django.db.models import Q
        
        # Obtener clave catastral y empresa
        cocata1 = bdcata1.cocata1
        empresa = bdcata1.empresa or empresa_codigo
        
        if not cocata1 or not empresa:
            logger.warning(f"No se puede calcular tasas: cocata1={cocata1}, empresa={empresa}")
            return {'calculadas': 0, 'errores': ['Clave catastral o empresa no válidos']}
        
        # Calcular el Avalúo total
        valor_tierra = bdcata1.bvl2tie or Decimal('0.00')
        edificaciones = bdcata1.mejoras or Decimal('0.00')
        detalles_adicionales = bdcata1.detalle or Decimal('0.00')
        cultivo_permanente = bdcata1.cultivo or Decimal('0.00')
        valor_declarado = bdcata1.declarado or Decimal('0.00')
        
        avaluo_total = valor_tierra + edificaciones + detalles_adicionales + cultivo_permanente + valor_declarado
        
        # Año predeterminado
        ano_predeterminado = int(datetime.now().year)

        def determinar_ano_tarifa(tarifa_obj):
            if tarifa_obj and tarifa_obj.ano is not None:
                try:
                    return int(tarifa_obj.ano)
                except (TypeError, ValueError, InvalidOperation):
                    return ano_predeterminado
            return ano_predeterminado

        def obtener_valor_plan_por_tipocat(tasa_municipal, tipocat_objetivo, ano_objetivo=None):
            planes = PlanArbitrio.objects.filter(
                empresa=empresa,
                rubro=tasa_municipal.rubro,
                cod_tarifa=tasa_municipal.cod_tarifa,
                ano=ano_objetivo if ano_objetivo is not None else ano_predeterminado,
                tipocat=str(tipocat_objetivo)
            ).order_by('codigo', 'id')

            if not planes.exists():
                return Decimal('0.00')

            valor_encontrado = Decimal('0.00')
            for plan in planes:
                minimo = plan.minimo if plan.minimo is not None else Decimal('0.00')
                maximo = plan.maximo if plan.maximo is not None else Decimal('9999999999.99')
                if minimo <= avaluo_total <= maximo:
                    valor_encontrado = plan.valor or Decimal('0.00')
                    break

            if valor_encontrado == Decimal('0.00'):
                for plan in planes:
                    minimo = plan.minimo if plan.minimo is not None else Decimal('0.00')
                    if minimo <= avaluo_total:
                        valor_encontrado = plan.valor or Decimal('0.00')
                        break

            if valor_encontrado == Decimal('0.00') and planes.exists():
                valor_encontrado = planes.first().valor or Decimal('0.00')

            return valor_encontrado
        
        # Obtener todos los rubros de tasasmunicipales que empiezan con "T"
        tasas_municipales = TasasMunicipales.objects.filter(
            empresa=empresa,
            clave=cocata1,
            rubro__startswith='T'
        )
        
        if not tasas_municipales.exists():
            # No hay tasas municipales para calcular, no es un error
            return {'calculadas': 0, 'errores': []}
        
        # Procesar cada tasa municipal
        tasas_calculadas = 0
        errores = []
        
        for tasa_municipal in tasas_municipales:
            try:
                # Buscar la tarifa correspondiente en la tabla tarifas
                tarifas_query = Tarifas.objects.filter(
                    empresa=empresa,
                    rubro=tasa_municipal.rubro,
                    cod_tarifa=tasa_municipal.cod_tarifa
                ).order_by('-ano')
                tarifa = tarifas_query.first()
                if not tarifa:
                    continue
                
                # Determinar el tipo de tarifa
                tipo_tarifa = (tarifa.tipo or '').strip().upper()
                nuevo_valor = Decimal('0.00')
                
                if tipo_tarifa == 'F':
                    # Tipo Fijo: Usar el valor directamente de la tarifa
                    nuevo_valor = tarifa.valor or Decimal('0.00')
                    
                elif tipo_tarifa == 'V':
                    try:
                        plan_ano = determinar_ano_tarifa(tarifa)
                        
                        # Detectar si es solar baldío:
                        # Valor terreno > 0, edificación = 0, viviendas = 0, cuartos = 0, apartamentos = 0
                        es_solar_baldio = (
                            valor_tierra > Decimal('0.00') and
                            edificaciones == Decimal('0.00') and
                            num_viviendas == Decimal('0') and
                            num_cuartos == Decimal('0') and
                            num_apartamentos == Decimal('0')
                        )
                        
                        if es_solar_baldio:
                            # Solar baldío: usar exclusivamente categoría 2 del plan de arbitrio
                            nuevo_valor_variable = obtener_valor_plan_por_tipocat(tasa_municipal, '2', plan_ano)
                            logger.info(f"Solar baldío detectado para Rubro {tasa_municipal.rubro}: usando categoría 2, valor={nuevo_valor_variable}")
                        else:
                            # Lógica existente: categorías 1, 3 y 4
                            valor_base = obtener_valor_plan_por_tipocat(tasa_municipal, '1', plan_ano)
                            nuevo_valor_variable = valor_base
                            if num_viviendas > 0:
                                nuevo_valor_variable = valor_base * num_viviendas

                            if num_cuartos > 0:
                                valor_cuartos = obtener_valor_plan_por_tipocat(tasa_municipal, '3', plan_ano)
                                nuevo_valor_variable += valor_cuartos * num_cuartos

                            if num_apartamentos > 0:
                                valor_apartamentos = obtener_valor_plan_por_tipocat(tasa_municipal, '4', plan_ano)
                                nuevo_valor_variable += valor_apartamentos * num_apartamentos

                        nuevo_valor = nuevo_valor_variable.quantize(Decimal('0.01'))
                    except Exception as e:
                        logger.warning(f"Error al buscar plan de arbitrio para Rubro {tasa_municipal.rubro}: {str(e)}")
                        continue
                else:
                    # Tipo desconocido, continuar con la siguiente
                    continue
                
                # Actualizar el valor en tasasmunicipales
                tasa_municipal.valor = nuevo_valor
                tasa_municipal.save(update_fields=['valor'])
                tasas_calculadas += 1
                
                logger.info(f"Tasa municipal calculada automáticamente: Empresa={empresa}, Clave={cocata1}, Rubro={tasa_municipal.rubro}, Tipo={tipo_tarifa}, Valor={nuevo_valor}")
                
            except Exception as e:
                error_msg = f"Error al procesar tasa Rubro {tasa_municipal.rubro}: {str(e)}"
                errores.append(error_msg)
                logger.warning(error_msg)
                continue
        
        return {'calculadas': tasas_calculadas, 'errores': errores}
        
    except Exception as e:
        logger.error(f"Error en calcular_tasas_municipales_automatico: {str(e)}", exc_info=True)
        return {'calculadas': 0, 'errores': [str(e)]}

def catastro_login_view(request):
    """
    El acceso a Catastro se gestiona desde el menú modular (usuario sin contraseña)
    y la pantalla de contraseña por módulo (`modules_core:acceso_modulo`).
    """
    from django.urls import reverse
    from urllib.parse import quote

    login_url = reverse('modules_core:login_principal')
    acceso_url = reverse('modules_core:acceso_modulo', kwargs={'codigo': 'catastro'})
    menu_catastro = reverse('catastro:catastro_menu_principal')
    if not request.session.get('user_id'):
        return redirect(f'{login_url}?next={quote(menu_catastro, safe="/")}')
    return redirect(f'{acceso_url}?next={quote(menu_catastro, safe="/")}')

def catastro_logout_view(request):
    """
    Sale del módulo Catastro: limpia claves del módulo y vuelve al menú modular
    manteniendo la sesión modular (privilegios según el usuario que inició sesión en /login/).
    No usar django.contrib.auth.logout aquí: haría flush de toda la sesión y perdería el login modular
    o mezclaría comportamiento con el usuario de Django.
    """
    for key in (
        'catastro_empresa',
        'catastro_municipio_descripcion',
        'catastro_usuario_id',
        'catastro_usuario_nombre',
    ):
        request.session.pop(key, None)
    request.session.modified = True
    return redirect('modules_core:menu_principal')

def catastro_require_auth(view_func):
    """
    Decorador personalizado para verificar autenticación en catastro
    """
    def wrapper(request, *args, **kwargs):
        # Verificar si el usuario está autenticado en catastro
        empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
        if not empresa_codigo:
            # Si es una petición AJAX, devolver JSON con error 403
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'error': 'No autenticado', 'mensaje': 'Debe iniciar sesión'}, status=403)
            return redirect('catastro:catastro_login')
        return view_func(request, *args, **kwargs)
    return wrapper

@catastro_require_auth
@catastro_require_permiso(CATASTRO_PERM_MENU_VER)
def catastro_menu_principal(request):
    """
    Menú principal del módulo de catastro
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro'
    }
    context.update(permisos_catastro_menu_context(request))
    return render(request, 'menu_principal.html', context)

@catastro_require_auth
@catastro_require_permiso((CATASTRO_PERM_BIENES_VER, CATASTRO_PERM_BIENES_EDITAR))
def catastro_bienes_inmuebles(request):
    """
    Gestión de bienes inmuebles
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Bienes Inmuebles'
    }
    return render(request, 'bienes_inmuebles.html', context)

@catastro_require_auth
@catastro_require_permiso(CATASTRO_PERM_MAPA_VER)
def mapa_georreferenciado_view(request):
    """
    Mapa georreferenciado - Muestra predios (bdcata1) y negocios en un mapa interactivo
    Con categorización por uso de la propiedad
    """
    from .models import BDCata1, Usos
    from decimal import Decimal
    import json
    
    def decimal_to_float(obj):
        """Convierte objetos Decimal y otros tipos a tipos serializables JSON"""
        from datetime import date, datetime
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, (date, datetime)):
            return obj.isoformat()
        elif hasattr(obj, '__dict__'):
            return str(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener todos los usos disponibles para el filtro
    usos_disponibles = []
    try:
        usos_qs = Usos.objects.all().order_by('uso')
        for uso_obj in usos_qs:
            usos_disponibles.append({
                'codigo': uso_obj.uso,
                'descripcion': uso_obj.desuso or uso_obj.uso
            })
    except Exception as e:
        logger.warning(f"Error al obtener usos: {str(e)}")
    
    # Obtener el uso seleccionado del request (si existe)
    uso_filtro = request.GET.get('uso', '')
    
    # Obtener predios georreferenciados de bdcata1 (cx != 0 y cy != 0)
    predios = []
    # Cache para descripciones de uso
    usos_cache = {}
    
    if empresa_codigo:
        # Obtener todos los predios de la empresa y filtrar manualmente
        predios_qs = BDCata1.objects.filter(empresa=empresa_codigo)
        
        # Aplicar filtro de uso si se especificó
        if uso_filtro:
            predios_qs = predios_qs.filter(uso=uso_filtro)
        
        for predio in predios_qs:
            try:
                # Verificar que existan coordenadas y no sean cero
                if predio.cx is None or predio.cy is None:
                    continue
                
                cx_val = float(predio.cx) if predio.cx is not None else 0.0
                cy_val = float(predio.cy) if predio.cy is not None else 0.0
                
                # Solo incluir si las coordenadas son válidas y diferentes de cero
                if cx_val != 0.0 and cy_val != 0.0 and not (cx_val == 0 and cy_val == 0):
                    # Convertir coordenadas UTM a lat/lng para mostrar en el mapa
                    lat = None
                    lng = None
                    # Asegurar que las utilidades estén disponibles
                    if not COORDENADAS_UTILS_AVAILABLE or utm_to_latlng is None:
                        _importar_utilidades_coordenadas()
                    
                    if COORDENADAS_UTILS_AVAILABLE and utm_to_latlng is not None:
                        try:
                            # Las coordenadas en BD están en UTM, convertir a lat/lng
                            lat, lng = utm_to_latlng(cx_val, cy_val)
                            if lat is None or lng is None:
                                # Si falla la conversión, usar las coordenadas originales
                                lat = cy_val
                                lng = cx_val
                                logger.warning(f"Error al convertir UTM a lat/lng para predio {predio.id}, usando coordenadas originales")
                        except Exception as e:
                            logger.error(f"Error al convertir coordenadas UTM a lat/lng: {str(e)}")
                            lat = cy_val
                            lng = cx_val
                    else:
                        # Si no hay utilidades disponibles, asumir que ya están en lat/lng
                        lat = cy_val
                        lng = cx_val
                    
                    # Obtener descripción del uso
                    uso_codigo = str(predio.uso) if predio.uso else '0'
                    uso_descripcion = 'Sin uso'
                    
                    if uso_codigo not in usos_cache:
                        try:
                            uso_obj = Usos.objects.filter(uso=uso_codigo).first()
                            if uso_obj:
                                uso_descripcion = uso_obj.desuso or uso_codigo
                                usos_cache[uso_codigo] = uso_descripcion
                            else:
                                usos_cache[uso_codigo] = uso_codigo
                        except Exception:
                            usos_cache[uso_codigo] = uso_codigo
                    
                    uso_descripcion = usos_cache.get(uso_codigo, uso_codigo)
                    
                    # Convertir valores Decimal a float
                    bvl2tie_val = 0.0
                    if predio.bvl2tie is not None:
                        try:
                            bvl2tie_val = float(predio.bvl2tie)
                        except (ValueError, TypeError):
                            bvl2tie_val = 0.0
                    
                    mejoras_val = 0.0
                    if predio.mejoras is not None:
                        try:
                            mejoras_val = float(predio.mejoras)
                        except (ValueError, TypeError):
                            mejoras_val = 0.0
                    
                    impuesto_val = 0.0
                    if predio.impuesto is not None:
                        try:
                            impuesto_val = float(predio.impuesto)
                        except (ValueError, TypeError):
                            impuesto_val = 0.0
                    
                    predios.append({
                        'id': int(predio.id),
                        'tipo': 'predio',
                        'cocata1': str(predio.cocata1) if predio.cocata1 else '',
                        'nombres': str(predio.nombres) if predio.nombres else '',
                        'apellidos': str(predio.apellidos) if predio.apellidos else '',
                        'ubicacion': str(predio.ubicacion) if predio.ubicacion else '',
                        'cx': float(lng),  # Longitud para el mapa
                        'cy': float(lat),  # Latitud para el mapa
                        'ficha': str(predio.ficha) if predio.ficha else '',
                        'bvl2tie': float(bvl2tie_val),
                        'mejoras': float(mejoras_val),
                        'impuesto': float(impuesto_val),
                        'uso': uso_codigo,
                        'uso_descripcion': uso_descripcion,
                    })
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Error procesando predio {predio.id}: {str(e)}")
                continue
    
    # Obtener negocios georreferenciados (cx != 0 y cy != 0)
    negocios = []
    if empresa_codigo:
        try:
            from tributario.models import Negocio
            negocios_qs = Negocio.objects.filter(
                empresa=empresa_codigo
            ).exclude(
                cx=Decimal('0.00'),
                cy=Decimal('0.00')
            ).exclude(
                cx__isnull=True,
                cy__isnull=True
            )
            
            for negocio in negocios_qs:
                try:
                    cx_val = float(negocio.cx) if negocio.cx else 0.0
                    cy_val = float(negocio.cy) if negocio.cy else 0.0
                    
                    if cx_val != 0.0 and cy_val != 0.0:
                        # Convertir coordenadas UTM a lat/lng para mostrar en el mapa
                        lat = None
                        lng = None
                        # Asegurar que las utilidades estén disponibles
                        if not COORDENADAS_UTILS_AVAILABLE or utm_to_latlng is None:
                            _importar_utilidades_coordenadas()
                        
                        if COORDENADAS_UTILS_AVAILABLE and utm_to_latlng is not None:
                            try:
                                # Las coordenadas en BD están en UTM, convertir a lat/lng
                                lat, lng = utm_to_latlng(cx_val, cy_val)
                                if lat is None or lng is None:
                                    # Si falla la conversión, usar las coordenadas originales
                                    lat = cy_val
                                    lng = cx_val
                                    logger.warning(f"Error al convertir UTM a lat/lng para negocio {negocio.id}, usando coordenadas originales")
                            except Exception as e:
                                logger.error(f"Error al convertir coordenadas UTM a lat/lng: {str(e)}")
                                lat = cy_val
                                lng = cx_val
                        else:
                            # Si no hay utilidades disponibles, asumir que ya están en lat/lng
                            lat = cy_val
                            lng = cx_val
                        
                        negocios.append({
                            'id': int(negocio.id),
                            'tipo': 'negocio',
                            'rtm': str(negocio.rtm) if negocio.rtm else '',
                            'expe': str(negocio.expe) if negocio.expe else '',
                            'nombrenego': str(negocio.nombrenego) if negocio.nombrenego else '',
                            'comerciante': str(negocio.comerciante) if negocio.comerciante else '',
                            'direccion': str(negocio.direccion) if negocio.direccion else '',
                            'catastral': str(negocio.catastral) if negocio.catastral else '',
                            'cx': float(lng),  # Longitud para el mapa
                            'cy': float(lat),  # Latitud para el mapa
                            'actividad': str(negocio.actividad) if negocio.actividad else '',
                            'estatus': str(negocio.estatus) if negocio.estatus else '',
                        })
                except (ValueError, TypeError, AttributeError) as e:
                    logger.warning(f"Error procesando negocio {negocio.id}: {str(e)}")
                    continue
        except ImportError:
            logger.warning("No se pudo importar el modelo Negocio desde tributario.models")
    
    # Buscar una coordenada válida en bdcata1 donde cx <> 0 y cy <> 0
    # Prioridad: 1) bdcata1, 2) negocios, 3) Promedio, 4) Geolocalización, 5) Por defecto
    centro_cx = None
    centro_cy = None
    usar_geolocalizacion = False  # Flag para usar geolocalización del navegador
    
    # PRIORIDAD 1: Buscar directamente en bdcata1 con coordenadas válidas (cx <> 0 y cy <> 0)
    if empresa_codigo:
        try:
            predio_base = BDCata1.objects.filter(
                empresa=empresa_codigo
            ).exclude(
                cx__isnull=True,
                cy__isnull=True
            ).exclude(
                cx=Decimal('0.00'),
                cy=Decimal('0.00')
            ).exclude(
                cx=0,
                cy=0
            ).first()
            
            if predio_base:
                try:
                    cx_val = float(predio_base.cx) if predio_base.cx is not None else None
                    cy_val = float(predio_base.cy) if predio_base.cy is not None else None
                    
                    # Verificar que las coordenadas sean válidas (no cero)
                    if (cx_val is not None and cy_val is not None and 
                        cx_val != 0.0 and cy_val != 0.0 and
                        abs(cx_val) > 0.0001 and abs(cy_val) > 0.0001):
                        # Convertir coordenadas UTM a lat/lng para el centro del mapa
                        # Asegurar que las utilidades estén disponibles
                        if not COORDENADAS_UTILS_AVAILABLE or utm_to_latlng is None:
                            _importar_utilidades_coordenadas()
                        
                        if COORDENADAS_UTILS_AVAILABLE and utm_to_latlng is not None:
                            try:
                                lat, lng = utm_to_latlng(cx_val, cy_val)
                                if lat is not None and lng is not None:
                                    centro_cx = lng
                                    centro_cy = lat
                                else:
                                    centro_cx = cx_val
                                    centro_cy = cy_val
                            except Exception as e:
                                logger.warning(f"Error al convertir coordenadas del predio base: {str(e)}")
                                centro_cx = cx_val
                                centro_cy = cy_val
                        else:
                            centro_cx = cx_val
                            centro_cy = cy_val
                        usar_geolocalizacion = False
                        logger.info(f"✓ Coordenada encontrada en bdcata1 - Predio: {predio_base.cocata1}, cx={centro_cx}, cy={centro_cy}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error al convertir coordenadas del predio base: {str(e)}")
        except Exception as e:
            logger.error(f"Error al buscar coordenada base en bdcata1: {str(e)}", exc_info=True)
    
    # PRIORIDAD 2: Si no se encontró en bdcata1, buscar en negocios
    if (centro_cx is None or centro_cy is None) and empresa_codigo:
        try:
            from tributario.models import Negocio
            negocio_base = Negocio.objects.filter(
                empresa=empresa_codigo
            ).exclude(
                cx__isnull=True,
                cy__isnull=True
            ).exclude(
                cx=Decimal('0.00'),
                cy=Decimal('0.00')
            ).exclude(
                cx=0,
                cy=0
            ).first()
            
            if negocio_base:
                try:
                    cx_val = float(negocio_base.cx) if negocio_base.cx is not None else None
                    cy_val = float(negocio_base.cy) if negocio_base.cy is not None else None
                    
                    if (cx_val is not None and cy_val is not None and 
                        cx_val != 0.0 and cy_val != 0.0 and
                        abs(cx_val) > 0.0001 and abs(cy_val) > 0.0001):
                        # Convertir coordenadas UTM a lat/lng para el centro del mapa
                        # Asegurar que las utilidades estén disponibles
                        if not COORDENADAS_UTILS_AVAILABLE or utm_to_latlng is None:
                            _importar_utilidades_coordenadas()
                        
                        if COORDENADAS_UTILS_AVAILABLE and utm_to_latlng is not None:
                            try:
                                lat, lng = utm_to_latlng(cx_val, cy_val)
                                if lat is not None and lng is not None:
                                    centro_cx = lng
                                    centro_cy = lat
                                else:
                                    centro_cx = cx_val
                                    centro_cy = cy_val
                            except Exception as e:
                                logger.warning(f"Error al convertir coordenadas del negocio base: {str(e)}")
                                centro_cx = cx_val
                                centro_cy = cy_val
                        else:
                            centro_cx = cx_val
                            centro_cy = cy_val
                        usar_geolocalizacion = False
                        logger.info(f"✓ Coordenada encontrada en negocios - Negocio: {negocio_base.rtm or negocio_base.id}, cx={centro_cx}, cy={centro_cy}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error al convertir coordenadas del negocio base: {str(e)}")
        except Exception as e:
            logger.warning(f"Error al buscar coordenada base en negocios: {str(e)}")
    
    # PRIORIDAD 3: Usar el promedio de puntos válidos ya procesados
    if centro_cx is None or centro_cy is None:
        puntos_validos = [p for p in predios + negocios 
                         if p.get('cx') and p.get('cy') and 
                         p.get('cx') != 0 and p.get('cy') != 0]
        if len(puntos_validos) > 0:
            suma_cx = sum(p['cx'] for p in puntos_validos)
            suma_cy = sum(p['cy'] for p in puntos_validos)
            centro_cx = suma_cx / len(puntos_validos)
            centro_cy = suma_cy / len(puntos_validos)
            usar_geolocalizacion = False
            logger.info(f"✓ Usando promedio de {len(puntos_validos)} puntos válidos: cx={centro_cx}, cy={centro_cy}")
    
    # PRIORIDAD 4: Si no se encontró coordenada válida, marcar para usar geolocalización
    if centro_cx is None or centro_cy is None or centro_cx == 0 or centro_cy == 0:
        usar_geolocalizacion = True
        # Coordenadas por defecto de Honduras (se usarán si la geolocalización falla)
        centro_cx = -86.2419
        centro_cy = 15.1999
        logger.warning(f"⚠ No se encontraron coordenadas válidas. Se intentará usar geolocalización del navegador.")
    
    # Validación final
    if abs(centro_cx) < 0.0001 or abs(centro_cy) < 0.0001:
        usar_geolocalizacion = True
        centro_cx = -86.2419
        centro_cy = 15.1999
        logger.warning(f"⚠ Coordenadas cercanas a cero. Se intentará usar geolocalización del navegador.")
    
    logger.info(f"📍 Coordenadas finales para el mapa: cx={centro_cx}, cy={centro_cy}, usar_geolocalizacion={usar_geolocalizacion}")
    
    # Serializar a JSON usando el encoder personalizado
    try:
        predios_json = json.dumps(predios, default=decimal_to_float, ensure_ascii=False)
    except TypeError as e:
        logger.error(f"Error serializando predios: {str(e)}")
        predios_json = "[]"
    
    try:
        negocios_json = json.dumps(negocios, default=decimal_to_float, ensure_ascii=False)
    except TypeError as e:
        logger.error(f"Error serializando negocios: {str(e)}")
        negocios_json = "[]"
    
    context = {
        'empresa': empresa_codigo,
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Mapa georreferenciado',
        'predios': predios_json,
        'negocios': negocios_json,
        'centro_cx': centro_cx,
        'centro_cy': centro_cy,
        'usar_geolocalizacion': usar_geolocalizacion,
        'total_predios': len(predios),
        'total_negocios': len(negocios),
        'usos_disponibles': usos_disponibles,
        'uso_filtro_seleccionado': uso_filtro,
    }
    return render(request, 'mapa_georreferenciado.html', context)

@catastro_require_auth
def catastro_miscelaneos(request):
    """
    Gestión de misceláneos - Réplica del formulario de tributario adaptado a catastro
    """
    from tributario.models import Oficina
    
    # Obtener el código de empresa de la sesión de catastro
    empresa = request.session.get('catastro_empresa')
    if not empresa:
        return redirect('catastro:catastro_login')
    
    # Obtener todas las oficinas ordenadas por código
    oficinas = Oficina.objects.all().order_by('codigo')
    
    context = {
        'oficinas': oficinas,
        'empresa': empresa,
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Misceláneos'
    }
    return render(request, 'miscelaneoscat.html', context)

@catastro_require_auth
def catastro_terrenos(request):
    """
    Gestión de terrenos
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Terrenos'
    }
    return render(request, 'terrenos.html', context)

@catastro_require_auth
def catastro_construcciones(request):
    """
    Gestión de construcciones
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Construcciones'
    }
    return render(request, 'construcciones.html', context)

@catastro_require_auth
def notificaciones_avaluo(request):
    """
    Notificaciones de avaluo
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Notificaciones de Avaluo'
    }
    return render(request, 'notificaciones_avaluo.html', context)

@catastro_require_auth
def notificacion_avaluo_catastral(request):
    """
    Formulario de notificación de avalúo catastral
    """
    from datetime import datetime
    from decimal import Decimal
    from .models import BDCata1, BDTerreno
    
    # Obtener fecha actual
    fecha_actual = datetime.now()
    
    # Si se busca por clave catastral
    cocata1 = request.GET.get('cocata1', '')
    datos_bien = None
    datos_terreno = None
    mejoras_detalles = Decimal('0.00')
    area_terreno = Decimal('0.00')
    
    if cocata1:
        empresa = request.session.get('catastro_empresa', '')
        try:
            datos_bien = BDCata1.objects.filter(empresa=empresa, cocata1=cocata1).first()
            if datos_bien:
                datos_terreno = BDTerreno.objects.filter(empresa=empresa, cocata1=cocata1).first()
                # Calcular mejoras + detalles
                mejoras = datos_bien.mejoras or Decimal('0.00')
                detalle = datos_bien.detalle or Decimal('0.00')
                mejoras_detalles = mejoras + detalle
                # Calcular área del terreno
                if datos_terreno:
                    baream21 = datos_terreno.baream21 or Decimal('0.00')
                    baream22 = datos_terreno.baream22 or Decimal('0.00')
                    area_terreno = baream21 + baream22
        except Exception as e:
            print(f"Error al buscar datos: {str(e)}")
    
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Notificación de Avaluo Catastral',
        'fecha_actual': fecha_actual,
        'dia': fecha_actual.day,
        'mes': fecha_actual.month,
        'ano': fecha_actual.year,
        'datos_bien': datos_bien,
        'datos_terreno': datos_terreno,
        'cocata1': cocata1,
        'mejoras_detalles': mejoras_detalles,
        'area_terreno': area_terreno
    }
    return render(request, 'notificacion_avaluo_catastral.html', context)

@catastro_require_auth
def generacion_graficos(request):
    """
    Formulario para generación de gráficos
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Generación Gráficos'
    }
    return render(request, 'generacion_graficos.html', context)

@catastro_require_auth
def generar_graficos(request):
    """
    Vista para generar gráficos de avalúos catastrales
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Naturaleza, Dominio, TipoDocumento, Legales, Barrios
    from django.db.models import Sum, Count, Q, F
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    import json
    
    if request.method != 'POST':
        messages.error(request, 'Método no permitido')
        return redirect('catastro:generacion_graficos')
    
    nivel_grafico = request.POST.get('nivel_grafico', '').strip()
    tipo_grafico = request.POST.get('tipo_grafico', '').strip()
    empresa = request.session.get('catastro_empresa', '')
    
    if not nivel_grafico:
        messages.error(request, 'Debe seleccionar el nivel del gráfico')
        return redirect('catastro:generacion_graficos')
    
    if not tipo_grafico:
        messages.error(request, 'Debe seleccionar un tipo de gráfico')
        return redirect('catastro:generacion_graficos')
    
    if not empresa:
        messages.error(request, 'No se ha seleccionado un municipio')
        return redirect('catastro:generacion_graficos')
    
    # Construir query base (similar a sumariales)
    query = Q(empresa=empresa)
    filtros_aplicados = []
    grupo_por = []
    codigo_uso_filtro = None  # Para usar en descripciones de subuso
    
    # Aplicar filtros según el tipo de gráfico (misma lógica que sumariales)
    if tipo_grafico == 'por_uso':
        grupo_por = ['uso']
    elif tipo_grafico == 'por_subuso':
        codigo_uso = request.POST.get('codigo_uso', '').strip()
        if codigo_uso:
            try:
                codigo_uso_normalizado = str(int(codigo_uso)).zfill(2)
            except ValueError:
                codigo_uso_normalizado = codigo_uso.zfill(2) if codigo_uso.isdigit() else codigo_uso
            codigo_uso_filtro = codigo_uso_normalizado
            query &= Q(uso=codigo_uso_normalizado)
            try:
                uso_obj = Usos.objects.filter(uso=codigo_uso_normalizado).first()
                if uso_obj and uso_obj.desuso:
                    filtros_aplicados.append(f'Uso: {uso_obj.desuso} ({codigo_uso_normalizado})')
                else:
                    filtros_aplicados.append(f'Uso: {codigo_uso_normalizado}')
            except:
                filtros_aplicados.append(f'Uso: {codigo_uso_normalizado}')
        grupo_por = ['subuso']
    elif tipo_grafico == 'por_barrio':
        grupo_por = ['barrio']
    elif tipo_grafico == 'usos_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        grupo_por = ['barrio', 'uso']
    elif tipo_grafico == 'usos_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
        grupo_por = ['ficha', 'uso']
    elif tipo_grafico == 'por_sexo':
        grupo_por = ['sexo']
    elif tipo_grafico == 'por_identidad':
        grupo_por = ['identidad']
    elif tipo_grafico == 'naturaleza_jur_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        grupo_por = ['naturaleza']
    elif tipo_grafico == 'naturaleza_jur_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
        grupo_por = ['naturaleza']
    elif tipo_grafico == 'clase_dom_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        grupo_por = ['dominio']
    elif tipo_grafico == 'clase_dom_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
        grupo_por = ['dominio']
    elif tipo_grafico == 'tipo_docto_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        grupo_por = ['tipo']
    elif tipo_grafico == 'tipo_docto_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
        grupo_por = ['tipo']
    elif tipo_grafico == 'por_estatus_tributario':
        grupo_por = ['st']
    elif tipo_grafico == 'por_zonificacion':
        grupo_por = ['zonificacion']
    elif tipo_grafico == 'por_naturaleza_juridica':
        grupo_por = ['naturaleza']
    elif tipo_grafico == 'por_clase_dominio':
        grupo_por = ['dominio']
    elif tipo_grafico == 'por_tipo_documento':
        grupo_por = ['tipo']
    
    # Determinar si necesita JOIN con legales
    necesita_join_legales = any(campo in grupo_por for campo in ['naturaleza', 'dominio', 'tipo'])
    
    # Obtener datos agrupados (similar a sumariales pero adaptado para gráficos)
    try:
        if necesita_join_legales:
            avaluos = BDCata1.objects.filter(query).extra(
                select={
                    'naturaleza_legales': 'legales.naturaleza',
                    'dominio_legales': 'legales.dominio',
                    'tipo_legales': 'legales.tipo'
                },
                tables=['legales'],
                where=['bdcata1.cocata1 = legales.colegal AND bdcata1.empresa = legales.empresa']
            )
        else:
            avaluos = BDCata1.objects.filter(query)
        
        # Agrupar datos manualmente
        datos_agrupados = {}
        
        for avaluo in avaluos:
            # Construir clave de agrupación
            clave_grupo = []
            grupo_descripcion = {}
            
            for campo in grupo_por:
                if campo == 'naturaleza' and necesita_join_legales:
                    valor = getattr(avaluo, 'naturaleza_legales', None)
                elif campo == 'dominio' and necesita_join_legales:
                    valor = getattr(avaluo, 'dominio_legales', None)
                elif campo == 'tipo' and necesita_join_legales:
                    valor = getattr(avaluo, 'tipo_legales', None)
                else:
                    valor = getattr(avaluo, campo, None)
                
                if valor is None:
                    valor = ''
                
                clave_grupo.append(str(valor))
                grupo_descripcion[campo] = valor
            
            clave = '|'.join(clave_grupo)
            
            if clave not in datos_agrupados:
                datos_agrupados[clave] = {
                    'grupo': grupo_descripcion,
                    'cantidad_predios': 0,
                    'avaluo_terreno': Decimal('0.00'),
                    'avaluo_edificacion': Decimal('0.00'),
                    'avaluo_detalles': Decimal('0.00'),
                    'avaluo_cultivo': Decimal('0.00'),
                    'exencion': Decimal('0.00'),
                    'valor_grabable': Decimal('0.00'),
                    'impuesto': Decimal('0.00'),
                    'total_area': Decimal('0.00')
                }
            
            datos_agrupados[clave]['cantidad_predios'] += 1
            
            # Obtener valores de terreno
            try:
                terreno = BDTerreno.objects.filter(
                    empresa=avaluo.empresa,
                    cocata1=avaluo.cocata1
                ).first()
                
                if terreno:
                    datos_agrupados[clave]['avaluo_terreno'] += terreno.bvl2tie or Decimal('0.00')
                    datos_agrupados[clave]['total_area'] += (terreno.baream21 or Decimal('0.00')) + (terreno.baream22 or Decimal('0.00'))
            except:
                pass
            
            # Valores de bdcata1
            datos_agrupados[clave]['avaluo_edificacion'] += avaluo.mejoras or Decimal('0.00')
            datos_agrupados[clave]['avaluo_detalles'] += avaluo.detalle or Decimal('0.00')
            datos_agrupados[clave]['avaluo_cultivo'] += avaluo.cultivo or Decimal('0.00')
            datos_agrupados[clave]['exencion'] += avaluo.exencion or Decimal('0.00')
            datos_agrupados[clave]['valor_grabable'] += avaluo.grabable or Decimal('0.00')
            datos_agrupados[clave]['impuesto'] += avaluo.impuesto or Decimal('0.00')
        
        # Obtener descripciones para las etiquetas del gráfico
        datos_grafico = []
        labels = []
        valores = []
        
        for clave, datos in sorted(datos_agrupados.items()):
            # Construir etiqueta
            etiqueta_parts = []
            for campo in grupo_por:
                valor = datos['grupo'].get(campo, '')
                descripcion = ''
                
                try:
                    if campo == 'uso':
                        uso_obj = Usos.objects.filter(uso=valor).first()
                        descripcion = uso_obj.desuso if uso_obj else valor
                    elif campo == 'subuso':
                        if codigo_uso_filtro:
                            subuso_obj = Subuso.objects.filter(uso=codigo_uso_filtro, codsubuso=valor).first()
                            descripcion = subuso_obj.descripcion if subuso_obj else valor
                        else:
                            descripcion = str(valor) if valor else '-'
                    elif campo == 'barrio':
                        barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=valor).first()
                        descripcion = barrio_obj.descripcion if barrio_obj else valor
                    elif campo == 'naturaleza':
                        naturaleza_obj = Naturaleza.objects.filter(empresa=empresa, codigo=Decimal(valor)).first()
                        descripcion = naturaleza_obj.descripcion if naturaleza_obj else valor
                    elif campo == 'dominio':
                        dominio_obj = Dominio.objects.filter(empresa=empresa, codigo=Decimal(valor)).first()
                        descripcion = dominio_obj.descripcion if dominio_obj else valor
                    elif campo == 'tipo':
                        tipo_obj = TipoDocumento.objects.filter(empresa=empresa, codigo=Decimal(valor)).first()
                        descripcion = tipo_obj.descripcion if tipo_obj else valor
                    elif campo == 'st':
                        st_descripciones = {
                            '1': 'Exento',
                            '2': 'Parcialmente Exento',
                            '3': 'Totalmente Tributario',
                            '4': 'Totalmente Exento sin Valores'
                        }
                        descripcion = st_descripciones.get(str(valor), valor)
                    elif campo == 'sexo':
                        descripcion = 'Masculino' if valor == 'M' else 'Femenino' if valor == 'F' else valor
                    elif campo == 'ficha':
                        descripcion = 'Urbana' if valor == '1' else 'Rural' if valor == '2' else valor
                    else:
                        descripcion = str(valor) if valor else '-'
                except:
                    descripcion = str(valor) if valor else '-'
                
                etiqueta_parts.append(descripcion)
            
            etiqueta = ' - '.join(etiqueta_parts) if etiqueta_parts else 'Sin clasificar'
            
            # Determinar valor según nivel del gráfico
            if nivel_grafico == 'predios':
                valor_grafico = datos['cantidad_predios']
            else:  # avaluo
                valor_grafico = float(
                    datos['avaluo_terreno'] + 
                    datos['avaluo_edificacion'] + 
                    datos['avaluo_detalles'] + 
                    datos['avaluo_cultivo']
                )
            
            labels.append(etiqueta)
            valores.append(valor_grafico)
        
        # Preparar datos para el template
        nombres_graficos = {
            'por_uso': 'Gráfico de Avaluos por Uso',
            'por_subuso': 'Gráfico de Avaluos por Sub Uso',
            'por_barrio': 'Gráfico de Avaluos por Barrio',
            'por_naturaleza_juridica': 'Gráfico por Naturaleza Juridica',
            'por_clase_dominio': 'Gráfico por Clase de Dominio',
            'por_tipo_documento': 'Gráfico por Tipo de Documento',
            'por_sexo': 'Gráfico de Avaluos Catastrales por Sexo',
            'por_identidad': 'Gráfico de Avaluos Catastrales por Numero de Identidad',
            'usos_por_barrio': 'Gráfico de Avaluos Catastrales Usos por Barrio',
            'naturaleza_jur_por_barrio': 'Gráfico de Avaluos Catastrales por Naturaleza Juridica por Barrio',
            'clase_dom_por_barrio': 'Gráfico de Avaluos Catastrales Clase de Dominio por Barrio',
            'tipo_docto_por_barrio': 'Gráfico de Avaluos Catastrales Tipo de Documento por Barrio',
            'usos_por_perimetro': 'Gráfico de Avaluos Catastrales Usos por Perímetro',
            'naturaleza_jur_por_perimetro': 'Gráfico de Avaluos Catastrales Naturaleza Juridica por Perímetro',
            'clase_dom_por_perimetro': 'Gráfico de Avaluos Catastrales Clase de Dominio por Perímetro',
            'tipo_docto_por_perimetro': 'Gráfico de Avaluos Catastrales Tipo de Documento por Perímetro',
            'por_estatus_tributario': 'Gráfico por Estatus Tributario',
            'por_zonificacion': 'Gráfico de Avaluos Catastrales por Zonificacion'
        }
        
        titulo_grafico = nombres_graficos.get(tipo_grafico, 'Gráfico de Avaluos Catastrales')
        nivel_texto = 'Predios' if nivel_grafico == 'predios' else 'Avalúo Catastral'
        
        context = {
            'empresa': empresa,
            'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
            'usuario_nombre': request.session.get('catastro_usuario_nombre'),
            'modulo': 'Catastro - Generación Gráficos',
            'titulo_grafico': titulo_grafico,
            'nivel_grafico': nivel_grafico,
            'nivel_texto': nivel_texto,
            'labels': json.dumps(labels),
            'valores': json.dumps(valores),
            'filtros_aplicados': filtros_aplicados,
            'tipo_grafico': tipo_grafico
        }
        
        return render(request, 'resultado_graficos.html', context)
        
    except Exception as e:
        logger.error(f"Error al generar gráfico: {str(e)}", exc_info=True)
        messages.error(request, f'Error al generar el gráfico: {str(e)}')
        return redirect('catastro:generacion_graficos')

@catastro_require_auth
def emision_documentos(request):
    """
    Emisión de documentos - Menú Principal
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Emisión de Documentos'
    }
    return render(request, 'emision_documentos.html', context)

@catastro_require_auth
def generar_constancia(request, tipo):
    """
    Vista unificada para generación de diferentes tipos de constancias
    """
    from .models import BDCata1, BDTerreno, Legales, Colindantes, Colindancias
    from decimal import Decimal
    from datetime import datetime
    
    empresa = request.session.get('catastro_empresa', '')
    usuario_nombre = request.session.get('catastro_usuario_nombre', 'Usuario')
    municipio_descripcion = request.session.get('catastro_municipio_descripcion', '')
    
    # Parámetros de búsqueda
    cocata1 = request.GET.get('cocata1', '').strip()
    identidad = request.GET.get('identidad', '').strip()
    nombre_solicitado = request.GET.get('nombre', '').strip()
    
    datos_bien = None
    bienes_lista = []
    mensaje_error = None
    
    # Lógica de búsqueda según el tipo de constancia y parámetros
    if tipo in ['poseer', 'no_poseer']:
        if identidad:
            bienes_query = BDCata1.objects.filter(empresa=empresa, identidad=identidad)
            for bien in bienes_query:
                # Adjuntar datos legales a cada bien para el reporte
                legal_info = Legales.objects.filter(empresa=empresa, colegal=bien.cocata1).first()
                if legal_info:
                    bien.legales_numero = legal_info.numero
                    bien.tomo = legal_info.tomo
                    bien.asiento = legal_info.asiento
                bienes_lista.append(bien)
            
            if not bienes_lista and tipo == 'poseer':
                mensaje_error = f"No se encontraron bienes para la identidad {identidad}"
        elif cocata1:
            # Si dan clave, buscar el dueño y luego todos sus bienes
            bien_inicial = BDCata1.objects.filter(empresa=empresa, cocata1=cocata1).first()
            if bien_inicial and bien_inicial.identidad:
                identidad = bien_inicial.identidad
                bienes_query = BDCata1.objects.filter(empresa=empresa, identidad=identidad)
                for bien in bienes_query:
                    legal_info = Legales.objects.filter(empresa=empresa, colegal=bien.cocata1).first()
                    if legal_info:
                        bien.legales_numero = legal_info.numero
                        bien.tomo = legal_info.tomo
                        bien.asiento = legal_info.asiento
                    bienes_lista.append(bien)
            else:
                mensaje_error = "No se encontró el bien o no tiene identidad asociada"
    else:
        # Constancias por Clave Catastral (Servicios, Colindantes, Escalafón, Avalúo)
        if cocata1:
            datos_bien = BDCata1.objects.filter(empresa=empresa, cocata1=cocata1).first()
            if not datos_bien:
                mensaje_error = f"No se encontró el predio con clave {cocata1}"
    
    # Preparar contexto común
    ahora = datetime.now()
    meses_es = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    context = {
        'empresa': empresa,
        'municipio_descripcion': municipio_descripcion,
        'usuario_nombre': usuario_nombre,
        'tipo': tipo,
        'tipo_titulo': tipo.replace('_', ' ').title(),
        'cocata1': cocata1,
        'identidad': identidad,
        'nombre_solicitado': nombre_solicitado,
        'datos_bien': datos_bien,
        'bienes_lista': bienes_lista,
        'mensaje_error': mensaje_error,
        'fecha_actual': ahora,
        'dia': ahora.day,
        'mes_nombre': meses_es[ahora.month - 1],
        'ano': ahora.year,
        'modulo': f'Catastro - Constancia de {tipo.replace("_", " ").title()}'
    }
    
    # Datos adicionales específicos por tipo
    if datos_bien:
        # Para Avalúo y Escalafón
        legales = Legales.objects.filter(empresa=empresa, colegal=datos_bien.cocata1).first()
        terreno = BDTerreno.objects.filter(empresa=empresa, cocata1=datos_bien.cocata1).first()
        
        context['legales'] = legales
        context['terreno'] = terreno
        
        area_total = Decimal('0.00')
        if terreno:
            area_total = (terreno.baream21 or Decimal('0.00')) + (terreno.baream22 or Decimal('0.00'))
        context['area_total'] = area_total
        
        # Valores de avalúo
        valor_terreno = datos_bien.bvl2tie or Decimal('0.00')
        valor_mejoras = datos_bien.mejoras or Decimal('0.00')
        valor_neto = valor_terreno + valor_mejoras
        
        context['valor_terreno'] = valor_terreno
        context['valor_mejoras'] = valor_mejoras
        context['valor_neto'] = valor_neto
        
        # Para Colindantes
        if tipo == 'colindantes':
            # Intentar obtener de la tabla Colindantes o Colindancias
            colindantes_list = Colindantes.objects.filter(empresa=empresa, cocata1=datos_bien.cocata1).order_by('tipo')
            context['colindantes_list'] = colindantes_list
            
        # Para Servicios Públicos
        if tipo == 'servicios':
            from .models import Complemento
            complemento = Complemento.objects.filter(empresa=empresa, cocomple=datos_bien.cocata1).first()
            context['complemento'] = complemento

    # Si no hay parámetros, mostrar página de búsqueda específica
    if not cocata1 and not identidad:
        return render(request, 'generar_constancia_busqueda.html', context)

    # Renderizar el template de la constancia (vista previa/impresión)
    template_name = f'constancia_{tipo}.html'
    return render(request, template_name, context)

@catastro_require_auth
@catastro_require_auth
def catastro_reportes(request):
    """
    Reportes de catastro
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Reportes'
    }
    return render(request, 'menu_reportes.html', context)

@catastro_require_auth
def catastro_menu_reportes(request):
    """
    Menú de reportes del módulo de catastro
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Reportes'
    }
    return render(request, 'menu_reportes.html', context)

@catastro_require_auth
def sumarial_avaluos_catastrales(request):
    """
    Vista para mostrar el formulario de sumariales de avalúos catastrales
    """
    usuario_nombre = request.session.get('catastro_usuario_nombre', 'Usuario')
    municipio_descripcion = request.session.get('catastro_municipio_descripcion', '')
    modulo = request.session.get('catastro_modulo', 'Catastro')
    
    context = {
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': municipio_descripcion,
        'modulo': modulo,
    }
    
    return render(request, 'sumarial_avaluos_catastrales.html', context)

@catastro_require_auth
def generar_sumarial_avaluos(request):
    """
    Vista para generar sumariales de avalúos catastrales
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Naturaleza, Dominio, TipoDocumento, Legales, Barrios
    from django.db.models import Sum, Count, Q, F
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    
    if request.method != 'POST':
        messages.error(request, 'Método no permitido')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    tipo_sumarial = request.POST.get('tipo_sumarial', '').strip()
    empresa = request.session.get('catastro_empresa', '')
    
    if not tipo_sumarial:
        messages.error(request, 'Debe seleccionar un tipo de sumarial')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    if not empresa:
        messages.error(request, 'No se ha seleccionado un municipio')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    # Construir query base
    query = Q(empresa=empresa)
    filtros_aplicados = []
    grupo_por = []
    
    # Aplicar filtros según el tipo de sumarial
    # Los sumariales básicos NO aplican filtros, agrupan TODOS los registros
    if tipo_sumarial == 'por_uso':
        # No aplicar filtro, mostrar todos los usos
        grupo_por = ['uso']
    elif tipo_sumarial == 'por_subuso':
        # Requiere seleccionar un uso, luego agrupa por subuso de ese uso
        codigo_uso = request.POST.get('codigo_uso', '').strip()
        if codigo_uso:
            # Normalizar el código de uso a 2 dígitos con ceros a la izquierda (01, 02, etc.)
            # Esto asegura que "1" se convierta en "01" para la comparación
            try:
                # Si es numérico, formatear con ceros a la izquierda
                codigo_uso_normalizado = str(int(codigo_uso)).zfill(2)
            except ValueError:
                # Si no es numérico, usar tal cual pero asegurar 2 caracteres
                codigo_uso_normalizado = codigo_uso.zfill(2) if codigo_uso.isdigit() else codigo_uso
            
            query &= Q(uso=codigo_uso_normalizado)
            # Obtener descripción del uso para mostrar en filtros
            try:
                uso_obj = Usos.objects.filter(uso=codigo_uso_normalizado).first()
                if uso_obj and uso_obj.desuso:
                    filtros_aplicados.append(f'Uso: {uso_obj.desuso} ({codigo_uso_normalizado})')
                else:
                    filtros_aplicados.append(f'Uso: {codigo_uso_normalizado}')
            except:
                filtros_aplicados.append(f'Uso: {codigo_uso_normalizado}')
        # Agrupar solo por subuso (el uso ya está filtrado)
        grupo_por = ['subuso']
    elif tipo_sumarial == 'por_barrio':
        # No aplicar filtro, mostrar todos los barrios
        grupo_por = ['barrio']
    elif tipo_sumarial == 'usos_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            # Obtener descripción del barrio para mostrar en filtros
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        # Agrupar por barrio y uso para mostrar todos los usos del barrio seleccionado
        grupo_por = ['barrio', 'uso']
    elif tipo_sumarial == 'usos_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
        # Agrupar por ficha y uso para mostrar todos los usos del perímetro seleccionado
        grupo_por = ['ficha', 'uso']
    elif tipo_sumarial == 'por_sexo':
        # No aplicar filtro, mostrar todos los sexos
        grupo_por = ['sexo']
    elif tipo_sumarial == 'por_identidad':
        # No aplicar filtro, agrupar por número de identidad (mostrará todos los diferentes)
        grupo_por = ['identidad']
    elif tipo_sumarial == 'naturaleza_jur_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            # Obtener descripción del barrio para mostrar en filtros
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        # Agrupar solo por naturaleza jurídica (el barrio ya está filtrado)
        # Esto mostrará todas las naturalezas jurídicas que existen en el barrio seleccionado
        grupo_por = ['naturaleza']
    elif tipo_sumarial == 'clase_dom_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            # Obtener descripción del barrio para mostrar en filtros
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        # Agrupar solo por dominio (el barrio ya está filtrado)
        # Esto mostrará todas las clases de dominio que existen en el barrio seleccionado
        grupo_por = ['dominio']
    elif tipo_sumarial == 'tipo_docto_por_barrio':
        codigo_barrio = request.POST.get('codigo_barrio', '').strip()
        if codigo_barrio:
            query &= Q(barrio=codigo_barrio)
            # Obtener descripción del barrio para mostrar en filtros
            try:
                barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
                if barrio_obj and barrio_obj.descripcion:
                    filtros_aplicados.append(f'Barrio: {barrio_obj.descripcion} ({codigo_barrio})')
                else:
                    filtros_aplicados.append(f'Barrio: {codigo_barrio}')
            except:
                filtros_aplicados.append(f'Barrio: {codigo_barrio}')
        # Agrupar solo por tipo (el barrio ya está filtrado)
        # El campo tipo está en la tabla legales
        grupo_por = ['tipo']
    elif tipo_sumarial == 'naturaleza_jur_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Perímetro: {tipo_ficha_nombre} (Ficha: {ficha})')
        # Agrupar por ficha y naturaleza para mostrar todas las naturalezas del perímetro seleccionado
        grupo_por = ['ficha', 'naturaleza']
    elif tipo_sumarial == 'clase_dom_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Perímetro: {tipo_ficha_nombre} (Ficha: {ficha})')
        # Agrupar por ficha y dominio para mostrar todas las clases de dominio del perímetro seleccionado
        grupo_por = ['ficha', 'dominio']
    elif tipo_sumarial == 'tipo_docto_por_perimetro':
        ficha = request.POST.get('ficha', '').strip()
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Perímetro: {tipo_ficha_nombre} (Ficha: {ficha})')
        # Agrupar solo por tipo (la ficha ya está filtrada)
        # El campo tipo está en la tabla legales
        grupo_por = ['tipo']
    elif tipo_sumarial == 'por_naturaleza_juridica':
        # No aplicar filtro, mostrar todas las naturalezas jurídicas
        # El campo naturaleza está en legales, se hará JOIN automáticamente
        grupo_por = ['naturaleza']
    elif tipo_sumarial == 'por_clase_dominio':
        # No aplicar filtro, mostrar todas las clases de dominio
        # El campo dominio está en legales, se hará JOIN automáticamente
        grupo_por = ['dominio']
    elif tipo_sumarial == 'por_tipo_documento':
        # No aplicar filtro, mostrar todos los tipos de documento
        # El campo tipo está en legales, se hará JOIN automáticamente
        grupo_por = ['tipo']
    elif tipo_sumarial == 'por_estatus_tributario':
        # No aplicar filtro, mostrar todos los estatus tributarios
        grupo_por = ['st']
    elif tipo_sumarial == 'por_zonificacion':
        # No aplicar filtro, mostrar todas las zonificaciones
        grupo_por = ['zonificacion']
    
    # Obtener datos agrupados
    # Si el grupo_por incluye 'naturaleza', 'dominio' o 'tipo', necesitamos hacer JOIN con la tabla legales
    # porque estos campos están en legales, no en bdcata1
    necesita_join_legales = 'naturaleza' in grupo_por or 'dominio' in grupo_por or 'tipo' in grupo_por
    
    if necesita_join_legales:
        # Hacer INNER JOIN con legales usando cocata1 = colegal
        # Los campos naturaleza, dominio y tipo están en la tabla legales, no en bdcata1
        select_dict = {}
        if 'naturaleza' in grupo_por:
            select_dict['naturaleza_legales'] = 'legales.naturaleza'
        if 'dominio' in grupo_por:
            select_dict['dominio_legales'] = 'legales.dominio'
        if 'tipo' in grupo_por:
            select_dict['tipo_legales'] = 'legales.tipo'
        
        avaluos = BDCata1.objects.filter(query).extra(
            select=select_dict,
            tables=['legales'],
            where=['bdcata1.cocata1 = legales.colegal AND bdcata1.empresa = legales.empresa']
        )
    else:
        # Obtener todos los registros que cumplen el query sin JOIN
        avaluos = BDCata1.objects.filter(query)
    
    # Agrupar manualmente ya que Django no soporta fácilmente múltiples agrupaciones con SUM de campos relacionados
    datos_agrupados = {}
    
    for avaluo in avaluos:
        # Crear clave de agrupación
        clave_grupo = []
        for campo in grupo_por:
            if campo == 'naturaleza' and necesita_join_legales:
                # Obtener naturaleza desde legales (viene del extra select)
                valor = getattr(avaluo, 'naturaleza_legales', None)
                if valor is None:
                    # Fallback: obtener desde legales directamente
                    try:
                        legal = Legales.objects.filter(empresa=avaluo.empresa, colegal=avaluo.cocata1).first()
                        valor = legal.naturaleza if legal else None
                    except:
                        valor = None
            elif campo == 'dominio' and necesita_join_legales:
                # Obtener dominio desde legales (viene del extra select)
                valor = getattr(avaluo, 'dominio_legales', None)
                if valor is None:
                    # Fallback: obtener desde legales directamente
                    try:
                        legal = Legales.objects.filter(empresa=avaluo.empresa, colegal=avaluo.cocata1).first()
                        valor = legal.dominio if legal else None
                    except:
                        valor = None
            elif campo == 'tipo' and necesita_join_legales:
                # Obtener tipo desde legales (viene del extra select)
                valor = getattr(avaluo, 'tipo_legales', None)
                if valor is None:
                    # Fallback: obtener desde legales directamente
                    try:
                        legal = Legales.objects.filter(empresa=avaluo.empresa, colegal=avaluo.cocata1).first()
                        valor = legal.tipo if legal else None
                    except:
                        valor = None
            else:
                valor = getattr(avaluo, campo, '')
            clave_grupo.append(str(valor) if valor else '')
        clave = '|'.join(clave_grupo)
        
        if clave not in datos_agrupados:
            datos_agrupados[clave] = {
                'grupo': {},
                'cantidad_predios': 0,
                'avaluo_terreno': Decimal('0.00'),
                'avaluo_edificacion': Decimal('0.00'),
                'avaluo_detalles': Decimal('0.00'),
                'avaluo_cultivo': Decimal('0.00'),
                'exencion': Decimal('0.00'),
                'valor_grabable': Decimal('0.00'),
                'impuesto': Decimal('0.00'),
                'total_area': Decimal('0.00'),
            }
            # Guardar valores del grupo
            for campo in grupo_por:
                if campo == 'naturaleza' and necesita_join_legales:
                    # Obtener naturaleza desde legales (viene del extra select)
                    valor = getattr(avaluo, 'naturaleza_legales', None)
                    if valor is None:
                        # Fallback: obtener desde legales directamente
                        try:
                            legal = Legales.objects.filter(empresa=avaluo.empresa, colegal=avaluo.cocata1).first()
                            valor = legal.naturaleza if legal else None
                        except:
                            valor = None
                elif campo == 'dominio' and necesita_join_legales:
                    # Obtener dominio desde legales (viene del extra select)
                    valor = getattr(avaluo, 'dominio_legales', None)
                    if valor is None:
                        # Fallback: obtener desde legales directamente
                        try:
                            legal = Legales.objects.filter(empresa=avaluo.empresa, colegal=avaluo.cocata1).first()
                            valor = legal.dominio if legal else None
                        except:
                            valor = None
                elif campo == 'tipo' and necesita_join_legales:
                    # Obtener tipo desde legales (viene del extra select)
                    valor = getattr(avaluo, 'tipo_legales', None)
                    if valor is None:
                        # Fallback: obtener desde legales directamente
                        try:
                            legal = Legales.objects.filter(empresa=avaluo.empresa, colegal=avaluo.cocata1).first()
                            valor = legal.tipo if legal else None
                        except:
                            valor = None
                else:
                    valor = getattr(avaluo, campo, '')
                datos_agrupados[clave]['grupo'][campo] = valor
        
        # Sumar valores
        datos_agrupados[clave]['cantidad_predios'] += 1
        datos_agrupados[clave]['avaluo_terreno'] += avaluo.bvl2tie or Decimal('0.00')
        datos_agrupados[clave]['avaluo_edificacion'] += avaluo.mejoras or Decimal('0.00')
        datos_agrupados[clave]['avaluo_detalles'] += avaluo.detalle or Decimal('0.00')
        datos_agrupados[clave]['avaluo_cultivo'] += avaluo.cultivo or Decimal('0.00')
        datos_agrupados[clave]['exencion'] += avaluo.exencion or Decimal('0.00')
        datos_agrupados[clave]['valor_grabable'] += avaluo.grabable or Decimal('0.00')
        datos_agrupados[clave]['impuesto'] += avaluo.impuesto or Decimal('0.00')
        
        # Obtener área del terreno
        try:
            terreno = BDTerreno.objects.filter(empresa=avaluo.empresa, cocata1=avaluo.cocata1).first()
            if terreno:
                area = (terreno.baream21 or Decimal('0.00')) + (terreno.baream22 or Decimal('0.00'))
                datos_agrupados[clave]['total_area'] += area
        except:
            pass
    
    # Convertir a lista y enriquecer con descripciones
    # Primero, obtener un ejemplo de avaluo para cada grupo para tener depto y municipio
    # Guardar el código de uso filtrado si existe (para sumarial por_subuso)
    codigo_uso_filtrado = None
    if tipo_sumarial == 'por_subuso':
        codigo_uso_post = request.POST.get('codigo_uso', '').strip()
        if codigo_uso_post:
            try:
                codigo_uso_filtrado = str(int(codigo_uso_post)).zfill(2)
            except:
                codigo_uso_filtrado = codigo_uso_post.zfill(2) if codigo_uso_post.isdigit() else codigo_uso_post
    
    datos_sumarial = []
    for clave, datos in datos_agrupados.items():
        fila = {
            'grupo': datos['grupo'],
            'grupo_descripcion': {},
            'cantidad_predios': datos['cantidad_predios'],
            'avaluo_terreno': datos['avaluo_terreno'],
            'avaluo_edificacion': datos['avaluo_edificacion'],
            'avaluo_detalles': datos['avaluo_detalles'],
            'avaluo_cultivo': datos['avaluo_cultivo'],
            'exencion': datos['exencion'],
            'valor_grabable': datos['valor_grabable'],
            'impuesto': datos['impuesto'],
            'total_area': datos['total_area'],
        }
        
        # Obtener un ejemplo de avaluo de este grupo para tener depto y municipio
        ejemplo_avaluo = None
        if 'barrio' in datos['grupo']:
            # Buscar un avaluo de este grupo para obtener depto y municipio
            grupo_filtro = Q(empresa=empresa)
            for campo, valor in datos['grupo'].items():
                grupo_filtro &= Q(**{campo: valor})
            ejemplo_avaluo = BDCata1.objects.filter(grupo_filtro).first()
        
        # Obtener descripciones
        if 'uso' in datos['grupo']:
            try:
                uso_obj = Usos.objects.filter(uso=datos['grupo']['uso']).first()
                fila['grupo_descripcion']['uso'] = uso_obj.desuso if uso_obj else datos['grupo']['uso']
            except:
                fila['grupo_descripcion']['uso'] = datos['grupo']['uso']
        
        if 'subuso' in datos['grupo']:
            try:
                # Para obtener la descripción del subuso, necesitamos el código de uso y el código de subuso
                codigo_uso = datos['grupo'].get('uso', '')
                codigo_subuso = datos['grupo']['subuso']
                
                # Si no hay código de uso en el grupo pero hay un filtro aplicado (sumarial por_subuso)
                if not codigo_uso and codigo_uso_filtrado:
                    codigo_uso = codigo_uso_filtrado
                
                if codigo_uso:
                    # Normalizar el código de uso a 2 dígitos
                    try:
                        codigo_uso_normalizado = str(int(codigo_uso)).zfill(2)
                    except:
                        codigo_uso_normalizado = codigo_uso.zfill(2) if codigo_uso.isdigit() else codigo_uso
                    
                    # Buscar en la tabla Subuso usando uso y codsubuso
                    subuso_obj = Subuso.objects.filter(uso=codigo_uso_normalizado, codsubuso=codigo_subuso).first()
                    if subuso_obj and subuso_obj.des_subuso:
                        fila['grupo_descripcion']['subuso'] = subuso_obj.des_subuso
                    else:
                        fila['grupo_descripcion']['subuso'] = codigo_subuso if codigo_subuso else '-'
                else:
                    # Si no hay código de uso, intentar buscar solo con codsubuso
                    subuso_obj = Subuso.objects.filter(codsubuso=codigo_subuso).first()
                    if subuso_obj and subuso_obj.des_subuso:
                        fila['grupo_descripcion']['subuso'] = subuso_obj.des_subuso
                    else:
                        fila['grupo_descripcion']['subuso'] = codigo_subuso if codigo_subuso else '-'
            except Exception as e:
                # En caso de error, mostrar el código del subuso
                fila['grupo_descripcion']['subuso'] = datos['grupo']['subuso'] if datos['grupo'].get('subuso') else '-'
        
        if 'ficha' in datos['grupo']:
            ficha_val = datos['grupo']['ficha']
            fila['grupo_descripcion']['ficha'] = 'Urbana' if ficha_val == '1' else 'Rural' if ficha_val == '2' else ficha_val
        
        if 'sexo' in datos['grupo']:
            sexo_val = datos['grupo']['sexo']
            fila['grupo_descripcion']['sexo'] = 'Masculino' if sexo_val == 'M' else 'Femenino' if sexo_val == 'F' else sexo_val
        
        if 'st' in datos['grupo']:
            # Mapeo de estatus tributario: 1=Exento, 2=Parcialmente Exento, 3=Totalmente Tributario, 4=Totalmente Exento sin Valores
            st_val = datos['grupo']['st']
            st_descripciones = {
                '1': 'Exento',
                '2': 'Parcialmente Exento',
                '3': 'Totalmente Tributario',
                '4': 'Totalmente Exento sin Valores',
            }
            # Convertir a string si es necesario
            st_str = str(st_val).strip() if st_val else ''
            fila['grupo_descripcion']['st'] = st_descripciones.get(st_str, st_str if st_str else '-')
        
        if 'naturaleza' in datos['grupo']:
            try:
                # Convertir el valor de naturaleza a Decimal para la búsqueda
                # El campo naturaleza en BDCata1 es Decimal, y codigo en Naturaleza también es Decimal
                naturaleza_val = datos['grupo']['naturaleza']
                if isinstance(naturaleza_val, Decimal):
                    codigo_nat = naturaleza_val
                else:
                    # Si es string u otro tipo, convertir a Decimal
                    try:
                        codigo_nat = Decimal(str(naturaleza_val))
                    except:
                        codigo_nat = Decimal('0')
                
                nat_obj = Naturaleza.objects.filter(codigo=codigo_nat).first()
                if nat_obj and nat_obj.descripcion:
                    fila['grupo_descripcion']['naturaleza'] = nat_obj.descripcion
                else:
                    # Si no se encuentra, mostrar el código
                    fila['grupo_descripcion']['naturaleza'] = str(naturaleza_val) if naturaleza_val else '-'
            except Exception as e:
                # En caso de error, mostrar el valor original
                fila['grupo_descripcion']['naturaleza'] = str(datos['grupo']['naturaleza']) if datos['grupo']['naturaleza'] else '-'
        
        if 'dominio' in datos['grupo']:
            try:
                # Convertir el valor de dominio a Decimal para la búsqueda
                # El campo dominio viene de legales y es Decimal, y codigo en Dominio también es Decimal
                dominio_val = datos['grupo']['dominio']
                if isinstance(dominio_val, Decimal):
                    codigo_dom = dominio_val
                else:
                    # Si es string u otro tipo, convertir a Decimal
                    try:
                        codigo_dom = Decimal(str(dominio_val))
                    except:
                        codigo_dom = Decimal('0')
                
                dom_obj = Dominio.objects.filter(codigo=codigo_dom).first()
                if dom_obj and dom_obj.descripcion:
                    fila['grupo_descripcion']['dominio'] = dom_obj.descripcion
                else:
                    # Si no se encuentra, mostrar el código
                    fila['grupo_descripcion']['dominio'] = str(dominio_val) if dominio_val else '-'
            except Exception as e:
                # En caso de error, mostrar el valor original
                fila['grupo_descripcion']['dominio'] = str(datos['grupo']['dominio']) if datos['grupo']['dominio'] else '-'
        
        if 'tipo' in datos['grupo']:
            try:
                # Convertir el valor de tipo a Decimal para la búsqueda
                # El campo tipo viene de legales y es Decimal, y codigo en TipoDocumento también es Decimal
                tipo_val = datos['grupo']['tipo']
                if isinstance(tipo_val, Decimal):
                    codigo_tipo = tipo_val
                else:
                    # Si es string u otro tipo, convertir a Decimal
                    try:
                        codigo_tipo = Decimal(str(tipo_val))
                    except:
                        codigo_tipo = Decimal('0')
                
                tipo_obj = TipoDocumento.objects.filter(codigo=codigo_tipo, empresa=empresa).first()
                if tipo_obj and tipo_obj.descripcion:
                    fila['grupo_descripcion']['tipo'] = tipo_obj.descripcion
                else:
                    # Si no se encuentra, mostrar el código
                    fila['grupo_descripcion']['tipo'] = str(tipo_val) if tipo_val else '-'
            except Exception as e:
                # En caso de error, mostrar el valor original
                fila['grupo_descripcion']['tipo'] = str(datos['grupo']['tipo']) if datos['grupo']['tipo'] else '-'
        
        if 'barrio' in datos['grupo']:
            try:
                codbarrio = datos['grupo']['barrio']
                # Obtener depto y municipio del ejemplo_avaluo si está disponible
                depto_val = None
                codmuni_val = None
                if ejemplo_avaluo:
                    depto_val = ejemplo_avaluo.depto
                    codmuni_val = ejemplo_avaluo.municipio
                
                # Buscar el barrio usando los campos correctos
                barrio_query = Barrios.objects.filter(empresa=empresa, codbarrio=codbarrio)
                
                # Si tenemos depto y municipio, usarlos para filtrar más precisamente
                if depto_val and codmuni_val:
                    # Convertir depto de CHAR(3) a CHAR(2) si es necesario (tomar últimos 2 caracteres)
                    depto_2 = depto_val[-2:] if len(depto_val) >= 2 else depto_val
                    barrio_query = barrio_query.filter(depto=depto_2, codmuni=codmuni_val)
                
                barrio_obj = barrio_query.first()
                
                if barrio_obj and barrio_obj.descripcion:
                    fila['grupo_descripcion']['barrio'] = barrio_obj.descripcion
                else:
                    # Si no se encuentra, intentar solo con empresa y codbarrio
                    barrio_obj_simple = Barrios.objects.filter(empresa=empresa, codbarrio=codbarrio).first()
                    fila['grupo_descripcion']['barrio'] = barrio_obj_simple.descripcion if barrio_obj_simple and barrio_obj_simple.descripcion else codbarrio
            except Exception as e:
                # En caso de error, mostrar el código del barrio
                fila['grupo_descripcion']['barrio'] = datos['grupo']['barrio']
        
        datos_sumarial.append(fila)
    
    # Ordenar los datos según el criterio de agrupación
    # Ordenar por los campos en grupo_por en el orden especificado
    def obtener_valor_ordenamiento(fila):
        valores_orden = []
        for campo in grupo_por:
            valor = fila['grupo'].get(campo, '')
            # Convertir a número si es posible para ordenamiento numérico
            try:
                if campo in ['uso', 'subuso', 'naturaleza', 'dominio', 'st', 'ficha']:
                    # Intentar convertir a número para ordenamiento numérico
                    valor_num = float(valor) if valor else 0
                    valores_orden.append((0, valor_num))  # 0 para orden numérico
                else:
                    valores_orden.append((1, str(valor).upper()))  # 1 para orden alfabético
            except:
                valores_orden.append((1, str(valor).upper()))
        return valores_orden
    
    datos_sumarial.sort(key=obtener_valor_ordenamiento)
    
    # Calcular totales generales (convertir Decimal a float para evitar problemas de serialización)
    totales = {
        'cantidad_predios': sum(d['cantidad_predios'] for d in datos_sumarial),
        'avaluo_terreno': float(sum(d['avaluo_terreno'] for d in datos_sumarial)),
        'avaluo_edificacion': float(sum(d['avaluo_edificacion'] for d in datos_sumarial)),
        'avaluo_detalles': float(sum(d['avaluo_detalles'] for d in datos_sumarial)),
        'avaluo_cultivo': float(sum(d['avaluo_cultivo'] for d in datos_sumarial)),
        'exencion': float(sum(d['exencion'] for d in datos_sumarial)),
        'valor_grabable': float(sum(d['valor_grabable'] for d in datos_sumarial)),
        'impuesto': float(sum(d['impuesto'] for d in datos_sumarial)),
        'total_area': float(sum(d['total_area'] for d in datos_sumarial)),
    }
    
    # Obtener nombre del tipo de sumarial
    nombres_sumariales = {
        'por_uso': 'Sumarial de Avaluos por Uso',
        'por_subuso': 'Sumarial de Avaluos por Sub Uso',
        'por_barrio': 'Sumarial de Avaluos por Barrio',
        'usos_por_barrio': 'Sumarial de Avaluos Catastrales Usos por Barrio',
        'usos_por_perimetro': 'Sumarial de Avaluos Catastrales Usos por Perímetro',
        'por_sexo': 'Sumarial de Avaluos Catastrales por Sexo',
        'por_identidad': 'Sumarial de Avaluos Catastrales por Numero de Identidad',
        'naturaleza_jur_por_barrio': 'Sumarial de Avaluos Catastrales por Naturaleza Juridica por Barrio',
        'clase_dom_por_barrio': 'Sumaria de Avaluos Catastrales Clase de Dominio por Barrio',
        'tipo_docto_por_barrio': 'Sumarial de Avaluos Catastrales Tipo de Documento por Barrio',
        'naturaleza_jur_por_perimetro': 'Sumarial de Avaluos Catastrales Naturaleza Juridica por Perímetro',
        'clase_dom_por_perimetro': 'Sumaria de Avaluos Catastrales Clase de Dominio por Perímetro',
        'tipo_docto_por_perimetro': 'Sumarial de Avaluos Catastrales Tipo de Documento por Perímetro',
        'por_naturaleza_juridica': 'Sumarial por Naturaleza Juridica',
        'por_clase_dominio': 'Sumarial por Clase de Dominio',
        'por_tipo_documento': 'Sumarial por Tipo de Documento',
        'por_estatus_tributario': 'Sumarial por Estatus Tributario',
        'por_zonificacion': 'Sumarial de Avaluos Catastrales por Zonificacion',
    }
    
    nombre_sumarial = nombres_sumariales.get(tipo_sumarial, 'Sumarial de Avaluos')
    
    # Convertir Decimal a float para serialización JSON (ANTES de crear el contexto)
    # Esto evita errores de serialización cuando Django intenta procesar el contexto
    def convertir_decimal_a_float(valor):
        """Convierte Decimal a float, deja otros tipos sin cambios"""
        if isinstance(valor, Decimal):
            return float(valor)
        return valor
    
    datos_sumarial_serializable = []
    for fila in datos_sumarial:
        # Convertir valores en 'grupo' si son Decimal
        grupo_serializable = {}
        for key, value in fila['grupo'].items():
            grupo_serializable[key] = convertir_decimal_a_float(value)
        
        # Convertir valores en 'grupo_descripcion' si son Decimal
        grupo_descripcion_serializable = {}
        for key, value in fila['grupo_descripcion'].items():
            grupo_descripcion_serializable[key] = convertir_decimal_a_float(value)
        
        fila_serializable = {
            'grupo': grupo_serializable,
            'grupo_descripcion': grupo_descripcion_serializable,
            'cantidad_predios': fila['cantidad_predios'],
            'avaluo_terreno': float(fila['avaluo_terreno']),
            'avaluo_edificacion': float(fila['avaluo_edificacion']),
            'avaluo_detalles': float(fila['avaluo_detalles']),
            'avaluo_cultivo': float(fila['avaluo_cultivo']),
            'exencion': float(fila['exencion']),
            'valor_grabable': float(fila['valor_grabable']),
            'impuesto': float(fila['impuesto']),
            'total_area': float(fila['total_area']),
        }
        datos_sumarial_serializable.append(fila_serializable)
    
    # Usar la versión serializable en el contexto para evitar errores de serialización
    context = {
        'nombre_sumarial': nombre_sumarial,
        'datos_sumarial': datos_sumarial_serializable,  # Usar versión serializable
        'totales': totales,
        'filtros_aplicados': filtros_aplicados,
        'grupo_por': grupo_por,
        'tipo_sumarial': tipo_sumarial,  # Pasar el tipo de sumarial al template
        'usuario_nombre': request.session.get('catastro_usuario_nombre', 'Usuario'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
        'modulo': request.session.get('catastro_modulo', 'Catastro'),
    }
    
    # Guardar parámetros en sesión para exportación
    
    # Convertir request.POST a diccionario serializable (evitar QueryDict)
    filtros_serializable = {}
    for key, value in request.POST.items():
        # Convertir listas a strings si es necesario
        if isinstance(value, list):
            filtros_serializable[key] = value[0] if value else ''
        else:
            filtros_serializable[key] = str(value) if value else ''
    
    request.session['sumarial_tipo'] = tipo_sumarial
    request.session['sumarial_filtros'] = filtros_serializable
    request.session['sumarial_datos'] = datos_sumarial_serializable
    # Los totales ya están convertidos a float arriba, solo guardarlos directamente
    request.session['sumarial_totales'] = totales
    request.session['sumarial_filtros_aplicados'] = filtros_aplicados
    request.session['sumarial_nombre'] = nombre_sumarial
    request.session['sumarial_grupo_por'] = grupo_por
    
    return render(request, 'resultado_sumarial.html', context)

@catastro_require_auth
def exportar_sumarial_excel(request):
    """
    Exportar sumarial de avalúos catastrales a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    from io import BytesIO
    from django.http import HttpResponse
    
    # Obtener datos de la sesión
    tipo_sumarial = request.session.get('sumarial_tipo')
    datos_sumarial = request.session.get('sumarial_datos', [])
    totales = request.session.get('sumarial_totales', {})
    filtros_aplicados = request.session.get('sumarial_filtros_aplicados', [])
    nombre_sumarial = request.session.get('sumarial_nombre', 'Sumarial de Avaluos')
    grupo_por = request.session.get('sumarial_grupo_por', [])
    
    if not datos_sumarial:
        messages.error(request, 'No hay datos para exportar')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sumarial"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    row = 1
    ws.merge_cells(f'A{row}:J{row}')
    cell = ws[f'A{row}']
    cell.value = nombre_sumarial
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Filtros aplicados
    if filtros_aplicados:
        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        cell = ws[f'A{row}']
        cell.value = 'Filtros Aplicados: ' + ', '.join(filtros_aplicados)
        cell.font = Font(size=10, italic=True)
        row += 1
    
    # Encabezados
    row += 1
    headers = ['Grupo', 'Cantidad de Predios', 'Avalúo Terreno', 'Avalúo Edificación', 
               'Avalúo Detalles Adicionales', 'Avalúo Cultivo', 'Exención', 
               'Valor Grabable', 'Impuesto', 'Total Área']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_style
    
    # Datos
    for fila_data in datos_sumarial:
        row += 1
        # Grupo
        grupo_texto = []
        for campo in grupo_por:
            valor = fila_data['grupo'].get(campo, '')
            if campo == 'uso' and fila_data['grupo_descripcion'].get('uso'):
                grupo_texto.append(f"Uso: {fila_data['grupo_descripcion']['uso']}")
            elif campo == 'barrio':
                grupo_texto.append(f"Barrio: {valor}")
            elif campo == 'ficha' and fila_data['grupo_descripcion'].get('ficha'):
                grupo_texto.append(f"Ficha: {fila_data['grupo_descripcion']['ficha']}")
            elif campo == 'sexo' and fila_data['grupo_descripcion'].get('sexo'):
                grupo_texto.append(f"Sexo: {fila_data['grupo_descripcion']['sexo']}")
            elif campo == 'identidad':
                grupo_texto.append(f"Identidad: {valor}")
            elif campo == 'naturaleza' and fila_data['grupo_descripcion'].get('naturaleza'):
                grupo_texto.append(f"Naturaleza Jurídica: {fila_data['grupo_descripcion']['naturaleza']}")
            elif campo == 'dominio' and fila_data['grupo_descripcion'].get('dominio'):
                grupo_texto.append(f"Clase de Dominio: {fila_data['grupo_descripcion']['dominio']}")
            elif campo == 'st':
                grupo_texto.append(f"Estatus Tributario: {valor}")
            elif campo == 'zonificacion':
                grupo_texto.append(f"Zonificación: {valor}")
            else:
                grupo_texto.append(f"{campo}: {valor}")
        
        ws.cell(row=row, column=1).value = '\n'.join(grupo_texto)
        ws.cell(row=row, column=2).value = fila_data['cantidad_predios']
        ws.cell(row=row, column=3).value = float(fila_data['avaluo_terreno'])
        ws.cell(row=row, column=4).value = float(fila_data['avaluo_edificacion'])
        ws.cell(row=row, column=5).value = float(fila_data['avaluo_detalles'])
        ws.cell(row=row, column=6).value = float(fila_data['avaluo_cultivo'])
        ws.cell(row=row, column=7).value = float(fila_data['exencion'])
        ws.cell(row=row, column=8).value = float(fila_data['valor_grabable'])
        ws.cell(row=row, column=9).value = float(fila_data['impuesto'])
        ws.cell(row=row, column=10).value = float(fila_data['total_area'])
        
        # Aplicar formato
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = border_style
            if col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col in [3, 4, 5, 6, 7, 8, 9]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Totales
    row += 1
    ws.cell(row=row, column=1).value = 'TOTALES:'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = totales.get('cantidad_predios', 0)
    ws.cell(row=row, column=3).value = totales.get('avaluo_terreno', 0)
    ws.cell(row=row, column=4).value = totales.get('avaluo_edificacion', 0)
    ws.cell(row=row, column=5).value = totales.get('avaluo_detalles', 0)
    ws.cell(row=row, column=6).value = totales.get('avaluo_cultivo', 0)
    ws.cell(row=row, column=7).value = totales.get('exencion', 0)
    ws.cell(row=row, column=8).value = totales.get('valor_grabable', 0)
    ws.cell(row=row, column=9).value = totales.get('impuesto', 0)
    ws.cell(row=row, column=10).value = totales.get('total_area', 0)
    
    # Formato de totales
    for col in range(1, 11):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = border_style
        if col in [3, 4, 5, 6, 7, 8, 9, 10]:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif col == 2:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sumarial_avaluos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response

@catastro_require_auth
def exportar_sumarial_pdf(request):
    """
    Exportar sumarial de avalúos catastrales a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import legal, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    from io import BytesIO
    from django.http import HttpResponse
    
    # Obtener datos de la sesión
    tipo_sumarial = request.session.get('sumarial_tipo')
    datos_sumarial = request.session.get('sumarial_datos', [])
    totales = request.session.get('sumarial_totales', {})
    filtros_aplicados = request.session.get('sumarial_filtros_aplicados', [])
    nombre_sumarial = request.session.get('sumarial_nombre', 'Sumarial de Avaluos')
    grupo_por = request.session.get('sumarial_grupo_por', [])
    
    if not datos_sumarial:
        messages.error(request, 'No hay datos para exportar')
        return redirect('catastro:sumarial_avaluos_catastrales')
    
    # Preparar respuesta PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=legal, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=1  # Centrado
    )
    
    # Contenido
    story = []
    
    # Título
    story.append(Paragraph(nombre_sumarial, title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Filtros aplicados
    if filtros_aplicados:
        filtros_texto = 'Filtros Aplicados: ' + ', '.join(filtros_aplicados)
        story.append(Paragraph(filtros_texto, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
    
    # Preparar datos de la tabla
    table_data = []
    
    # Encabezados
    headers = ['Grupo', 'Cant. Predios', 'Avalúo Terreno', 'Avalúo Edificación', 
               'Avalúo Detalles', 'Avalúo Cultivo', 'Exención', 
               'Valor Grabable', 'Impuesto', 'Total Área']
    table_data.append(headers)
    
    # Datos
    for fila_data in datos_sumarial:
        grupo_texto = []
        for campo in grupo_por:
            valor = fila_data['grupo'].get(campo, '')
            if campo == 'uso' and fila_data['grupo_descripcion'].get('uso'):
                grupo_texto.append(f"Uso: {fila_data['grupo_descripcion']['uso']}")
            elif campo == 'barrio':
                grupo_texto.append(f"Barrio: {valor}")
            elif campo == 'ficha' and fila_data['grupo_descripcion'].get('ficha'):
                grupo_texto.append(f"Ficha: {fila_data['grupo_descripcion']['ficha']}")
            elif campo == 'sexo' and fila_data['grupo_descripcion'].get('sexo'):
                grupo_texto.append(f"Sexo: {fila_data['grupo_descripcion']['sexo']}")
            elif campo == 'identidad':
                grupo_texto.append(f"Identidad: {valor}")
            elif campo == 'naturaleza' and fila_data['grupo_descripcion'].get('naturaleza'):
                grupo_texto.append(f"Nat. Jur.: {fila_data['grupo_descripcion']['naturaleza']}")
            elif campo == 'dominio' and fila_data['grupo_descripcion'].get('dominio'):
                grupo_texto.append(f"Clase Dom.: {fila_data['grupo_descripcion']['dominio']}")
            elif campo == 'st':
                grupo_texto.append(f"Est. Trib.: {valor}")
            elif campo == 'zonificacion':
                grupo_texto.append(f"Zonif.: {valor}")
            else:
                grupo_texto.append(f"{campo}: {valor}")
        
        row = [
            '\n'.join(grupo_texto),
            str(fila_data['cantidad_predios']),
            f"L. {fila_data['avaluo_terreno']:.2f}",
            f"L. {fila_data['avaluo_edificacion']:.2f}",
            f"L. {fila_data['avaluo_detalles']:.2f}",
            f"L. {fila_data['avaluo_cultivo']:.2f}",
            f"L. {fila_data['exencion']:.2f}",
            f"L. {fila_data['valor_grabable']:.2f}",
            f"L. {fila_data['impuesto']:.2f}",
            f"{fila_data['total_area']:.2f}",
        ]
        table_data.append(row)
    
    # Totales
    table_data.append([
        'TOTALES:',
        str(totales.get('cantidad_predios', 0)),
        f"L. {totales.get('avaluo_terreno', 0):.2f}",
        f"L. {totales.get('avaluo_edificacion', 0):.2f}",
        f"L. {totales.get('avaluo_detalles', 0):.2f}",
        f"L. {totales.get('avaluo_cultivo', 0):.2f}",
        f"L. {totales.get('exencion', 0):.2f}",
        f"L. {totales.get('valor_grabable', 0):.2f}",
        f"L. {totales.get('impuesto', 0):.2f}",
        f"{totales.get('total_area', 0):.2f}",
    ])
    
    # Crear tabla
    table = Table(table_data, repeatRows=1)
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (0, -2), 'LEFT'),  # Columna Grupo
        ('ALIGN', (1, 1), (-1, -2), 'RIGHT'),  # Resto de columnas
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # Fila de totales
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E7E6E6')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (0, -1), 'LEFT'),
        ('ALIGN', (1, -1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 8),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ])
    
    table.setStyle(table_style)
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sumarial_avaluos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response

@catastro_require_auth
def listado_avaluos_catastrales(request):
    """
    Formulario para generar listados de avalúos catastrales
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Listado de Avalúos'
    }
    return render(request, 'listado_avaluos_catastrales.html', context)

@catastro_require_auth
def listados_catastrales(request):
    """
    Formulario para generar diferentes tipos de listados catastrales
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Listados'
    }
    return render(request, 'listados_catastrales.html', context)

@catastro_require_auth
def generar_listado_catastral(request):
    """
    Vista para generar los diferentes tipos de listados catastrales
    """
    if request.method == 'POST':
        tipo_listado = request.POST.get('tipo_listado')
        # Obtener empresa de la sesión (código del municipio seleccionado en el login)
        empresa = request.session.get('catastro_empresa')
        
        if not empresa:
            messages.error(request, 'No se encontró la empresa en la sesión. Por favor, inicie sesión nuevamente.')
            return redirect('catastro:catastro_login')
        
        if not tipo_listado:
            messages.error(request, 'Debe seleccionar un tipo de listado')
            return redirect('catastro:listados_catastrales')
        
        try:
            # Obtener filtros del POST
            filtros_dict = {}
            for key in request.POST.keys():
                if key != 'csrfmiddlewaretoken' and key != 'tipo_listado':
                    value = request.POST.get(key, '').strip()
                    if isinstance(value, list) and len(value) > 0:
                        value = value[0].strip()
                    filtros_dict[key] = value
            
            logger.info(f"Tipo de listado: {tipo_listado}")
            logger.info(f"Empresa (municipio): {empresa}")
            logger.info(f"Filtros recibidos: {filtros_dict}")
            
            # Obtener datos según el tipo de listado
            if tipo_listado == 'listado_general':
                datos_informe, filtros_aplicados = obtener_listado_general(request, empresa)
            elif tipo_listado == 'listado_barrio':
                codigo_barrio = filtros_dict.get('codigo_barrio', '').strip()
                if not codigo_barrio:
                    messages.error(request, 'Debe seleccionar un barrio')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_barrio(request, empresa, codigo_barrio)
            elif tipo_listado == 'listado_uso':
                codigo_uso = filtros_dict.get('codigo_uso', '').strip()
                if not codigo_uso:
                    messages.error(request, 'Debe seleccionar un uso')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_uso(request, empresa, codigo_uso)
            elif tipo_listado == 'listado_dni':
                numero_dni = filtros_dict.get('numero_dni', '').strip()
                if not numero_dni:
                    messages.error(request, 'Debe ingresar el número de DNI')
                    return redirect('catastro:listados_catastrales')
                if len(numero_dni) < 5:
                    messages.error(request, 'El número de DNI debe tener al menos 5 caracteres')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_dni(request, empresa, numero_dni)
            elif tipo_listado == 'listado_propietario':
                nombre_propietario = filtros_dict.get('nombre_propietario', '').strip()
                if not nombre_propietario:
                    messages.error(request, 'Debe ingresar el nombre del propietario')
                    return redirect('catastro:listados_catastrales')
                if len(nombre_propietario) < 3:
                    messages.error(request, 'El nombre del propietario debe tener al menos 3 caracteres')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_propietario(request, empresa, nombre_propietario)
            elif tipo_listado == 'listado_sexo':
                codigo_sexo = filtros_dict.get('codigo_sexo', '').strip()
                if not codigo_sexo:
                    messages.error(request, 'Debe seleccionar un sexo')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_sexo(request, empresa, codigo_sexo)
            elif tipo_listado == 'listado_estatus':
                codigo_estatus = filtros_dict.get('codigo_estatus', '').strip()
                if not codigo_estatus:
                    messages.error(request, 'Debe seleccionar un estatus tributario')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_estatus(request, empresa, codigo_estatus)
            elif tipo_listado == 'listado_perimetro':
                tipo_ficha = filtros_dict.get('tipo_ficha_perimetro', '').strip()
                if not tipo_ficha:
                    messages.error(request, 'Debe seleccionar el tipo de perímetro (Urbano o Rural)')
                    return redirect('catastro:listados_catastrales')
                if tipo_ficha not in ['1', '2']:
                    messages.error(request, 'Tipo de perímetro no válido')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_perimetro(request, empresa, tipo_ficha)
            elif tipo_listado == 'listado_rango_avaluo':
                avaluo_minimo_str = filtros_dict.get('avaluo_minimo', '').strip()
                avaluo_maximo_str = filtros_dict.get('avaluo_maximo', '').strip()
                if not avaluo_minimo_str:
                    messages.error(request, 'Debe ingresar el valor mínimo del avalúo')
                    return redirect('catastro:listados_catastrales')
                if not avaluo_maximo_str:
                    messages.error(request, 'Debe ingresar el valor máximo del avalúo')
                    return redirect('catastro:listados_catastrales')
                try:
                    avaluo_minimo = Decimal(avaluo_minimo_str)
                    avaluo_maximo = Decimal(avaluo_maximo_str)
                    if avaluo_minimo < 0:
                        messages.error(request, 'El valor mínimo del avalúo debe ser mayor o igual a cero')
                        return redirect('catastro:listados_catastrales')
                    if avaluo_maximo < 0:
                        messages.error(request, 'El valor máximo del avalúo debe ser mayor o igual a cero')
                        return redirect('catastro:listados_catastrales')
                    if avaluo_minimo > avaluo_maximo:
                        messages.error(request, 'El valor mínimo no puede ser mayor que el valor máximo')
                        return redirect('catastro:listados_catastrales')
                except (ValueError, InvalidOperation):
                    messages.error(request, 'Los valores del rango de avalúo no son válidos')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_rango_avaluo(request, empresa, avaluo_minimo, avaluo_maximo)
            elif tipo_listado == 'listado_dominio':
                codigo_dominio = filtros_dict.get('codigo_dominio', '').strip()
                if not codigo_dominio:
                    messages.error(request, 'Debe seleccionar una clase de dominio')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_dominio(request, empresa, codigo_dominio)
            elif tipo_listado == 'listado_tipo_documento':
                codigo_tipo_documento = filtros_dict.get('codigo_tipo_documento', '').strip()
                if not codigo_tipo_documento:
                    messages.error(request, 'Debe seleccionar un tipo de documento')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_tipo_documento(request, empresa, codigo_tipo_documento)
            elif tipo_listado == 'listado_naturaleza':
                codigo_naturaleza = filtros_dict.get('codigo_naturaleza', '').strip()
                if not codigo_naturaleza:
                    messages.error(request, 'Debe seleccionar una naturaleza jurídica')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_naturaleza(request, empresa, codigo_naturaleza)
            elif tipo_listado == 'listado_codigo_habitacional':
                codigo_habitacional = filtros_dict.get('codigo_habitacional', '').strip()
                if not codigo_habitacional:
                    messages.error(request, 'Debe seleccionar un código habitacional')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_codigo_habitacional(request, empresa, codigo_habitacional)
            elif tipo_listado == 'listado_dominio_barrio':
                codigo_barrio = filtros_dict.get('codigo_barrio_dominio', '').strip()
                codigo_dominio = filtros_dict.get('codigo_dominio_barrio', '').strip()
                if not codigo_barrio:
                    messages.error(request, 'Debe seleccionar un barrio')
                    return redirect('catastro:listados_catastrales')
                if not codigo_dominio:
                    messages.error(request, 'Debe seleccionar una clase de dominio')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_dominio_barrio(request, empresa, codigo_barrio, codigo_dominio)
            elif tipo_listado == 'listado_naturaleza_barrio':
                codigo_barrio = filtros_dict.get('codigo_barrio_naturaleza', '').strip()
                codigo_naturaleza = filtros_dict.get('codigo_naturaleza_barrio', '').strip()
                if not codigo_barrio:
                    messages.error(request, 'Debe seleccionar un barrio')
                    return redirect('catastro:listados_catastrales')
                if not codigo_naturaleza:
                    messages.error(request, 'Debe seleccionar una naturaleza jurídica')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_naturaleza_barrio(request, empresa, codigo_barrio, codigo_naturaleza)
            elif tipo_listado == 'listado_tipo_documento_barrio':
                codigo_barrio = filtros_dict.get('codigo_barrio_tipo_documento', '').strip()
                codigo_tipo_documento = filtros_dict.get('codigo_tipo_documento_barrio', '').strip()
                if not codigo_barrio:
                    messages.error(request, 'Debe seleccionar un barrio')
                    return redirect('catastro:listados_catastrales')
                if not codigo_tipo_documento:
                    messages.error(request, 'Debe seleccionar un tipo de documento')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_tipo_documento_barrio(request, empresa, codigo_barrio, codigo_tipo_documento)
            elif tipo_listado == 'listado_codigo_habitacional_barrio':
                codigo_barrio = filtros_dict.get('codigo_barrio_habitacional', '').strip()
                codigo_habitacional = filtros_dict.get('codigo_habitacional_barrio', '').strip()
                if not codigo_barrio:
                    messages.error(request, 'Debe seleccionar un barrio')
                    return redirect('catastro:listados_catastrales')
                if not codigo_habitacional:
                    messages.error(request, 'Debe seleccionar un código habitacional')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_codigo_habitacional_barrio(request, empresa, codigo_barrio, codigo_habitacional)
            elif tipo_listado == 'listado_dominio_perimetro':
                tipo_ficha = filtros_dict.get('tipo_ficha_dominio', '').strip()
                codigo_dominio = filtros_dict.get('codigo_dominio_perimetro', '').strip()
                if not tipo_ficha:
                    messages.error(request, 'Debe seleccionar el tipo de perímetro')
                    return redirect('catastro:listados_catastrales')
                if not codigo_dominio:
                    messages.error(request, 'Debe seleccionar una clase de dominio')
                    return redirect('catastro:listados_catastrales')
                if tipo_ficha not in ['1', '2']:
                    messages.error(request, 'Tipo de perímetro no válido')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_dominio_perimetro(request, empresa, tipo_ficha, codigo_dominio)
            elif tipo_listado == 'listado_naturaleza_perimetro':
                tipo_ficha = filtros_dict.get('tipo_ficha_naturaleza', '').strip()
                codigo_naturaleza = filtros_dict.get('codigo_naturaleza_perimetro', '').strip()
                if not tipo_ficha:
                    messages.error(request, 'Debe seleccionar el tipo de perímetro')
                    return redirect('catastro:listados_catastrales')
                if not codigo_naturaleza:
                    messages.error(request, 'Debe seleccionar una naturaleza jurídica')
                    return redirect('catastro:listados_catastrales')
                if tipo_ficha not in ['1', '2']:
                    messages.error(request, 'Tipo de perímetro no válido')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_naturaleza_perimetro(request, empresa, tipo_ficha, codigo_naturaleza)
            elif tipo_listado == 'listado_tipo_documento_perimetro':
                tipo_ficha = filtros_dict.get('tipo_ficha_tipo_documento', '').strip()
                codigo_tipo_documento = filtros_dict.get('codigo_tipo_documento_perimetro', '').strip()
                if not tipo_ficha:
                    messages.error(request, 'Debe seleccionar el tipo de perímetro')
                    return redirect('catastro:listados_catastrales')
                if not codigo_tipo_documento:
                    messages.error(request, 'Debe seleccionar un tipo de documento')
                    return redirect('catastro:listados_catastrales')
                if tipo_ficha not in ['1', '2']:
                    messages.error(request, 'Tipo de perímetro no válido')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_tipo_documento_perimetro(request, empresa, tipo_ficha, codigo_tipo_documento)
            elif tipo_listado == 'listado_codigo_habitacional_perimetro':
                tipo_ficha = filtros_dict.get('tipo_ficha_habitacional', '').strip()
                codigo_habitacional = filtros_dict.get('codigo_habitacional_perimetro', '').strip()
                if not tipo_ficha:
                    messages.error(request, 'Debe seleccionar el tipo de perímetro')
                    return redirect('catastro:listados_catastrales')
                if not codigo_habitacional:
                    messages.error(request, 'Debe seleccionar un código habitacional')
                    return redirect('catastro:listados_catastrales')
                if tipo_ficha not in ['1', '2']:
                    messages.error(request, 'Tipo de perímetro no válido')
                    return redirect('catastro:listados_catastrales')
                datos_informe, filtros_aplicados = obtener_listado_por_codigo_habitacional_perimetro(request, empresa, tipo_ficha, codigo_habitacional)
            else:
                messages.error(request, 'Tipo de listado no válido')
                return redirect('catastro:listados_catastrales')
            
            # Calcular totales
            if datos_informe:
                totales = {
                    'bvl2tie': sum(Decimal(str(d.get('bvl2tie', 0))) for d in datos_informe),
                    'mejoras': sum(Decimal(str(d.get('mejoras', 0))) for d in datos_informe),
                    'detalle': sum(Decimal(str(d.get('detalle', 0))) for d in datos_informe),
                    'cultivo': sum(Decimal(str(d.get('cultivo', 0))) for d in datos_informe),
                    'exencion': sum(Decimal(str(d.get('exencion', 0))) for d in datos_informe),
                    'grabable': sum(Decimal(str(d.get('grabable', 0))) for d in datos_informe),
                    'impuesto': sum(Decimal(str(d.get('impuesto', 0))) for d in datos_informe),
                    'valor_area': sum(Decimal(str(d.get('valor_area', 0))) for d in datos_informe),
                }
            else:
                totales = {
                    'bvl2tie': Decimal('0.00'),
                    'mejoras': Decimal('0.00'),
                    'detalle': Decimal('0.00'),
                    'cultivo': Decimal('0.00'),
                    'exencion': Decimal('0.00'),
                    'grabable': Decimal('0.00'),
                    'impuesto': Decimal('0.00'),
                    'valor_area': Decimal('0.00'),
                }
            
            # Guardar en sesión para exportaciones
            request.session['listado_tipo'] = tipo_listado
            request.session['listado_filtros'] = filtros_dict
            
            context = {
                'empresa': empresa or '',
                'municipio_descripcion': request.session.get('catastro_municipio_descripcion') or '',
                'usuario_nombre': request.session.get('catastro_usuario_nombre') or 'Usuario',
                'modulo': 'Catastro - Listados',
                'datos_informe': datos_informe if datos_informe else [],
                'tipo_listado': tipo_listado or '',
                'filtros_aplicados': filtros_aplicados if filtros_aplicados else [],
                'total_registros': len(datos_informe) if datos_informe else 0,
                'totales': totales,
                'filtros_dict': filtros_dict,
            }
            
            logger.info(f"Listado generado. Total registros: {context['total_registros']}")
            
            return render(request, 'informe_avaluos_catastrales.html', context)
            
        except Exception as e:
            logger.error(f"Error al generar listado: {str(e)}", exc_info=True)
            messages.error(request, f'Error al generar el listado: {str(e)}')
            return redirect('catastro:listados_catastrales')
    
    return redirect('catastro:listados_catastrales')

def obtener_listado_general(request, empresa):
    """
    Obtiene el listado general de avalúos catastrales
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa (código de municipio del login)
        avaluos_qs = BDCata1.objects.filter(empresa=empresa).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado general. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], ['Listado General de Avalúos Catastrales', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, ['Listado General de Avalúos Catastrales']
        
    except Exception as e:
        logger.error(f"Error al obtener listado general: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_barrio(request, empresa, codigo_barrio):
    """
    Obtiene el listado de avalúos catastrales por barrio
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Barrios
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Obtener descripción del barrio (filtrado por empresa)
        barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
        barrio_desc = barrio_obj.descripcion if barrio_obj else codigo_barrio
        
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa (código de municipio del login) y barrio
        avaluos_qs = BDCata1.objects.filter(empresa=empresa, barrio=codigo_barrio).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por barrio {codigo_barrio}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Barrio: {codigo_barrio} - {barrio_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Barrio: {codigo_barrio} - {barrio_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por barrio: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_uso(request, empresa, codigo_uso):
    """
    Obtiene el listado de avalúos catastrales por uso
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Obtener descripción del uso
        uso_obj = Usos.objects.filter(uso=codigo_uso).first()
        uso_desc = uso_obj.desuso if uso_obj else codigo_uso
        
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa (código de municipio del login) y uso
        avaluos_qs = BDCata1.objects.filter(empresa=empresa, uso=codigo_uso).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por uso {codigo_uso}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Uso: {codigo_uso} - {uso_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Uso: {codigo_uso} - {uso_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por uso: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_dni(request, empresa, numero_dni):
    """
    Obtiene el listado de avalúos catastrales por DNI (identidad)
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa (código de municipio del login) y identidad (DNI)
        # Usar icontains para búsqueda parcial (puede contener guiones o espacios)
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            identidad__icontains=numero_dni
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por DNI {numero_dni}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por DNI: {numero_dni}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por DNI: {numero_dni}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por DNI: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_propietario(request, empresa, nombre_propietario):
    """
    Obtiene el listado de avalúos catastrales por nombre o apellido del propietario
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa y buscar en nombres o apellidos (búsqueda parcial)
        from django.db.models import Q
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa
        ).filter(
            Q(nombres__icontains=nombre_propietario) | Q(apellidos__icontains=nombre_propietario)
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por propietario '{nombre_propietario}'. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Propietario: {nombre_propietario}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario_completo = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario_completo,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Propietario: {nombre_propietario}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por propietario: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_sexo(request, empresa, codigo_sexo):
    """
    Obtiene el listado de avalúos catastrales por sexo del propietario
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Obtener descripción del sexo
        sexo_desc = 'Masculino' if codigo_sexo == 'M' else 'Femenino' if codigo_sexo == 'F' else codigo_sexo
        
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa (código de municipio del login) y sexo
        avaluos_qs = BDCata1.objects.filter(empresa=empresa, sexo=codigo_sexo).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por sexo {codigo_sexo}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Sexo: {sexo_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Sexo: {sexo_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por sexo: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_estatus(request, empresa, codigo_estatus):
    """
    Obtiene el listado de avalúos catastrales por estatus tributario (campo ST)
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Mapeo de códigos de estatus tributario a descripciones
        estatus_descripciones = {
            '1': 'Exento',
            '2': 'Parcialmente Exento',
            '3': 'Totalmente Tributario',
            '4': 'Totalmente Exento sin Valores'
        }
        
        estatus_desc = estatus_descripciones.get(codigo_estatus, codigo_estatus)
        
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Filtrar por empresa (código de municipio del login) y estatus tributario (st)
        avaluos_qs = BDCata1.objects.filter(empresa=empresa, st=codigo_estatus).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por estatus tributario {codigo_estatus} ({estatus_desc}). Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Estatus Tributario: {codigo_estatus} - {estatus_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Calcular área total desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Estatus Tributario: {codigo_estatus} - {estatus_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por estatus: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_perimetro(request, empresa, tipo_ficha):
    """
    Obtiene el listado de avalúos catastrales por perímetro
    El perímetro se refiere al campo ficha en bdcata1:
    - ficha = 1 → Perímetro Urbano
    - ficha = 2 → Perímetro Rural
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    Filtra por empresa (código de municipio) según la sesión del login
    Filtra por ficha (1=Urbano, 2=Rural)
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Determinar el nombre del perímetro
        tipo_perimetro_nombre = 'Urbano' if tipo_ficha == '1' else 'Rural' if tipo_ficha == '2' else tipo_ficha
        
        # Filtrar por empresa y ficha (perímetro)
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Obtener área desde bdterreno
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            ficha=tipo_ficha
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por perímetro {tipo_perimetro_nombre} (ficha={tipo_ficha}). Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Perímetro {tipo_perimetro_nombre} (Ficha: {tipo_ficha})', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Obtener valores desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            
            # Calcular área total
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Perímetro {tipo_perimetro_nombre} (Ficha: {tipo_ficha})']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por perímetro: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']


def obtener_listado_por_rango_avaluo(request, empresa, avaluo_minimo, avaluo_maximo):
    """
    Obtiene el listado de avalúos catastrales por rango de avalúo
    Filtra por empresa (código de municipio) según la sesión del login
    Filtra por el campo grabable (valor grabable) dentro del rango especificado
    Usa LEFT JOIN entre bdcata1 (maestra) y bdterreno para obtener el área
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Filtrar por empresa y rango de avalúo (campo grabable)
        # LEFT JOIN usando extra() para unir bdcata1 con bdterreno
        # bdcata1 es la tabla maestra
        # Obtener área desde bdterreno
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            grabable__gte=avaluo_minimo,
            grabable__lte=avaluo_maximo
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por rango de avalúo [{avaluo_minimo} - {avaluo_maximo}]. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Rango de Avalúo: L. {avaluo_minimo:,.2f} - L. {avaluo_maximo:,.2f}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            # Obtener valores desde el LEFT JOIN
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            
            # Calcular área total
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Rango de Avalúo: L. {avaluo_minimo:,.2f} - L. {avaluo_maximo:,.2f}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por rango de avalúo: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_dominio(request, empresa, codigo_dominio):
    """
    Obtiene el listado de avalúos catastrales por clase de dominio
    Usa JOIN con la tabla legales para obtener el dominio
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, Dominio
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de dominio a Decimal
        try:
            codigo_dom_decimal = Decimal(str(codigo_dominio))
        except:
            codigo_dom_decimal = Decimal('0')
        
        # Obtener descripción del dominio
        dominio_obj = Dominio.objects.filter(codigo=codigo_dom_decimal).first()
        dominio_desc = dominio_obj.descripcion if dominio_obj else codigo_dominio
        
        # Obtener claves catastrales que tienen el dominio especificado en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            dominio=codigo_dom_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa y claves que están en legales con el dominio especificado
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por dominio {codigo_dominio}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Clase de Dominio: {codigo_dominio} - {dominio_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Clase de Dominio: {codigo_dominio} - {dominio_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por dominio: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_tipo_documento(request, empresa, codigo_tipo_documento):
    """
    Obtiene el listado de avalúos catastrales por tipo de documento
    Usa JOIN con la tabla legales para obtener el tipo de documento
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, TipoDocumento
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de tipo de documento a Decimal
        try:
            codigo_tipo_decimal = Decimal(str(codigo_tipo_documento))
        except:
            codigo_tipo_decimal = Decimal('0')
        
        # Obtener descripción del tipo de documento
        tipo_doc_obj = TipoDocumento.objects.filter(empresa=empresa, codigo=codigo_tipo_decimal).first()
        tipo_doc_desc = tipo_doc_obj.descripcion if tipo_doc_obj else codigo_tipo_documento
        
        # Obtener claves catastrales que tienen el tipo de documento especificado en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            tipo=codigo_tipo_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa y claves que están en legales con el tipo especificado
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por tipo de documento {codigo_tipo_documento}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Tipo de Documento: {codigo_tipo_documento} - {tipo_doc_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Tipo de Documento: {codigo_tipo_documento} - {tipo_doc_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por tipo de documento: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_naturaleza(request, empresa, codigo_naturaleza):
    """
    Obtiene el listado de avalúos catastrales por naturaleza jurídica
    Usa JOIN con la tabla legales para obtener la naturaleza
    Filtra por empresa (código de municipio) según la sesión del login
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, Naturaleza
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de naturaleza a Decimal
        try:
            codigo_nat_decimal = Decimal(str(codigo_naturaleza))
        except:
            codigo_nat_decimal = Decimal('0')
        
        # Obtener descripción de la naturaleza
        naturaleza_obj = Naturaleza.objects.filter(codigo=codigo_nat_decimal).first()
        naturaleza_desc = naturaleza_obj.descripcion if naturaleza_obj else codigo_naturaleza
        
        # Obtener claves catastrales que tienen la naturaleza especificada en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            naturaleza=codigo_nat_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa y claves que están en legales con la naturaleza especificada
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por naturaleza {codigo_naturaleza}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Naturaleza Jurídica: {codigo_naturaleza} - {naturaleza_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Naturaleza Jurídica: {codigo_naturaleza} - {naturaleza_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por naturaleza: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_codigo_habitacional(request, empresa, codigo_habitacional):
    """
    Obtiene el listado de avalúos catastrales por código habitacional
    Filtra por empresa (código de municipio) según la sesión del login
    El campo codhab está directamente en bdcata1
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Habitacional
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Obtener descripción del código habitacional
        habitacional_obj = Habitacional.objects.filter(cohabit=codigo_habitacional).first()
        habitacional_desc = habitacional_obj.bdeshabit if habitacional_obj else codigo_habitacional
        
        # Filtrar bdcata1 por empresa y código habitacional
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            codhab=codigo_habitacional
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por código habitacional {codigo_habitacional}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Código Habitacional: {codigo_habitacional} - {habitacional_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Código Habitacional: {codigo_habitacional} - {habitacional_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por código habitacional: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_dominio_barrio(request, empresa, codigo_barrio, codigo_dominio):
    """
    Obtiene el listado de avalúos catastrales por clase de dominio y barrio
    Usa JOIN con la tabla legales para obtener el dominio
    Filtra por empresa, barrio y dominio
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, Dominio, Barrios
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de dominio a Decimal
        try:
            codigo_dom_decimal = Decimal(str(codigo_dominio))
        except:
            codigo_dom_decimal = Decimal('0')
        
        # Obtener descripciones
        barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
        barrio_desc = barrio_obj.descripcion if barrio_obj else codigo_barrio
        
        dominio_obj = Dominio.objects.filter(codigo=codigo_dom_decimal).first()
        dominio_desc = dominio_obj.descripcion if dominio_obj else codigo_dominio
        
        # Obtener claves catastrales que tienen el dominio especificado en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            dominio=codigo_dom_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa, barrio y claves que están en legales con el dominio especificado
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            barrio=codigo_barrio,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por dominio {codigo_dominio} y barrio {codigo_barrio}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Clase de Dominio / Barrio: {codigo_dominio} - {dominio_desc} / {codigo_barrio} - {barrio_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Clase de Dominio / Barrio: {codigo_dominio} - {dominio_desc} / {codigo_barrio} - {barrio_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por dominio y barrio: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_naturaleza_barrio(request, empresa, codigo_barrio, codigo_naturaleza):
    """
    Obtiene el listado de avalúos catastrales por naturaleza jurídica y barrio
    Usa JOIN con la tabla legales para obtener la naturaleza
    Filtra por empresa, barrio y naturaleza
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, Naturaleza, Barrios
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de naturaleza a Decimal
        try:
            codigo_nat_decimal = Decimal(str(codigo_naturaleza))
        except:
            codigo_nat_decimal = Decimal('0')
        
        # Obtener descripciones
        barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
        barrio_desc = barrio_obj.descripcion if barrio_obj else codigo_barrio
        
        naturaleza_obj = Naturaleza.objects.filter(codigo=codigo_nat_decimal).first()
        naturaleza_desc = naturaleza_obj.descripcion if naturaleza_obj else codigo_naturaleza
        
        # Obtener claves catastrales que tienen la naturaleza especificada en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            naturaleza=codigo_nat_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa, barrio y claves que están en legales con la naturaleza especificada
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            barrio=codigo_barrio,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por naturaleza {codigo_naturaleza} y barrio {codigo_barrio}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Naturaleza Jurídica / Barrio: {codigo_naturaleza} - {naturaleza_desc} / {codigo_barrio} - {barrio_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Naturaleza Jurídica / Barrio: {codigo_naturaleza} - {naturaleza_desc} / {codigo_barrio} - {barrio_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por naturaleza y barrio: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_tipo_documento_barrio(request, empresa, codigo_barrio, codigo_tipo_documento):
    """
    Obtiene el listado de avalúos catastrales por tipo de documento y barrio
    Usa JOIN con la tabla legales para obtener el tipo de documento
    Filtra por empresa, barrio y tipo de documento
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, TipoDocumento, Barrios
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de tipo de documento a Decimal
        try:
            codigo_tipo_decimal = Decimal(str(codigo_tipo_documento))
        except:
            codigo_tipo_decimal = Decimal('0')
        
        # Obtener descripciones
        barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
        barrio_desc = barrio_obj.descripcion if barrio_obj else codigo_barrio
        
        tipo_doc_obj = TipoDocumento.objects.filter(empresa=empresa, codigo=codigo_tipo_decimal).first()
        tipo_doc_desc = tipo_doc_obj.descripcion if tipo_doc_obj else codigo_tipo_documento
        
        # Obtener claves catastrales que tienen el tipo de documento especificado en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            tipo=codigo_tipo_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa, barrio y claves que están en legales con el tipo especificado
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            barrio=codigo_barrio,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por tipo de documento {codigo_tipo_documento} y barrio {codigo_barrio}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Tipo de Documento / Barrio: {codigo_tipo_documento} - {tipo_doc_desc} / {codigo_barrio} - {barrio_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Tipo de Documento / Barrio: {codigo_tipo_documento} - {tipo_doc_desc} / {codigo_barrio} - {barrio_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por tipo de documento y barrio: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_codigo_habitacional_barrio(request, empresa, codigo_barrio, codigo_habitacional):
    """
    Obtiene el listado de avalúos catastrales por código habitacional y barrio
    Filtra por empresa, barrio y código habitacional
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Habitacional, Barrios
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Obtener descripciones
        barrio_obj = Barrios.objects.filter(empresa=empresa, codbarrio=codigo_barrio).first()
        barrio_desc = barrio_obj.descripcion if barrio_obj else codigo_barrio
        
        habitacional_obj = Habitacional.objects.filter(cohabit=codigo_habitacional).first()
        habitacional_desc = habitacional_obj.bdeshabit if habitacional_obj else codigo_habitacional
        
        # Filtrar bdcata1 por empresa, barrio y código habitacional
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            barrio=codigo_barrio,
            codhab=codigo_habitacional
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por código habitacional {codigo_habitacional} y barrio {codigo_barrio}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Código Habitacional / Barrio: {codigo_habitacional} - {habitacional_desc} / {codigo_barrio} - {barrio_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Código Habitacional / Barrio: {codigo_habitacional} - {habitacional_desc} / {codigo_barrio} - {barrio_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por código habitacional y barrio: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_dominio_perimetro(request, empresa, tipo_ficha, codigo_dominio):
    """
    Obtiene el listado de avalúos catastrales por clase de dominio y perímetro
    Usa JOIN con la tabla legales para obtener el dominio
    Filtra por empresa, perímetro (ficha) y dominio
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, Dominio
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de dominio a Decimal
        try:
            codigo_dom_decimal = Decimal(str(codigo_dominio))
        except:
            codigo_dom_decimal = Decimal('0')
        
        # Obtener descripción del dominio
        dominio_obj = Dominio.objects.filter(codigo=codigo_dom_decimal).first()
        dominio_desc = dominio_obj.descripcion if dominio_obj else codigo_dominio
        
        perimetro_desc = 'Urbano' if tipo_ficha == '1' else 'Rural'
        
        # Obtener claves catastrales que tienen el dominio especificado en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            dominio=codigo_dom_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa, perímetro (ficha) y claves que están en legales con el dominio especificado
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            ficha=tipo_ficha,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por dominio {codigo_dominio} y perímetro {tipo_ficha}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Clase de Dominio / Perímetro: {codigo_dominio} - {dominio_desc} / {perimetro_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Clase de Dominio / Perímetro: {codigo_dominio} - {dominio_desc} / {perimetro_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por dominio y perímetro: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_naturaleza_perimetro(request, empresa, tipo_ficha, codigo_naturaleza):
    """
    Obtiene el listado de avalúos catastrales por naturaleza jurídica y perímetro
    Usa JOIN con la tabla legales para obtener la naturaleza
    Filtra por empresa, perímetro (ficha) y naturaleza
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, Naturaleza
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de naturaleza a Decimal
        try:
            codigo_nat_decimal = Decimal(str(codigo_naturaleza))
        except:
            codigo_nat_decimal = Decimal('0')
        
        # Obtener descripciones
        naturaleza_obj = Naturaleza.objects.filter(codigo=codigo_nat_decimal).first()
        naturaleza_desc = naturaleza_obj.descripcion if naturaleza_obj else codigo_naturaleza
        
        perimetro_desc = 'Urbano' if tipo_ficha == '1' else 'Rural'
        
        # Obtener claves catastrales que tienen la naturaleza especificada en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            naturaleza=codigo_nat_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa, perímetro (ficha) y claves que están en legales con la naturaleza especificada
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            ficha=tipo_ficha,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por naturaleza {codigo_naturaleza} y perímetro {tipo_ficha}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Naturaleza Jurídica / Perímetro: {codigo_naturaleza} - {naturaleza_desc} / {perimetro_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Naturaleza Jurídica / Perímetro: {codigo_naturaleza} - {naturaleza_desc} / {perimetro_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por naturaleza y perímetro: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_tipo_documento_perimetro(request, empresa, tipo_ficha, codigo_tipo_documento):
    """
    Obtiene el listado de avalúos catastrales por tipo de documento y perímetro
    Usa JOIN con la tabla legales para obtener el tipo de documento
    Filtra por empresa, perímetro (ficha) y tipo de documento
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Legales, TipoDocumento
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Convertir código de tipo de documento a Decimal
        try:
            codigo_tipo_decimal = Decimal(str(codigo_tipo_documento))
        except:
            codigo_tipo_decimal = Decimal('0')
        
        # Obtener descripciones
        tipo_doc_obj = TipoDocumento.objects.filter(empresa=empresa, codigo=codigo_tipo_decimal).first()
        tipo_doc_desc = tipo_doc_obj.descripcion if tipo_doc_obj else codigo_tipo_documento
        
        perimetro_desc = 'Urbano' if tipo_ficha == '1' else 'Rural'
        
        # Obtener claves catastrales que tienen el tipo de documento especificado en legales
        legales_ids = Legales.objects.filter(
            empresa=empresa,
            tipo=codigo_tipo_decimal
        ).values_list('colegal', flat=True)
        
        # Filtrar bdcata1 por empresa, perímetro (ficha) y claves que están en legales con el tipo especificado
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            ficha=tipo_ficha,
            cocata1__in=legales_ids
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por tipo de documento {codigo_tipo_documento} y perímetro {tipo_ficha}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Tipo de Documento / Perímetro: {codigo_tipo_documento} - {tipo_doc_desc} / {perimetro_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Tipo de Documento / Perímetro: {codigo_tipo_documento} - {tipo_doc_desc} / {perimetro_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por tipo de documento y perímetro: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

def obtener_listado_por_codigo_habitacional_perimetro(request, empresa, tipo_ficha, codigo_habitacional):
    """
    Obtiene el listado de avalúos catastrales por código habitacional y perímetro
    Filtra por empresa, perímetro (ficha) y código habitacional
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso, Habitacional
    
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión']
    
    try:
        # Obtener descripciones
        habitacional_obj = Habitacional.objects.filter(cohabit=codigo_habitacional).first()
        habitacional_desc = habitacional_obj.bdeshabit if habitacional_obj else codigo_habitacional
        
        perimetro_desc = 'Urbano' if tipo_ficha == '1' else 'Rural'
        
        # Filtrar bdcata1 por empresa, perímetro (ficha) y código habitacional
        avaluos_qs = BDCata1.objects.filter(
            empresa=empresa,
            ficha=tipo_ficha,
            codhab=codigo_habitacional
        ).extra(
            select={
                'baream21': 'SELECT baream21 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
                'baream22': 'SELECT baream22 FROM bdterreno WHERE bdterreno.empresa = bdcata1.empresa AND bdterreno.cocata1 = bdcata1.cocata1 LIMIT 1',
            }
        ).order_by('cocata1')
        
        total_bd = avaluos_qs.count()
        logger.info(f"Listado por código habitacional {codigo_habitacional} y perímetro {tipo_ficha}. Total de registros en BD: {total_bd}")
        
        if total_bd == 0:
            return [], [f'Listado por Código Habitacional / Perímetro: {codigo_habitacional} - {habitacional_desc} / {perimetro_desc}', 'No se encontraron registros']
        
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            baream21 = getattr(avaluo, 'baream21', None)
            baream22 = getattr(avaluo, 'baream22', None)
            area_total = (Decimal(str(baream21)) if baream21 is not None else Decimal('0.00')) + \
                        (Decimal(str(baream22)) if baream22 is not None else Decimal('0.00'))
            
            desc_uso = avaluo.uso or ''
            if avaluo.uso:
                if avaluo.uso not in usos_cache:
                    try:
                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                    except:
                        usos_cache[avaluo.uso] = avaluo.uso
                desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
            
            desc_subuso = avaluo.subuso or ''
            if avaluo.subuso and avaluo.uso:
                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                if cache_key not in subusos_cache:
                    try:
                        subuso_obj = Subuso.objects.filter(
                            uso=avaluo.uso,
                            codsubuso=avaluo.subuso
                        ).first()
                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                    except:
                        subusos_cache[cache_key] = avaluo.subuso
                desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
            
            nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
            
            datos_informe.append({
                'clave_catastral': avaluo.cocata1 or '',
                'nombre_propietario': nombre_propietario,
                'ubicacion': avaluo.ubicacion or '',
                'uso': avaluo.uso or '',
                'desc_uso': desc_uso,
                'subuso': avaluo.subuso or '',
                'desc_subuso': desc_subuso,
                'st': avaluo.st or '',
                'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
                'mejoras': avaluo.mejoras or Decimal('0.00'),
                'detalle': avaluo.detalle or Decimal('0.00'),
                'cultivo': avaluo.cultivo or Decimal('0.00'),
                'exencion': avaluo.exencion or Decimal('0.00'),
                'grabable': avaluo.grabable or Decimal('0.00'),
                'impuesto': avaluo.impuesto or Decimal('0.00'),
                'valor_area': area_total,
            })
        
        return datos_informe, [f'Listado por Código Habitacional / Perímetro: {codigo_habitacional} - {habitacional_desc} / {perimetro_desc}']
        
    except Exception as e:
        logger.error(f"Error al obtener listado por código habitacional y perímetro: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}']

@catastro_require_auth
def generar_informe_avaluos(request):
    """
    Vista para generar el informe de avalúos catastrales según los parámetros
    """
    if request.method == 'POST':
        tipo_informe = request.POST.get('tipo_informe')
        empresa = request.session.get('catastro_empresa')
        
        if not tipo_informe:
            messages.error(request, 'Debe seleccionar un tipo de informe')
            return redirect('catastro:listado_avaluos_catastrales')
        
        try:
            # Obtener filtros del POST - usar .get() para evitar problemas con listas
            filtros_dict = {}
            for key in request.POST.keys():
                if key != 'csrfmiddlewaretoken':
                    value = request.POST.get(key, '').strip()
                    # Si es una lista, tomar el primer valor
                    if isinstance(value, list) and len(value) > 0:
                        value = value[0].strip()
                    filtros_dict[key] = value
            
            logger.info(f"Tipo de informe: {tipo_informe}")
            logger.info(f"Filtros recibidos: {filtros_dict}")
            
            # Obtener datos usando la función auxiliar
            datos_informe, filtros_aplicados, query = obtener_datos_informe_avaluos(
                request, tipo_informe, filtros_dict
            )
            
            logger.info(f"Total de registros encontrados: {len(datos_informe)}")
            
            # Calcular totales - manejar casos donde no hay datos
            if datos_informe:
                totales = {
                    'bvl2tie': sum(Decimal(str(d.get('bvl2tie', 0))) for d in datos_informe),
                    'mejoras': sum(Decimal(str(d.get('mejoras', 0))) for d in datos_informe),
                    'detalle': sum(Decimal(str(d.get('detalle', 0))) for d in datos_informe),
                    'cultivo': sum(Decimal(str(d.get('cultivo', 0))) for d in datos_informe),
                    'exencion': sum(Decimal(str(d.get('exencion', 0))) for d in datos_informe),
                    'grabable': sum(Decimal(str(d.get('grabable', 0))) for d in datos_informe),
                    'impuesto': sum(Decimal(str(d.get('impuesto', 0))) for d in datos_informe),
                    'valor_area': sum(Decimal(str(d.get('valor_area', 0))) for d in datos_informe),
                }
            else:
                totales = {
                    'bvl2tie': Decimal('0.00'),
                    'mejoras': Decimal('0.00'),
                    'detalle': Decimal('0.00'),
                    'cultivo': Decimal('0.00'),
                    'exencion': Decimal('0.00'),
                    'grabable': Decimal('0.00'),
                    'impuesto': Decimal('0.00'),
                    'valor_area': Decimal('0.00'),
                }
            
            # Guardar filtros en sesión para exportaciones
            request.session['informe_avaluos_tipo'] = tipo_informe
            request.session['informe_avaluos_filtros'] = filtros_dict
            
            # Asegurar que todas las variables del contexto tengan valores válidos
            context = {
                'empresa': empresa or '',
                'municipio_descripcion': request.session.get('catastro_municipio_descripcion') or '',
                'usuario_nombre': request.session.get('catastro_usuario_nombre') or 'Usuario',
                'modulo': 'Catastro - Listado de Avalúos',
                'datos_informe': datos_informe if datos_informe else [],
                'tipo_informe': tipo_informe or '',
                'filtros_aplicados': filtros_aplicados if filtros_aplicados else [],
                'total_registros': len(datos_informe) if datos_informe else 0,
                'totales': totales,
                'filtros_dict': filtros_dict,  # Para pasar a las exportaciones
            }
            
            logger.info(f"Contexto preparado. Total registros: {context['total_registros']}, Filtros: {len(context['filtros_aplicados'])}")
            
            return render(request, 'informe_avaluos_catastrales.html', context)
            
        except Exception as e:
            logger.error(f"Error al generar informe de avalúos: {str(e)}", exc_info=True)
            messages.error(request, f'Error al generar el informe: {str(e)}')
            return redirect('catastro:listado_avaluos_catastrales')
    
    return redirect('catastro:listado_avaluos_catastrales')

def obtener_datos_informe_avaluos(request, tipo_informe, filtros_dict):
    """
    Función auxiliar para obtener los datos del informe de avalúos
    Retorna: (datos_informe, filtros_aplicados, query)
    """
    from .models import BDCata1, BDTerreno, Usos, Subuso
    
    empresa = request.session.get('catastro_empresa')
    if not empresa:
        logger.warning("No se encontró empresa en la sesión")
        return [], ['Error: No se encontró empresa en la sesión'], Q()
    
    query = Q(empresa=empresa)
    filtros_aplicados = []
    
    logger.info(f"Procesando tipo_informe: {tipo_informe}, empresa: {empresa}")
    
    if tipo_informe == 'listado_completo':
        filtros_aplicados.append('Listado Completo de Avalúos Catastrales')
    elif tipo_informe == 'general':
        filtros_aplicados.append('Informe General (Urbano y Rural)')
    elif tipo_informe == 'clave_catastral':
        clave = filtros_dict.get('clave_catastral', '').strip()
        if clave:
            query &= Q(cocata1__icontains=clave)
            filtros_aplicados.append(f'Clave Catastral: {clave}')
        else:
            logger.warning("Tipo informe 'clave_catastral' pero no se proporcionó clave")
            filtros_aplicados.append('Clave Catastral: (no especificada)')
    elif tipo_informe == 'ficha':
        tipo_ficha = filtros_dict.get('tipo_ficha')
        numero_ficha = filtros_dict.get('numero_ficha', '').strip()
        if tipo_ficha:
            query &= Q(ficha=tipo_ficha)
            tipo_ficha_nombre = 'Urbana' if tipo_ficha == '1' else 'Rural'
            filtros_aplicados.append(f'Tipo de Ficha: {tipo_ficha_nombre}')
        if numero_ficha:
            query &= Q(nofichas=numero_ficha)
            filtros_aplicados.append(f'Número de Ficha: {numero_ficha}')
    elif tipo_informe == 'propietario':
        nombre = filtros_dict.get('nombre_propietario', '').strip()
        apellido = filtros_dict.get('apellido_propietario', '').strip()
        if nombre:
            query &= Q(nombres__icontains=nombre)
            filtros_aplicados.append(f'Nombre: {nombre}')
        if apellido:
            query &= Q(apellidos__icontains=apellido)
            filtros_aplicados.append(f'Apellido: {apellido}')
    elif tipo_informe == 'dni':
        dni = filtros_dict.get('dni', '').strip()
        if dni:
            query &= Q(identidad__icontains=dni)
            filtros_aplicados.append(f'DNI: {dni}')
    elif tipo_informe == 'uso':
        codigo_uso = filtros_dict.get('codigo_uso', '').strip()
        if codigo_uso:
            query &= Q(uso=codigo_uso)
            try:
                uso_obj = Usos.objects.filter(uso=codigo_uso).first()
                uso_desc = uso_obj.desuso if uso_obj else codigo_uso
                filtros_aplicados.append(f'Uso: {codigo_uso} - {uso_desc}')
            except:
                filtros_aplicados.append(f'Uso: {codigo_uso}')
    elif tipo_informe == 'codigo_habitacion':
        codigo_hab = filtros_dict.get('codigo_habitacion', '').strip()
        if codigo_hab:
            query &= Q(codhab=codigo_hab)
            filtros_aplicados.append(f'Código Habitación: {codigo_hab}')
    elif tipo_informe == 'sexo':
        sexo = filtros_dict.get('sexo', '').strip()
        if sexo:
            query &= Q(sexo=sexo)
            sexo_desc = 'Masculino' if sexo == 'M' else 'Femenino' if sexo == 'F' else sexo
            filtros_aplicados.append(f'Sexo: {sexo_desc}')
    elif tipo_informe == 'subuso':
        codigo_uso = filtros_dict.get('codigo_uso_subuso', '').strip()
        codigo_subuso = filtros_dict.get('codigo_subuso', '').strip()
        if codigo_uso and codigo_subuso:
            query &= Q(uso=codigo_uso, subuso=codigo_subuso)
            try:
                uso_obj = Usos.objects.filter(uso=codigo_uso).first()
                subuso_obj = Subuso.objects.filter(uso=codigo_uso, codsubuso=codigo_subuso).first()
                uso_desc = uso_obj.desuso if uso_obj else codigo_uso
                subuso_desc = subuso_obj.des_subuso if subuso_obj else codigo_subuso
                filtros_aplicados.append(f'Uso: {codigo_uso} - {uso_desc}')
                filtros_aplicados.append(f'Sub Uso: {codigo_subuso} - {subuso_desc}')
            except:
                filtros_aplicados.append(f'Uso: {codigo_uso}, Sub Uso: {codigo_subuso}')
    elif tipo_informe == 'barrio_uso':
        barrio = filtros_dict.get('barrio_barrio_uso', '').strip()
        codigo_uso = filtros_dict.get('codigo_uso_barrio_uso', '').strip()
        if barrio:
            query &= Q(barrio=barrio)
            filtros_aplicados.append(f'Barrio: {barrio}')
        if codigo_uso:
            query &= Q(uso=codigo_uso)
            try:
                uso_obj = Usos.objects.filter(uso=codigo_uso).first()
                uso_desc = uso_obj.desuso if uso_obj else codigo_uso
                filtros_aplicados.append(f'Uso: {codigo_uso} - {uso_desc}')
            except:
                filtros_aplicados.append(f'Uso: {codigo_uso}')
    elif tipo_informe == 'rango_avaluo':
        tipo_valor = filtros_dict.get('tipo_valor_rango', '').strip()
        valor_min = filtros_dict.get('valor_minimo', '').strip()
        valor_max = filtros_dict.get('valor_maximo', '').strip()
        if tipo_valor and (valor_min or valor_max):
            try:
                from decimal import Decimal
                if valor_min:
                    valor_min_decimal = Decimal(valor_min)
                else:
                    valor_min_decimal = Decimal('0')
                if valor_max:
                    valor_max_decimal = Decimal(valor_max)
                else:
                    valor_max_decimal = Decimal('999999999')
                
                if tipo_valor == 'total':
                    # Valor total = bvl2tie + mejoras + detalle + cultivo
                    # Filtrar después de calcular el total en el procesamiento de datos
                    # Por ahora, aplicar filtros básicos que se refinarán después
                    pass  # Se filtrará en el procesamiento de datos después de calcular el total
                elif tipo_valor == 'grabable':
                    query &= Q(grabable__gte=valor_min_decimal, grabable__lte=valor_max_decimal)
                elif tipo_valor == 'impuesto':
                    query &= Q(impuesto__gte=valor_min_decimal, impuesto__lte=valor_max_decimal)
                elif tipo_valor == 'terreno':
                    query &= Q(bvl2tie__gte=valor_min_decimal, bvl2tie__lte=valor_max_decimal)
                elif tipo_valor == 'edificacion':
                    query &= Q(mejoras__gte=valor_min_decimal, mejoras__lte=valor_max_decimal)
                
                tipo_valor_desc = {
                    'total': 'Valor Total',
                    'grabable': 'Valor Grabable',
                    'impuesto': 'Impuesto',
                    'terreno': 'Avalúo Terreno',
                    'edificacion': 'Avalúo Edificación'
                }.get(tipo_valor, tipo_valor)
                filtros_aplicados.append(f'{tipo_valor_desc}: L. {valor_min} - L. {valor_max}')
            except Exception as e:
                logger.error(f"Error al procesar rango de avalúo: {str(e)}")
    elif tipo_informe == 'naturaleza_juridica_perimetro':
        naturaleza = filtros_dict.get('naturaleza_juridica_perimetro', '').strip()
        ficha = filtros_dict.get('ficha_naturaleza', '').strip()
        if naturaleza:
            # Convertir naturaleza a Decimal para la búsqueda y el query
            try:
                naturaleza_decimal = Decimal(str(naturaleza))
            except:
                naturaleza_decimal = Decimal('0')
            
            query &= Q(naturaleza=naturaleza_decimal)
            try:
                from .models import Naturaleza
                # Buscar en la tabla Naturaleza usando el código (Decimal)
                nat_obj = Naturaleza.objects.filter(codigo=naturaleza_decimal).first()
                nat_desc = nat_obj.descripcion if nat_obj and nat_obj.descripcion else str(naturaleza)
                filtros_aplicados.append(f'Naturaleza Jurídica: {naturaleza} - {nat_desc}')
            except Exception as e:
                filtros_aplicados.append(f'Naturaleza Jurídica: {naturaleza}')
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
    elif tipo_informe == 'clase_dominio_perimetro':
        dominio = filtros_dict.get('clase_dominio_perimetro', '').strip()
        ficha = filtros_dict.get('ficha_dominio', '').strip()
        if dominio:
            query &= Q(dominio=dominio)
            try:
                from .models import Dominio
                dom_obj = Dominio.objects.filter(codigo=dominio).first()
                dom_desc = dom_obj.descripcion if dom_obj else dominio
                filtros_aplicados.append(f'Clase de Dominio: {dominio} - {dom_desc}')
            except:
                filtros_aplicados.append(f'Clase de Dominio: {dominio}')
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
    elif tipo_informe == 'tipo_documento_perimetro':
        tipo_doc = filtros_dict.get('tipo_documento_perimetro', '').strip()
        ficha = filtros_dict.get('ficha_documento', '').strip()
        if tipo_doc:
            # El tipo de documento está en la tabla legales (campo tipo)
            # colegal se relaciona con cocata1
            from .models import Legales
            legales_ids = Legales.objects.filter(tipo=tipo_doc, empresa=empresa).values_list('colegal', flat=True)
            query &= Q(cocata1__in=legales_ids)
            try:
                from .models import TipoDocumento
                tipo_obj = TipoDocumento.objects.filter(codigo=tipo_doc, empresa=empresa).first()
                tipo_desc = tipo_obj.descripcion if tipo_obj else tipo_doc
                filtros_aplicados.append(f'Tipo de Documento: {tipo_doc} - {tipo_desc}')
            except:
                filtros_aplicados.append(f'Tipo de Documento: {tipo_doc}')
        if ficha:
            query &= Q(ficha=ficha)
            tipo_ficha_nombre = 'Urbana' if ficha == '1' else 'Rural' if ficha == '2' else ficha
            filtros_aplicados.append(f'Ficha: {tipo_ficha_nombre} ({ficha})')
    
    try:
        avaluos_qs = BDCata1.objects.filter(query).order_by('cocata1')
        total_bd = avaluos_qs.count()
        logger.info(f"Query aplicado. Total de registros en BD: {total_bd}")
        if total_bd == 0:
            logger.warning(f"No se encontraron registros con los filtros aplicados. Query: {query}")
            return [], filtros_aplicados + ['No se encontraron registros con los criterios especificados'], query
    except Exception as e:
        logger.error(f"Error al consultar BDCata1: {str(e)}", exc_info=True)
        return [], [f'Error al consultar la base de datos: {str(e)}'], query
    
    datos_informe = []
    
    # Si es rango de avalúo con valor total, necesitamos filtrar después de calcular
    tipo_valor_rango = filtros_dict.get('tipo_valor_rango', '').strip() if tipo_informe == 'rango_avaluo' else None
    valor_min_rango = None
    valor_max_rango = None
    if tipo_valor_rango == 'total':
        try:
            from decimal import Decimal
            valor_min_str = filtros_dict.get('valor_minimo', '').strip()
            valor_max_str = filtros_dict.get('valor_maximo', '').strip()
            valor_min_rango = Decimal(valor_min_str) if valor_min_str else Decimal('0')
            valor_max_rango = Decimal(valor_max_str) if valor_max_str else Decimal('999999999')
        except:
            pass
    
    usos_cache = {}
    subusos_cache = {}
    
    for avaluo in avaluos_qs:
        terreno = None
        area_total = Decimal('0.00')
        try:
            terreno = BDTerreno.objects.filter(
                empresa=avaluo.empresa,
                cocata1=avaluo.cocata1
            ).first()
            if terreno:
                area_total = (terreno.baream21 or Decimal('0.00')) + (terreno.baream22 or Decimal('0.00'))
        except Exception as e:
            logger.error(f"Error al obtener terreno para {avaluo.cocata1}: {str(e)}")
        
        desc_uso = avaluo.uso or ''
        if avaluo.uso:
            if avaluo.uso not in usos_cache:
                try:
                    uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                    usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                except:
                    usos_cache[avaluo.uso] = avaluo.uso
            desc_uso = usos_cache.get(avaluo.uso, avaluo.uso)
        
        desc_subuso = avaluo.subuso or ''
        if avaluo.subuso and avaluo.uso:
            cache_key = f"{avaluo.uso}_{avaluo.subuso}"
            if cache_key not in subusos_cache:
                try:
                    subuso_obj = Subuso.objects.filter(
                        uso=avaluo.uso,
                        codsubuso=avaluo.subuso
                    ).first()
                    subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                except:
                    subusos_cache[cache_key] = avaluo.subuso
            desc_subuso = subusos_cache.get(cache_key, avaluo.subuso)
        
        # Calcular valor total para filtrado por rango
        if tipo_valor_rango == 'total':
            valor_total = (avaluo.bvl2tie or Decimal('0.00')) + (avaluo.mejoras or Decimal('0.00')) + (avaluo.detalle or Decimal('0.00')) + (avaluo.cultivo or Decimal('0.00'))
            # Filtrar por rango de valor total
            if valor_min_rango is not None and valor_max_rango is not None:
                if valor_total < valor_min_rango or valor_total > valor_max_rango:
                    continue  # Saltar este registro si no está en el rango
        
        nombre_propietario = f"{(avaluo.nombres or '').strip()} {(avaluo.apellidos or '').strip()}".strip()
        
        datos_informe.append({
            'clave_catastral': avaluo.cocata1 or '',
            'nombre_propietario': nombre_propietario,
            'ubicacion': avaluo.ubicacion or '',
            'uso': avaluo.uso or '',
            'desc_uso': desc_uso,
            'subuso': avaluo.subuso or '',
            'desc_subuso': desc_subuso,
            'st': avaluo.st or '',
            'bvl2tie': avaluo.bvl2tie or Decimal('0.00'),
            'mejoras': avaluo.mejoras or Decimal('0.00'),
            'detalle': avaluo.detalle or Decimal('0.00'),
            'cultivo': avaluo.cultivo or Decimal('0.00'),
            'exencion': avaluo.exencion or Decimal('0.00'),
            'grabable': avaluo.grabable or Decimal('0.00'),
            'impuesto': avaluo.impuesto or Decimal('0.00'),
            'valor_area': area_total,
        })
    
    logger.info(f"Total de registros procesados para el informe: {len(datos_informe)}")
    return datos_informe, filtros_aplicados, query

@catastro_require_auth
def exportar_informe_avaluos_excel(request):
    """
    Exportar informe de avalúos catastrales a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:listado_avaluos_catastrales')
    
    try:
        # Obtener filtros desde GET, POST o sesión
        tipo_informe = request.GET.get('tipo_informe') or request.POST.get('tipo_informe') or request.session.get('informe_avaluos_tipo')
        if not tipo_informe:
            messages.error(request, 'No se especificó el tipo de informe')
            return redirect('catastro:listado_avaluos_catastrales')
        
        # Obtener filtros
        filtros_dict = {}
        if request.method == 'GET' and request.GET:
            filtros_dict = dict(request.GET.items())
        elif request.method == 'POST' and request.POST:
            filtros_dict = dict(request.POST.items())
        else:
            # Usar filtros de la sesión
            filtros_dict = request.session.get('informe_avaluos_filtros', {})
        
        filtros_dict.pop('csrfmiddlewaretoken', None)
        
        # Obtener datos usando la función auxiliar
        datos_informe, filtros_aplicados, query = obtener_datos_informe_avaluos(
            request, tipo_informe, filtros_dict
        )
        
        empresa = request.session.get('catastro_empresa', '')
        municipio_desc = request.session.get('catastro_municipio_descripcion', '')
        
        # Crear libro de trabajo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado Avalúos"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.append(['LISTADO DE AVALÚOS CATASTRALES'])
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.append([])
        
        # Información
        ws.append(['Municipio:', municipio_desc])
        ws.merge_cells('B2:N2')
        if filtros_aplicados:
            ws.append(['Filtros aplicados:', ', '.join(filtros_aplicados)])
            ws.merge_cells('B3:N3')
            ws.append(['Total de registros:', len(datos_informe)])
            ws.merge_cells('B4:N4')
            ws.append([])
            start_row = 6
        else:
            ws.append(['Total de registros:', len(datos_informe)])
            ws.merge_cells('B3:N3')
            ws.append([])
            start_row = 5
        
        # Encabezados
        headers = [
            'Clave Catastral', 'Nombre del Propietario', 'Ubicación', 'Uso',
            'Sub Uso', 'St', 'Avaluo Terreno', 'Avaluo Edificacion',
            'Avaluo Detalles Adicionales', 'Avaluo Cultivos Permanente',
            'Exención', 'Valor Grabable', 'Impuesto', 'Valor Area'
        ]
        ws.append(headers)
        
        # Aplicar estilo a encabezados
        header_row = ws[start_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_style
        
        # Agregar datos
        row_num = start_row + 1
        for dato in datos_informe:
            ws.append([
                dato['clave_catastral'],
                dato['nombre_propietario'],
                dato['ubicacion'],
                dato['desc_uso'] if dato.get('desc_uso') else dato.get('uso', ''),
                dato['desc_subuso'] if dato.get('desc_subuso') else dato.get('subuso', ''),
                dato['st'],
                float(dato['bvl2tie']),
                float(dato['mejoras']),
                float(dato['detalle']),
                float(dato['cultivo']),
                float(dato['exencion']),
                float(dato['grabable']),
                float(dato['impuesto']),
                float(dato['valor_area']),
            ])
            
            # Aplicar bordes y formato
            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border_style
                if col >= 7 and col <= 13:  # Columnas numéricas (montos)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 14:  # Valor Area
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            row_num += 1
        
        # Ajustar ancho de columnas
        column_widths = [20, 30, 35, 15, 15, 8, 15, 15, 15, 15, 15, 15, 15, 12]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'listado_avaluos_catastrales_{empresa}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar informe a Excel: {str(e)}", exc_info=True)
        messages.error(request, f'Error al exportar a Excel: {str(e)}')
        return redirect('catastro:listado_avaluos_catastrales')

@catastro_require_auth
def exportar_informe_avaluos_pdf(request):
    """
    Exportar informe de avalúos catastrales a PDF con tamaño de hoja legal
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import legal, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:listado_avaluos_catastrales')
    
    from io import BytesIO
    
    try:
        # Obtener filtros desde GET, POST o sesión
        tipo_informe = request.GET.get('tipo_informe') or request.POST.get('tipo_informe') or request.session.get('informe_avaluos_tipo')
        if not tipo_informe:
            messages.error(request, 'No se especificó el tipo de informe')
            return redirect('catastro:listado_avaluos_catastrales')
        
        # Obtener filtros
        filtros_dict = {}
        if request.method == 'GET' and request.GET:
            filtros_dict = dict(request.GET.items())
        elif request.method == 'POST' and request.POST:
            filtros_dict = dict(request.POST.items())
        else:
            # Usar filtros de la sesión
            filtros_dict = request.session.get('informe_avaluos_filtros', {})
        
        filtros_dict.pop('csrfmiddlewaretoken', None)
        
        # Obtener datos usando la función auxiliar
        datos_informe, filtros_aplicados, query = obtener_datos_informe_avaluos(
            request, tipo_informe, filtros_dict
        )
        
        empresa = request.session.get('catastro_empresa', '')
        municipio_desc = request.session.get('catastro_municipio_descripcion', '')
        
        # Crear buffer para el PDF
        buffer = BytesIO()
        # Usar tamaño de hoja legal en modo landscape
        doc = SimpleDocTemplate(buffer, pagesize=landscape(legal))
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            alignment=1,  # Centrado
            spaceAfter=20
        )
        
        # Título
        title = Paragraph('LISTADO DE AVALÚOS CATASTRALES', title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Información
        info_data = [
            ['Municipio:', municipio_desc],
            ['Total de registros:', str(len(datos_informe))],
        ]
        if filtros_aplicados:
            info_data.insert(1, ['Filtros aplicados:', ', '.join(filtros_aplicados)])
        
        info_table = Table(info_data, colWidths=[2*inch, 6*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Preparar datos de la tabla
        data = [[
            'Clave Catastral', 'Nombre del Propietario', 'Ubicación', 'Uso', 'Sub Uso',
            'St', 'Avaluo Terreno', 'Avaluo Edificacion', 'Avaluo Detalles Adicionales',
            'Avaluo Cultivos Permanente', 'Exención', 'Valor Grabable', 'Impuesto', 'Valor Area'
        ]]
        
        for dato in datos_informe:
            data.append([
                dato['clave_catastral'],
                dato['nombre_propietario'],
                dato['ubicacion'],
                dato['desc_uso'] if dato.get('desc_uso') else dato.get('uso', ''),
                dato['desc_subuso'] if dato.get('desc_subuso') else dato.get('subuso', ''),
                dato['st'],
                f"L. {float(dato['bvl2tie']):,.2f}",
                f"L. {float(dato['mejoras']):,.2f}",
                f"L. {float(dato['detalle']):,.2f}",
                f"L. {float(dato['cultivo']):,.2f}",
                f"L. {float(dato['exencion']):,.2f}",
                f"L. {float(dato['grabable']):,.2f}",
                f"L. {float(dato['impuesto']):,.2f}",
                f"{float(dato['valor_area']):,.2f}",
            ])
        
        # Crear tabla con ancho de columnas ajustado para hoja legal landscape
        col_widths = [0.8*inch, 1.2*inch, 1.2*inch, 0.6*inch, 0.6*inch, 0.3*inch,
                      0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.6*inch, 0.6*inch,
                      0.6*inch, 0.6*inch]
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            # Datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (6, 1), (13, -1), 'RIGHT'),  # Alinear números a la derecha
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Filas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f'listado_avaluos_catastrales_{empresa}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar informe a PDF: {str(e)}", exc_info=True)
        messages.error(request, f'Error al exportar a PDF: {str(e)}')
        return redirect('catastro:listado_avaluos_catastrales')

@catastro_require_auth
def informes_dinamicos(request):
    """
    Vista para mostrar el formulario de informes dinámicos
    """
    from .models import BDCata1, BDTerreno, Edificacion, DetalleAdicionales, CultivoPermanente
    
    # Definir campos disponibles por tabla
    campos_disponibles = {
        'bdcata1': {
            'nombre': 'Datos Principales',
            'campos': [
                ('cocata1', 'Clave Catastral'),
                ('claveant', 'Clave Anterior'),
                ('ficha', 'Ficha (1=Urbana, 2=Rural)'),
                ('mapa', 'Mapa'),
                ('bloque', 'Bloque'),
                ('predio', 'Predio'),
                ('depto', 'Departamento'),
                ('municipio', 'Municipio'),
                ('barrio', 'Barrio'),
                ('caserio', 'Caserío'),
                ('sitio', 'Sitio'),
                ('lote', 'Lote'),
                ('bloquecol', 'Bloque-Col'),
                ('nombres', 'Nombres'),
                ('apellidos', 'Apellidos'),
                ('identidad', 'Número de Identidad'),
                ('rtn', 'RTN'),
                ('sexo', 'Género'),
                ('nacionalidad', 'Nacionalidad'),
                ('ubicacion', 'Ubicación'),
                ('telefono', 'Teléfono'),
                ('uso', 'Uso'),
                ('subuso', 'Sub Uso'),
                ('codhab', 'Código Habitacional'),
                ('codprop', 'Código de Propiedad'),
                ('st', 'Estatus Tributario'),
                ('estado', 'Estado del Registro'),
                ('zonificacion', 'Zonificación'),
                ('constru', 'Construcción'),
                ('nofichas', 'Número de Fichas'),
                ('vivienda', 'Número de Viviendas'),
                ('apartamentos', 'Número de Apartamentos'),
                ('cuartos', 'Número de Cuartos Adicionales'),
                ('tipopropiedad', 'Tipo de Propiedad'),
                ('bvl2tie', 'Avaluo Terreno'),
                ('mejoras', 'Avaluo Edificacion'),
                ('detalle', 'Avaluo Detalles Adicionales'),
                ('cultivo', 'Avaluo Cultivos Permanente'),
                ('declarado', 'Valor Declarado'),
                ('exencion', 'Exención'),
                ('bexenc', 'Porcentaje Exención'),
                ('grabable', 'Valor Grabable'),
                ('impuesto', 'Impuesto'),
                ('tasaimpositiva', 'Tasa Impositiva'),
                ('declaimpto', 'Declaración de Impuesto'),
                ('conedi', 'Conservación de Edificación'),
                ('cedif', 'Clase de Edificación'),
                ('condetalle', 'Condición de Detalle'),
                ('clavesure', 'Clave SURE'),
                ('cx', 'Coordenada X'),
                ('cy', 'Coordenada Y'),
                ('usuario', 'Usuario'),
                ('fechasys', 'Fecha de Registro'),
            ]
        },
        'bdterreno': {
            'nombre': 'Datos del Terreno',
            'campos': [
                ('baream21', 'Área 2.1'),
                ('baream22', 'Área 2.2'),
                ('bvlbas1', 'Valor Básico 1'),
                ('bvlbas2', 'Valor Básico 2'),
                ('bfrente', 'Frente'),
                ('bfrente2', 'Frente 2'),
                ('btopogra', 'Tipo de Topografía'),
                ('bfactopo', 'Factor de Topografía'),
            ]
        },
        'edificacion': {
            'nombre': 'Edificaciones',
            'campos': [
                ('edifino', 'No. Edificación'),
                ('piso', 'Piso'),
                ('area', 'Área'),
                ('uso', 'Uso'),
                ('clase', 'Clase'),
                ('calidad', 'Calidad'),
                ('costo', 'Costo'),
                ('totedi', 'Total Edificación'),
            ]
        },
        'detalleadicionales': {
            'nombre': 'Detalles Adicionales',
            'campos': [
                ('codigo', 'Código'),
                ('area', 'Área'),
                ('porce', 'Porcentaje'),
                ('unit', 'Valor Unitario'),
                ('total', 'Total'),
                ('descripcion', 'Descripción'),
            ]
        },
        'cultivopermanente': {
            'nombre': 'Cultivos Permanentes',
            'campos': [
                ('clase', 'Clase de Cultivo'),
                ('variedad', 'Variedad'),
                ('area', 'Área'),
                ('factor', 'Factor'),
                ('valor', 'Valor'),
            ]
        },
    }
    
    # Verificar que todos los campos estén correctamente definidos
    logger.info(f"Total de tablas en campos_disponibles: {len(campos_disponibles)}")
    for tabla_key, tabla_info in campos_disponibles.items():
        logger.info(f"Tabla {tabla_key}: {len(tabla_info.get('campos', []))} campos definidos")
        if tabla_key == 'bdcata1':
            logger.info(f"Campos de bdcata1: {[campo[0] for campo in tabla_info.get('campos', [])]}")
    
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Informes Dinámicos',
        'campos_disponibles': campos_disponibles,
    }
    
    return render(request, 'informes_dinamicos.html', context)

@catastro_require_auth
def generar_informe_dinamico(request):
    """
    Vista para generar el informe dinámico con los campos seleccionados
    """
    from .models import BDCata1, BDTerreno, Edificacion, DetalleAdicionales, CultivoPermanente, Usos, Subuso
    
    if request.method == 'POST':
        empresa = request.session.get('catastro_empresa')
        campos_seleccionados = request.POST.getlist('campos_seleccionados')
        
        if not campos_seleccionados:
            messages.error(request, 'Debe seleccionar al menos un campo para el informe')
            return redirect('catastro:informes_dinamicos')
        
        # Obtener filtros opcionales
        filtro_clave = request.POST.get('filtro_clave', '').strip()
        filtro_ficha = request.POST.get('filtro_ficha', '')
        filtro_uso = request.POST.get('filtro_uso', '').strip()
        
        try:
            # Crear mapeo de campos seleccionados a etiquetas (debe estar antes de procesar los datos)
            campos_disponibles = {
                'bdcata1': {
                    'cocata1': 'Clave Catastral',
                    'claveant': 'Clave Anterior',
                    'ficha': 'Ficha',
                    'mapa': 'Mapa',
                    'bloque': 'Bloque',
                    'predio': 'Predio',
                    'depto': 'Departamento',
                    'municipio': 'Municipio',
                    'barrio': 'Barrio',
                    'caserio': 'Caserío',
                    'sitio': 'Sitio',
                    'lote': 'Lote',
                    'bloquecol': 'Bloque-Col',
                    'nombres': 'Nombres',
                    'apellidos': 'Apellidos',
                    'identidad': 'Número de Identidad',
                    'rtn': 'RTN',
                    'sexo': 'Género',
                    'nacionalidad': 'Nacionalidad',
                    'ubicacion': 'Ubicación',
                    'telefono': 'Teléfono',
                    'uso': 'Uso',
                    'subuso': 'Sub Uso',
                    'codhab': 'Código Habitacional',
                    'codprop': 'Código de Propiedad',
                    'st': 'Estatus Tributario',
                    'estado': 'Estado del Registro',
                    'zonificacion': 'Zonificación',
                    'constru': 'Construcción',
                    'nofichas': 'Número de Fichas',
                    'vivienda': 'Número de Viviendas',
                    'apartamentos': 'Número de Apartamentos',
                    'cuartos': 'Número de Cuartos Adicionales',
                    'tipopropiedad': 'Tipo de Propiedad',
                    'bvl2tie': 'Avaluo Terreno',
                    'mejoras': 'Avaluo Edificacion',
                    'detalle': 'Avaluo Detalles Adicionales',
                    'cultivo': 'Avaluo Cultivos Permanente',
                    'declarado': 'Valor Declarado',
                    'exencion': 'Exención',
                    'bexenc': 'Porcentaje Exención',
                    'grabable': 'Valor Grabable',
                    'impuesto': 'Impuesto',
                    'tasaimpositiva': 'Tasa Impositiva',
                    'declaimpto': 'Declaración de Impuesto',
                    'conedi': 'Conservación de Edificación',
                    'cedif': 'Clase de Edificación',
                    'condetalle': 'Condición de Detalle',
                    'clavesure': 'Clave Segura',
                    'cx': 'Coordenada X',
                    'cy': 'Coordenada Y',
                    'usuario': 'Usuario',
                    'fechasys': 'Fecha de Registro',
                },
                'bdterreno': {
                    'baream21': 'Área 2.1',
                    'baream22': 'Área 2.2',
                    'bvlbas1': 'Valor Básico 1',
                    'bvlbas2': 'Valor Básico 2',
                    'bfrente': 'Frente',
                    'bfrente2': 'Frente 2',
                    'btopogra': 'Tipo de Topografía',
                    'bfactopo': 'Factor de Topografía',
                },
                'edificacion': {
                    'edifino': 'No. Edificación',
                    'piso': 'Piso',
                    'area': 'Área Edificación',
                    'uso': 'Uso Edificación',
                    'clase': 'Clase Edificación',
                    'calidad': 'Calidad',
                    'costo': 'Costo Edificación',
                    'totedi': 'Total Edificación',
                },
                'detalleadicionales': {
                    'codigo': 'Código Detalle',
                    'area': 'Área Detalle',
                    'porce': 'Porcentaje',
                    'unit': 'Valor Unitario',
                    'total': 'Total Detalles',
                    'descripcion': 'Descripción Detalle',
                },
                'cultivopermanente': {
                    'clase': 'Clase Cultivo',
                    'variedad': 'Variedad',
                    'area': 'Área Cultivo',
                    'factor': 'Factor Cultivo',
                    'valor': 'Valor Cultivo',
                },
            }
            
            # Construir query base
            query = Q(empresa=empresa) if empresa else Q()
            
            if filtro_clave:
                query &= Q(cocata1__icontains=filtro_clave)
            if filtro_ficha:
                query &= Q(ficha=filtro_ficha)
            if filtro_uso:
                query &= Q(uso=filtro_uso)
            
            # Obtener registros principales
            avaluos_qs = BDCata1.objects.filter(query).order_by('cocata1')
            
            # Preparar datos del informe
            datos_informe = []
            
            # Cache para descripciones
            usos_cache = {}
            subusos_cache = {}
            
            for avaluo in avaluos_qs:
                registro = {}
                
                # Procesar cada campo seleccionado
                for campo_completo in campos_seleccionados:
                    tabla, campo = campo_completo.split('__', 1) if '__' in campo_completo else ('bdcata1', campo_completo)
                    
                    # Obtener la etiqueta del campo
                    if tabla in campos_disponibles and campo in campos_disponibles[tabla]:
                        etiqueta = campos_disponibles[tabla][campo]
                    else:
                        etiqueta = campo
                    
                    # Log para debugging (solo para los primeros registros)
                    if len(datos_informe) < 2:
                        logger.info(f"Procesando campo: {campo_completo}, tabla: {tabla}, campo: {campo}, etiqueta: {etiqueta}")
                    
                    if tabla == 'bdcata1':
                        # Campos de BDCata1 - Usar la variable etiqueta en lugar de hardcodear
                        if campo == 'cocata1':
                            valor = avaluo.cocata1 or ''
                            registro[etiqueta] = valor
                            if len(datos_informe) < 2:
                                logger.info(f"  -> cocata1: valor='{valor}', etiqueta='{etiqueta}', registro[{etiqueta}]='{registro.get(etiqueta)}'")
                        elif campo == 'claveant':
                            registro[etiqueta] = avaluo.claveant or ''
                        elif campo == 'ficha':
                            registro[etiqueta] = 'Urbana' if avaluo.ficha == 1 else 'Rural' if avaluo.ficha == 2 else str(avaluo.ficha)
                        elif campo == 'mapa':
                            registro[etiqueta] = avaluo.mapa or ''
                        elif campo == 'bloque':
                            registro[etiqueta] = avaluo.bloque or ''
                        elif campo == 'predio':
                            registro[etiqueta] = avaluo.predio or ''
                        elif campo == 'depto':
                            registro[etiqueta] = avaluo.depto or ''
                        elif campo == 'municipio':
                            registro[etiqueta] = avaluo.municipio or ''
                        elif campo == 'barrio':
                            registro[etiqueta] = avaluo.barrio or ''
                        elif campo == 'caserio':
                            registro[etiqueta] = avaluo.caserio or ''
                        elif campo == 'sitio':
                            registro[etiqueta] = avaluo.sitio or ''
                        elif campo == 'lote':
                            registro[etiqueta] = avaluo.lote or ''
                        elif campo == 'bloquecol':
                            registro[etiqueta] = avaluo.bloquecol or ''
                        elif campo == 'nombres':
                            valor = avaluo.nombres or ''
                            registro[etiqueta] = valor
                            if len(datos_informe) < 2:
                                logger.info(f"  -> nombres: valor='{valor}', etiqueta='{etiqueta}', registro[{etiqueta}]='{registro.get(etiqueta)}'")
                        elif campo == 'apellidos':
                            valor = avaluo.apellidos or ''
                            registro[etiqueta] = valor
                            if len(datos_informe) < 2:
                                logger.info(f"  -> apellidos: valor='{valor}', etiqueta='{etiqueta}', registro[{etiqueta}]='{registro.get(etiqueta)}'")
                        elif campo == 'identidad':
                            valor = avaluo.identidad or ''
                            registro[etiqueta] = valor
                            if len(datos_informe) < 2:
                                logger.info(f"  -> identidad: valor='{valor}', etiqueta='{etiqueta}', registro[{etiqueta}]='{registro.get(etiqueta)}'")
                        elif campo == 'rtn':
                            registro[etiqueta] = avaluo.rtn or ''
                        elif campo == 'sexo':
                            registro[etiqueta] = dict(avaluo.SEXO_CHOICES).get(avaluo.sexo, avaluo.sexo or '')
                        elif campo == 'nacionalidad':
                            registro[etiqueta] = avaluo.nacionalidad or ''
                        elif campo == 'ubicacion':
                            registro[etiqueta] = avaluo.ubicacion or ''
                        elif campo == 'telefono':
                            registro[etiqueta] = avaluo.telefono or ''
                        elif campo == 'uso':
                            if avaluo.uso:
                                if avaluo.uso not in usos_cache:
                                    try:
                                        uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                                        usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                                    except:
                                        usos_cache[avaluo.uso] = avaluo.uso
                                registro[etiqueta] = usos_cache.get(avaluo.uso, avaluo.uso)
                            else:
                                registro[etiqueta] = ''
                        elif campo == 'subuso':
                            if avaluo.subuso and avaluo.uso:
                                cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                                if cache_key not in subusos_cache:
                                    try:
                                        subuso_obj = Subuso.objects.filter(
                                            uso=avaluo.uso,
                                            codsubuso=avaluo.subuso
                                        ).first()
                                        subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                                    except:
                                        subusos_cache[cache_key] = avaluo.subuso
                                registro[etiqueta] = subusos_cache.get(cache_key, avaluo.subuso)
                            else:
                                registro[etiqueta] = ''
                        elif campo == 'codhab':
                            registro[etiqueta] = avaluo.codhab or ''
                        elif campo == 'codprop':
                            registro[etiqueta] = avaluo.codprop or ''
                        elif campo == 'st':
                            registro[etiqueta] = avaluo.st or ''
                        elif campo == 'estado':
                            registro[etiqueta] = dict(avaluo.ESTADO_CHOICES).get(avaluo.estado, avaluo.estado or '')
                        elif campo == 'zonificacion':
                            registro[etiqueta] = avaluo.zonificacion or ''
                        elif campo == 'constru':
                            registro[etiqueta] = int(avaluo.constru or 0)
                        elif campo == 'nofichas':
                            registro[etiqueta] = int(avaluo.nofichas or 0)
                        elif campo == 'vivienda':
                            registro[etiqueta] = int(avaluo.vivienda or 0)
                        elif campo == 'apartamentos':
                            registro[etiqueta] = int(avaluo.apartamentos or 0)
                        elif campo == 'cuartos':
                            registro[etiqueta] = int(avaluo.cuartos or 0)
                        elif campo == 'tipopropiedad':
                            registro[etiqueta] = float(avaluo.tipopropiedad or 0)
                        elif campo == 'bvl2tie':
                            registro[etiqueta] = float(avaluo.bvl2tie or 0)
                        elif campo == 'mejoras':
                            registro[etiqueta] = float(avaluo.mejoras or 0)
                        elif campo == 'detalle':
                            registro[etiqueta] = float(avaluo.detalle or 0)
                        elif campo == 'cultivo':
                            registro[etiqueta] = float(avaluo.cultivo or 0)
                        elif campo == 'declarado':
                            registro[etiqueta] = float(avaluo.declarado or 0)
                        elif campo == 'exencion':
                            registro[etiqueta] = float(avaluo.exencion or 0)
                        elif campo == 'bexenc':
                            registro[etiqueta] = float(avaluo.bexenc or 0)
                        elif campo == 'grabable':
                            registro[etiqueta] = float(avaluo.grabable or 0)
                        elif campo == 'impuesto':
                            registro[etiqueta] = float(avaluo.impuesto or 0)
                        elif campo == 'tasaimpositiva':
                            registro[etiqueta] = float(avaluo.tasaimpositiva or 0)
                        elif campo == 'declaimpto':
                            registro[etiqueta] = int(avaluo.declaimpto or 0)
                        elif campo == 'conedi':
                            registro[etiqueta] = int(avaluo.conedi or 0)
                        elif campo == 'cedif':
                            registro[etiqueta] = int(avaluo.cedif or 0)
                        elif campo == 'condetalle':
                            registro[etiqueta] = int(avaluo.condetalle or 0)
                        elif campo == 'clavesure':
                            registro[etiqueta] = avaluo.clavesure or ''
                        elif campo == 'cx':
                            registro[etiqueta] = float(avaluo.cx or 0)
                        elif campo == 'cy':
                            registro[etiqueta] = float(avaluo.cy or 0)
                        elif campo == 'usuario':
                            registro[etiqueta] = avaluo.usuario or ''
                        elif campo == 'fechasys':
                            registro[etiqueta] = avaluo.fechasys.strftime('%Y-%m-%d %H:%M:%S') if avaluo.fechasys else ''
                    
                    elif tabla == 'bdterreno':
                        # Campos de BDTerreno
                        try:
                            terreno = BDTerreno.objects.filter(
                                empresa=avaluo.empresa,
                                cocata1=avaluo.cocata1
                            ).first()
                            
                            if terreno:
                                if campo == 'baream21':
                                    registro['Área 2.1'] = float(terreno.baream21 or 0)
                                elif campo == 'baream22':
                                    registro['Área 2.2'] = float(terreno.baream22 or 0)
                                elif campo == 'bvlbas1':
                                    registro['Valor Básico 1'] = float(terreno.bvlbas1 or 0)
                                elif campo == 'bvlbas2':
                                    registro['Valor Básico 2'] = float(terreno.bvlbas2 or 0)
                                elif campo == 'bfrente':
                                    registro['Frente'] = float(terreno.bfrente or 0)
                                elif campo == 'bfrente2':
                                    registro['Frente 2'] = float(terreno.bfrente2 or 0)
                                elif campo == 'btopogra':
                                    registro['Tipo de Topografía'] = terreno.btopogra or ''
                                elif campo == 'bfactopo':
                                    registro['Factor de Topografía'] = float(terreno.bfactopo or 0)
                            else:
                                # Si no hay terreno, poner valores vacíos
                                if campo in ['baream21', 'baream22', 'bvlbas1', 'bvlbas2', 'bfrente', 'bfrente2', 'bfactopo']:
                                    registro[campo] = 0
                                else:
                                    registro[campo] = ''
                        except Exception as e:
                            logger.error(f"Error al obtener terreno para {avaluo.cocata1}: {str(e)}")
                    
                    elif tabla == 'edificacion':
                        # Campos de Edificacion (puede haber múltiples por clave)
                        try:
                            edificaciones = Edificacion.objects.filter(
                                empresa=avaluo.empresa,
                                clave=avaluo.cocata1
                            )
                            
                            if edificaciones.exists():
                                # Si hay múltiples edificaciones, concatenar o tomar la primera
                                edif = edificaciones.first()
                                if campo == 'edifino':
                                    registro['No. Edificación'] = int(edif.edifino or 0)
                                elif campo == 'piso':
                                    registro['Piso'] = int(edif.piso or 0) if edif.piso else ''
                                elif campo == 'area':
                                    registro['Área Edificación'] = float(edif.area or 0)
                                elif campo == 'uso':
                                    registro['Uso Edificación'] = edif.uso or ''
                                elif campo == 'clase':
                                    registro['Clase Edificación'] = edif.clase or ''
                                elif campo == 'calidad':
                                    registro['Calidad'] = edif.calidad or ''
                                elif campo == 'costo':
                                    registro['Costo Edificación'] = float(edif.costo or 0)
                                elif campo == 'totedi':
                                    registro['Total Edificación'] = float(edif.totedi or 0)
                            else:
                                # Si no hay edificaciones, poner valores vacíos
                                if campo in ['edifino', 'piso', 'area', 'costo', 'totedi']:
                                    registro[campo] = 0
                                else:
                                    registro[campo] = ''
                        except Exception as e:
                            logger.error(f"Error al obtener edificaciones para {avaluo.cocata1}: {str(e)}")
                    
                    elif tabla == 'detalleadicionales':
                        # Campos de DetalleAdicionales (puede haber múltiples por clave)
                        try:
                            detalles = DetalleAdicionales.objects.filter(
                                empresa=avaluo.empresa,
                                clave=avaluo.cocata1
                            )
                            
                            if detalles.exists():
                                # Sumar totales o tomar el primero
                                det = detalles.first()
                                if campo == 'codigo':
                                    registro['Código Detalle'] = det.codigo or ''
                                elif campo == 'area':
                                    registro['Área Detalle'] = float(det.area or 0)
                                elif campo == 'porce':
                                    registro['Porcentaje'] = float(det.porce or 0)
                                elif campo == 'unit':
                                    registro['Valor Unitario'] = float(det.unit or 0)
                                elif campo == 'total':
                                    # Sumar todos los totales de detalles adicionales
                                    total_detalles = sum(float(d.total or 0) for d in detalles)
                                    registro['Total Detalles'] = total_detalles
                                elif campo == 'descripcion':
                                    registro['Descripción Detalle'] = det.descripcion or ''
                            else:
                                if campo in ['area', 'porce', 'unit', 'total']:
                                    registro[campo] = 0
                                else:
                                    registro[campo] = ''
                        except Exception as e:
                            logger.error(f"Error al obtener detalles adicionales para {avaluo.cocata1}: {str(e)}")
                    
                    elif tabla == 'cultivopermanente':
                        # Campos de CultivoPermanente (puede haber múltiples por clave)
                        try:
                            cultivos = CultivoPermanente.objects.filter(
                                empresa=avaluo.empresa,
                                cocata1=avaluo.cocata1
                            )
                            
                            if cultivos.exists():
                                cult = cultivos.first()
                                if campo == 'clase':
                                    registro['Clase Cultivo'] = cult.clase or ''
                                elif campo == 'variedad':
                                    registro['Variedad'] = cult.variedad or ''
                                elif campo == 'area':
                                    registro['Área Cultivo'] = float(cult.area or 0)
                                elif campo == 'factor':
                                    registro['Factor Cultivo'] = float(cult.factor or 0)
                                elif campo == 'valor':
                                    registro['Valor Cultivo'] = float(cult.valor or 0)
                            else:
                                if campo in ['area', 'factor', 'valor']:
                                    registro[campo] = 0
                                else:
                                    registro[campo] = ''
                        except Exception as e:
                            logger.error(f"Error al obtener cultivos para {avaluo.cocata1}: {str(e)}")
                
                datos_informe.append(registro)
            
            # Crear lista de etiquetas en el orden de campos_seleccionados
            etiquetas_columnas = []
            for campo_completo in campos_seleccionados:
                if '__' in campo_completo:
                    tabla, campo = campo_completo.split('__', 1)
                    if tabla in campos_disponibles and campo in campos_disponibles[tabla]:
                        etiquetas_columnas.append(campos_disponibles[tabla][campo])
                    else:
                        etiquetas_columnas.append(campo)
                else:
                    # Si no tiene __, es un campo de bdcata1 por defecto
                    campo = campo_completo
                    if 'bdcata1' in campos_disponibles and campo in campos_disponibles['bdcata1']:
                        etiquetas_columnas.append(campos_disponibles['bdcata1'][campo])
                    else:
                        etiquetas_columnas.append(campo)
            
            # Reorganizar datos_informe para que sea una lista de listas en el mismo orden que etiquetas_columnas
            datos_informe_ordenados = []
            logger.info(f"Total de registros a procesar: {len(datos_informe)}")
            logger.info(f"Etiquetas de columnas ({len(etiquetas_columnas)}): {etiquetas_columnas}")
            for idx, registro in enumerate(datos_informe):
                fila = []
                if idx == 0:  # Log solo para el primer registro
                    logger.info(f"Primer registro - Claves disponibles ({len(registro.keys())}): {list(registro.keys())}")
                    logger.info(f"Primer registro - Valores: {list(registro.values())}")
                
                for etiqueta_idx, etiqueta in enumerate(etiquetas_columnas):
                    # Buscar el valor en el registro usando la etiqueta exacta
                    valor = registro.get(etiqueta)
                    
                    # Si no se encuentra, intentar diferentes variaciones
                    if valor is None:
                        # Intentar sin espacios al inicio/final
                        valor = registro.get(etiqueta.strip())
                        if valor is None:
                            # Intentar buscar por coincidencia parcial (último recurso)
                            for key in registro.keys():
                                if key.strip() == etiqueta.strip():
                                    valor = registro[key]
                                    logger.warning(f"Valor encontrado usando coincidencia parcial: '{key}' -> '{etiqueta}'")
                                    break
                    
                    # Si aún no se encuentra, intentar buscar el campo original
                    if valor is None:
                        # Obtener el campo original de campos_seleccionados
                        campo_completo = campos_seleccionados[etiqueta_idx] if etiqueta_idx < len(campos_seleccionados) else None
                        if campo_completo:
                            tabla_orig, campo_orig = campo_completo.split('__', 1) if '__' in campo_completo else ('bdcata1', campo_completo)
                            # Intentar buscar usando el nombre del campo directamente
                            if campo_orig in registro:
                                valor = registro[campo_orig]
                                logger.warning(f"Valor encontrado usando nombre de campo original: '{campo_orig}' para etiqueta '{etiqueta}'")
                    
                    # Si aún no se encuentra, registrar warning
                    if valor is None and idx == 0:
                        logger.warning(f"[Columna {etiqueta_idx}] Etiqueta '{etiqueta}' no encontrada en registro. Claves disponibles: {list(registro.keys())}")
                        logger.warning(f"  Campo original: {campos_seleccionados[etiqueta_idx] if etiqueta_idx < len(campos_seleccionados) else 'N/A'}")
                    
                    # Normalizar valores None y vacíos
                    # IMPORTANTE: Preservar valores de texto tal como están
                    if valor is None:
                        valor = ''
                    elif isinstance(valor, (int, float)):
                        # Mantener números como están (incluyendo 0)
                        # No convertir a string para que el template pueda formatearlos
                        pass
                    elif isinstance(valor, str):
                        # Para cadenas, mantener el valor original
                        # Si está vacío o solo tiene espacios, mantenerlo como cadena vacía
                        if valor.strip() == '':
                            valor = ''  # Normalizar cadenas vacías o solo espacios
                        # Si tiene contenido, mantenerlo tal cual (no hacer strip)
                    elif isinstance(valor, bool):
                        # Convertir booleanos a string
                        valor = 'Sí' if valor else 'No'
                    else:
                        # Para otros tipos (date, etc.), convertir a cadena
                        valor = str(valor) if valor is not None else ''
                    
                    fila.append(valor)
                    if idx == 0:  # Log solo para el primer registro
                        logger.info(f"[Columna {etiqueta_idx}] Etiqueta: '{etiqueta}' -> Valor: '{valor}' (Tipo: {type(valor).__name__}, Repr: {repr(valor)})")
                
                datos_informe_ordenados.append(fila)
                if idx == 0:  # Log solo para el primer registro
                    logger.info(f"Primera fila ordenada ({len(fila)} valores): {fila}")
            
            # Guardar configuración en sesión para exportaciones
            request.session['informe_dinamico_campos'] = campos_seleccionados
            request.session['informe_dinamico_filtros'] = {
                'filtro_clave': filtro_clave,
                'filtro_ficha': filtro_ficha,
                'filtro_uso': filtro_uso,
            }
            
            # Log final para verificar estructura de datos
            logger.info(f"=== RESUMEN FINAL ===")
            logger.info(f"Total registros procesados: {len(datos_informe)}")
            logger.info(f"Total filas ordenadas: {len(datos_informe_ordenados)}")
            logger.info(f"Total etiquetas: {len(etiquetas_columnas)}")
            if datos_informe_ordenados:
                logger.info(f"Primera fila ordenada tiene {len(datos_informe_ordenados[0])} valores")
                logger.info(f"Primera fila: {datos_informe_ordenados[0]}")
            if datos_informe:
                logger.info(f"Primer registro original tiene {len(datos_informe[0])} claves")
                logger.info(f"Claves del primer registro: {list(datos_informe[0].keys())}")
            
            context = {
                'empresa': empresa,
                'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
                'usuario_nombre': request.session.get('catastro_usuario_nombre'),
                'modulo': 'Catastro - Informes Dinámicos',
                'datos_informe': datos_informe_ordenados,
                'datos_informe_original': datos_informe,  # Mantener original para debug
                'campos_seleccionados': campos_seleccionados,
                'etiquetas_columnas': etiquetas_columnas,
                'total_registros': len(datos_informe),
                'filtros_aplicados': {
                    'clave': filtro_clave,
                    'ficha': filtro_ficha,
                    'uso': filtro_uso,
                },
            }
            
            return render(request, 'informe_dinamico_resultado.html', context)
            
        except Exception as e:
            logger.error(f"Error al generar informe dinámico: {str(e)}", exc_info=True)
            messages.error(request, f'Error al generar el informe: {str(e)}')
            return redirect('catastro:informes_dinamicos')
    
    return redirect('catastro:informes_dinamicos')

@catastro_require_auth
def generar_informe_dinamico_excel(request):
    """
    Exportar informe dinámico a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:informes_dinamicos')
    
    try:
        from .models import BDCata1, BDTerreno, Edificacion, DetalleAdicionales, CultivoPermanente, Usos, Subuso
        from django.db.models import Q
        from datetime import datetime
        from django.http import HttpResponse
        
        empresa = request.session.get('catastro_empresa')
        campos_seleccionados = request.session.get('informe_dinamico_campos', [])
        filtros = request.session.get('informe_dinamico_filtros', {})
        
        if not campos_seleccionados:
            messages.error(request, 'No hay configuración de informe dinámico guardada')
            return redirect('catastro:informes_dinamicos')
        
        # Reconstruir query
        query = Q(empresa=empresa) if empresa else Q()
        
        if filtros.get('filtro_clave'):
            query &= Q(cocata1__icontains=filtros['filtro_clave'])
        if filtros.get('filtro_ficha'):
            query &= Q(ficha=filtros['filtro_ficha'])
        if filtros.get('filtro_uso'):
            query &= Q(uso=filtros['filtro_uso'])
        
        # Obtener registros
        avaluos_qs = BDCata1.objects.filter(query).order_by('cocata1')
        
        # Crear mapeo de campos a etiquetas
        campos_disponibles = {
            'bdcata1': {
                'cocata1': 'Clave Catastral',
                'claveant': 'Clave Anterior',
                'ficha': 'Ficha',
                'mapa': 'Mapa',
                'bloque': 'Bloque',
                'predio': 'Predio',
                'depto': 'Departamento',
                'municipio': 'Municipio',
                'barrio': 'Barrio',
                'caserio': 'Caserío',
                'sitio': 'Sitio',
                'lote': 'Lote',
                'bloquecol': 'Bloque-Col',
                'nombres': 'Nombres',
                'apellidos': 'Apellidos',
                'identidad': 'Número de Identidad',
                'rtn': 'RTN',
                'sexo': 'Género',
                'nacionalidad': 'Nacionalidad',
                'ubicacion': 'Ubicación',
                'telefono': 'Teléfono',
                'uso': 'Uso',
                'subuso': 'Sub Uso',
                'codhab': 'Código Habitacional',
                'codprop': 'Código de Propiedad',
                'st': 'Estatus Tributario',
                'estado': 'Estado del Registro',
                'zonificacion': 'Zonificación',
                'constru': 'Construcción',
                'nofichas': 'Número de Fichas',
                'vivienda': 'Número de Viviendas',
                'apartamentos': 'Número de Apartamentos',
                'cuartos': 'Número de Cuartos Adicionales',
                'tipopropiedad': 'Tipo de Propiedad',
                'bvl2tie': 'Avaluo Terreno',
                'mejoras': 'Avaluo Edificacion',
                'detalle': 'Avaluo Detalles Adicionales',
                'cultivo': 'Avaluo Cultivos Permanente',
                'declarado': 'Valor Declarado',
                'exencion': 'Exención',
                'bexenc': 'Porcentaje Exención',
                'grabable': 'Valor Grabable',
                'impuesto': 'Impuesto',
                'tasaimpositiva': 'Tasa Impositiva',
                'declaimpto': 'Declaración de Impuesto',
                'conedi': 'Conservación de Edificación',
                'cedif': 'Clase de Edificación',
                'condetalle': 'Condición de Detalle',
                'clavesure': 'Clave Segura',
                'cx': 'Coordenada X',
                'cy': 'Coordenada Y',
                'usuario': 'Usuario',
                'fechasys': 'Fecha de Registro',
            },
            'bdterreno': {
                'baream21': 'Área 2.1',
                'baream22': 'Área 2.2',
                'bvlbas1': 'Valor Básico 1',
                'bvlbas2': 'Valor Básico 2',
                'bfrente': 'Frente',
                'bfrente2': 'Frente 2',
                'btopogra': 'Tipo de Topografía',
                'bfactopo': 'Factor de Topografía',
            },
            'edificacion': {
                'edifino': 'No. Edificación',
                'piso': 'Piso',
                'area': 'Área Edificación',
                'uso': 'Uso Edificación',
                'clase': 'Clase Edificación',
                'calidad': 'Calidad',
                'costo': 'Costo Edificación',
                'totedi': 'Total Edificación',
            },
            'detalleadicionales': {
                'codigo': 'Código Detalle',
                'area': 'Área Detalle',
                'porce': 'Porcentaje',
                'unit': 'Valor Unitario',
                'total': 'Total Detalles',
                'descripcion': 'Descripción Detalle',
            },
            'cultivopermanente': {
                'clase': 'Clase Cultivo',
                'variedad': 'Variedad',
                'area': 'Área Cultivo',
                'factor': 'Factor Cultivo',
                'valor': 'Valor Cultivo',
            },
        }
        
        # Crear lista de etiquetas (debe coincidir con generar_informe_dinamico)
        etiquetas_columnas = []
        for campo_completo in campos_seleccionados:
            if '__' in campo_completo:
                tabla, campo = campo_completo.split('__', 1)
                if tabla in campos_disponibles and campo in campos_disponibles[tabla]:
                    etiquetas_columnas.append(campos_disponibles[tabla][campo])
                else:
                    etiquetas_columnas.append(campo)
            else:
                # Si no tiene __, es un campo de bdcata1 por defecto
                campo = campo_completo
                if 'bdcata1' in campos_disponibles and campo in campos_disponibles['bdcata1']:
                    etiquetas_columnas.append(campos_disponibles['bdcata1'][campo])
                else:
                    etiquetas_columnas.append(campo)
        
        # Generar datos (reutilizar lógica de generar_informe_dinamico)
        datos_informe = []
        usos_cache = {}
        subusos_cache = {}
        
        for avaluo in avaluos_qs:
            registro = {}
            for campo_completo in campos_seleccionados:
                tabla, campo = campo_completo.split('__', 1) if '__' in campo_completo else ('bdcata1', campo_completo)
                etiqueta = campos_disponibles.get(tabla, {}).get(campo, campo)
                
                # Procesar campo (simplificado, reutilizar lógica de generar_informe_dinamico)
                if tabla == 'bdcata1':
                    if campo == 'cocata1':
                        registro[etiqueta] = avaluo.cocata1 or ''
                    elif campo == 'ficha':
                        registro[etiqueta] = 'Urbana' if avaluo.ficha == 1 else 'Rural' if avaluo.ficha == 2 else str(avaluo.ficha)
                    elif campo == 'nombres':
                        registro[etiqueta] = avaluo.nombres or ''
                    elif campo == 'apellidos':
                        registro[etiqueta] = avaluo.apellidos or ''
                    elif campo == 'identidad':
                        registro[etiqueta] = avaluo.identidad or ''
                    elif campo == 'ubicacion':
                        registro[etiqueta] = avaluo.ubicacion or ''
                    elif campo == 'uso':
                        if avaluo.uso:
                            if avaluo.uso not in usos_cache:
                                try:
                                    uso_obj = Usos.objects.filter(uso=avaluo.uso).first()
                                    usos_cache[avaluo.uso] = uso_obj.desuso if uso_obj else avaluo.uso
                                except:
                                    usos_cache[avaluo.uso] = avaluo.uso
                            registro[etiqueta] = usos_cache.get(avaluo.uso, avaluo.uso)
                        else:
                            registro[etiqueta] = ''
                    elif campo == 'subuso':
                        if avaluo.subuso and avaluo.uso:
                            cache_key = f"{avaluo.uso}_{avaluo.subuso}"
                            if cache_key not in subusos_cache:
                                try:
                                    subuso_obj = Subuso.objects.filter(
                                        uso=avaluo.uso,
                                        codsubuso=avaluo.subuso
                                    ).first()
                                    subusos_cache[cache_key] = subuso_obj.des_subuso if subuso_obj else avaluo.subuso
                                except:
                                    subusos_cache[cache_key] = avaluo.subuso
                            registro[etiqueta] = subusos_cache.get(cache_key, avaluo.subuso)
                        else:
                            registro[etiqueta] = ''
                    elif campo in ['st', 'telefono', 'barrio', 'caserio']:
                        registro[etiqueta] = getattr(avaluo, campo, '') or ''
                    elif campo in ['bvl2tie', 'mejoras', 'detalle', 'cultivo', 'exencion', 'grabable', 'impuesto']:
                        registro[etiqueta] = float(getattr(avaluo, campo, 0) or 0)
                    elif campo in ['cx', 'cy']:
                        registro[etiqueta] = float(getattr(avaluo, campo, 0) or 0)
                # ... (simplificado para otros tipos de tablas)
            
            datos_informe.append(registro)
        
        # Crear libro Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe Dinámico"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.append(['INFORME DINÁMICO - CATASTRO'])
        ws.merge_cells(f'A1:{get_column_letter(len(etiquetas_columnas))}1')
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.append([])
        
        # Información
        municipio_desc = request.session.get('catastro_municipio_descripcion', '')
        ws.append(['Municipio:', municipio_desc])
        ws.merge_cells(f'B2:{get_column_letter(len(etiquetas_columnas))}2')
        ws.append(['Total de registros:', len(datos_informe)])
        ws.merge_cells(f'B3:{get_column_letter(len(etiquetas_columnas))}3')
        ws.append([])
        
        # Encabezados
        ws.append(etiquetas_columnas)
        header_row = ws[5]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_style
        
        # Datos
        row_num = 6
        for dato in datos_informe:
            row_data = []
            for etiqueta in etiquetas_columnas:
                valor = dato.get(etiqueta, '')
                row_data.append(valor)
            ws.append(row_data)
            
            # Aplicar formato
            for col_idx, etiqueta in enumerate(etiquetas_columnas, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border = border_style
                valor = dato.get(etiqueta, '')
                if isinstance(valor, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            row_num += 1
        
        # Ajustar ancho de columnas
        for idx in range(1, len(etiquetas_columnas) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'informe_dinamico_{empresa}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar informe dinámico a Excel: {str(e)}", exc_info=True)
        messages.error(request, f'Error al exportar a Excel: {str(e)}')
        return redirect('catastro:informes_dinamicos')

@catastro_require_auth
def generar_informe_dinamico_pdf(request):
    """
    Exportar informe dinámico a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import legal, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:informes_dinamicos')
    
    try:
        from .models import BDCata1, BDTerreno, Edificacion, DetalleAdicionales, CultivoPermanente, Usos, Subuso
        from django.db.models import Q
        from datetime import datetime
        from django.http import HttpResponse
        from io import BytesIO
        
        empresa = request.session.get('catastro_empresa')
        campos_seleccionados = request.session.get('informe_dinamico_campos', [])
        filtros = request.session.get('informe_dinamico_filtros', {})
        
        if not campos_seleccionados:
            messages.error(request, 'No hay configuración de informe dinámico guardada')
            return redirect('catastro:informes_dinamicos')
        
        # Similar a Excel, generar datos y crear PDF
        # (Implementación simplificada, similar a exportar_informe_avaluos_pdf)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=legal, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            alignment=1,  # Center
            spaceAfter=30,
        )
        
        elements = []
        
        # Título
        elements.append(Paragraph('INFORME DINÁMICO - CATASTRO', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Información
        municipio_desc = request.session.get('catastro_municipio_descripcion', '')
        info_text = f"<b>Municipio:</b> {municipio_desc}<br/><b>Total de registros:</b> {len(campos_seleccionados)}"
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabla (simplificada)
        # ... (implementar tabla similar a exportar_informe_avaluos_pdf)
        
        doc.build(elements)
        
        response = HttpResponse(content_type='application/pdf')
        filename = f'informe_dinamico_{empresa}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        buffer.close()
        
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar informe dinámico a PDF: {str(e)}", exc_info=True)
        messages.error(request, f'Error al exportar a PDF: {str(e)}')
        return redirect('catastro:informes_dinamicos')

@catastro_require_auth
@catastro_require_permiso(CATASTRO_PERM_CONFIG_VER)
def catastro_configuracion(request):
    """
    Menú de configuración del módulo de catastro
    """
    context = {
        'empresa': request.session.get('catastro_empresa'),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion'),
        'usuario_nombre': request.session.get('catastro_usuario_nombre'),
        'modulo': 'Catastro - Configuración'
    }
    return render(request, 'configuracion_menu.html', context)

@catastro_require_auth
@catastro_require_permiso(CATASTRO_PERM_BIENES_EDITAR)
def registrar_bien_inmueble(request):
    """
    Vista para registrar un nuevo bien inmueble
    Ruta del template: C:\\simafiweb\\venv\\Scripts\\catastro\\templates\\bienes_inmuebles_form.html
    """
    logger.info("=" * 100)
    logger.info("=" * 100)
    logger.info("=== INICIO registrar_bien_inmueble ===")
    logger.info("=" * 100)
    logger.info(f"Método: {request.method}")
    logger.info(f"Usuario: {request.session.get('catastro_usuario_nombre')}")
    logger.info(f"Empresa: {request.session.get('catastro_empresa')}")
    logger.info(f"Path: {request.path}")
    logger.info(f"GET params: {dict(request.GET)}")
    logger.info(f"POST params count: {len(request.POST)}")
    
    # Importar el formulario y modelo desde el módulo local
    try:
        from .forms import BDCata1Form
        from .models import BDCata1
        logger.info("✓ Formulario y modelo importados correctamente desde módulo local")
    except ImportError as e:
        logger.error(f"Error al importar: {str(e)}", exc_info=True)
        messages.error(request, f'Error: No se pudo cargar el formulario. {str(e)}')
        return redirect('catastro:catastro_menu_principal')
    
    current_user = request.session.get('catastro_usuario_nombre', '')
    municipio_codigo_sesion = request.session.get('catastro_empresa', '')
    existing_instance = None
    
    # Inicializar el diccionario initial para usar en caso de errores
    initial = {}
    if municipio_codigo_sesion:
        # Ejemplo: si empresa = "0301", entonces:
        # - Depto = "03" (primeros dos dígitos)
        # - Municipio = "01" (últimos dos dígitos)
        if len(municipio_codigo_sesion) >= 4:
            initial['depto'] = municipio_codigo_sesion[:2]  # Primeros dos dígitos
            initial['municipio'] = municipio_codigo_sesion[-2:]  # Últimos dos dígitos
        elif len(municipio_codigo_sesion) >= 2:
            # Si tiene menos de 4 dígitos, solo asignar municipio
            initial['municipio'] = municipio_codigo_sesion[-2:]  # Últimos dos dígitos
        else:
            initial['municipio'] = municipio_codigo_sesion
        initial['empresa'] = municipio_codigo_sesion
    
    initial['fechasys'] = timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')
    
    # Obtener el porcentaje de concertación desde la tabla municipio (para ambos GET y POST)
    municipio_por_concer = None
    codigo_municipio = None
    if municipio_codigo_sesion:
        if len(municipio_codigo_sesion) >= 2:
            codigo_municipio = municipio_codigo_sesion[-2:]  # Últimos dos dígitos
        else:
            codigo_municipio = municipio_codigo_sesion
        
        # Obtener el porcentaje de concertación desde la tabla municipio para mostrar (NO se guarda en bdcata1)
        try:
            # Intentar importar el modelo Municipio desde core
            try:
                from core.models import Municipio
            except ImportError:
                # Si no está disponible, intentar desde otro lugar
                import sys
                import os
                core_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'core', 'models.py')
                if os.path.exists(core_path):
                    spec = importlib.util.spec_from_file_location("core.models", core_path)
                    core_models = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(core_models)
                    Municipio = core_models.Municipio
                else:
                    raise ImportError("No se pudo importar el modelo Municipio")
            
            # Buscar el municipio por código para obtener por_concer (solo para mostrar)
            if codigo_municipio:
                municipio_obj = Municipio.objects.filter(codigo=codigo_municipio).first()
                if municipio_obj and municipio_obj.por_concer is not None:
                    municipio_por_concer = municipio_obj.por_concer
                    logger.info(f"Porcentaje de concertación obtenido desde municipio: {municipio_por_concer}")
                else:
                    logger.warning(f"No se encontró municipio o por_concer es None para código: {codigo_municipio}")
        except Exception as e:
            logger.error(f"Error al obtener porcentaje de concertación desde municipio: {str(e)}")

    if request.method == 'POST':
        logger.info("=" * 80)
        logger.info("=== PROCESANDO POST ===")
        logger.info("=" * 80)
        logger.info(f"Total de campos POST: {len(request.POST)}")
        logger.info(f"Content-Type: {request.META.get('CONTENT_TYPE', 'N/A')}")
        logger.info(f"Request path: {request.path}")
        logger.info(f"Request method: {request.method}")
        logger.info(f"User-Agent: {request.META.get('HTTP_USER_AGENT', 'N/A')[:100]}")
        logger.info(f"Referer: {request.META.get('HTTP_REFERER', 'N/A')}")
        
        # IMPORTANTE: Verificar si hay datos POST
        if len(request.POST) == 0:
            logger.error("=" * 100)
            logger.error("⚠️⚠️⚠️ ADVERTENCIA CRÍTICA: El POST llegó pero NO tiene campos ⚠️⚠️⚠️")
            logger.error("=" * 100)
            logger.error("Esto puede indicar que:")
            logger.error("1. El formulario no se está enviando correctamente")
            logger.error("2. Hay un problema con el CSRF token")
            logger.error("3. El formulario está siendo bloqueado por JavaScript")
            logger.error(f"Headers recibidos: {dict(request.META)}")
            messages.error(request, 'Error: El formulario se envió vacío. Por favor intente nuevamente.')
            # Continuar para renderizar el formulario de nuevo
        else:
            logger.info("=" * 100)
            logger.info("✓✓✓ POST TIENE DATOS - CONTINUANDO CON EL PROCESAMIENTO ✓✓✓")
            logger.info("=" * 100)
            logger.info(f"Total de campos POST recibidos: {len(request.POST)}")
            
            # Mostrar TODOS los campos que se están enviando
            campos_enviados = {}
            for key, value in request.POST.items():
                # Limitar el valor a 100 caracteres para no saturar los logs
                valor_display = str(value)[:100] + ('...' if len(str(value)) > 100 else '')
                campos_enviados[key] = valor_display
                logger.info(f"  Campo '{key}': '{valor_display}'")
            
            logger.info(f"Total de campos procesados: {len(campos_enviados)}")
            
            # Verificar específicamente los campos críticos: bexenc, zonificacion, uso, subuso, barrio, caserio
            logger.info("=" * 100)
            logger.info("=== VERIFICACIÓN DE CAMPOS CRÍTICOS ===")
            logger.info("=" * 100)
            campos_criticos = ['bexenc', 'zonificacion', 'uso', 'subuso', 'barrio', 'caserio']
            for campo in campos_criticos:
                valor = request.POST.get(campo, '(NO ENVIADO)')
                logger.info(f"  Campo '{campo}': '{valor}'")
            logger.info("=" * 100)
            
            cocata_codigo = request.POST.get('cocata1', '').strip()
            logger.info("=" * 100)
            logger.info(f"Código catastral recibido: '{cocata_codigo}'")
            logger.info(f"Empresa de sesión: '{municipio_codigo_sesion}'")
            logger.info("=" * 100)
            
            # Validar que cocata1 no esté vacío o sea solo un guion
            if not cocata_codigo or cocata_codigo == '-' or cocata_codigo == '':
                logger.error(f"❌ ERROR: cocata1 está vacío o inválido: '{cocata_codigo}'")
                messages.error(request, 'Error: La clave catastral (cocata1) es obligatoria y no puede estar vacía.')
                # Continuar para renderizar el formulario con el error
                form = BDCata1Form(initial=initial)
                context = {
                    'titulo': 'Registrar Bien Inmueble',
                    'form': form,
                    'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
                    'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
                    'empresa': municipio_codigo_sesion or '',
                    'empresa_codigo': municipio_codigo_sesion or '',
                    'municipio_por_concer': municipio_por_concer,
                    'municipio_por_concer_str': str(municipio_por_concer) if municipio_por_concer is not None else '0.00',
                }
                return render(request, 'bienes_inmuebles_form.html', context)
            
            if cocata_codigo and municipio_codigo_sesion:
                try:
                    # Buscar por empresa y cocata1 (clave única compuesta)
                    existing_instance = BDCata1.objects.filter(
                        empresa=municipio_codigo_sesion,
                        cocata1=cocata_codigo
                    ).first()
                    if existing_instance:
                        logger.info(f"✓ Registro existente encontrado: ID={existing_instance.id}, cocata1={existing_instance.cocata1}")
                    else:
                        logger.info("✓ Nuevo registro (no existe en BD)")
                except Exception as e:
                    logger.error(f"Error al buscar bien inmueble: {str(e)}", exc_info=True)
                    existing_instance = None
            else:
                logger.warning(f"No hay código catastral o empresa. cocata1='{cocata_codigo}', empresa='{municipio_codigo_sesion}'")
                existing_instance = None
            
            logger.info("=" * 100)
            logger.info("Creando formulario BDCata1Form...")
            
            # Crear una copia de POST para modificar si es necesario
            post_data = request.POST.copy()
            
            # Asegurar que campos personalizados se incluyan en el formulario
            # Los campos barrio, caserio, subuso vienen de comboboxes personalizados
            # Verificar que estén presentes en POST
            campos_personalizados = ['barrio', 'caserio', 'subuso']
            for campo in campos_personalizados:
                valor = post_data.get(campo, '')
                logger.info(f"  Campo personalizado '{campo}' en POST: '{valor}'")
                if not valor and campo in request.POST:
                    logger.warning(f"    ⚠️ Campo '{campo}' está vacío en POST")
            
            form = BDCata1Form(post_data, instance=existing_instance)
            logger.info(f"Formulario creado. Instance: {existing_instance}")
            
            # Validar el formulario
            es_valido = form.is_valid()
            logger.info("=" * 100)
            logger.info(f"Formulario validado. Válido: {es_valido}")
            logger.info("=" * 100)
            
            # Validación del campo terceraedad
            terceraedad = request.POST.get('terceraedad', '').strip()
            if terceraedad == 'S':
                st = request.POST.get('st', '').strip()
                codhab = request.POST.get('codhab', '').strip()
                codprop = request.POST.get('codprop', '').strip()
                
                requisitos_no_cumplidos = []
                
                # Validar que st = '1' (Exento)
                if st != '1':
                    requisitos_no_cumplidos.append('El predio debe estar exento (Estatus Tributario = 1)')
                
                # Validar que codhab = '0' o vacío
                if codhab and codhab != '0':
                    requisitos_no_cumplidos.append('El propietario debe habitar la propiedad (Código Habitacional = 0)')
                
                # Validar que codprop = '01'
                if codprop != '01':
                    requisitos_no_cumplidos.append('Debe ser uso particular (Código Propietario = 01)')
                
                if requisitos_no_cumplidos:
                    # No bloquear el guardado aquí, la validación se hace en el frontend con el modal
                    # Solo registrar en el log para auditoría
                    logger.info(f"Validación tercera edad - Requisitos no cumplidos: {', '.join(requisitos_no_cumplidos)}")
                    # Permitir que el formulario se valide normalmente, el frontend ya mostró el modal
            
            # Verificar campos personalizados antes de validar
            logger.info("=== VERIFICACIÓN DE CAMPOS PERSONALIZADOS EN CLEANED_DATA ===")
            campos_personalizados_check = ['barrio', 'caserio', 'subuso']
            if es_valido:
                for campo in campos_personalizados_check:
                    valor_post = request.POST.get(campo, '')
                    valor_cleaned = form.cleaned_data.get(campo, 'NO EN CLEANED_DATA')
                    logger.info(f"  {campo}: POST='{valor_post}' -> cleaned_data='{valor_cleaned}'")
            else:
                logger.warning("Formulario inválido, no se puede verificar cleaned_data")
            logger.info("=" * 100)
            
            # Mostrar errores de validación si los hay
            if not es_valido:
                logger.error("=" * 80)
                logger.error("=== ERRORES DE VALIDACIÓN ===")
                logger.error("=" * 80)
                logger.error(f"Total de campos con errores: {len(form.errors)}")
                logger.error(f"Campos con errores: {list(form.errors.keys())}")
                
                for field, errors in form.errors.items():
                    for error in errors:
                        logger.error(f"  Campo '{field}': {error}")
                        # Mostrar también el valor que causó el error
                        valor_campo = request.POST.get(field, '(vacío)')
                        logger.error(f"    Valor enviado en POST: '{valor_campo[:100]}'")
                
                # Mostrar mensaje de error al usuario
                error_messages = []
                for field, errors in form.errors.items():
                    for error in errors:
                        error_messages.append(f"{field}: {error}")
                
                if error_messages:
                    mensaje_error = f'Errores de validación ({len(error_messages)}): ' + '; '.join(error_messages[:5])
                    if len(error_messages) > 5:
                        mensaje_error += f' ... y {len(error_messages) - 5} más'
                    messages.error(request, mensaje_error)
                    logger.error(f"Mensaje de error mostrado al usuario: {mensaje_error}")
                else:
                    messages.error(request, 'Por favor corrija los errores en el formulario.')
                    logger.error("Formulario inválido pero no se encontraron errores específicos")
                
                # IMPORTANTE: Renderizar el formulario con errores para que el usuario los vea
                logger.info("Renderizando formulario con errores de validación")
                # Continuar al final para renderizar el formulario con errores
            else:
                logger.info("=" * 100)
                logger.info("✓✓✓ FORMULARIO VÁLIDO - PROCEDIENDO A GUARDAR ✓✓✓")
                logger.info("=" * 100)
                
                # Lista de TODOS los campos del modelo BDCata1
                todos_los_campos = [
                        'cocata1', 'empresa', 'ficha', 'claveant', 'mapa', 'bloque', 'predio',
                        'depto', 'municipio', 'barrio', 'caserio', 'sitio', 'nombres', 'apellidos',
                        'identidad', 'rtn', 'telefono', 'ubicacion', 'nacionalidad', 'sexo', 'uso',
                        'subuso', 'constru', 'nofichas', 'bvl2tie', 'conedi', 'mejoras',
                        'cedif', 'detalle', 'impuesto', 'grabable', 'cultivo', 'declarado', 'exencion',
                        'condetalle', 'st', 'codhab', 'codprop', 'tasaimpositiva', 'declaimpto', 'bexenc',
                        'tipopropiedad', 'estado', 'clavesure', 'cx', 'cy', 'zonificacion', 'vivienda',
                        'apartamentos', 'cuartos', 'lote', 'bloquecol', 'usuario', 'fechasys'
                    ]
                
                # Verificar que todos los campos estén en cleaned_data
                logger.info("=== VERIFICACIÓN DE CLEANED_DATA COMPLETO ===")
                campos_faltantes = []
                for campo_nombre in todos_los_campos:
                    if campo_nombre not in ['usuario', 'fechasys']:  # Estos se asignan manualmente
                        if campo_nombre not in form.cleaned_data:
                            campos_faltantes.append(campo_nombre)
                            logger.warning(f"  ⚠️ Campo '{campo_nombre}' NO está en cleaned_data")
                        else:
                            valor_cleaned = form.cleaned_data[campo_nombre]
                            valor_post = request.POST.get(campo_nombre, '')
                            if valor_post != str(valor_cleaned):
                                logger.info(f"  {campo_nombre}: POST='{valor_post}' -> cleaned='{valor_cleaned}' (diferente)")
                if campos_faltantes:
                    logger.error(f"⚠️⚠️⚠️ CAMPOS FALTANTES EN CLEANED_DATA: {campos_faltantes} ⚠️⚠️⚠️")
                logger.info("=" * 100)
                
                try:
                    bien = form.save(commit=False)
                    
                    # Validar nuevamente que cocata1 no esté vacío antes de guardar
                    if not bien.cocata1 or bien.cocata1.strip() == '' or bien.cocata1.strip() == '-':
                        logger.error(f"❌ ERROR: cocata1 está vacío o inválido en el objeto: '{bien.cocata1}'")
                        messages.error(request, 'Error: La clave catastral (cocata1) es obligatoria y no puede estar vacía.')
                        # Continuar para renderizar el formulario con el error
                        context = {
                            'titulo': 'Registrar Bien Inmueble',
                            'form': form,
                            'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
                            'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
                            'empresa': municipio_codigo_sesion or '',
                            'empresa_codigo': municipio_codigo_sesion or '',
                            'municipio_por_concer': municipio_por_concer,
                            'municipio_por_concer_str': str(municipio_por_concer) if municipio_por_concer is not None else '0.00',
                        }
                        return render(request, 'bienes_inmuebles_form.html', context)
                    
                    logger.info("=== TODOS LOS CAMPOS ANTES DE ASIGNAR VALORES ADICIONALES ===")
                    campos_con_valor_antes = 0
                    for campo_nombre in todos_los_campos:
                        try:
                            valor = getattr(bien, campo_nombre, None)
                            if valor is not None and str(valor).strip() != '':
                                campos_con_valor_antes += 1
                                valor_str = str(valor)[:80] + ('...' if len(str(valor)) > 80 else '')
                                logger.info(f"  {campo_nombre}: {valor_str}")
                        except Exception:
                            pass
                    logger.info(f"Total de campos con valor (antes): {campos_con_valor_antes}")
                    
                    # Asignar usuario y fecha si están disponibles
                    if current_user:
                        # Usuario es CharField, guardar el nombre del usuario directamente
                        bien.usuario = current_user[:50]  # Limitar a 50 caracteres
                        logger.info(f"  usuario asignado: {bien.usuario}")
                    
                    bien.fechasys = timezone.now()
                    logger.info(f"  fechasys asignado: {bien.fechasys}")
                    
                    # Asegurar que cocata1 tenga un valor válido
                    bien.cocata1 = cocata_codigo.strip()
                    logger.info(f"  cocata1 asignado: '{bien.cocata1}'")
                    
                    if municipio_codigo_sesion:
                        bien.empresa = municipio_codigo_sesion
                        # Ejemplo: si empresa = "0301", entonces:
                        # - Depto = "03" (primeros dos dígitos)
                        # - Municipio = "01" (últimos dos dígitos)
                        if len(municipio_codigo_sesion) >= 4:
                            bien.depto = municipio_codigo_sesion[:2]  # Primeros dos dígitos
                            bien.municipio = municipio_codigo_sesion[-2:]  # Últimos dos dígitos
                        elif len(municipio_codigo_sesion) >= 2:
                            # Si tiene menos de 4 dígitos, solo asignar municipio
                            bien.municipio = municipio_codigo_sesion[-2:]  # Últimos dos dígitos
                        else:
                            bien.municipio = municipio_codigo_sesion
                        logger.info(f"  empresa asignada: {bien.empresa}")
                        logger.info(f"  depto asignado: {bien.depto}")
                        logger.info(f"  municipio asignado: {bien.municipio}")
                        
                        # Nota: bexenc ahora es un campo independiente en el formulario, no se asigna desde municipio.por_concer
                    
                    logger.info("=== TODOS LOS CAMPOS DESPUÉS DE ASIGNAR VALORES ADICIONALES ===")
                    campos_con_valor_despues = 0
                    for campo_nombre in todos_los_campos:
                        try:
                            valor = getattr(bien, campo_nombre, None)
                            if valor is not None and str(valor).strip() != '':
                                campos_con_valor_despues += 1
                                valor_str = str(valor)[:80] + ('...' if len(str(valor)) > 80 else '')
                                logger.info(f"  {campo_nombre}: {valor_str}")
                        except Exception:
                            pass
                    logger.info(f"Total de campos con valor (después): {campos_con_valor_despues}")
                    
                    # Verificar específicamente los campos críticos antes de guardar
                    logger.info("=" * 100)
                    logger.info("=== VERIFICACIÓN DE CAMPOS CRÍTICOS ANTES DE GUARDAR ===")
                    logger.info("=" * 100)
                    campos_criticos = ['bexenc', 'zonificacion', 'uso', 'subuso', 'barrio', 'caserio']
                    for campo in campos_criticos:
                        try:
                            valor = getattr(bien, campo, None)
                            valor_str = str(valor) if valor is not None else '(None)'
                            logger.info(f"  bien.{campo}: {valor_str}")
                        except Exception as e:
                            logger.error(f"  Error al obtener bien.{campo}: {str(e)}")
                    logger.info("=" * 100)
                    
                    logger.info("=== INTENTANDO GUARDAR EN BD ===")
                    logger.info(f"Objeto a guardar: cocata1={bien.cocata1}, empresa={bien.empresa}")
                    
                    # Validar que cocata1 y empresa no estén vacíos antes de guardar
                    if not bien.cocata1 or not bien.cocata1.strip() or bien.cocata1.strip() == '-' or not bien.empresa:
                        logger.error(f"❌ ERROR: cocata1 o empresa están vacíos o inválidos. cocata1='{bien.cocata1}', empresa='{bien.empresa}'")
                        messages.error(request, f'Error: La clave catastral (cocata1) y la empresa son obligatorias. cocata1="{bien.cocata1}", empresa="{bien.empresa}"')
                        context = {
                            'titulo': 'Registrar Bien Inmueble',
                            'form': form,
                            'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
                            'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
                            'empresa': municipio_codigo_sesion or '',
                            'empresa_codigo': municipio_codigo_sesion or '',
                            'municipio_por_concer': municipio_por_concer,
                            'municipio_por_concer_str': str(municipio_por_concer) if municipio_por_concer is not None else '0.00',
                        }
                        return render(request, 'bienes_inmuebles_form.html', context)
                    
                    # Verificar TODOS los campos antes de guardar
                    logger.info("=" * 100)
                    logger.info("=== VERIFICACIÓN COMPLETA DE TODOS LOS CAMPOS ANTES DE GUARDAR ===")
                    logger.info("=" * 100)
                    campos_no_guardados = []
                    for campo_nombre in todos_los_campos:
                        try:
                            valor_en_post = request.POST.get(campo_nombre, '(NO EN POST)')
                            valor_en_bien = getattr(bien, campo_nombre, None)
                            valor_en_bien_str = str(valor_en_bien) if valor_en_bien is not None else '(None)'
                            
                            # Verificar si el campo tiene un valor significativo
                            tiene_valor = valor_en_bien is not None and str(valor_en_bien).strip() != '' and str(valor_en_bien).strip() != '0' and str(valor_en_bien).strip() != '0.00' and str(valor_en_bien).strip() != '0.0'
                            
                            logger.info(f"  {campo_nombre}:")
                            logger.info(f"    POST: '{valor_en_post}'")
                            logger.info(f"    Bien: '{valor_en_bien_str}'")
                            logger.info(f"    Tiene valor: {tiene_valor}")
                            
                            # Campos que deberían tener valor si están en POST
                            if valor_en_post != '(NO EN POST)' and valor_en_post != '' and not tiene_valor:
                                campos_no_guardados.append(campo_nombre)
                                logger.warning(f"    ⚠️ ADVERTENCIA: Campo '{campo_nombre}' está en POST pero no tiene valor en bien")
                        except Exception as e:
                            logger.error(f"  Error al verificar campo '{campo_nombre}': {str(e)}")
                    
                    if campos_no_guardados:
                        logger.warning(f"Campos que pueden no haberse guardado correctamente: {campos_no_guardados}")
                    logger.info("=" * 100)
                    
                    # Variable para determinar si el registro fue creado o actualizado
                    registro_creado = False
                    
                    # Usar update_or_create para evitar errores de duplicado
                    try:
                        # Asegurar que cocata1 tenga un valor válido (no vacío ni solo guion)
                        bien.cocata1 = cocata_codigo.strip()
                        bien.empresa = municipio_codigo_sesion
                        
                        # Crear un diccionario con todos los campos excepto id, empresa y cocata1
                        defaults_dict = {}
                        for field in bien._meta.fields:
                            if field.name not in ['id', 'empresa', 'cocata1']:
                                try:
                                    defaults_dict[field.name] = getattr(bien, field.name)
                                except Exception:
                                    pass
                        
                        # Usar update_or_create para manejar correctamente la creación/actualización
                        bien_guardado, creado = BDCata1.objects.update_or_create(
                            empresa=bien.empresa,
                            cocata1=bien.cocata1,
                            defaults=defaults_dict
                        )
                        bien = bien_guardado
                        logger.info("=" * 100)
                        logger.info(f"✓✓✓ BIEN INMUEBLE {'CREADO' if creado else 'ACTUALIZADO'} EXITOSAMENTE - ID: {bien.id} ✓✓✓")
                        logger.info(f"  cocata1: {bien.cocata1}")
                        logger.info(f"  empresa: {bien.empresa}")
                        logger.info("=" * 100)
                        
                        # Guardar el estado de creación para usarlo después
                        registro_creado = creado
                    except Exception as e:
                        logger.error(f"❌ Error al guardar con update_or_create: {str(e)}", exc_info=True)
                        # Si update_or_create falla, intentar guardar normalmente
                        try:
                            # Verificar si es nuevo o existente antes de guardar
                            es_nuevo = bien.pk is None
                            bien.save()
                            registro_creado = es_nuevo
                            logger.info("=" * 100)
                            logger.info(f"✓✓✓ GUARDADO EXITOSO (método alternativo) - ID: {bien.id} ✓✓✓")
                            logger.info("=" * 100)
                        except Exception as e2:
                            logger.error(f"❌ Error crítico al guardar bien inmueble: {str(e2)}", exc_info=True)
                            messages.error(request, f'Error al guardar el bien inmueble: {str(e2)}')
                            context = {
                                'titulo': 'Registrar Bien Inmueble',
                                'form': form,
                                'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
                                'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
                                'empresa': municipio_codigo_sesion or '',
                                'empresa_codigo': municipio_codigo_sesion or '',
                                'municipio_por_concer': municipio_por_concer,
                                'municipio_por_concer_str': str(municipio_por_concer) if municipio_por_concer is not None else '0.00',
                            }
                            return render(request, 'bienes_inmuebles_form.html', context)
                    
                    # Verificar que se guardó correctamente leyendo de la BD
                    try:
                        bien_verificado = BDCata1.objects.get(id=bien.id)
                        logger.info("=== VERIFICACIÓN POST-GUARDADO COMPLETA ===")
                        logger.info(f"  ID: {bien_verificado.id}")
                        logger.info(f"  cocata1: {bien_verificado.cocata1}")
                        logger.info(f"  empresa: {bien_verificado.empresa}")
                        
                        # Verificar TODOS los campos después de guardar
                        logger.info("=" * 100)
                        logger.info("=== VERIFICACIÓN COMPLETA DE TODOS LOS CAMPOS DESPUÉS DE GUARDAR ===")
                        logger.info("=" * 100)
                        campos_no_guardados_final = []
                        for campo_nombre in todos_los_campos:
                            try:
                                valor_en_post = request.POST.get(campo_nombre, '(NO EN POST)')
                                valor_en_bd = getattr(bien_verificado, campo_nombre, None)
                                valor_en_bd_str = str(valor_en_bd) if valor_en_bd is not None else '(None)'
                                
                                tiene_valor = valor_en_bd is not None and str(valor_en_bd).strip() != '' and str(valor_en_bd).strip() != '0' and str(valor_en_bd).strip() != '0.00' and str(valor_en_bd).strip() != '0.0'
                                
                                logger.info(f"  {campo_nombre}:")
                                logger.info(f"    POST: '{valor_en_post}'")
                                logger.info(f"    BD: '{valor_en_bd_str}'")
                                logger.info(f"    Guardado: {tiene_valor}")
                                
                                # Verificar si el valor en POST coincide con el valor en BD
                                if valor_en_post != '(NO EN POST)' and valor_en_post != '':
                                    valor_post_limpio = str(valor_en_post).strip()
                                    valor_bd_limpio = str(valor_en_bd).strip() if valor_en_bd is not None else ''
                                    
                                    # Comparar valores (considerando conversiones de tipo)
                                    if valor_post_limpio != valor_bd_limpio:
                                        # Intentar comparación numérica
                                        try:
                                            if float(valor_post_limpio) != float(valor_bd_limpio):
                                                campos_no_guardados_final.append(campo_nombre)
                                                logger.warning(f"    ⚠️ VALOR NO COINCIDE: POST='{valor_post_limpio}' vs BD='{valor_bd_limpio}'")
                                        except (ValueError, TypeError):
                                            campos_no_guardados_final.append(campo_nombre)
                                            logger.warning(f"    ⚠️ VALOR NO COINCIDE: POST='{valor_post_limpio}' vs BD='{valor_bd_limpio}'")
                            except Exception as e:
                                logger.error(f"  Error al verificar campo '{campo_nombre}': {str(e)}")
                        
                        if campos_no_guardados_final:
                            logger.error(f"⚠️⚠️⚠️ CAMPOS QUE NO SE GUARDARON CORRECTAMENTE: {campos_no_guardados_final} ⚠️⚠️⚠️")
                        else:
                            logger.info("✓✓✓ TODOS LOS CAMPOS SE GUARDARON CORRECTAMENTE ✓✓✓")
                        logger.info("=" * 100)
                        
                        logger.info(f"  usuario: {bien_verificado.usuario}")
                        logger.info(f"  fechasys: {bien_verificado.fechasys}")
                        logger.info(f"  nombres: {bien_verificado.nombres}")
                        logger.info(f"  apellidos: {bien_verificado.apellidos}")
                        logger.info(f"  barrio: {bien_verificado.barrio}")
                        logger.info(f"  caserio: {bien_verificado.caserio}")
                        logger.info(f"  subuso: {bien_verificado.subuso}")
                        logger.info(f"  zonificacion: {bien_verificado.zonificacion}")
                        logger.info(f"  bexenc: {bien_verificado.bexenc}")
                        logger.info(f"  uso: {bien_verificado.uso}")
                        logger.info(f"  vivienda: {bien_verificado.vivienda}")
                        logger.info(f"  apartamentos: {bien_verificado.apartamentos}")
                        logger.info(f"  cuartos: {bien_verificado.cuartos}")
                    except Exception as e:
                        logger.error(f"Error al verificar el guardado: {str(e)}", exc_info=True)
                    
                    # Determinar si fue creado o actualizado usando la variable registro_creado
                    accion = "registrado" if registro_creado else "actualizado"
                    mensaje = f'¡Bien inmueble {accion} correctamente! Código catastral: {bien.cocata1}'
                    messages.success(request, mensaje)
                    logger.info(f"✓✓✓ BIEN INMUEBLE {accion.upper()}: {bien.cocata1} (ID: {bien.id}) ✓✓✓")
                    logger.info(f"Mensaje de éxito: {mensaje}")

                    # Recalcular impuesto siempre en backend antes de actualizar rubros/tasas.
                    # Esto evita que el rubro de impuesto quede en cero cuando el frontend
                    # envía un valor desactualizado o vacío.
                    try:
                        impuesto_calculado = calcular_impuesto_bdcata1(bien, municipio_codigo_sesion)
                        bien.impuesto = impuesto_calculado
                        bien.save(update_fields=['impuesto'])
                        logger.info(
                            f"Impuesto recalculado al guardar bien inmueble: "
                            f"empresa={bien.empresa}, clave={bien.cocata1}, impuesto={impuesto_calculado}"
                        )
                    except Exception as e:
                        logger.warning(
                            f"No se pudo recalcular/guardar impuesto al registrar bien inmueble: {str(e)}"
                        )
                    
                    # ========================================================================
                    # CREAR RUBRO EN TASASMUNICIPALES SEGÚN PERÍMETRO (URBANO/RURAL)
                    # ========================================================================
                    try:
                        # Importar modelos necesarios
                        # TasasMunicipales ya está importado con "from .models import *" al inicio
                        # pero lo importamos explícitamente por si acaso
                        try:
                            from .models import TasasMunicipales
                        except ImportError:
                            try:
                                from modules.catastro.models import TasasMunicipales
                            except ImportError:
                                # Si no está disponible, intentar desde el import global
                                TasasMunicipales = globals().get('TasasMunicipales')
                                if not TasasMunicipales:
                                    raise ImportError("No se pudo importar TasasMunicipales")
                        
                        # Rubro está en tributario.models
                        try:
                            from tributario.models import Rubro
                        except ImportError:
                            # Intentar alternativa
                            try:
                                from tributario.models import Rubro
                            except ImportError:
                                raise ImportError("No se pudo importar Rubro")
                        
                        # Obtener ficha para determinar si es urbano (1) o rural (2)
                        ficha_valor = int(bien.ficha) if bien.ficha else 1
                        
                        # Determinar el rubro según el perímetro
                        if ficha_valor == 1:  # Urbano
                            codigo_rubro = 'B0001'
                            codigo_rubro_eliminar = 'B0002'  # Rubro a eliminar si existe
                        else:  # Rural
                            codigo_rubro = 'B0002'
                            codigo_rubro_eliminar = 'B0001'  # Rubro a eliminar si existe
                        
                        logger.info(f"Procesando rubro para perímetro {'URBANO' if ficha_valor == 1 else 'RURAL'}: {codigo_rubro}")
                        
                        # Eliminar el rubro opuesto si existe para evitar duplicados
                        try:
                            rubros_opuestos = TasasMunicipales.objects.filter(
                                empresa=bien.empresa,
                                clave=bien.cocata1,
                                rubro=codigo_rubro_eliminar
                            )
                            cantidad_eliminados = rubros_opuestos.count()
                            if cantidad_eliminados > 0:
                                rubros_opuestos.delete()
                                logger.info(f"✓ Eliminados {cantidad_eliminados} registro(s) de rubro {codigo_rubro_eliminar} para evitar duplicados")
                        except Exception as e:
                            logger.warning(f"Error al eliminar rubro opuesto {codigo_rubro_eliminar}: {str(e)}")
                        
                        # Obtener datos del rubro desde la tabla Rubros
                        try:
                            rubro_obj = Rubro.objects.get(
                                empresa=bien.empresa,
                                codigo=codigo_rubro
                            )
                            cuenta_rubro = rubro_obj.cuenta or ''
                            cuentarez_rubro = rubro_obj.cuentarez or ''
                            descripcion_rubro = rubro_obj.descripcion or ''
                            logger.info(f"✓ Rubro encontrado en tabla Rubros: {codigo_rubro} - {descripcion_rubro}")
                        except Rubro.DoesNotExist:
                            logger.warning(f"⚠ Rubro {codigo_rubro} no existe en tabla Rubros para empresa {bien.empresa}")
                            cuenta_rubro = ''
                            cuentarez_rubro = ''
                        except Exception as e:
                            logger.error(f"Error al obtener rubro desde tabla Rubros: {str(e)}")
                            cuenta_rubro = ''
                            cuentarez_rubro = ''
                        
                        # Obtener valor del impuesto desde bdcata1
                        valor_impuesto = bien.impuesto or Decimal('0.00')
                        
                        # Verificar si ya existe un registro en tasasmunicipales con la misma empresa, clave y rubro
                        try:
                            tasa_existente = TasasMunicipales.objects.get(
                                empresa=bien.empresa,
                                clave=bien.cocata1,
                                rubro=codigo_rubro
                            )
                            # Actualizar el registro existente
                            tasa_existente.cod_tarifa = '01'
                            tasa_existente.valor = valor_impuesto
                            tasa_existente.cuenta = cuenta_rubro
                            tasa_existente.cuentarez = cuentarez_rubro
                            tasa_existente.save()
                            logger.info(f"✓ Registro en TasasMunicipales actualizado: empresa={bien.empresa}, clave={bien.cocata1}, rubro={codigo_rubro}")
                            
                            # Calcular tasas municipales (rubros que empiezan con T) después de actualizar el rubro
                            calcular_tasas_municipales_automatico(bien, municipio_codigo_sesion)
                        except TasasMunicipales.DoesNotExist:
                            # Crear nuevo registro en tasasmunicipales
                            nueva_tasa = TasasMunicipales.objects.create(
                                empresa=bien.empresa,
                                clave=bien.cocata1,
                                rubro=codigo_rubro,
                                cod_tarifa='01',
                                valor=valor_impuesto,
                                cuenta=cuenta_rubro,
                                cuentarez=cuentarez_rubro
                            )
                            logger.info(f"✓ Nuevo registro creado en TasasMunicipales: ID={nueva_tasa.id}, empresa={bien.empresa}, clave={bien.cocata1}, rubro={codigo_rubro}")
                        except Exception as e:
                            logger.error(f"Error al crear/actualizar registro en TasasMunicipales: {str(e)}", exc_info=True)
                        
                        # Calcular tasas municipales (rubros que empiezan con T) después de crear/actualizar el rubro
                        try:
                            calcular_tasas_municipales_automatico(bien, municipio_codigo_sesion)
                        except Exception as e:
                            logger.warning(f"Error al calcular tasas municipales automáticamente: {str(e)}")
                            
                    except Exception as e:
                        logger.error(f"Error al procesar rubro en TasasMunicipales: {str(e)}", exc_info=True)
                        # No interrumpir el flujo si hay error, solo loguear
                    
                    # Redirigir al mismo formulario manteniendo el código catastral en la URL
                    return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={bien.cocata1}")
                except Exception as e:
                    logger.error("=" * 100)
                    logger.error(f"ERROR AL GUARDAR BIEN INMUEBLE: {str(e)}")
                    logger.error("=" * 100)
                    logger.error(f"Traceback completo:", exc_info=True)
                    messages.error(request, f'Error al guardar el bien inmueble: {str(e)}')
    else:
        # En GET, actualizar initial si hay cocata1 en la URL
        # (initial ya fue inicializado arriba)
        
        # TESTEO: Verificar si hay un código catastral en GET para cargar datos existentes
        cocata_test = request.GET.get('cocata1', '').strip()
        bien_test = None
        if cocata_test and municipio_codigo_sesion:
            try:
                bien_test = BDCata1.objects.filter(
                    empresa=municipio_codigo_sesion,
                    cocata1=cocata_test
                ).first()
                if bien_test:
                    logger.info("=" * 100)
                    logger.info("=== TESTEO: REGISTRO ENCONTRADO PARA CARGAR EN FORMULARIO ===")
                    logger.info("=" * 100)
                    logger.info(f"ID: {bien_test.id}")
                    logger.info(f"cocata1: {bien_test.cocata1}")
                    logger.info(f"empresa: {bien_test.empresa}")
                    
                    # Listar TODOS los campos del registro
                    todos_los_campos_test = [
                        'cocata1', 'empresa', 'ficha', 'claveant', 'mapa', 'bloque', 'predio',
                        'depto', 'municipio', 'barrio', 'caserio', 'sitio', 'nombres', 'apellidos',
                        'identidad', 'rtn', 'telefono', 'ubicacion', 'nacionalidad', 'sexo', 'uso',
                        'subuso', 'constru', 'nofichas', 'bvl2tie', 'conedi', 'mejoras',
                        'cedif', 'detalle', 'impuesto', 'grabable', 'cultivo', 'declarado', 'exencion',
                        'condetalle', 'st', 'codhab', 'codprop', 'tasaimpositiva', 'declaimpto', 'bexenc',
                        'tipopropiedad', 'estado', 'clavesure', 'cx', 'cy', 'zonificacion', 'vivienda',
                        'apartamentos', 'cuartos', 'lote', 'bloquecol', 'usuario', 'fechasys'
                    ]
                    
                    logger.info("=== TODOS LOS CAMPOS DEL REGISTRO EN BD ===")
                    for campo_nombre in todos_los_campos_test:
                        try:
                            valor = getattr(bien_test, campo_nombre, None)
                            valor_str = str(valor) if valor is not None else '(None)'
                            tiene_valor = valor is not None and str(valor).strip() != '' and str(valor).strip() != '0' and str(valor).strip() != '0.00' and str(valor).strip() != '0.0'
                            logger.info(f"  {campo_nombre}: '{valor_str}' (tiene valor: {tiene_valor})")
                        except Exception as e:
                            logger.error(f"  Error al obtener {campo_nombre}: {str(e)}")
                    logger.info("=" * 100)
                    
                    # Crear formulario con la instancia para que se carguen los valores
                    form = BDCata1Form(instance=bien_test)
                    logger.info("✓ Formulario creado con instancia existente")
                else:
                    # Si no existe el registro pero hay cocata1 en GET, establecerlo como valor inicial
                    initial['cocata1'] = cocata_test
                    form = BDCata1Form(initial=initial)
                    logger.info(f"✓ Formulario creado con cocata1 inicial: {cocata_test}")
            except Exception as e:
                logger.error(f"Error en testeo de carga de datos: {str(e)}", exc_info=True)
                # Si hay error pero hay cocata1 en GET, establecerlo como valor inicial
                if cocata_test:
                    initial['cocata1'] = cocata_test
                form = BDCata1Form(initial=initial)
        else:
            form = BDCata1Form(initial=initial)

    # Obtener el registro si existe para pasar al template
    registro = None
    if form.instance and form.instance.pk:
        registro = form.instance
    elif 'cocata1' in request.GET:
        try:
            registro = BDCata1.objects.get(cocata1=request.GET.get('cocata1'))
        except BDCata1.DoesNotExist:
            pass
    
    context = {
        'titulo': 'Registrar Bien Inmueble',
        'form': form,
        'registro': registro,  # Pasar el registro al template para verificar ficha
        'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
        'empresa': municipio_codigo_sesion or '',
        'empresa_codigo': municipio_codigo_sesion or '',
        'municipio_por_concer': municipio_por_concer,
        'municipio_por_concer_str': str(municipio_por_concer) if municipio_por_concer is not None else '0.00',
    }
    logger.info(f"Renderizando template: bienes_inmuebles_form.html")
    logger.info(f"Template path: C:\\simafiweb\\venv\\Scripts\\catastro\\templates\\bienes_inmuebles_form.html")
    logger.info("=== FIN registrar_bien_inmueble ===")
    return render(request, 'bienes_inmuebles_form.html', context)

@catastro_require_auth
def eliminar_bien_inmueble(request):
    """
    Vista para eliminar un bien inmueble
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    cocata1 = request.GET.get('cocata1', '').strip()
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    if not cocata1:
        messages.error(request, 'No se especificó el código catastral a eliminar.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    try:
        from .models import BDCata1
        bien_inmueble = BDCata1.objects.get(cocata1=cocata1, empresa=empresa_codigo)
        
        # Eliminar registros relacionados primero
        from .models import Edificacion, AreasRurales, DetalleAdicionales, BDTerreno
        
        # Eliminar edificaciones
        Edificacion.objects.filter(clave=cocata1, empresa=empresa_codigo).delete()
        
        # Eliminar áreas rurales
        AreasRurales.objects.filter(cocata1=cocata1, empresa=empresa_codigo).delete()
        
        # Eliminar detalles adicionales
        DetalleAdicionales.objects.filter(clave=cocata1, empresa=empresa_codigo).delete()
        
        # Eliminar terreno
        BDTerreno.objects.filter(cocata1=cocata1, empresa=empresa_codigo).delete()
        
        # Eliminar el bien inmueble
        bien_inmueble.delete()
        
        messages.success(request, f'Bien inmueble con código catastral {cocata1} eliminado correctamente.')
    except BDCata1.DoesNotExist:
        messages.error(request, 'No se encontró el bien inmueble especificado.')
    except Exception as e:
        logger.error(f"Error al eliminar bien inmueble: {str(e)}")
        messages.error(request, f'Error al eliminar el bien inmueble: {str(e)}')
    
    # Mantener el código catastral en la redirección si estaba disponible
    cocata1_param = request.GET.get('cocata1', '').strip()
    if cocata1_param:
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1_param}")
    return redirect('catastro:bienes_inmuebles_registrar')

@catastro_require_auth
@catastro_require_auth
def api_buscar_bien_inmueble(request):
    """
    API endpoint para búsqueda interactiva de código catastral
    Busca en la tabla bdcata1 según empresa y cocata1
    """
    try:
        from .models import BDCata1
    except ImportError as e:
        logger.error(f"Error al importar BDCata1: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    cocata1 = request.GET.get('cocata1', '').strip()
    # Obtener empresa de la URL o de la sesión
    empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
    
    if not cocata1:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe ingresar un código catastral'
        })
    
    try:
        # Buscar el registro por empresa y cocata1
        if empresa:
            bien = BDCata1.objects.filter(empresa=empresa, cocata1=cocata1).first()
            logger.info(f"Buscando bien: empresa={empresa}, cocata1={cocata1}")
        else:
            bien = BDCata1.objects.filter(cocata1=cocata1).first()
            logger.info(f"Buscando bien sin empresa: cocata1={cocata1}")
        
        if bien:
            # Preparar los datos del registro encontrado
            campos = {
                'cocata1': bien.cocata1 or '',
                'ficha': str(int(bien.ficha)) if bien.ficha else '1',
                'claveant': bien.claveant or '',
                'nombres': bien.nombres or '',
                'apellidos': bien.apellidos or '',
                'identidad': bien.identidad or '',
                'rtn': bien.rtn or '',
                'telefono': bien.telefono or '',
                'ubicacion': bien.ubicacion or '',
                'municipio': bien.municipio or '',
                'barrio': bien.barrio or '',
                'caserio': bien.caserio or '',
                'mapa': bien.mapa or '',
                'bloque': bien.bloque or '',
                'predio': bien.predio or '',
                'depto': bien.depto or '',
                'sitio': bien.sitio or '',
                'nacionalidad': bien.nacionalidad or '',
                'sexo': bien.sexo or '',
                'uso': str(bien.uso) if bien.uso else '',
                'subuso': bien.subuso or '',
                'constru': str(bien.constru) if bien.constru else '',
                'nofichas': str(bien.nofichas) if bien.nofichas else '',
                'bvl2tie': str(bien.bvl2tie) if bien.bvl2tie else '',
                'conedi': str(bien.conedi) if bien.conedi else '',
                'mejoras': str(bien.mejoras) if bien.mejoras else '',
                'cedif': str(bien.cedif) if bien.cedif else '',
                'detalle': str(bien.detalle) if bien.detalle else '',
                'impuesto': str(bien.impuesto) if bien.impuesto else '',
                'grabable': str(bien.grabable) if bien.grabable else '',
                'cultivo': str(bien.cultivo) if bien.cultivo else '',
                'declarado': str(bien.declarado) if bien.declarado else '',
                'exencion': str(bien.exencion) if bien.exencion else '',
                'condetalle': str(bien.condetalle) if bien.condetalle else '',
                'st': bien.st or '',
                'codhab': bien.codhab or '',
                'codprop': bien.codprop or '',
                'bexenc': str(bien.bexenc) if bien.bexenc else '',
                'tasaimpositiva': str(bien.tasaimpositiva) if bien.tasaimpositiva else '',
                'declaimpto': str(bien.declaimpto) if bien.declaimpto else '',
                'tipopropiedad': str(int(float(bien.tipopropiedad))) if bien.tipopropiedad else '1',
                'clavesure': bien.clavesure or '',
                # Mostrar coordenadas en formato UTM (no convertir a lat/lng)
                'cx': str(bien.cx) if bien.cx else '',
                'cy': str(bien.cy) if bien.cy else '',
                'zonificacion': bien.zonificacion or '',
                'estado': bien.estado or 'A',
                'vivienda': str(int(bien.vivienda)) if bien.vivienda is not None else '0',
                'apartamentos': str(int(bien.apartamentos)) if bien.apartamentos is not None else '0',
                'cuartos': str(int(bien.cuartos)) if bien.cuartos is not None else '0',
                'lote': bien.lote or '',
                'bloquecol': bien.bloquecol or '',
                'usuario': bien.usuario or '',
                'fechasys': bien.fechasys.strftime('%Y-%m-%d %H:%M:%S') if bien.fechasys else '',
            }
            
            logger.info("=== CAMPOS RETORNADOS EN API BUSCAR BIEN INMUEBLE ===")
            logger.info(f"ID: {bien.id}")
            logger.info(f"cocata1: {campos.get('cocata1')}")
            logger.info(f"Total de campos retornados: {len(campos)}")
            for campo, valor in campos.items():
                logger.info(f"  {campo}: '{valor}'")
            logger.info("=" * 100)
            
            return JsonResponse({
                'encontrado': True,
                'campos': campos,
                'mensaje': f'Registro encontrado: {bien.apellidos or ""} {bien.nombres or ""}'.strip() or f'Código: {cocata1}'
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': f'No se encontró registro con código catastral: {cocata1}'
            })
            
    except Exception as e:
        logger.error(f"Error en búsqueda de bien inmueble: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'error': str(e),
            'mensaje': 'Error al buscar el registro'
        }, status=500)

@catastro_require_auth
def buscar_bien_inmueble(request):
    """
    Vista para buscar bienes inmuebles
    """
    from .forms import BusquedaBienInmuebleForm
    from .models import BDCata1
    
    resultados = []
    total_resultados = 0
    texto_busqueda = ''
    return_url = request.GET.get('return_url', '')
    
    # Si no hay return_url, usar la URL del formulario principal
    if not return_url:
        return_url = reverse('catastro:bienes_inmuebles_registrar')
    
    if request.method == 'GET':
        form = BusquedaBienInmuebleForm(request.GET)
        if form.is_valid():
            texto_busqueda = form.cleaned_data.get('busqueda', '').strip()
            
            # Obtener empresa/municipio de la sesión (siempre requerido)
            empresa = request.session.get('catastro_empresa', '')
            
            # Verificar si hay algún criterio de búsqueda
            if texto_busqueda:
                # Construir query siempre filtrando por empresa/municipio
                query = Q()
                
                # SIEMPRE filtrar por empresa/municipio de la sesión
                if empresa:
                    query &= Q(empresa=empresa)
                    
                    # Búsqueda en múltiples campos de texto
                    busqueda_query = (
                        Q(cocata1__icontains=texto_busqueda) |
                        Q(identidad__icontains=texto_busqueda) |
                        Q(nombres__icontains=texto_busqueda) |
                        Q(apellidos__icontains=texto_busqueda) |
                        Q(clavesure__icontains=texto_busqueda) |
                        Q(ubicacion__icontains=texto_busqueda)
                    )
                    query &= busqueda_query
                    
                    # Obtener resultados (siempre filtrados por empresa)
                    bienes = BDCata1.objects.filter(query).order_by('-fechasys')[:50]  # Limitar a 50 resultados
                    
                    resultados = []
                    for bien in bienes:
                        resultados.append({
                            'cocata1': bien.cocata1 or '',
                            'nombres': bien.nombres or '',
                            'apellidos': bien.apellidos or '',
                            'identidad': bien.identidad or '',
                            'clavesure': bien.clavesure or '',
                            'ubicacion': bien.ubicacion or '',
                        })
                    
                    total_resultados = len(resultados)
                else:
                    # Si no hay empresa en sesión, mostrar advertencia
                    messages.warning(request, 'No se ha seleccionado un municipio. Por favor, inicie sesión nuevamente.')
    else:
        form = BusquedaBienInmuebleForm()
    
    context = {
        'titulo': 'Búsqueda de Bienes Inmuebles',
        'form': form,
        'resultados': resultados,
        'total_resultados': total_resultados,
        'texto_busqueda': texto_busqueda,
        'return_url': return_url,
    }
    
    return render(request, 'busqueda_bien_inmueble.html', context)

@catastro_require_auth
def buscar_bien_inmueble_export_excel(request):
    """
    Exportar resultados de búsqueda de bienes inmuebles a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.contrib import messages
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:buscar_bien_inmueble')
    
    from .models import BDCata1
    from django.contrib import messages
    
    empresa = request.session.get('catastro_empresa', '')
    texto_busqueda = request.GET.get('busqueda', '').strip()
    
    if not empresa:
        messages.warning(request, 'No se ha seleccionado un municipio. Por favor, inicie sesión nuevamente.')
        return redirect('catastro:buscar_bien_inmueble')
    
    # Construir query
    query = Q(empresa=empresa)
    
    if texto_busqueda:
        busqueda_query = (
            Q(cocata1__icontains=texto_busqueda) |
            Q(identidad__icontains=texto_busqueda) |
            Q(nombres__icontains=texto_busqueda) |
            Q(apellidos__icontains=texto_busqueda) |
            Q(clavesure__icontains=texto_busqueda) |
            Q(ubicacion__icontains=texto_busqueda)
        )
        query &= busqueda_query
    
    # Obtener resultados
    bienes = BDCata1.objects.filter(query).order_by('-fechasys')[:500]  # Limitar a 500 para Excel
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Búsqueda Bienes Inmuebles"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.append(['BÚSQUEDA DE BIENES INMUEBLES'])
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Información de búsqueda
    if texto_busqueda:
        ws.append(['Criterio de búsqueda:', texto_busqueda])
        ws.merge_cells('B2:G2')
    ws.append(['Total de resultados:', len(bienes)])
    ws.merge_cells('B3:G3')
    ws.append([])
    
    # Encabezados
    headers = ['Clave Catastral', 'Nombres', 'Apellidos', 'Identidad', 'Clave SURE', 'Ubicación']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    row_num = 6
    for bien in bienes:
        ws.append([
            bien.cocata1 or '',
            bien.nombres or '',
            bien.apellidos or '',
            bien.identidad or '',
            bien.clavesure or '',
            bien.ubicacion or '',
        ])
        
        # Aplicar bordes
        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border_style
        
        row_num += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 40
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'busqueda_bienes_inmuebles_{empresa}_{texto_busqueda[:20] or "all"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def buscar_bien_inmueble_export_pdf(request):
    """
    Exportar resultados de búsqueda de bienes inmuebles a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        from django.contrib import messages
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:buscar_bien_inmueble')
    
    from .models import BDCata1
    from django.contrib import messages
    from io import BytesIO
    
    empresa = request.session.get('catastro_empresa', '')
    texto_busqueda = request.GET.get('busqueda', '').strip()
    
    if not empresa:
        messages.warning(request, 'No se ha seleccionado un municipio. Por favor, inicie sesión nuevamente.')
        return redirect('catastro:buscar_bien_inmueble')
    
    # Construir query
    query = Q(empresa=empresa)
    
    if texto_busqueda:
        busqueda_query = (
            Q(cocata1__icontains=texto_busqueda) |
            Q(identidad__icontains=texto_busqueda) |
            Q(nombres__icontains=texto_busqueda) |
            Q(apellidos__icontains=texto_busqueda) |
            Q(clavesure__icontains=texto_busqueda) |
            Q(ubicacion__icontains=texto_busqueda)
        )
        query &= busqueda_query
    
    # Obtener resultados
    bienes = BDCata1.objects.filter(query).order_by('-fechasys')[:500]  # Limitar a 500 para PDF
    
    # Crear buffer para el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Centrado
        spaceAfter=30
    )
    
    # Título
    title = Paragraph('BÚSQUEDA DE BIENES INMUEBLES', title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información de búsqueda
    info_data = []
    if texto_busqueda:
        info_data.append(['Criterio de búsqueda:', texto_busqueda])
    info_data.append(['Total de resultados:', str(len(bienes))])
    info_data.append(['Empresa/Municipio:', empresa])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Preparar datos de la tabla
    data = [['Clave Catastral', 'Nombres', 'Apellidos', 'Identidad', 'Clave SURE', 'Ubicación']]
    
    for bien in bienes:
        data.append([
            bien.cocata1 or '',
            bien.nombres or '',
            bien.apellidos or '',
            bien.identidad or '',
            bien.clavesure or '',
            bien.ubicacion or '',
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch, 2*inch])
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y crear respuesta
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    filename = f'busqueda_bienes_inmuebles_{empresa}_{texto_busqueda[:20] or "all"}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response

@catastro_require_auth
def api_barrios(request):
    """
    API endpoint para obtener barrios filtrados por depto y codmuni
    Filtra según los valores de depto y municipio de la tabla bdcata1
    """
    try:
        from .models import Barrios
    except ImportError as e:
        logger.error(f"Error al importar Barrios: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    depto = request.GET.get('depto', '').strip()
    codmuni = request.GET.get('codmuni', '').strip() or request.GET.get('municipio', '').strip()
    
    # Si no se proporcionan parámetros, intentar obtenerlos de la sesión
    if not depto or not codmuni:
        municipio_codigo_sesion = request.session.get('catastro_empresa', '')
        if municipio_codigo_sesion and len(municipio_codigo_sesion) >= 4:
            if not depto:
                depto = municipio_codigo_sesion[:2]  # Posiciones 1-2
            if not codmuni:
                codmuni = municipio_codigo_sesion[2:4]  # Posiciones 3-4
    
    if not depto or not codmuni:
        return JsonResponse({
            'barrios': [],
            'mensaje': 'Debe proporcionar depto y codmuni'
        })
    
    try:
        # Filtrar barrios por depto y codmuni
        barrios = Barrios.objects.filter(depto=depto, codmuni=codmuni).order_by('codbarrio')
        
        barrios_list = []
        for barrio in barrios:
            barrios_list.append({
                'id': barrio.id,
                'codbarrio': barrio.codbarrio,
                'descripcion': barrio.descripcion or '',
                'tipica': str(barrio.tipica) if barrio.tipica else '0.00',
                'display': f"{barrio.codbarrio} - {barrio.descripcion or ''}".strip()
            })
        
        logger.info(f"Barrios encontrados para depto={depto}, codmuni={codmuni}: {len(barrios_list)}")
        
        return JsonResponse({
            'barrios': barrios_list,
            'depto': depto,
            'codmuni': codmuni,
            'total': len(barrios_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de barrios: {str(e)}", exc_info=True)
        return JsonResponse({
            'barrios': [],
            'error': str(e),
            'mensaje': 'Error al buscar barrios'
        }, status=500)

@catastro_require_auth
def api_identificacion(request):
    """
    API endpoint para buscar identificación por número de identidad (DNI)
    Busca en la tabla identificacion donde identidad = identidad de bdcata1
    """
    try:
        from .models import Identificacion
    except ImportError as e:
        logger.error(f"Error al importar Identificacion: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    identidad = request.GET.get('identidad', '').strip()
    
    if not identidad:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar un número de identidad'
        })
    
    try:
        # Buscar identificación por número de identidad
        identificacion = Identificacion.objects.filter(identidad=identidad).first()
        
        if identificacion:
            logger.info(f"Identificación encontrada para identidad={identidad}: {identificacion.nombres} {identificacion.apellidos}")
            
            return JsonResponse({
                'encontrado': True,
                'identidad': identificacion.identidad,
                'nombres': identificacion.nombres or '',
                'apellidos': identificacion.apellidos or '',
                'fechanac': identificacion.fechanac.strftime('%Y-%m-%d') if identificacion.fechanac else '',
                'nombre_completo': f"{identificacion.nombres or ''} {identificacion.apellidos or ''}".strip(),
                'mensaje': f'Identificación encontrada: {identificacion.nombres or ""} {identificacion.apellidos or ""}'.strip()
            })
        else:
            logger.info(f"No se encontró identificación para identidad={identidad}")
            return JsonResponse({
                'encontrado': False,
                'mensaje': f'No se encontró identificación con número: {identidad}'
            })
            
    except Exception as e:
        logger.error(f"Error en búsqueda de identificación: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'error': str(e),
            'mensaje': 'Error al buscar la identificación'
        }, status=500)

@catastro_require_auth
def api_tipos_sexo(request):
    """
    API endpoint para obtener todos los tipos de sexo disponibles
    """
    try:
        from .models import TipoSexo
    except ImportError as e:
        logger.error(f"Error al importar TipoSexo: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        # Obtener todos los tipos de sexo ordenados por código
        tipos_sexo = TipoSexo.objects.all().order_by('codigo')
        
        tipos_list = []
        for tipo in tipos_sexo:
            tipos_list.append({
                'id': tipo.id,
                'codigo': tipo.codigo,
                'descripcion': tipo.descripcion or '',
                'display': f"{tipo.codigo} - {tipo.descripcion or ''}".strip()
            })
        
        logger.info(f"Tipos de sexo encontrados: {len(tipos_list)}")
        
        return JsonResponse({
            'tipos_sexo': tipos_list,
            'total': len(tipos_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de tipos de sexo: {str(e)}", exc_info=True)
        return JsonResponse({
            'tipos_sexo': [],
            'error': str(e),
            'mensaje': 'Error al buscar tipos de sexo'
        }, status=500)

@catastro_require_auth
def api_naturalezas(request):
    """
    API endpoint para obtener todas las naturalezas jurídicas disponibles
    """
    try:
        from .models import Naturaleza
    except ImportError as e:
        logger.error(f"Error al importar Naturaleza: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        naturalezas = Naturaleza.objects.all().order_by('codigo')
        
        naturalezas_list = []
        for nat in naturalezas:
            naturalezas_list.append({
                'id': nat.id,
                'codigo': str(nat.codigo),
                'descripcion': nat.descripcion or '',
                'display': f"{nat.codigo} - {nat.descripcion or ''}".strip()
            })
        
        logger.info(f"Naturalezas jurídicas encontradas: {len(naturalezas_list)}")
        
        return JsonResponse({
            'naturalezas': naturalezas_list,
            'total': len(naturalezas_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de naturalezas jurídicas: {str(e)}", exc_info=True)
        return JsonResponse({
            'naturalezas': [],
            'error': str(e),
            'mensaje': 'Error al buscar naturalezas jurídicas'
        }, status=500)

@catastro_require_auth
def api_dominios(request):
    """
    API endpoint para obtener todas las clases de dominio disponibles
    """
    try:
        from .models import Dominio
    except ImportError as e:
        logger.error(f"Error al importar Dominio: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        dominios = Dominio.objects.all().order_by('codigo')
        
        dominios_list = []
        for dom in dominios:
            dominios_list.append({
                'id': dom.id,
                'codigo': str(dom.codigo),
                'descripcion': dom.descripcion or '',
                'display': f"{dom.codigo} - {dom.descripcion or ''}".strip()
            })
        
        logger.info(f"Clases de dominio encontradas: {len(dominios_list)}")
        
        return JsonResponse({
            'dominios': dominios_list,
            'total': len(dominios_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de clases de dominio: {str(e)}", exc_info=True)
        return JsonResponse({
            'dominios': [],
            'error': str(e),
            'mensaje': 'Error al buscar clases de dominio'
        }, status=500)

@catastro_require_auth
def api_tipos_documento(request):
    """
    API endpoint para obtener todos los tipos de documento disponibles
    Filtrados por empresa si está disponible en la sesión
    """
    try:
        from .models import TipoDocumento
    except ImportError as e:
        logger.error(f"Error al importar TipoDocumento: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        empresa = request.session.get('catastro_empresa', '')
        
        if empresa:
            tipos_doc = TipoDocumento.objects.filter(empresa=empresa).order_by('codigo')
        else:
            tipos_doc = TipoDocumento.objects.all().order_by('codigo')
        
        tipos_list = []
        for tipo in tipos_doc:
            tipos_list.append({
                'id': tipo.id,
                'codigo': str(tipo.codigo),
                'descripcion': tipo.descripcion or '',
                'display': f"{tipo.codigo} - {tipo.descripcion or ''}".strip()
            })
        
        logger.info(f"Tipos de documento encontrados: {len(tipos_list)}")
        
        return JsonResponse({
            'tipos_documento': tipos_list,
            'total': len(tipos_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de tipos de documento: {str(e)}", exc_info=True)
        return JsonResponse({
            'tipos_documento': [],
            'error': str(e),
            'mensaje': 'Error al buscar tipos de documento'
        }, status=500)

@catastro_require_auth
def api_usos(request):
    """
    API endpoint para obtener todos los usos disponibles
    """
    try:
        from .models import Usos
    except ImportError as e:
        logger.error(f"Error al importar Usos: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        # Obtener todos los usos ordenados por código
        usos = Usos.objects.all().order_by('uso')
        
        usos_list = []
        for uso in usos:
            # Normalizar el código de uso a 2 dígitos con ceros a la izquierda (01, 02, etc.)
            codigo_uso = str(uso.uso) if uso.uso else ''
            if codigo_uso and codigo_uso.isdigit():
                codigo_uso = codigo_uso.zfill(2)
            
            usos_list.append({
                'id': uso.id,
                'uso': codigo_uso,
                'desuso': uso.desuso or '',
                'display': f"{codigo_uso} - {uso.desuso or ''}".strip()
            })
        
        logger.info(f"Usos encontrados: {len(usos_list)}")
        
        return JsonResponse({
            'usos': usos_list,
            'total': len(usos_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de usos: {str(e)}", exc_info=True)
        return JsonResponse({
            'usos': [],
            'error': str(e),
            'mensaje': 'Error al buscar usos'
        }, status=500)

@catastro_require_auth
def api_subusos(request):
    """
    API endpoint para obtener subusos filtrados por uso
    Filtra según el campo uso de la tabla subuso que debe coincidir con el campo USO de la tabla usos
    """
    try:
        from .models import Subuso
    except ImportError as e:
        logger.error(f"Error al importar Subuso: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    uso = request.GET.get('uso', '').strip()
    
    if not uso:
        return JsonResponse({
            'subusos': [],
            'mensaje': 'Debe proporcionar un código de uso'
        })
    
    try:
        # Filtrar subusos por uso (donde subuso.uso = usos.USO)
        subusos = Subuso.objects.filter(uso=uso).order_by('codsubuso')
        
        subusos_list = []
        for subuso in subusos:
            subusos_list.append({
                'id': subuso.id,
                'codsubuso': subuso.codsubuso.strip() if subuso.codsubuso else '',
                'des_subuso': subuso.des_subuso or '',
                'uso': subuso.uso,
                'display': f"{subuso.codsubuso.strip() if subuso.codsubuso else ''} - {subuso.des_subuso or ''}".strip()
            })
        
        logger.info(f"Subusos encontrados para uso={uso}: {len(subusos_list)}")
        
        return JsonResponse({
            'subusos': subusos_list,
            'uso': uso,
            'total': len(subusos_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de subusos: {str(e)}", exc_info=True)
        return JsonResponse({
            'subusos': [],
            'error': str(e),
            'mensaje': 'Error al buscar subusos'
        }, status=500)

@catastro_require_auth
def api_habitacional(request):
    """
    API endpoint para obtener todos los códigos habitacionales disponibles
    """
    try:
        from .models import Habitacional
    except ImportError as e:
        logger.error(f"Error al importar Habitacional: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        # Obtener todos los códigos habitacionales ordenados por código
        habitacionales = Habitacional.objects.all().order_by('cohabit')
        
        habitacionales_list = []
        for habitacional in habitacionales:
            habitacionales_list.append({
                'id': habitacional.id,
                'cohabit': habitacional.cohabit,
                'bdeshabit': habitacional.bdeshabit or '',
                'display': f"{habitacional.cohabit} - {habitacional.bdeshabit or ''}".strip()
            })
        
        logger.info(f"Códigos habitacionales encontrados: {len(habitacionales_list)}")
        
        return JsonResponse({
            'habitacionales': habitacionales_list,
            'total': len(habitacionales_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de códigos habitacionales: {str(e)}", exc_info=True)
        return JsonResponse({
            'habitacionales': [],
            'error': str(e),
            'mensaje': 'Error al buscar códigos habitacionales'
        }, status=500)

@catastro_require_auth
def api_propietarios(request):
    """
    API endpoint para obtener todos los códigos de propietarios disponibles
    """
    try:
        from .models import Propietarios
    except ImportError as e:
        logger.error(f"Error al importar Propietarios: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        # Obtener todos los códigos de propietarios ordenados por código
        propietarios = Propietarios.objects.all().order_by('copropi')
        
        propietarios_list = []
        for propietario in propietarios:
            propietarios_list.append({
                'id': propietario.id,
                'copropi': propietario.copropi,
                'bdespro': propietario.bdespro or '',
                'display': f"{propietario.copropi} - {propietario.bdespro or ''}".strip()
            })
        
        logger.info(f"Códigos de propietarios encontrados: {len(propietarios_list)}")
        
        return JsonResponse({
            'propietarios': propietarios_list,
            'total': len(propietarios_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de códigos de propietarios: {str(e)}", exc_info=True)
        return JsonResponse({
            'propietarios': [],
            'error': str(e),
            'mensaje': 'Error al buscar códigos de propietarios'
        }, status=500)

@catastro_require_auth
def api_zonasusos(request):
    """
    API endpoint para obtener todas las zonas de uso disponibles
    """
    try:
        from .models import Zonasusos
    except ImportError as e:
        logger.error(f"Error al importar Zonasusos: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        # Obtener todas las zonas de uso ordenadas por tipo de zona
        zonasusos = Zonasusos.objects.all().order_by('tipozona')
        
        zonas_list = []
        for zona in zonasusos:
            zonas_list.append({
                'id': zona.id,
                'tipozona': zona.tipozona or '',
                'descripcion': zona.descripcion or '',
                'display': f"{zona.tipozona or ''} - {zona.descripcion or ''}".strip()
            })
        
        logger.info(f"Zonas de uso encontradas: {len(zonas_list)}")
        
        return JsonResponse({
            'zonasusos': zonas_list,
            'total': len(zonas_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de zonas de uso: {str(e)}", exc_info=True)
        return JsonResponse({
            'zonasusos': [],
            'error': str(e),
            'mensaje': 'Error al buscar zonas de uso'
        }, status=500)

@catastro_require_auth
def api_nacionalidad(request):
    """
    API endpoint para obtener todas las nacionalidades disponibles
    """
    try:
        from .models import Nacionalidad
    except ImportError as e:
        logger.error(f"Error al importar Nacionalidad: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    try:
        # Obtener todas las nacionalidades ordenadas por código
        nacionalidades = Nacionalidad.objects.all().order_by('codigo')
        
        nacionalidades_list = []
        for nacionalidad in nacionalidades:
            nacionalidades_list.append({
                'id': nacionalidad.id,
                'codigo': nacionalidad.codigo,
                'descripcion': nacionalidad.descripcion or '',
                'display': f"{nacionalidad.codigo} - {nacionalidad.descripcion or ''}".strip()
            })
        
        logger.info(f"Nacionalidades encontradas: {len(nacionalidades_list)}")
        
        return JsonResponse({
            'nacionalidades': nacionalidades_list,
            'total': len(nacionalidades_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de nacionalidades: {str(e)}", exc_info=True)
        return JsonResponse({
            'nacionalidades': [],
            'error': str(e),
            'mensaje': 'Error al buscar nacionalidades'
        }, status=500)

@catastro_require_auth
def api_caserio(request):
    """
    API endpoint para obtener caseríos filtrados por depto, codmuni y codbarrio
    Filtra según los valores de depto, municipio y barrio de la tabla bdcata1:
    - caserio.depto = bdcata1.depto
    - caserio.codmuni = bdcata1.municipio
    - caserio.codbarrio = bdcata1.barrio
    """
    try:
        from .models import Caserio
    except ImportError as e:
        logger.error(f"Error al importar Caserio: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    depto = request.GET.get('depto', '').strip()
    codmuni = request.GET.get('codmuni', '').strip() or request.GET.get('municipio', '').strip()
    codbarrio = request.GET.get('codbarrio', '').strip() or request.GET.get('barrio', '').strip()
    
    # Si no se proporcionan parámetros, intentar obtenerlos de la sesión
    if not depto or not codmuni:
        municipio_codigo_sesion = request.session.get('catastro_empresa', '')
        if municipio_codigo_sesion and len(municipio_codigo_sesion) >= 4:
            if not depto:
                depto = municipio_codigo_sesion[:2]  # Posiciones 1-2
            if not codmuni:
                codmuni = municipio_codigo_sesion[2:4]  # Posiciones 3-4
    
    if not depto or not codmuni:
        return JsonResponse({
            'caserios': [],
            'mensaje': 'Debe proporcionar depto y codmuni'
        })
    
    try:
        # Filtrar caseríos por depto, codmuni y codbarrio
        # Si codbarrio está vacío, también buscar caseríos sin codbarrio específico
        if codbarrio:
            caserios = Caserio.objects.filter(depto=depto, codmuni=codmuni, codbarrio=codbarrio).order_by('codigo')
        else:
            # Si no hay barrio seleccionado, buscar caseríos sin barrio específico o con barrio vacío
            caserios = Caserio.objects.filter(depto=depto, codmuni=codmuni, codbarrio='').order_by('codigo')
        
        caserios_list = []
        for caserio in caserios:
            caserios_list.append({
                'id': caserio.id,
                'codigo': caserio.codigo.strip() if caserio.codigo else '',
                'descripcion': caserio.descripcion or '',
                'depto': caserio.depto,
                'codmuni': caserio.codmuni,
                'codbarrio': caserio.codbarrio.strip() if caserio.codbarrio else '',
                'display': f"{caserio.codigo.strip() if caserio.codigo else ''} - {caserio.descripcion or ''}".strip()
            })
        
        logger.info(f"Caseríos encontrados para depto={depto}, codmuni={codmuni}, codbarrio={codbarrio}: {len(caserios_list)}")
        
        return JsonResponse({
            'caserios': caserios_list,
            'depto': depto,
            'codmuni': codmuni,
            'codbarrio': codbarrio,
            'total': len(caserios_list)
        })
        
    except Exception as e:
        logger.error(f"Error en búsqueda de caseríos: {str(e)}", exc_info=True)
        return JsonResponse({
            'caserios': [],
            'error': str(e),
            'mensaje': 'Error al buscar caseríos'
        }, status=500)


def test_guardado_datos(request):
    """
    Vista de prueba para simular el guardado con datos reales
    Solo para testing - NO usar en producción
    """
    logger.info("=" * 80)
    logger.info("=== TEST DE GUARDADO CON DATOS REALES ===")
    logger.info("=" * 80)
    
    try:
        from .forms import BDCata1Form
        from .models import BDCata1
        from decimal import Decimal
        from datetime import datetime
        
        # Datos de prueba proporcionados por el usuario
        datos_post = {
            'empresa': '0301',
            'cocata1': 'CL11-0B-01-10-TEST',
            'ficha': '2',
            'claveant': '',
            'mapa': 'CL11-0B',
            'bloque': '01',
            'predio': '10',
            'depto': '03',
            'municipio': '01',
            'barrio': '',
            'caserio': '',
            'sitio': 'EF',
            'nombres': 'GERMAN',
            'apellidos': 'SANCHEZ',
            'identidad': '0401-1974-00258',
            'rtn': '',
            'ubicacion': 'ALDEA BELEN',
            'nacionalidad': '205',
            'uso': '0',
            'subuso': '1',
            'constru': '0',
            'nofichas': '0',
            'bvl2tie': '653976.25',
            'conedi': '0',
            'mejoras': '348504.5',
            'cedif': '1',
            'detalle': '13601.75',
            'impuesto': '2607.98',
            'grabable': '1133905.91',
            'cultivo': '0',
            'declarado': '0',
            'condetalle': '1',
            'exencion': '20000',
            'st': '0',
            'codhab': '0',
            'codprop': '01',
            'bexenc': '1',
            'tasaimpositiva': '0',
            'declaimpto': '0',
            'sexo': '',
            'telefono': '0',
            'tipopropiedad': '1',
            'estado': 'A',
            'clavesure': '',
            'cx': '0',
            'cy': '0',
            'zonificacion': '',
        }
        
        logger.info("1. Creando formulario con datos de prueba...")
        form = BDCata1Form(datos_post)
        
        logger.info(f"2. Formulario válido: {form.is_valid()}")
        
        if not form.is_valid():
            logger.error("=== ERRORES DE VALIDACIÓN ===")
            for field, errors in form.errors.items():
                for error in errors:
                    logger.error(f"  {field}: {error}")
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'message': 'Errores de validación en el formulario'
            })
        
        logger.info("3. Formulario válido, guardando...")
        bien = form.save(commit=False)
        bien.usuario = 'TEST_USER'
        bien.fechasys = timezone.now()
        
        logger.info(f"4. Objeto a guardar: cocata1={bien.cocata1}, empresa={bien.empresa}")
        bien.save()
        
        logger.info(f"✓✓✓ GUARDADO EXITOSO - ID: {bien.id} ✓✓✓")
        
        # Verificar
        bien_verificado = BDCata1.objects.get(id=bien.id)
        
        return JsonResponse({
            'success': True,
            'message': f'¡Bien inmueble guardado correctamente! ID: {bien.id}',
            'data': {
                'id': bien_verificado.id,
                'cocata1': bien_verificado.cocata1,
                'empresa': bien_verificado.empresa,
                'nombres': bien_verificado.nombres,
                'apellidos': bien_verificado.apellidos,
            }
        })
        
    except Exception as e:
        logger.error(f"ERROR EN TEST: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}',
            'error': str(e)
        })

@catastro_require_auth
def terreno_urbano_form(request, cocata1: str):
    """
    Formulario dedicado para Avalúo de Terreno Urbano vinculado a una clave catastral.
    - Carga el BDTerreno existente (si existe) o prepara uno nuevo enlazado a la clave.
    - Usa BDTerrenoForm con los parámetros urbanos.
    """
    import sys
    import os
    import importlib.util
    
    # La ruta correcta del módulo catastro es C:\simafiweb\venv\Scripts\catastro
    # Los modelos BDCata1 y BDTerreno ahora están en C:\simafiweb\venv\Scripts\catastro\models.py
    current_file_dir = os.path.dirname(os.path.abspath(__file__))
    # current_file_dir debería ser C:\simafiweb\venv\Scripts\catastro
    catastro_module_dir = current_file_dir  # C:\simafiweb\venv\Scripts\catastro
    
    # Agregar el directorio del módulo catastro al path
    if catastro_module_dir not in sys.path:
        sys.path.insert(0, catastro_module_dir)
    
    # Estrategia de importación: usar la ruta correcta del módulo catastro
    try:
        # Los modelos BDCata1 y BDTerreno están en C:\simafiweb\venv\Scripts\catastro\models.py
        catastro_models_path = os.path.join(catastro_module_dir, 'models.py')
        # Los formularios también están en la misma ubicación
        catastro_forms_path = os.path.join(catastro_module_dir, 'forms.py')
        
        # Normalizar las rutas para evitar problemas
        catastro_models_path = os.path.normpath(catastro_models_path)
        catastro_forms_path = os.path.normpath(catastro_forms_path)
        
        logger.info(f"Intentando cargar modelos desde: {catastro_models_path}")
        logger.info(f"Intentando cargar formularios desde: {catastro_forms_path}")
        
        if os.path.exists(catastro_models_path) and os.path.exists(catastro_forms_path):
            # Verificar si el módulo cargado es el correcto antes de reutilizarlo
            should_reload = False
            if 'catastro.models' in sys.modules:
                cached_models = sys.modules['catastro.models']
                # Verificar que tenga los modelos correctos
                if hasattr(cached_models, 'BDCata1') and hasattr(cached_models, 'BDTerreno'):
                    catastro_models = cached_models
                    logger.info("✓ Reutilizando módulo catastro.models correcto ya cargado")
                else:
                    # El módulo en caché es incorrecto, necesitamos recargarlo
                    logger.warning("⚠ Módulo catastro.models en caché es incorrecto, recargando...")
                    should_reload = True
            
            if should_reload or 'catastro.models' not in sys.modules:
                # Limpiar el módulo incorrecto si existe
                if 'catastro.models' in sys.modules:
                    del sys.modules['catastro.models']
                
                # Cargar modelos usando importación relativa simple
                from . import models as catastro_models
                # Registrar inmediatamente para evitar cargas múltiples
                sys.modules['catastro.models'] = catastro_models
                logger.info("✓ Módulo catastro.models cargado y registrado desde: " + catastro_models_path)
            
            # Verificar que los modelos existan después de ejecutar el módulo
            if not hasattr(catastro_models, 'BDCata1'):
                # Listar todos los atributos para debugging
                attrs = [x for x in dir(catastro_models) if not x.startswith('_')]
                logger.error(f"Ruta del módulo cargado: {getattr(catastro_models, '__file__', 'desconocida')}")
                raise AttributeError(f"El módulo catastro.models no tiene el atributo BDCata1. Atributos disponibles: {attrs}")
            if not hasattr(catastro_models, 'BDTerreno'):
                attrs = [x for x in dir(catastro_models) if not x.startswith('_')]
                logger.error(f"Ruta del módulo cargado: {getattr(catastro_models, '__file__', 'desconocida')}")
                raise AttributeError(f"El módulo catastro.models no tiene el atributo BDTerreno. Atributos disponibles: {attrs}")
            
            BDCata1 = catastro_models.BDCata1
            BDTerreno = catastro_models.BDTerreno
            
            # Asegurar que el paquete catastro exista en sys.modules para los imports relativos
            if 'catastro' not in sys.modules:
                import types
                catastro_pkg = types.ModuleType('catastro')
                catastro_pkg.__path__ = [os.path.join(root_dir, 'catastro')]
                sys.modules['catastro'] = catastro_pkg
            
            # Registrar catastro.models en sys.modules y como atributo del paquete (si no está ya registrado)
            if 'catastro.models' not in sys.modules:
                sys.modules['catastro.models'] = catastro_models
            if not hasattr(sys.modules['catastro'], 'models'):
                sys.modules['catastro'].models = catastro_models
            
            # Cargar formularios - verificar si el módulo cargado tiene el formulario correcto
            should_reload_forms = False
            if 'catastro.forms' in sys.modules:
                cached_forms = sys.modules['catastro.forms']
                # Verificar que tenga el formulario correcto
                if hasattr(cached_forms, 'BDTerrenoForm'):
                    catastro_forms = cached_forms
                    logger.info("✓ Reutilizando módulo catastro.forms correcto ya cargado")
                else:
                    logger.warning("⚠ Módulo catastro.forms en caché no tiene BDTerrenoForm, recargando...")
                    should_reload_forms = True
            
            if should_reload_forms or 'catastro.forms' not in sys.modules:
                # Limpiar el módulo incorrecto si existe
                if 'catastro.forms' in sys.modules:
                    del sys.modules['catastro.forms']
                
                # Asegurar que catastro.models esté completamente configurado antes de cargar formularios
                # Esto es crítico para que el import relativo funcione
                if 'catastro.models' not in sys.modules:
                    sys.modules['catastro.models'] = catastro_models
                if 'catastro' not in sys.modules:
                    import types
                    catastro_pkg = types.ModuleType('catastro')
                    catastro_pkg.__path__ = [os.path.join(root_dir, 'catastro')]
                    sys.modules['catastro'] = catastro_pkg
                if not hasattr(sys.modules['catastro'], 'models'):
                    sys.modules['catastro'].models = catastro_models
                
                # Cargar formularios - el import relativo debería funcionar ahora
                # Cargar formularios usando importación relativa simple
                from . import forms as catastro_forms
                
                # Registrar catastro.forms en sys.modules
                sys.modules['catastro.forms'] = catastro_forms
                logger.info("✓ Módulo catastro.forms cargado y registrado desde: " + catastro_forms_path)
                logger.info(f"Archivo real cargado: {getattr(catastro_forms, '__file__', 'desconocida')}")
            
            # Verificar que el formulario exista
            if not hasattr(catastro_forms, 'BDTerrenoForm'):
                attrs = [x for x in dir(catastro_forms) if not x.startswith('_')]
                file_loaded = getattr(catastro_forms, '__file__', 'desconocida')
                logger.error(f"Ruta del módulo de formularios cargado: {file_loaded}")
                logger.error(f"Ruta esperada: {catastro_forms_path}")
                logger.error(f"¿Coinciden las rutas? {os.path.normpath(file_loaded) == os.path.normpath(catastro_forms_path) if file_loaded != 'desconocida' else 'N/A'}")
                
                # Intentar verificar si BDTerreno está disponible en el módulo de formularios
                if hasattr(catastro_forms, 'BDTerreno'):
                    logger.info("BDTerreno está disponible en catastro.forms, pero BDTerrenoForm no")
                    # Si BDTerreno está disponible pero BDTerrenoForm no, puede ser un problema de ejecución
                    # Intentar definir BDTerrenoForm manualmente como último recurso
                    try:
                        from django import forms
                        class BDTerrenoFormTemp(forms.ModelForm):
                            class Meta:
                                model = catastro_forms.BDTerreno
                                fields = '__all__'
                                exclude = ['cocata1', 'usuario', 'fechasys']
                        catastro_forms.BDTerrenoForm = BDTerrenoFormTemp
                        logger.warning("⚠ BDTerrenoForm definido manualmente como solución temporal")
                    except Exception as temp_error:
                        logger.error(f"No se pudo definir BDTerrenoForm manualmente: {str(temp_error)}")
                
                if not hasattr(catastro_forms, 'BDTerrenoForm'):
                    raise AttributeError(f"El módulo catastro.forms no tiene el atributo BDTerrenoForm. Archivo cargado: {file_loaded}. Atributos disponibles: {attrs}")
            
            BDTerrenoForm = catastro_forms.BDTerrenoForm
            
            logger.info("✓ Modelos y formularios cargados dinámicamente desde catastro")
            logger.info(f"BDCata1: {BDCata1}, BDTerreno: {BDTerreno}, BDTerrenoForm: {BDTerrenoForm}")
        else:
            # Fallback: importación normal
            from catastro.models import BDCata1, BDTerreno
            from catastro.forms import BDTerrenoForm
            logger.info("✓ Modelos y formularios importados normalmente desde catastro")
            
    except (ImportError, AttributeError, Exception) as e:
        logger.error(f"Error al importar modelos/formularios: {str(e)}", exc_info=True)
        logger.error(f"root_dir: {root_dir}")
        logger.error(f"sys.path contiene root_dir: {root_dir in sys.path}")
        logger.error(f"catastro/models.py existe: {os.path.exists(os.path.join(root_dir, 'catastro', 'models.py'))}")
        logger.error(f"catastro/forms.py existe: {os.path.exists(os.path.join(root_dir, 'catastro', 'forms.py'))}")
        
        # Intentar verificar qué atributos tiene el módulo si se cargó
        try:
            if 'catastro_models' in locals():
                attrs = [x for x in dir(catastro_models) if not x.startswith('_')]
                logger.error(f"Atributos disponibles en catastro_models: {attrs}")
        except:
            pass
        
        messages.error(request, f'Error: No se pudo cargar el formulario de terreno urbano. Los modelos BDCata1 y BDTerreno no están disponibles. Error: {str(e)}')
        # Intentar obtener cocata1 de la URL si está disponible
        cocata1_from_url = request.path.split('/')[-2] if len(request.path.split('/')) >= 3 else ''
        if cocata1_from_url and cocata1_from_url != 'urbano':
            return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1_from_url}")
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Obtener empresa de la sesión
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Buscar el registro por cocata1 y empresa
    try:
        if empresa_codigo:
            registro = BDCata1.objects.get(cocata1=cocata1, empresa=empresa_codigo)
        else:
            registro = BDCata1.objects.get(cocata1=cocata1)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con código catastral: {cocata1}')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1}")
    
    # Verificar que la ficha sea urbana (1)
    if registro.ficha and int(registro.ficha) != 1:
        messages.warning(request, f'Advertencia: La ficha actual es {int(registro.ficha)} (rural). El avalúo de terreno urbano solo aplica para ficha 1.')
    
    # Obtener o crear terreno
    try:
        terreno = BDTerreno.objects.get(cocata1=registro.cocata1)
    except BDTerreno.DoesNotExist:
        terreno = BDTerreno(cocata1=registro.cocata1, empresa=empresa_codigo)
    
    if request.method == 'POST':
        form = BDTerrenoForm(request.POST, instance=terreno, prefix='terreno')
        if form.is_valid():
            obj = form.save(commit=False)
            obj.cocata1 = registro.cocata1
            obj.empresa = empresa_codigo
            current_user = request.session.get('catastro_usuario_nombre', '') or ''
            if not obj.usuario:
                obj.usuario = current_user[:50] if current_user else ''
            obj.fechasys = timezone.now()
            obj.save()
            
            # Actualizar valor de tierra en registro principal si cambió
            try:
                valor_terreno = obj.calcular_valor_terreno()
                if valor_terreno != (registro.bvl2tie or Decimal('0.00')):
                    registro.bvl2tie = valor_terreno
                
                # Recalcular el impuesto usando la misma lógica del frontend
                impuesto_calculado = calcular_impuesto_bdcata1(registro, empresa_codigo)
                
                # Actualizar el campo impuesto en bdcata1
                registro.impuesto = impuesto_calculado
                
                # Guardar ambos campos en una sola operación
                registro.save(update_fields=['bvl2tie', 'impuesto'])
                logger.info(f'Impuesto calculado y actualizado en bdcata1 después de guardar terreno urbano: {impuesto_calculado} para clave {cocata1}')
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro, empresa_codigo)
            except Exception as e:
                logger.error(f"Error al actualizar valor del terreno e impuesto: {str(e)}", exc_info=True)
            
            messages.success(request, 'Avalúo de Terreno Urbano guardado correctamente.')
            # Redirigir manteniendo el código catastral en la URL
            return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1}")
    else:
        form = BDTerrenoForm(instance=terreno, prefix='terreno')
    
    context = {
        'registro': registro,
        'form': form,
        'titulo': f'Avalúo de Terreno Urbano - {registro.cocata1}',
        'cocata1': registro.cocata1,
        'empresa': registro.empresa or empresa_codigo,
        'es_urbano': registro.ficha and int(registro.ficha) == 1,
        'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
    }
    return render(request, 'terreno_urbano_form.html', context)


@csrf_exempt
@catastro_require_auth
def api_avaluo_terreno(request):
    """
    API endpoint para cargar, guardar y eliminar datos de avalúo de terreno urbano
    GET: Carga los datos del terreno existente
    POST: Guarda o elimina datos del terreno
    """
    from django.http import JsonResponse
    from decimal import Decimal
    
    try:
        from .models import BDCata1, BDTerreno
        from .forms import BDTerrenoForm
    except ImportError as e:
        logger.error(f"Error al importar modelos/formularios: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo', 'mensaje': str(e)}, status=500)
    
    cocata1 = request.GET.get('cocata1', '').strip() or request.POST.get('cocata1', '').strip()
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if not cocata1:
        return JsonResponse({
            'error': 'Debe proporcionar cocata1',
            'mensaje': 'Código catastral requerido'
        }, status=400)
    
    # Buscar el registro principal
    try:
        if empresa_codigo:
            registro = BDCata1.objects.get(cocata1=cocata1, empresa=empresa_codigo)
        else:
            registro = BDCata1.objects.get(cocata1=cocata1)
    except BDCata1.DoesNotExist:
        return JsonResponse({
            'error': 'Registro no encontrado',
            'mensaje': f'No se encontró un registro con código catastral: {cocata1}'
        }, status=404)
    
    if request.method == 'GET':
        # Cargar datos del terreno existente
        try:
            terreno = BDTerreno.objects.get(cocata1=cocata1)
            campos = {}
            # Convertir todos los campos del modelo a diccionario
            for field in terreno._meta.get_fields():
                if hasattr(terreno, field.name):
                    value = getattr(terreno, field.name)
                    if value is None:
                        campos[field.name] = ''
                    elif isinstance(value, Decimal):
                        campos[field.name] = str(value)
                    else:
                        campos[field.name] = str(value)
            
            # Incluir valor calculado del terreno
            try:
                valor_terreno = terreno.calcular_valor_terreno()
                campos['valor_terreno_calculado'] = str(valor_terreno)
            except:
                campos['valor_terreno_calculado'] = '0.00'
            
            return JsonResponse({
                'exito': True,
                'campos': campos,
                'mensaje': 'Datos cargados correctamente'
            })
        except BDTerreno.DoesNotExist:
            # No existe terreno, retornar campos vacíos
            return JsonResponse({
                'exito': True,
                'campos': {},
                'mensaje': 'No hay datos de terreno guardados'
            })
    
    elif request.method == 'POST':
        accion = request.POST.get('accion', 'guardar').strip().lower()
        
        if accion == 'eliminar':
            # Eliminar terreno
            try:
                terreno = BDTerreno.objects.get(cocata1=cocata1, empresa=empresa_codigo)
                terreno.delete()
                
                # Actualizar valor de tierra en registro principal a 0
                registro.bvl2tie = Decimal('0.00')
                
                # Recalcular el impuesto usando la misma lógica del frontend
                impuesto_calculado = calcular_impuesto_bdcata1(registro, empresa_codigo)
                
                # Actualizar el campo impuesto en bdcata1
                registro.impuesto = impuesto_calculado
                
                # Guardar ambos campos en una sola operación
                registro.save(update_fields=['bvl2tie', 'impuesto'])
                logger.info(f'Campo bvl2tie actualizado a 0 e impuesto recalculado en bdcata1 después de eliminar terreno urbano: {impuesto_calculado} para clave {cocata1}')
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro, empresa_codigo)
                
                return JsonResponse({
                    'exito': True,
                    'mensaje': 'Avalúo de terreno eliminado correctamente'
                })
            except BDTerreno.DoesNotExist:
                return JsonResponse({
                    'exito': False,
                    'mensaje': 'No existe un avalúo de terreno para eliminar'
                }, status=404)
        
        else:
            # Guardar terreno
            try:
                terreno, created = BDTerreno.objects.get_or_create(
                    cocata1=cocata1,
                    defaults={'empresa': empresa_codigo}
                )
                
                form = BDTerrenoForm(request.POST, instance=terreno)
                if form.is_valid():
                    obj = form.save(commit=False)
                    obj.cocata1 = cocata1
                    obj.empresa = empresa_codigo
                    current_user = request.session.get('catastro_usuario_nombre', '') or ''
                    if not obj.usuario:
                        obj.usuario = current_user[:50] if current_user else ''
                    obj.fechasys = timezone.now()
                    obj.save()
                    
                    # Actualizar valor de tierra en registro principal
                    try:
                        valor_terreno = obj.calcular_valor_terreno()
                        if valor_terreno != (registro.bvl2tie or Decimal('0.00')):
                            registro.bvl2tie = valor_terreno
                        
                        # Recalcular el impuesto usando la misma lógica del frontend
                        impuesto_calculado = calcular_impuesto_bdcata1(registro, empresa_codigo)
                        
                        # Actualizar el campo impuesto en bdcata1
                        registro.impuesto = impuesto_calculado
                        
                        # Guardar ambos campos en una sola operación
                        registro.save(update_fields=['bvl2tie', 'impuesto'])
                        logger.info(f'Impuesto calculado y actualizado en bdcata1 después de guardar terreno urbano (API): {impuesto_calculado} para clave {cocata1}')
                        
                        # Actualizar el impuesto en tasasmunicipales
                        actualizar_impuesto_tasas_municipales(registro, impuesto_calculado, empresa_codigo)
                        
                        # Calcular tasas municipales (rubros que empiezan con T)
                        calcular_tasas_municipales_automatico(registro, empresa_codigo)
                    except Exception as e:
                        logger.error(f"Error al actualizar valor del terreno e impuesto: {str(e)}", exc_info=True)
                    
                    return JsonResponse({
                        'exito': True,
                        'mensaje': 'Avalúo de terreno guardado correctamente',
                        'creado': created
                    })
                else:
                    # Errores de validación
                    errores = {}
                    for field, errors in form.errors.items():
                        errores[field] = [str(e) for e in errors]
                    
                    return JsonResponse({
                        'exito': False,
                        'mensaje': 'Errores de validación',
                        'errores': errores
                    }, status=400)
            except Exception as e:
                logger.error(f"Error al guardar terreno: {str(e)}", exc_info=True)
                return JsonResponse({
                    'exito': False,
                    'mensaje': f'Error al guardar: {str(e)}'
                }, status=500)
    
    return JsonResponse({
        'error': 'Método no permitido',
        'mensaje': 'Solo se permiten GET y POST'
    }, status=405)


@catastro_require_auth
def terreno_rural_form(request, cocata1):
    """
    Vista para el formulario de avalúo de terreno rural
    Similar a terreno_urbano_form pero para ficha=2 (rural)
    """
    import sys
    import os
    
    try:
        from .models import BDCata1, BDTerreno
        from .forms import BDTerrenoForm
    except ImportError as e:
        logger.error(f"Error al importar modelos/formularios: {str(e)}", exc_info=True)
        messages.error(request, f'Error: No se pudo cargar el formulario de terreno rural. Los modelos BDCata1 y BDTerreno no están disponibles. Error: {str(e)}')
        # Intentar obtener cocata1 de la URL si está disponible
        cocata1_from_url = request.path.split('/')[-2] if len(request.path.split('/')) >= 3 else ''
        if cocata1_from_url and cocata1_from_url != 'rural':
            return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1_from_url}")
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Obtener empresa de la sesión
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Buscar el registro por cocata1 y empresa
    try:
        if empresa_codigo:
            registro = BDCata1.objects.get(cocata1=cocata1, empresa=empresa_codigo)
        else:
            registro = BDCata1.objects.get(cocata1=cocata1)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con código catastral: {cocata1}')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1}")
    
    # Verificar que la ficha sea rural (2)
    if registro.ficha and int(registro.ficha) != 2:
        messages.warning(request, f'Advertencia: La ficha actual es {int(registro.ficha)} (urbana). El avalúo de terreno rural solo aplica para ficha 2.')
    
    # Obtener o crear terreno
    try:
        terreno = BDTerreno.objects.get(cocata1=registro.cocata1)
    except BDTerreno.DoesNotExist:
        terreno = BDTerreno(cocata1=registro.cocata1, empresa=empresa_codigo)
    
    if request.method == 'POST':
        logger.info(f"=== POST recibido en terreno_rural_form para cocata1={cocata1} ===")
        logger.info(f"Empresa del login: {empresa_codigo}")
        logger.info(f"Total de claves en POST: {len(request.POST)}")
        logger.info(f"Primeras 30 claves en POST: {list(request.POST.keys())[:30]}")
        
        # Verificar algunos campos clave
        campos_clave = ['terreno-fcarea', 'terreno-fcubic', 'terreno-fcservi', 'terreno-fcacceso', 'terreno-fcagua', 
                       'terreno-bfacmodi', 'terreno-fcarea2', 'terreno-fcservi2', 'terreno-fctopo', 'terreno-fcconfi',
                       'terreno-bfacmod2', 'terreno-bvlbas2', 'baream21', 'terreno-bvlbas1', 'valor_terreno']
        for campo in campos_clave:
            valor = request.POST.get(campo, 'NO ENCONTRADO')
            logger.info(f"POST[{campo}]: {valor}")
        
        form = BDTerrenoForm(request.POST, instance=terreno, prefix='terreno')
        logger.info(f"Formulario creado. ¿Es válido? {form.is_valid()}")
        if not form.is_valid():
            logger.error(f"=== ERROR: Formulario inválido para cocata1={cocata1} ===")
            logger.error(f"Errores del formulario: {form.errors}")
            for campo, errores in form.errors.items():
                logger.error(f"  Campo '{campo}': {errores}")
        
        if form.is_valid():
            logger.info(f"✓ Formulario válido para cocata1={cocata1}")
            obj = form.save(commit=False)
            obj.cocata1 = registro.cocata1
            obj.empresa = empresa_codigo
            
            # Obtener usuario del login
            current_user = request.session.get('catastro_usuario_nombre', '') or request.session.get('catastro_usuario_id', '') or ''
            obj.usuario = current_user[:50] if current_user else ''
            obj.fechasys = timezone.now()
            logger.info(f"Usuario asignado: {obj.usuario}, Fecha: {obj.fechasys}")
            
            # Actualizar bvlbas1 con el Valor Total de areasrurales antes de guardar
            try:
                from django.db.models import Sum
                from .models import AreasRurales
                suma_monto = AreasRurales.objects.filter(
                    cocata1=registro.cocata1,
                    empresa=empresa_codigo
                ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
                obj.bvlbas1 = suma_monto
                logger.info(f"bvlbas1 actualizado desde formulario completo para cocata1={registro.cocata1}: {suma_monto}")
            except Exception as e:
                logger.warning(f"Error al actualizar bvlbas1 desde formulario completo: {str(e)}")
                # Si hay error, intentar obtener el valor del POST
                try:
                    bvlbas1_post = request.POST.get('terreno-bvlbas1', '').strip()
                    if bvlbas1_post:
                        obj.bvlbas1 = Decimal(bvlbas1_post.replace(',', '.'))
                except:
                    pass
            
            # Procesar campo baream21 que viene sin prefijo 'terreno-'
            try:
                baream21_post = request.POST.get('baream21', '').strip()
                if baream21_post:
                    obj.baream21 = Decimal(baream21_post.replace(',', '.'))
                    logger.info(f"baream21 actualizado desde formulario para cocata1={registro.cocata1}: {obj.baream21}")
            except Exception as e:
                    logger.warning(f"Error al actualizar baream21 desde formulario: {str(e)}")
            
            logger.info(f"Guardando objeto BDTerreno para cocata1={cocata1}...")
            logger.info(f"Valores antes de guardar - bvlbas1: {obj.bvlbas1}, fcarea: {obj.fcarea}, fcubic: {obj.fcubic}, bfacmodi: {obj.bfacmodi}")
            try:
                obj.save()
                logger.info(f"✓ Objeto BDTerreno guardado exitosamente. ID: {obj.id}")
            except Exception as e:
                logger.error(f"✗ Error al guardar objeto BDTerreno: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                messages.error(request, f'Error al guardar el avalúo: {str(e)}')
                return redirect('catastro:terreno_rural_form', cocata1=cocata1)
            
            # Procesar datos de clasificación del suelo y guardar en AreasRurales
            try:
                from .models import AreasRurales, FactoresRiego
                from decimal import InvalidOperation
                
                # Obtener datos de clasificación del suelo del POST
                codfacs = request.POST.getlist('suelo-codfac[]')
                facs = request.POST.getlist('suelo-fac[]')
                facareas = request.POST.getlist('suelo-facarea[]')
                montos = request.POST.getlist('suelo-monto[]')
                
                # Procesar la primera fila (ya que AreasRurales solo permite un registro por cocata1)
                if codfacs and len(codfacs) > 0:
                    codfac_valor = codfacs[0].strip() if codfacs[0] else ''
                    fac_codigo = facs[0].strip() if facs and len(facs) > 0 and facs[0] else ''
                    facarea = facareas[0] if facareas and len(facareas) > 0 else '0.00'
                    monto = montos[0] if montos and len(montos) > 0 else '0.00'
                    
                    # Solo guardar si hay datos válidos
                    if codfac_valor or fac_codigo or (facarea and facarea != '0.00') or (monto and monto != '0.00'):
                        # Obtener el valor numérico del factor de riego
                        fac_valor = Decimal('0.000')
                        if codfac_valor:
                            try:
                                fac_valor = Decimal(codfac_valor)
                            except (ValueError, InvalidOperation):
                                if fac_codigo:
                                    try:
                                        factor_riego_obj = FactoresRiego.objects.get(codigo=fac_codigo, empresa=empresa_codigo)
                                        fac_valor = factor_riego_obj.valor
                                    except FactoresRiego.DoesNotExist:
                                        pass
                        elif fac_codigo:
                            try:
                                factor_riego_obj = FactoresRiego.objects.get(codigo=fac_codigo, empresa=empresa_codigo)
                                fac_valor = factor_riego_obj.valor
                            except FactoresRiego.DoesNotExist:
                                pass
                        
                        # Crear NUEVO registro en AreasRurales (siempre agregar, nunca actualizar)
                        AreasRurales.objects.create(
                            empresa=empresa_codigo,
                            cocata1=cocata1,
                            fac=fac_valor,
                            codfac=fac_codigo if fac_codigo else None,
                            facarea=Decimal(facarea) if facarea else Decimal('0.00'),
                            monto=Decimal(monto) if monto else Decimal('0.00'),
                            usuario=current_user[:50] if current_user else '',
                            fechasys=timezone.now()
                        )
                        logger.info(f"AreasRurales creado desde formulario completo para cocata1={cocata1}")
            except Exception as e:
                logger.error(f"Error al guardar AreasRurales desde formulario: {str(e)}")
                # No interrumpir el flujo si hay error, solo loguear
            
            # Actualizar valor de tierra en registro principal
            # Primero intentar usar el valor calculado del formulario (valor_terreno)
            try:
                valor_terreno_form = request.POST.get('valor_terreno', '').strip()
                if valor_terreno_form:
                    try:
                        valor_terreno_decimal = Decimal(valor_terreno_form.replace(',', '.'))
                        registro.bvl2tie = valor_terreno_decimal
                        logger.info(f"Valor de terreno actualizado desde formulario para cocata1={cocata1}: {valor_terreno_decimal}")
                    except (ValueError, InvalidOperation) as e:
                        logger.warning(f"Error al parsear valor_terreno del formulario: {str(e)}")
                        # Si falla, intentar calcular con el método del modelo
                        try:
                            valor_terreno = obj.calcular_valor_terreno()
                            if valor_terreno != (registro.bvl2tie or Decimal('0.00')):
                                registro.bvl2tie = valor_terreno
                        except Exception as e2:
                            logger.warning(f"No se pudo calcular el valor del terreno: {str(e2)}")
                else:
                    # Si no viene valor_terreno del formulario, calcular con el método del modelo
                    try:
                        valor_terreno = obj.calcular_valor_terreno()
                        if valor_terreno != (registro.bvl2tie or Decimal('0.00')):
                            registro.bvl2tie = valor_terreno
                    except Exception as e:
                        logger.warning(f"No se pudo calcular el valor del terreno: {str(e)}")
                
                # Recalcular el impuesto usando la misma lógica del frontend
                impuesto_calculado = calcular_impuesto_bdcata1(registro, empresa_codigo)
                
                # Actualizar el campo impuesto en bdcata1
                registro.impuesto = impuesto_calculado
                
                # Guardar ambos campos en una sola operación
                registro.save(update_fields=['bvl2tie', 'impuesto'])
                logger.info(f'Impuesto calculado y actualizado en bdcata1 después de guardar terreno rural: {impuesto_calculado} para clave {cocata1}')
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro, empresa_codigo)
            except Exception as e:
                logger.error(f"Error general al actualizar valor del terreno e impuesto: {str(e)}", exc_info=True)
            
            # bvlbas1 ya fue actualizado antes de guardar obj, así que no es necesario actualizarlo nuevamente aquí
            # El objeto obj ya tiene el valor correcto de bvlbas1
            
            messages.success(request, 'Avalúo de Terreno Rural guardado correctamente.')
            logger.info(f"Avalúo de terreno rural guardado exitosamente para cocata1={cocata1}")
            # Redirigir al formulario principal manteniendo el código catastral
            return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1}")
        else:
            logger.error(f"Formulario inválido para cocata1={cocata1}. Errores: {form.errors}")
            logger.error(f"Datos del formulario: {form.data}")
            messages.error(request, f'Error al guardar el avalúo. Por favor, revise los datos. Errores: {form.errors}')
    else:
        form = BDTerrenoForm(instance=terreno, prefix='terreno')
    
    # Obtener datos registrados en AreasRurales filtrados por empresa y cocata1
    areas_rurales_registradas = []
    clasificaciones_suelo = []  # Datos para la tabla de clasificación del suelo
    factores_riego = []  # Factores de riego para el combobox
    suma_monto_areas_rurales = Decimal('0.00')  # Sumatoria del campo monto
    
    try:
        from .models import AreasRurales, FactoresRiego
        from django.db.models import Sum
        
        # Filtrar por empresa y cocata1
        empresa_filtro = registro.empresa or empresa_codigo
        
        # Obtener factores de riego para el combobox
        try:
            factores_riego = list(FactoresRiego.objects.filter(empresa=empresa_filtro).order_by('codigo'))
        except Exception as e:
            logger.warning(f"Error al obtener FactoresRiego: {str(e)}")
            factores_riego = []
        
        # Obtener datos de AreasRurales para el grid y la tabla de clasificación
        areas_rurales_list = AreasRurales.objects.filter(
            cocata1=cocata1,
            empresa=empresa_filtro
        ).order_by('-fechasys')
        
        # Preparar datos para el grid y la tabla de clasificación
        for area in areas_rurales_list:
            # Obtener descripción del factor de riego si existe codfac
            descripcion_factor = ''
            if area.codfac:
                try:
                    factor = FactoresRiego.objects.filter(
                        codigo=area.codfac,
                        empresa=empresa_filtro
                    ).first()
                    if factor:
                        descripcion_factor = factor.descripcion
                except:
                    pass
            
            area_data = {
                'id': area.id,
                'empresa': area.empresa,
                'cocata1': area.cocata1,
                'fac': area.fac,
                'fac_formatted': f"{area.fac:.3f}" if area.fac else "0.000",
                'codfac': area.codfac,
                'codigo_factor': area.codfac,  # Alias para compatibilidad con template
                'descripcion_factor': descripcion_factor,
                'facarea': area.facarea,
                'monto': area.monto,
                'usuario': area.usuario,
                'fechasys': area.fechasys,
            }
            
            areas_rurales_registradas.append(area_data)
        
        # Calcular la sumatoria del campo monto de areasrurales filtrado por empresa y cocata1
        try:
            suma_monto_areas_rurales = AreasRurales.objects.filter(
                cocata1=registro.cocata1,
                empresa=empresa_codigo
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            logger.info(f"Sumatoria de monto calculada para cocata1={registro.cocata1}, empresa={empresa_codigo}: {suma_monto_areas_rurales}")
        except Exception as e:
            logger.error(f"Error al calcular sumatoria de monto: {str(e)}")
            suma_monto_areas_rurales = Decimal('0.00')
        
        # Ya no se necesita clasificaciones_suelo porque ahora es una fila única para entrada
        # Se mantiene vacío para compatibilidad con el template
        clasificaciones_suelo = []
            
    except Exception as e:
        logger.error(f"Error al obtener AreasRurales: {str(e)}")
        areas_rurales_registradas = []
        clasificaciones_suelo = []
        suma_monto_areas_rurales = Decimal('0.00')
    
    context = {
        'registro': registro,
        'form': form,
        'terreno': terreno,  # Objeto BDTerreno para acceder a usuario y fechasys
        'titulo': f'Avalúo de Terreno Rural - {registro.cocata1}',
        'cocata1': registro.cocata1,
        'empresa': registro.empresa or empresa_codigo,
        'es_rural': registro.ficha and int(registro.ficha) == 2,
        'usuario_nombre': request.session.get('catastro_usuario_nombre', ''),
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
        'areas_rurales_registradas': areas_rurales_registradas,  # Datos para el grid
        'clasificaciones_suelo': clasificaciones_suelo,  # Datos para la tabla de clasificación
        'factores_riego': factores_riego,  # Factores de riego para el combobox
        'suma_monto_areas_rurales': suma_monto_areas_rurales,  # Sumatoria del campo monto
    }
    return render(request, 'terreno_rural_form.html', context)

def guardar_areas_rurales(request):
    """
    Vista AJAX para guardar datos de CLASIFICACION DEL SUELO en la tabla areasrurales.
    Toma los datos de la primera fila de clasificación del suelo y los guarda en AreasRurales.
    """
    from decimal import InvalidOperation
    
    # Verificar autenticación usando sesiones de catastro (no request.user)
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    usuario_id = request.session.get('catastro_usuario_id')
    
    logger.info(f"guardar_areas_rurales llamada - Método: {request.method}, Empresa: {empresa_codigo}, Usuario ID: {usuario_id}")
    
    # Verificar autenticación para solicitudes AJAX usando sesiones de catastro
    if not empresa_codigo or not usuario_id:
        logger.warning(f"Usuario no autenticado en guardar_areas_rurales - Empresa: {empresa_codigo}, Usuario: {usuario_id}")
        # Para solicitudes AJAX, devolver JSON en lugar de redirigir
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autenticado. Por favor, inicie sesión en catastro.'}, status=401)
        # Para solicitudes normales, redirigir al login de catastro
        return redirect('catastro:catastro_login')
    
    if request.method != 'POST':
        logger.warning(f"Método no permitido en guardar_areas_rurales: {request.method}")
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        cocata1 = request.POST.get('cocata1', '').strip()
        # Usar la empresa de la sesión de catastro (ya verificada arriba)
        empresa_login = empresa_codigo
        
        if not cocata1:
            return JsonResponse({'success': False, 'error': 'Clave catastral no proporcionada'}, status=400)
        
        # Obtener datos de TODAS las filas de clasificación del suelo
        # Mapeo de campos:
        # - suelo-codfac[] = Factor Riego (valor numérico) -> fac
        # - suelo-fac[] = Codigo (código del factor de riego) -> codfac
        # - suelo-facarea[] = AREA(HAS) -> facarea
        # - suelo-monto[] = Monto -> monto
        codfacs = request.POST.getlist('suelo-codfac[]')
        facs = request.POST.getlist('suelo-fac[]')
        facareas = request.POST.getlist('suelo-facarea[]')
        montos = request.POST.getlist('suelo-monto[]')
        
        if not codfacs or len(codfacs) == 0:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron datos de clasificación del suelo'}, status=400)
        
        # Procesar cada fila y crear nuevos registros
        registros_creados = []
        registros_con_error = []
        
        for i in range(len(codfacs)):
            codfac_valor = codfacs[i].strip() if i < len(codfacs) and codfacs[i] else ''
            fac_codigo = facs[i].strip() if i < len(facs) and facs[i] else ''
            facarea = facareas[i] if i < len(facareas) and facareas[i] else '0.00'
            monto = montos[i] if i < len(montos) and montos[i] else '0.00'
            
            # Validar que la fila tenga al menos un dato
            if not codfac_valor and not fac_codigo and (not facarea or facarea == '0.00') and (not monto or monto == '0.00'):
                continue  # Saltar filas vacías
            
            # Obtener el valor numérico del factor de riego (fac)
            # Prioridad: 1) valor del campo Factor Riego, 2) valor del código seleccionado en combobox
            fac_valor = Decimal('0.000')
            if codfac_valor:
                try:
                    fac_valor = Decimal(codfac_valor)
                except (ValueError, InvalidOperation):
                    # Si no es válido, intentar obtener el valor del código seleccionado
                    if fac_codigo:
                        try:
                            factor_riego_obj = FactoresRiego.objects.get(codigo=fac_codigo, empresa=empresa_login)
                            fac_valor = factor_riego_obj.valor
                        except FactoresRiego.DoesNotExist:
                            pass
            elif fac_codigo:
                try:
                    factor_riego_obj = FactoresRiego.objects.get(codigo=fac_codigo, empresa=empresa_login)
                    fac_valor = factor_riego_obj.valor
                except FactoresRiego.DoesNotExist:
                    pass
            
            # Crear NUEVO registro en AreasRurales (siempre agregar, nunca actualizar)
            # Mapeo final:
            # - empresa = Municipio (del login)
            # - cocata1 = Clave Catastral
            # - fac = Factor Riego (valor numérico)
            # - codfac = Codigo (código del factor de riego)
            # - facarea = AREA(HAS)
            # - monto = Monto
            # - usuario = login de usuario
            # - fechasys = fecha actual al grabar
            try:
                area_rural = AreasRurales.objects.create(
                    empresa=empresa_login,  # Municipio (del login)
                    cocata1=cocata1,  # Clave Catastral
                    fac=fac_valor,  # Factor Riego (valor numérico)
                    codfac=fac_codigo if fac_codigo else None,  # Codigo (código del factor de riego)
                    facarea=Decimal(facarea) if facarea else Decimal('0.00'),  # AREA(HAS)
                    monto=Decimal(monto) if monto else Decimal('0.00'),  # Monto
                    usuario=request.session.get('catastro_usuario_nombre', '') or request.session.get('catastro_usuario_id', ''),  # login de usuario desde sesión
                    fechasys=timezone.now()  # fecha actual al grabar
                )
                registros_creados.append(area_rural.id)
                logger.info(f"AreasRurales creado (ID: {area_rural.id}) para cocata1={cocata1}, empresa={empresa_login}")
            except Exception as e:
                logger.error(f"Error al crear AreasRurales para fila {i+1}: {str(e)}")
                registros_con_error.append(f"Fila {i+1}: {str(e)}")
        
        if len(registros_creados) == 0:
            return JsonResponse({
                'success': False,
                'error': 'No se pudo crear ningún registro. Verifique que los datos sean válidos.'
            }, status=400)
        
        # Calcular la sumatoria del campo monto de todos los registros en areasrurales
        # filtrados por empresa y cocata1, y actualizar bvlbas1 en bdterreno
        try:
            from .models import BDTerreno
            from django.db.models import Sum
            
            # Calcular la sumatoria de monto
            suma_monto = AreasRurales.objects.filter(
                cocata1=cocata1,
                empresa=empresa_login
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            
            # Obtener o crear el registro de BDTerreno
            try:
                terreno = BDTerreno.objects.get(cocata1=cocata1)
            except BDTerreno.DoesNotExist:
                terreno = BDTerreno(cocata1=cocata1, empresa=empresa_login)
                # Guardar primero el registro recién creado
                terreno.save()
                logger.info(f"BDTerreno creado para cocata1={cocata1}, empresa={empresa_login}")
            
            # Actualizar el campo bvlbas1 con la sumatoria
            terreno.bvlbas1 = suma_monto
            terreno.save(update_fields=['bvlbas1'])
            
            logger.info(f"bvlbas1 actualizado en BDTerreno para cocata1={cocata1}: {suma_monto}")
            
        except Exception as e:
            logger.error(f"Error al actualizar bvlbas1 en BDTerreno: {str(e)}")
            # No interrumpir el flujo si hay error, solo loguear
        
        mensaje = f'Se crearon {len(registros_creados)} registro(s) exitosamente en Areas Rurales'
        if registros_con_error:
            mensaje += f'. Errores: {", ".join(registros_con_error)}'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'created_count': len(registros_creados),
            'ids': registros_creados,
            'errors': registros_con_error,
            'suma_monto': str(suma_monto) if 'suma_monto' in locals() else '0.00'
        })
        
    except Exception as e:
        import traceback
        logger.error(f"Error al guardar en AreasRurales: {str(e)}\n{traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': f'Error al guardar: {str(e)}'
        }, status=500)

def eliminar_avaluo_terreno(request):
    """
    Vista AJAX para eliminar un avalúo de terreno (BDTerreno).
    """
    # Verificar autenticación usando sesiones de catastro
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    usuario_id = request.session.get('catastro_usuario_id') or request.session.get('usuario_id')
    
    if not empresa_codigo or not usuario_id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autenticado'}, status=401)
        return redirect('catastro:catastro_login')
    
    if request.method != 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
        # Intentar obtener cocata1 de POST o GET
        cocata1_param = request.POST.get('cocata1', '') or request.GET.get('cocata1', '')
        if cocata1_param:
            return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={cocata1_param}")
        return redirect('catastro:bienes_inmuebles_registrar')
    
    try:
        from .models import BDTerreno
        import json
        
        data = json.loads(request.body)
        cocata1 = data.get('cocata1', '').strip()
        
        if not cocata1:
            return JsonResponse({'success': False, 'error': 'Clave catastral no proporcionada'}, status=400)
        
        # Buscar y eliminar el avalúo del terreno
        try:
            terreno = BDTerreno.objects.get(cocata1=cocata1, empresa=empresa_codigo)
            terreno.delete()
            logger.info(f"Avalúo de terreno eliminado para cocata1={cocata1}, empresa={empresa_codigo}")
            
            # Eliminar registros de AreasRurales relacionados
            try:
                from .models import AreasRurales
                areas_eliminadas = AreasRurales.objects.filter(cocata1=cocata1, empresa=empresa_codigo).delete()
                logger.info(f"Registros de AreasRurales eliminados para cocata1={cocata1}, empresa={empresa_codigo}. Total eliminado: {areas_eliminadas[0]}")
            except Exception as e:
                logger.warning(f"Error al eliminar registros de AreasRurales: {str(e)}")
            
            # Actualizar bvl2tie a 0 y recalcular el impuesto en bdcata1
            try:
                from .models import BDCata1
                registro_bdcata1 = BDCata1.objects.get(cocata1=cocata1, empresa=empresa_codigo)
                
                # Establecer bvl2tie a 0
                registro_bdcata1.bvl2tie = Decimal('0.00')
                
                # Recalcular el impuesto usando la misma lógica del frontend
                impuesto_calculado = calcular_impuesto_bdcata1(registro_bdcata1, empresa_codigo)
                
                # Actualizar el campo impuesto en bdcata1
                registro_bdcata1.impuesto = impuesto_calculado
                
                # Guardar ambos campos en una sola operación
                registro_bdcata1.save(update_fields=['bvl2tie', 'impuesto'])
                logger.info(f'Campo bvl2tie actualizado a 0 e impuesto recalculado en bdcata1 después de eliminar terreno rural: {impuesto_calculado} para clave {cocata1}')
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro_bdcata1, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro_bdcata1, empresa_codigo)
            except BDCata1.DoesNotExist:
                logger.warning(f"No se encontró registro en bdcata1 para clave={cocata1}, empresa={empresa_codigo} al eliminar terreno")
            except Exception as e:
                logger.error(f"Error al actualizar bvl2tie e impuesto después de eliminar terreno: {str(e)}", exc_info=True)
            
            return JsonResponse({
                'success': True,
                'message': f'Avalúo del terreno eliminado correctamente para la clave catastral {cocata1}'
            })
        except BDTerreno.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'No se encontró un avalúo de terreno para la clave catastral {cocata1}'
            }, status=404)
        except Exception as e:
            logger.error(f"Error al eliminar avalúo de terreno: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Error al eliminar el avalúo: {str(e)}'
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        logger.error(f"Error inesperado al eliminar avalúo de terreno: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error inesperado: {str(e)}'
        }, status=500)

@catastro_require_auth
def edificaciones_form(request, clave: str):
    """
    Formulario para gestionar edificaciones vinculadas a una clave catastral.
    Permite crear, editar y listar edificaciones.
    """
    # Obtener empresa del login
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    usuario_nombre = request.session.get('catastro_usuario_nombre', '') or request.session.get('catastro_usuario_id', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    # Verificar que existe el registro de bien inmueble
    try:
        registro = BDCata1.objects.get(cocata1=clave, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave}")
    
    # Obtener todas las edificaciones para esta clave y empresa
    edificaciones = Edificacion.objects.filter(clave=clave, empresa=empresa_codigo).order_by('edifino', 'piso')
    
    # Agrupar edificaciones por edifino para mostrar los pisos desglosados
    edificaciones_por_edifino = {}
    for edificacion in edificaciones:
        edifino_key = str(edificacion.edifino) if edificacion.edifino else '0'
        if edifino_key not in edificaciones_por_edifino:
            edificaciones_por_edifino[edifino_key] = {
                'edifino': edificacion.edifino,
                'pisos': [],
                'total_edificacion': Decimal('0.00'),
                'total_area': Decimal('0.00'),
            }
        edificaciones_por_edifino[edifino_key]['pisos'].append(edificacion)
        edificaciones_por_edifino[edifino_key]['total_edificacion'] += edificacion.totedi or Decimal('0.00')
        edificaciones_por_edifino[edifino_key]['total_area'] += edificacion.area or Decimal('0.00')
    
    # Convertir a lista ordenada por edifino para facilitar el renderizado en el template
    edificaciones_por_edifino_lista = sorted(
        edificaciones_por_edifino.items(),
        key=lambda x: int(x[1]['edifino']) if x[1]['edifino'] and str(x[1]['edifino']).isdigit() else 0
    )
    
    # Calcular totales generales para el resumen
    total_general_edificaciones = Decimal('0.00')
    total_area_general = Decimal('0.00')
    for edifino_key, datos_edif in edificaciones_por_edifino_lista:
        total_general_edificaciones += datos_edif['total_edificacion']
        total_area_general += datos_edif['total_area']
    
    # Calcular número de edificaciones (grupos únicos por edifino)
    num_edificaciones = len(edificaciones_por_edifino_lista)
    
    # Si hay un ID de edificación en GET, cargar esa edificación para editar
    edifino_editar = request.GET.get('edifino', None)
    piso_editar = request.GET.get('piso', None)
    edificacion_editar = None
    if edifino_editar:
        try:
            filtro = {'clave': clave, 'edifino': edifino_editar, 'empresa': empresa_codigo}
            if piso_editar:
                filtro['piso'] = piso_editar
            edificacion_editar = Edificacion.objects.get(**filtro)
        except Edificacion.DoesNotExist:
            pass
    
    if request.method == 'POST':
        # Determinar si es crear o actualizar
        edifino_post = request.POST.get('edifino', '').strip()
        piso_post = request.POST.get('piso', '').strip() or None
        
        # Convertir piso a Decimal si tiene valor
        piso_decimal = None
        if piso_post and piso_post.strip():
            try:
                piso_decimal = Decimal(piso_post.strip())
            except (ValueError, TypeError):
                piso_decimal = None
        
        if edifino_post:
            try:
                # Intentar obtener edificación existente
                # Validación según empresa, clave, edifino y piso
                if piso_decimal is not None:
                    # Si piso tiene valor, buscar exactamente ese piso
                    edificacion_existente = Edificacion.objects.get(
                        empresa=empresa_codigo,
                        clave=clave,
                        edifino=edifino_post,
                        piso=piso_decimal
                    )
                else:
                    # Si piso es None o vacío, buscar registros donde piso es NULL
                    edificacion_existente = Edificacion.objects.filter(
                        empresa=empresa_codigo,
                        clave=clave,
                        edifino=edifino_post
                    ).filter(Q(piso__isnull=True) | Q(piso=0)).get()
                
                form = EdificacionForm(request.POST, instance=edificacion_existente)
                es_nuevo = False
            except Edificacion.DoesNotExist:
                form = EdificacionForm(request.POST)
                es_nuevo = True
            except Edificacion.MultipleObjectsReturned:
                # Si hay múltiples registros, tomar el primero
                if piso_decimal is not None:
                    edificacion_existente = Edificacion.objects.filter(
                        empresa=empresa_codigo,
                        clave=clave,
                        edifino=edifino_post,
                        piso=piso_decimal
                    ).first()
                else:
                    edificacion_existente = Edificacion.objects.filter(
                        empresa=empresa_codigo,
                        clave=clave,
                        edifino=edifino_post
                    ).filter(Q(piso__isnull=True) | Q(piso=0)).first()
                
                if edificacion_existente:
                    form = EdificacionForm(request.POST, instance=edificacion_existente)
                    es_nuevo = False
                else:
                    form = EdificacionForm(request.POST)
                    es_nuevo = True
        else:
            form = EdificacionForm(request.POST)
            es_nuevo = True
        
        if form.is_valid():
            edificacion = form.save(commit=False)
            edificacion.empresa = empresa_codigo
            edificacion.clave = clave
            edificacion.usuario = usuario_nombre[:50] if usuario_nombre else ''
            edificacion.fechasys = timezone.now()
            
            # Calcular totedi = area * costo * (bueno / 100)
            try:
                area_val = Decimal(str(edificacion.area or 0))
                costo_val = Decimal(str(edificacion.costo or 0))
                bueno_val = Decimal(str(edificacion.bueno or 0))
                edificacion.totedi = (area_val * costo_val * (bueno_val / Decimal('100'))).quantize(Decimal('0.01'))
            except Exception as e:
                logger.warning(f"Error al calcular totedi: {str(e)}")
                edificacion.totedi = Decimal('0.00')
            
            edificacion.save()
            
            # Calcular el total general de todas las edificaciones para esta clave y empresa
            try:
                total_general = Edificacion.objects.filter(clave=clave, empresa=empresa_codigo).aggregate(
                    total=Sum('totedi')
                )['total'] or Decimal('0.00')
                
                # Contar el número de edificaciones únicas (por edifino)
                # Una edificación puede tener varios pisos, pero cuenta como una sola edificación
                num_edificaciones = Edificacion.objects.filter(clave=clave, empresa=empresa_codigo).values('edifino').distinct().count()
                
                # Actualizar el campo mejoras y cedif en la tabla bdcata1
                try:
                    registro_bdcata1 = BDCata1.objects.get(cocata1=clave, empresa=empresa_codigo)
                    registro_bdcata1.mejoras = total_general.quantize(Decimal('0.01'))
                    registro_bdcata1.cedif = Decimal(str(num_edificaciones)).quantize(Decimal('1'))
                    
                    # Recalcular el impuesto usando la misma lógica del frontend
                    impuesto_calculado = calcular_impuesto_bdcata1(registro_bdcata1, empresa_codigo)
                    
                    # Actualizar el campo impuesto en bdcata1
                    registro_bdcata1.impuesto = impuesto_calculado
                    
                    # Guardar todos los campos en una sola operación
                    registro_bdcata1.save(update_fields=['mejoras', 'cedif', 'impuesto'])
                    logger.info(f"Campos actualizados en bdcata1 para clave={clave}, empresa={empresa_codigo}, mejoras={total_general}, cedif={num_edificaciones}, impuesto={impuesto_calculado}")
                    
                    # Actualizar el impuesto en tasasmunicipales
                    actualizar_impuesto_tasas_municipales(registro_bdcata1, impuesto_calculado, empresa_codigo)
                    
                    # Calcular tasas municipales (rubros que empiezan con T)
                    calcular_tasas_municipales_automatico(registro_bdcata1, empresa_codigo)
                except BDCata1.DoesNotExist:
                    logger.warning(f"No se encontró registro en bdcata1 para clave={clave}, empresa={empresa_codigo}")
                except Exception as e:
                    logger.error(f"Error al actualizar campos en bdcata1: {str(e)}")
            except Exception as e:
                logger.error(f"Error al calcular total general de edificaciones: {str(e)}")
            
            if es_nuevo:
                messages.success(request, f'Edificación {edificacion.edifino} creada correctamente.')
            else:
                messages.success(request, f'Edificación {edificacion.edifino} actualizada correctamente.')
            
            return redirect('catastro:edificaciones_form', clave=clave)
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        if edificacion_editar:
            form = EdificacionForm(instance=edificacion_editar)
            # Formatear fecha para mostrar en el campo si existe
            if edificacion_editar.fechasys:
                form.fields['fechasys'].initial = edificacion_editar.fechasys.strftime('%Y-%m-%d %H:%M:%S')
            else:
                form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
            # Asegurar que el usuario se muestre
            if edificacion_editar.usuario:
                form.fields['usuario'].initial = edificacion_editar.usuario
            else:
                form.fields['usuario'].initial = usuario_nombre[:50] if usuario_nombre else ''
        else:
            form = EdificacionForm()
            # Establecer valores iniciales para nuevo registro
            form.fields['usuario'].initial = usuario_nombre[:50] if usuario_nombre else ''
            form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Obtener lista de usos de edificación para el datalist
    try:
        from .models import UsoEdifica
        usoedifica_list = UsoEdifica.objects.all().order_by('codigo')
    except Exception:
        usoedifica_list = []
    
    context = {
        'titulo': 'Gestión de Edificaciones',
        'form': form,
        'clave': clave,
        'registro': registro,
        'edificaciones': edificaciones,
        'edificaciones_por_edifino': edificaciones_por_edifino,
        'edificaciones_por_edifino_lista': edificaciones_por_edifino_lista,
        'total_general_edificaciones': total_general_edificaciones,
        'total_area_general': total_area_general,
        'num_edificaciones': num_edificaciones,
        'edificacion_editar': edificacion_editar,
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
        'empresa': empresa_codigo,
        'usoedifica_list': usoedifica_list,
    }
    
    return render(request, 'edificaciones_form.html', context)

@catastro_require_auth
def edificaciones_export_excel(request, clave: str):
    """
    Exportar edificaciones a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:edificaciones_form', clave=clave)
    
    empresa_codigo = request.session.get('catastro_empresa', '') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    # Obtener las edificaciones
    edificaciones = Edificacion.objects.filter(clave=clave, empresa=empresa_codigo).order_by('edifino', 'piso')
    
    # Calcular total general
    suma_total = edificaciones.aggregate(Sum('totedi'))['totedi__sum'] or Decimal('0.00')
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Edificaciones"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Edif. No.', 'Piso', 'Área', 'Uso', 'Clase', 'Calidad', 'Costo', 'Bueno (%)', 'Total', 'Usuario', 'Fecha']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    for edif in edificaciones:
        row = [
            edif.edifino or '',
            str(edif.piso) if edif.piso and edif.piso != 0 else '',
            float(edif.area) if edif.area else 0.00,
            edif.uso or '',
            edif.clase or '',
            edif.calidad or '',
            float(edif.costo) if edif.costo else 0.00,
            int(edif.bueno) if edif.bueno else 0,
            float(edif.totedi) if edif.totedi else 0.00,
            edif.usuario or '',
            edif.fechasys.strftime('%d/%m/%Y %H:%M') if edif.fechasys else ''
        ]
        ws.append(row)
    
    # Aplicar estilo a datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border_style
            if cell.column in [3, 7, 9]:  # Columnas numéricas (Área, Costo, Total)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            elif cell.column == 8:  # Bueno (%)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '0'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Agregar fila de totales
    total_row = [''] * 8 + [float(suma_total)]
    ws.append(total_row)
    total_cell = ws.cell(row=ws.max_row, column=9)
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = '#,##0.00'
    total_cell.border = border_style
    
    # Agregar información adicional
    ws.append([])
    info_row = ['Clave Catastral:', clave, 'Total de Edificaciones:', edificaciones.count()]
    ws.append(info_row)
    
    # Ajustar ancho de columnas
    column_widths = [12, 10, 12, 10, 10, 10, 12, 12, 15, 15, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'edificaciones_{clave}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def edificaciones_export_pdf(request, clave: str):
    """
    Exportar edificaciones a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:edificaciones_form', clave=clave)
    
    empresa_codigo = request.session.get('catastro_empresa', '') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    # Obtener las edificaciones
    edificaciones = Edificacion.objects.filter(clave=clave, empresa=empresa_codigo).order_by('edifino', 'piso')
    
    # Calcular total general
    suma_total = edificaciones.aggregate(Sum('totedi'))['totedi__sum'] or Decimal('0.00')
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = f"Avalúo de Edificaciones<br/>Clave Catastral: {clave}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Preparar datos de la tabla
    data = []
    
    # Encabezados
    headers = ['Edif. No.', 'Piso', 'Área', 'Uso', 'Clase', 'Calidad', 'Costo', 'Bueno (%)', 'Total', 'Usuario', 'Fecha']
    data.append(headers)
    
    # Datos
    for edif in edificaciones:
        row = [
            str(edif.edifino) if edif.edifino else '-',
            str(edif.piso) if edif.piso and edif.piso != 0 else '-',
            f"{edif.area:.2f}" if edif.area else "0.00",
            edif.uso or '-',
            edif.clase or '-',
            edif.calidad or '-',
            f"{edif.costo:.2f}" if edif.costo else "0.00",
            str(int(edif.bueno)) if edif.bueno else '0',
            f"{edif.totedi:.2f}" if edif.totedi else "0.00",
            edif.usuario or '-',
            edif.fechasys.strftime('%d/%m/%Y') if edif.fechasys else '-'
        ]
        data.append(row)
    
    # Fila de totales
    total_row = [''] * 8 + [f"{suma_total:.2f}"] + [''] * 2
    data.append(total_row)
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('ALIGN', (2, 1), (2, -2), 'RIGHT'),  # Área
        ('ALIGN', (6, 1), (8, -2), 'RIGHT'),  # Costo, Bueno, Total
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        
        # Fila de totales
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (7, -1), 'RIGHT'),
        ('ALIGN', (8, -1), (8, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Información adicional
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        alignment=0  # Izquierda
    )
    info_text = f"<b>Total de edificaciones:</b> {edificaciones.count()}<br/>"
    info_text += f"<b>Suma total:</b> L. {suma_total:.2f}<br/>"
    info_text += f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'edificaciones_{clave}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response

@catastro_require_auth
def eliminar_edificacion(request, clave: str):
    """
    Vista para eliminar una edificación
    """
    # Obtener empresa del login
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    # Obtener parámetros de la edificación a eliminar
    edifino = request.GET.get('edifino', None)
    piso = request.GET.get('piso', None)
    
    if not edifino:
        messages.error(request, 'No se especificó la edificación a eliminar.')
        return redirect('catastro:edificaciones_form', clave=clave)
    
    # Convertir piso a Decimal si tiene valor
    piso_decimal = None
    if piso and piso.strip():
        try:
            from decimal import Decimal
            piso_decimal = Decimal(piso.strip())
        except (ValueError, TypeError):
            piso_decimal = None
    
    try:
        # Buscar la edificación
        # Validación según empresa, clave, edifino y piso
        if piso_decimal is not None:
            # Si piso tiene valor, buscar exactamente ese piso
            edificacion = Edificacion.objects.get(
                empresa=empresa_codigo,
                clave=clave,
                edifino=edifino,
                piso=piso_decimal
            )
        else:
            # Si piso es None o vacío, buscar registros donde piso es NULL o 0
            edificacion = Edificacion.objects.filter(
                empresa=empresa_codigo,
                clave=clave,
                edifino=edifino
            ).filter(Q(piso__isnull=True) | Q(piso=0)).get()
        
        # Verificar si es una edificación especial (uso='E')
        es_edificacion_especial = (edificacion.uso == 'E')
        
        # Guardar el valor de totedi antes de eliminar (para edificaciones especiales)
        valor_totedi = edificacion.totedi or Decimal('0.00')
        descripcion_edificacion = edificacion.descripcion or f'Edificación {edifino}'
        
        # Eliminar la edificación
        edificacion.delete()
        
        if es_edificacion_especial:
            messages.success(request, f'Edificación especial eliminada correctamente. Descripción: {descripcion_edificacion}')
        else:
            messages.success(request, f'Edificación {edifino} eliminada correctamente.')
        
        # Recalcular el total general de todas las edificaciones para esta clave y empresa
        try:
            total_general = Edificacion.objects.filter(clave=clave, empresa=empresa_codigo).aggregate(
                total=Sum('totedi')
            )['total'] or Decimal('0.00')
            
            # Contar el número de edificaciones únicas (por edifino)
            # Excluir edificaciones especiales (uso='E') del conteo de cedif
            num_edificaciones = Edificacion.objects.filter(
                clave=clave, 
                empresa=empresa_codigo
            ).exclude(uso='E').values('edifino').distinct().count()
            
            # Actualizar el campo mejoras y cedif en la tabla bdcata1
            try:
                registro_bdcata1 = BDCata1.objects.get(cocata1=clave, empresa=empresa_codigo)
                
                # Usar la sumatoria total de todas las edificaciones restantes (ya calculada arriba)
                # Esto funciona tanto para edificaciones normales como especiales
                registro_bdcata1.mejoras = total_general.quantize(Decimal('0.01'))
                registro_bdcata1.cedif = Decimal(str(num_edificaciones)).quantize(Decimal('1'))
                
                # Recalcular el impuesto usando la misma lógica del frontend
                impuesto_calculado = calcular_impuesto_bdcata1(registro_bdcata1, empresa_codigo)
                
                # Actualizar el campo impuesto en bdcata1
                registro_bdcata1.impuesto = impuesto_calculado
                
                # Guardar todos los campos en una sola operación
                registro_bdcata1.save(update_fields=['mejoras', 'cedif', 'impuesto'])
                logger.info(f"Campos actualizados en bdcata1 después de eliminar edificación: clave={clave}, empresa={empresa_codigo}, mejoras={registro_bdcata1.mejoras}, cedif={num_edificaciones}, impuesto={impuesto_calculado}")
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro_bdcata1, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro_bdcata1, empresa_codigo)
            except BDCata1.DoesNotExist:
                logger.warning(f"No se encontró registro en bdcata1 para clave={clave}, empresa={empresa_codigo}")
            except Exception as e:
                logger.error(f"Error al actualizar campos en bdcata1: {str(e)}")
        except Exception as e:
            logger.error(f"Error al calcular total general de edificaciones: {str(e)}")
        
    except Edificacion.DoesNotExist:
        messages.error(request, 'No se encontró la edificación especificada.')
    except Exception as e:
        logger.error(f"Error al eliminar edificación: {str(e)}")
        messages.error(request, f'Error al eliminar la edificación: {str(e)}')
    
    return redirect('catastro:edificaciones_form', clave=clave)

# ============================================================================
# GESTIÓN DE COSTOS BÁSICOS UNITARIOS
# ============================================================================

@catastro_require_auth
def costos_list(request):
    """
    Lista de costos básicos unitarios agrupados por uso
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    costos = _costos_por_municipio_qs(request).order_by('uso', 'clase', 'calidad')
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        costos = costos.filter(
            Q(uso__icontains=search) |
            Q(clase__icontains=search) |
            Q(calidad__icontains=search)
        )
    
    # Obtener descripciones de usos desde la tabla usoedifica
    from .models import UsoEdifica
    usos_edifica = UsoEdifica.objects.all().order_by('codigo')
    
    # Crear diccionario de descripciones de uso (código de 2 caracteres -> descripción)
    descripciones_uso = {}
    for uso_edifica in usos_edifica:
        codigo_uso = uso_edifica.codigo[:2]  # Tomar primeros 2 caracteres
        if codigo_uso not in descripciones_uso:
            descripciones_uso[codigo_uso] = uso_edifica.descripcion
    
    # Diccionario de descripciones de clase
    descripciones_clase = {
        '1': 'Madera',
        '2': 'Ladrillo o Bloque',
        '3': 'Loza o Terraza',
        '4': 'Bajareque o Adobe',
        '5': 'Acero Estructural',
        '6': 'Panelit',
    }
    
    # Agrupar costos por uso y luego por clase
    costos_por_uso = {}
    for costo in costos:
        uso = costo.uso or ''
        clase = costo.clase or ''
        
        if uso not in costos_por_uso:
            costos_por_uso[uso] = {
                'descripcion': descripciones_uso.get(uso, f'Uso {uso}'),
                'clases': {},
                'suma_costo': Decimal('0.00')
            }
        
        if clase not in costos_por_uso[uso]['clases']:
            costos_por_uso[uso]['clases'][clase] = {
                'descripcion': descripciones_clase.get(clase, f'Clase {clase}'),
                'costos': [],
                'suma_costo': Decimal('0.00')
            }
        
        costos_por_uso[uso]['clases'][clase]['costos'].append(costo)
        costos_por_uso[uso]['clases'][clase]['suma_costo'] += costo.costo or Decimal('0.00')
        costos_por_uso[uso]['suma_costo'] += costo.costo or Decimal('0.00')
    
    # Calcular sumatorias totales
    suma_costo = costos.aggregate(Sum('costo'))['costo__sum'] or Decimal('0.00')
    
    context = {
        'titulo': 'Costos Básicos Unitarios - Catastro',
        'costos_por_uso': costos_por_uso,
        'costos': costos,  # Mantener para compatibilidad
        'search': search,
        'total_registros': costos.count(),
        'suma_costo': suma_costo,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'costos_list.html', context)

@catastro_require_auth
def costo_create(request):
    """
    Crear nuevo costo básico unitario
    Si ya existe un registro con la misma combinación (empresa, uso, clase, calidad),
    muestra los datos existentes para permitir modificarlos
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        editar_existente = request.POST.get('editar_existente', '') == '1'
        
        # Obtener valores del formulario
        uso = request.POST.get('uso', '').strip()
        clase = request.POST.get('clase', '').strip()
        calidad = request.POST.get('calidad', '').strip()
        
        form = CostosForm(request.POST)
        
        if form.is_valid():
            # Verificar si ya existe un registro con la misma combinación
            if not editar_existente:
                emp_dup = codigos_empresa_equivalentes(empresa_codigo) or ([empresa_codigo] if empresa_codigo else [])
                costo_duplicado = Costos.objects.filter(
                    empresa__in=emp_dup,
                    uso=uso[:2] if len(uso) >= 2 else uso,
                    clase=clase[:1] if len(clase) >= 1 else clase,
                    calidad=calidad[:3] if len(calidad) >= 3 else calidad
                ).first() if emp_dup else None
                
                if costo_duplicado:
                    # Si existe, actualizar el registro existente con los nuevos valores
                    costo_duplicado.costo = form.cleaned_data.get('costo', costo_duplicado.costo)
                    costo_duplicado.rango1 = form.cleaned_data.get('rango1', costo_duplicado.rango1)
                    costo_duplicado.rango2 = form.cleaned_data.get('rango2', costo_duplicado.rango2)
                    
                    try:
                        costo_duplicado.save()
                        messages.success(request, 'Costo básico unitario actualizado exitosamente.')
                        return redirect('catastro:costos_list')
                    except Exception as e:
                        error_msg = str(e)
                        messages.error(request, f'Error al actualizar: {error_msg}')
                        context = {
                            'titulo': 'Editar Costo Básico Unitario Existente - Catastro',
                            'form': form,
                            'costo': costo_duplicado,
                            'empresa': empresa_codigo,
                        }
                        return render(request, 'costo_form.html', context)
            
            # Si no hay duplicado, crear nuevo registro
            costo = form.save(commit=False)
            costo.empresa = empresa_codigo
            
            try:
                costo.save()
                messages.success(request, 'Costo básico unitario creado exitosamente.')
                return redirect('catastro:costos_list')
            except Exception as e:
                error_msg = str(e)
                if 'Duplicate entry' in error_msg or 'UNIQUE constraint' in error_msg:
                    messages.error(request, 'Ya existe un costo con esta combinación de Uso, Clase y Calidad.')
                else:
                    messages.error(request, f'Error al guardar: {error_msg}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CostosForm()
    
    context = {
        'titulo': 'Nuevo Costo Básico Unitario - Catastro',
        'form': form,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'costo_form.html', context)

@catastro_require_auth
def costo_update(request, pk):
    """
    Actualizar costo básico unitario
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    costo = get_object_or_404(_costos_por_municipio_qs(request), pk=pk)
    
    if request.method == 'POST':
        form = CostosForm(request.POST, instance=costo)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Costo básico unitario actualizado exitosamente.')
                return redirect('catastro:costos_list')
            except Exception as e:
                error_msg = str(e)
                if 'Duplicate entry' in error_msg or 'UNIQUE constraint' in error_msg:
                    messages.error(request, 'Ya existe un costo con esta combinación de Uso, Clase y Calidad.')
                else:
                    messages.error(request, f'Error al actualizar: {error_msg}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = CostosForm(instance=costo)
    
    context = {
        'titulo': f'Editar Costo Básico Unitario - Catastro',
        'form': form,
        'costo': costo,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'costo_form.html', context)

@catastro_require_auth
def costos_clasificacion_pesos(request, uso, clase):
    """
    Vista para la clasificación de pesos por uso y clase
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    try:
        
        # Costos del municipio en sesión para este uso y clase
        costos = _costos_por_municipio_qs(request).filter(
            uso=uso,
            clase=clase
        ).order_by('calidad')
        
        # Obtener descripción del uso
        from .models import UsoEdifica
        uso_obj = UsoEdifica.objects.filter(codigo__startswith=uso).first()
        descripcion_uso = uso_obj.descripcion if uso_obj else f'Uso {uso}'
        
        # Diccionario de descripciones de clase
        descripciones_clase = {
            '1': 'Madera',
            '2': 'Ladrillo o Bloque',
            '3': 'Loza o Terraza',
            '4': 'Bajareque o Adobe',
            '5': 'Acero Estructural',
            '6': 'Panelit',
        }
        descripcion_clase = descripciones_clase.get(clase, f'Clase {clase}')
        
        # Obtener tipos de material para el combobox
        tipos_material = TipoMaterial.objects.all().order_by('No')
        
        # Crear diccionario de descripciones de tipos de material
        tipos_material_dict = {tm.No: tm.descripcion for tm in tipos_material}
        
        # Obtener configuraciones de tipología para este uso y clase (FILTRADO)
        # Ordenar numéricamente por tipo (para que 10 venga después de 9)
        # Coalesce maneja valores NULL convirtiéndolos a '0' antes de convertir a entero
        configuraciones_tipologia = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase
        ).annotate(
            tipo_numeric=Cast(Coalesce('tipo', Value('0')), IntegerField())
        ).order_by('tipo_numeric', 'categoria')
        
        # Agrupar configuraciones por tipo (clasificación)
        configuraciones_por_tipo = {}
        for confi in configuraciones_tipologia:
            tipo_key = confi.tipo or 'Sin Clasificación'
            if tipo_key not in configuraciones_por_tipo:
                configuraciones_por_tipo[tipo_key] = {
                    'descripcion': tipos_material_dict.get(confi.tipo, f'Clasificación {confi.tipo}') if confi.tipo else 'Sin Clasificación',
                    'tipo_numeric': confi.tipo_numeric if hasattr(confi, 'tipo_numeric') else (int(confi.tipo) if confi.tipo and confi.tipo.isdigit() else 0),
                    'registros': []
                }
            configuraciones_por_tipo[tipo_key]['registros'].append(confi)
        
        # Ordenar las clasificaciones numéricamente
        configuraciones_por_tipo_ordenado = dict(sorted(
            configuraciones_por_tipo.items(),
            key=lambda x: (x[1]['tipo_numeric'] if x[0] != 'Sin Clasificación' else 999999, x[0])
        ))
        
        # Obtener categorías únicas con sus datos
        categorias_dict = {}
        for confi in configuraciones_tipologia:
            if confi.categoria and confi.categoria not in categorias_dict:
                categorias_dict[confi.categoria] = {
                    'descripcion': confi.descripcion or '',
                    'peso': confi.peso or 0,
                    'tipo': confi.tipo or '',
                }
        
        context = {
            'titulo': 'Clasificación Pesos - Catastro',
            'uso': uso,
            'clase': clase,
            'descripcion_uso': descripcion_uso,
            'descripcion_clase': descripcion_clase,
            'costos': costos,
            'empresa_codigo': empresa_codigo,
            'tipos_material': tipos_material,
            'configuraciones_tipologia': configuraciones_tipologia,
            'configuraciones_por_tipo': configuraciones_por_tipo_ordenado,
            'tipos_material_dict': tipos_material_dict,
            'categorias_dict': categorias_dict,
            'tipos_disponibles': ConfiTipologia.objects.filter(uso=uso, clase=clase).values_list('tipo', flat=True).distinct().exclude(tipo__isnull=True).exclude(tipo=''),
        }
        return render(request, 'costos_clasificacion_pesos.html', context)
    except Exception as e:
        logger.error(f"Error al cargar clasificación de pesos: {str(e)}", exc_info=True)
        messages.error(request, f'Error al cargar clasificación de pesos: {str(e)}')
        return redirect('catastro:costos_list')

@catastro_require_auth
def costo_delete(request, pk):
    """
    Eliminar costo básico unitario
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    costo = get_object_or_404(_costos_por_municipio_qs(request), pk=pk)
    
    if request.method == 'POST':
        costo.delete()
        messages.success(request, 'Costo básico unitario eliminado exitosamente.')
        return redirect('catastro:costos_list')
    
    context = {
        'titulo': f'Eliminar Costo Básico Unitario - Catastro',
        'costo': costo,
    }
    
    return render(request, 'costo_confirm_delete.html', context)

@catastro_require_auth
def costos_export_excel(request):
    """
    Exportar costos básicos unitarios a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:costos_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    costos = _costos_por_municipio_qs(request).order_by('uso', 'clase', 'calidad')
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        costos = costos.filter(
            Q(uso__icontains=search) |
            Q(clase__icontains=search) |
            Q(calidad__icontains=search)
        )
    
    # Obtener descripciones de usos
    from .models import UsoEdifica
    usos_edifica = UsoEdifica.objects.all().order_by('codigo')
    descripciones_uso = {}
    for uso_edifica in usos_edifica:
        codigo_uso = uso_edifica.codigo[:2]
        if codigo_uso not in descripciones_uso:
            descripciones_uso[codigo_uso] = uso_edifica.descripcion
    
    # Diccionario de descripciones de clase
    descripciones_clase = {
        '1': 'Madera',
        '2': 'Ladrillo o Bloque',
        '3': 'Loza o Terraza',
        '4': 'Bajareque o Adobe',
        '5': 'Acero Estructural',
        '6': 'Panelit',
    }
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costos Básicos Unitarios"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    uso_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    clase_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.append(['COSTOS BÁSICOS UNITARIOS'])
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Encabezados
    headers = ['ID', 'Uso', 'Descripción Uso', 'Clase', 'Descripción Clase', 'Calidad', 'Costo', 'Rango 1', 'Rango 2']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agrupar y agregar datos
    current_uso = None
    current_clase = None
    row_num = 4
    suma_total = Decimal('0.00')
    
    for costo in costos:
        uso = costo.uso or ''
        clase = costo.clase or ''
        
        # Agregar fila de uso si cambia
        if uso != current_uso:
            if current_uso is not None:
                ws.append([])  # Línea en blanco entre usos
                row_num += 1
            uso_desc = descripciones_uso.get(uso, f'Uso {uso}')
            uso_row = ['USO ' + uso, uso_desc, '', '', '', '', '', '', '']
            ws.append(uso_row)
            row_num += 1
            # Aplicar estilo a fila de uso
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = uso_fill
                cell.font = Font(bold=True, size=11)
                cell.border = border_style
                if col <= 2:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            current_uso = uso
            current_clase = None
        
        # Agregar fila de clase si cambia
        if clase != current_clase:
            if current_clase is not None:
                ws.append([])  # Línea en blanco entre clases
                row_num += 1
            clase_desc = descripciones_clase.get(clase, f'Clase {clase}')
            clase_row = ['', '', 'CLASE ' + clase, clase_desc, '', '', '', '', '']
            ws.append(clase_row)
            row_num += 1
            # Aplicar estilo a fila de clase
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = clase_fill
                cell.font = Font(bold=True, size=10)
                cell.border = border_style
                if col in [3, 4]:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            current_clase = clase
        
        # Agregar fila de costo
        row = [
            costo.id,
            '',
            '',
            '',
            '',
            costo.calidad or '',
            float(costo.costo) if costo.costo else 0.00,
            int(costo.rango1) if costo.rango1 else 0,
            int(costo.rango2) if costo.rango2 else 0
        ]
        ws.append(row)
        row_num += 1
        suma_total += costo.costo or Decimal('0.00')
        
        # Aplicar estilo a fila de datos
        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border_style
            if col == 1:  # ID
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 6:  # Calidad
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col in [7, 8, 9]:  # Costo, Rango 1, Rango 2
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if col == 7:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Agregar fila de totales
    ws.append([])
    row_num += 1
    total_row = ['', '', '', '', '', 'TOTAL:', float(suma_total), '', '']
    ws.append(total_row)
    row_num += 1
    
    # Aplicar estilo a fila de totales
    for col in range(1, 10):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border_style
        if col == 6:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif col == 7:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '#,##0.00'
    
    # Ajustar ancho de columnas
    column_widths = [8, 8, 30, 8, 25, 10, 15, 12, 12]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'costos_basicos_unitarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def costos_export_pdf(request):
    """
    Exportar costos básicos unitarios a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:costos_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    costos = _costos_por_municipio_qs(request).order_by('uso', 'clase', 'calidad')
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        costos = costos.filter(
            Q(uso__icontains=search) |
            Q(clase__icontains=search) |
            Q(calidad__icontains=search)
        )
    
    # Obtener descripciones de usos
    from .models import UsoEdifica
    usos_edifica = UsoEdifica.objects.all().order_by('codigo')
    descripciones_uso = {}
    for uso_edifica in usos_edifica:
        codigo_uso = uso_edifica.codigo[:2]
        if codigo_uso not in descripciones_uso:
            descripciones_uso[codigo_uso] = uso_edifica.descripcion
    
    # Diccionario de descripciones de clase
    descripciones_clase = {
        '1': 'Madera',
        '2': 'Ladrillo o Bloque',
        '3': 'Loza o Terraza',
        '4': 'Bajareque o Adobe',
        '5': 'Acero Estructural',
        '6': 'Panelit',
    }
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = "COSTOS BÁSICOS UNITARIOS"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Agrupar costos por uso y clase primero
    costos_por_uso_clase = {}
    suma_total = Decimal('0.00')
    
    for costo in costos:
        uso = costo.uso or ''
        clase = costo.clase or ''
        
        if uso not in costos_por_uso_clase:
            costos_por_uso_clase[uso] = {}
        if clase not in costos_por_uso_clase[uso]:
            costos_por_uso_clase[uso][clase] = []
        
        costos_por_uso_clase[uso][clase].append(costo)
        suma_total += costo.costo or Decimal('0.00')
    
    # Generar contenido PDF agrupado
    for uso, clases in costos_por_uso_clase.items():
        uso_desc = descripciones_uso.get(uso, f'Uso {uso}')
        uso_style = ParagraphStyle(
            'UsoStyle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#4A90E2'),
            spaceAfter=10,
            leftIndent=0
        )
        uso_text = f"<b>USO {uso}</b> - {uso_desc}"
        elements.append(Paragraph(uso_text, uso_style))
        
        for clase, costos_clase in clases.items():
            clase_desc = descripciones_clase.get(clase, f'Clase {clase}')
            clase_style = ParagraphStyle(
                'ClaseStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                spaceAfter=5,
                leftIndent=20
            )
            clase_text = f"<b>Clase {clase}</b> - {clase_desc}"
            elements.append(Paragraph(clase_text, clase_style))
            
            # Encabezados de tabla para esta clase
            headers = ['ID', 'Calidad', 'Costo', 'Rango 1', 'Rango 2']
            data = [headers]
            
            # Agregar costos de esta clase
            for c in costos_clase:
                row = [
                    str(c.id),
                    c.calidad or '-',
                    f"{float(c.costo):,.2f}" if c.costo else '0.00',
                    str(int(c.rango1)) if c.rango1 else '0',
                    str(int(c.rango2)) if c.rango2 else '0'
                ]
                data.append(row)
            
            # Crear tabla
            table = Table(data, colWidths=[1*inch, 1.2*inch, 1.5*inch, 1.2*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Costo alineado a la derecha
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Rangos alineados a la derecha
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.1*inch))
        
        elements.append(Spacer(1, 0.2*inch))
    
    # Agregar total general
    elements.append(Spacer(1, 0.2*inch))
    total_style = ParagraphStyle(
        'TotalStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#006600'),
        spaceAfter=10,
        alignment=2  # Derecha
    )
    total_text = f"<b>TOTAL GENERAL: {float(suma_total):,.2f}</b>"
    elements.append(Paragraph(total_text, total_style))
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'costos_basicos_unitarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    return response

@catastro_require_auth
def costos_clasificacion_pesos_export_excel(request, uso, clase):
    """
    Exportar clasificación de pesos a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:costos_clasificacion_pesos', uso=uso, clase=clase)
    
    # Obtener descripción del uso
    from .models import UsoEdifica
    uso_obj = UsoEdifica.objects.filter(codigo__startswith=uso).first()
    descripcion_uso = uso_obj.descripcion if uso_obj else f'Uso {uso}'
    
    # Diccionario de descripciones de clase
    descripciones_clase = {
        '1': 'Madera',
        '2': 'Ladrillo o Bloque',
        '3': 'Loza o Terraza',
        '4': 'Bajareque o Adobe',
        '5': 'Acero Estructural',
        '6': 'Panelit',
    }
    descripcion_clase = descripciones_clase.get(clase, f'Clase {clase}')
    
    # Obtener configuraciones de tipología para este uso y clase
    # Ordenar numéricamente por tipo (para que 10 venga después de 9)
    # Coalesce maneja valores NULL convirtiéndolos a '0' antes de convertir a entero
    configuraciones_tipologia = ConfiTipologia.objects.filter(
        uso=uso,
        clase=clase
    ).annotate(
        tipo_numeric=Cast(Coalesce('tipo', Value('0')), IntegerField())
    ).order_by('tipo_numeric', 'categoria')
    
    # Obtener tipos de material para mostrar descripciones
    tipos_material = TipoMaterial.objects.all()
    tipos_dict = {tm.No: tm.descripcion for tm in tipos_material}
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clasificación de Pesos"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.append(['CLASIFICACIÓN DE PESOS'])
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Información de uso y clase
    ws.append(['Uso:', uso, descripcion_uso])
    ws.append(['Clase:', clase, descripcion_clase])
    ws.append([])
    
    # Encabezados
    headers = ['ID', 'Uso', 'Clase', 'Clasificación', 'Categoría', 'Descripción', 'Peso']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    for confi in configuraciones_tipologia:
        tipo_desc = tipos_dict.get(confi.tipo, confi.tipo or '-')
        row = [
            confi.id,
            confi.uso or '',
            confi.clase or '',
            tipo_desc,
            confi.categoria or '',
            confi.descripcion or '',
            int(confi.peso) if confi.peso else 0
        ]
        ws.append(row)
    
    # Aplicar estilo a datos
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for cell in row:
            cell.border = border_style
            if cell.column == 1:  # ID
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.column == 7:  # Peso (numérico)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '0'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar ancho de columnas
    column_widths = [8, 10, 10, 20, 12, 40, 15]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'clasificacion_pesos_{uso}_{clase}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def costos_clasificacion_pesos_export_pdf(request, uso, clase):
    """
    Exportar clasificación de pesos a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:costos_clasificacion_pesos', uso=uso, clase=clase)
    
    # Obtener descripción del uso
    from .models import UsoEdifica
    uso_obj = UsoEdifica.objects.filter(codigo__startswith=uso).first()
    descripcion_uso = uso_obj.descripcion if uso_obj else f'Uso {uso}'
    
    # Diccionario de descripciones de clase
    descripciones_clase = {
        '1': 'Madera',
        '2': 'Ladrillo o Bloque',
        '3': 'Loza o Terraza',
        '4': 'Bajareque o Adobe',
        '5': 'Acero Estructural',
        '6': 'Panelit',
    }
    descripcion_clase = descripciones_clase.get(clase, f'Clase {clase}')
    
    # Obtener configuraciones de tipología para este uso y clase
    # Ordenar numéricamente por tipo (para que 10 venga después de 9)
    # Coalesce maneja valores NULL convirtiéndolos a '0' antes de convertir a entero
    configuraciones_tipologia = ConfiTipologia.objects.filter(
        uso=uso,
        clase=clase
    ).annotate(
        tipo_numeric=Cast(Coalesce('tipo', Value('0')), IntegerField())
    ).order_by('tipo_numeric', 'categoria')
    
    # Obtener tipos de material para mostrar descripciones
    tipos_material = TipoMaterial.objects.all()
    tipos_dict = {tm.No: tm.descripcion for tm in tipos_material}
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = "CLASIFICACIÓN DE PESOS"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información de uso y clase
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=10
    )
    info_text = f"<b>Uso:</b> {uso} - {descripcion_uso}<br/><b>Clase:</b> {clase} - {descripcion_clase}"
    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Encabezados de tabla
    headers = ['ID', 'Uso', 'Clase', 'Clasificación', 'Categoría', 'Descripción', 'Peso']
    data = [headers]
    
    # Agregar datos
    for confi in configuraciones_tipologia:
        tipo_desc = tipos_dict.get(confi.tipo, confi.tipo or '-')
        row = [
            str(confi.id),
            confi.uso or '-',
            confi.clase or '-',
            tipo_desc,
            confi.categoria or '-',
            confi.descripcion or '-',
            str(int(confi.peso)) if confi.peso else '0'
        ]
        data.append(row)
    
    # Crear tabla
    table = Table(data, colWidths=[0.8*inch, 0.8*inch, 0.8*inch, 1.5*inch, 1*inch, 3*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),  # Peso alineado a la derecha
        ('ALIGN', (5, 1), (5, -1), 'LEFT'),  # Descripción alineada a la izquierda
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'clasificacion_pesos_{uso}_{clase}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    return response

@csrf_exempt
@catastro_require_auth
def api_buscar_edificacion(request):
    """
    API endpoint para buscar una edificación por empresa, clave, edifino y piso
    """
    try:
        from .models import Edificacion
    except ImportError:
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    # Obtener empresa de la sesión
    empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '') or request.session.get('empresa', '')
    clave = request.GET.get('clave', '').strip()
    edifino = request.GET.get('edifino', '').strip()
    piso = request.GET.get('piso', '').strip() or None
    
    if not empresa:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'No se encontró información de empresa en la sesión'
        })
    
    if not clave or not edifino:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar clave catastral y número de edificación'
        })
    
    try:
        # Construir filtro con empresa, clave y edifino (siempre requeridos)
        # La validación debe ser según empresa, clave, edifino y piso
        
        # Convertir piso a Decimal si tiene valor
        piso_decimal = None
        if piso and piso.strip():
            try:
                from decimal import Decimal
                piso_decimal = Decimal(piso.strip())
            except (ValueError, TypeError):
                piso_decimal = None
        
        if piso_decimal is not None:
            # Si piso tiene valor, buscar exactamente ese piso
            edificacion = Edificacion.objects.filter(
                empresa=empresa,
                clave=clave,
                edifino=edifino,
                piso=piso_decimal
            ).first()
        else:
            # Si piso es None o vacío, buscar registros donde piso es NULL
            edificacion = Edificacion.objects.filter(
                empresa=empresa,
                clave=clave,
                edifino=edifino
            ).filter(Q(piso__isnull=True) | Q(piso=0)).first()
        
        if edificacion:
            return JsonResponse({
                'encontrado': True,
                'edificacion': {
                    'id': edificacion.id,
                    'edifino': str(edificacion.edifino),
                    'piso': str(edificacion.piso) if edificacion.piso and edificacion.piso != 0 else '',
                    'area': str(edificacion.area or '0.00'),
                    'uso': edificacion.uso or '',
                    'clase': edificacion.clase or '',
                    'calidad': edificacion.calidad or '',
                    'costo': str(edificacion.costo or '0.00'),
                    'bueno': str(edificacion.bueno or '0'),
                    'totedi': str(edificacion.totedi or '0.00'),
                    'usuario': edificacion.usuario or '',
                    'fechasys': edificacion.fechasys.strftime('%Y-%m-%d %H:%M:%S') if edificacion.fechasys else '',
                },
                'mensaje': 'Edificación encontrada'
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': f'No se encontró edificación para empresa={empresa}, clave={clave}, edifino={edifino}, piso={piso or "N/A"}'
            })
            
    except Exception as e:
        logger.error(f"Error al buscar edificación: {str(e)}")
        return JsonResponse({
            'encontrado': False,
            'error': str(e),
            'mensaje': 'Error al buscar la edificación'
        }, status=500)

@csrf_exempt
def api_buscar_costo(request):
    """
    API endpoint para buscar costo por empresa, uso, clase y calidad
    """
    try:
        from .models import Costos
    except ImportError:
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    # Obtener empresa de la sesión
    empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '') or request.session.get('empresa', '')
    uso = request.GET.get('uso', '').strip()
    clase = request.GET.get('clase', '').strip()
    calidad = request.GET.get('calidad', '').strip()
    
    if not empresa:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'No se encontró información de empresa en la sesión'
        })
    
    if not uso or not clase or not calidad:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar uso, clase y calidad'
        })
    
    try:
        empresas_mun = codigos_empresa_equivalentes(empresa)
        if not empresas_mun:
            empresas_mun = [empresa]
        costo_obj = Costos.objects.filter(
            empresa__in=empresas_mun,
            uso=uso,
            clase=clase,
            calidad=calidad
        ).first()
        
        if costo_obj:
            return JsonResponse({
                'encontrado': True,
                'costo': str(costo_obj.costo),
                'rango1': str(costo_obj.rango1) if costo_obj.rango1 else '0',
                'rango2': str(costo_obj.rango2) if costo_obj.rango2 else '0',
                'mensaje': 'Costo encontrado'
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': f'No se encontró costo para empresa={empresa}, uso={uso}, clase={clase}, calidad={calidad}'
            })
            
    except Exception as e:
        logger.error(f"Error al buscar costo: {str(e)}")
        return JsonResponse({
            'encontrado': False,
            'error': str(e),
            'mensaje': 'Error al buscar el costo'
        }, status=500)

def eliminar_area_rural(request):
    """
    Vista AJAX para eliminar un registro de AreasRurales.
    """
    # Verificar autenticación usando sesiones de catastro
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    usuario_id = request.session.get('catastro_usuario_id')
    
    logger.info(f"eliminar_area_rural llamada - Método: {request.method}, Empresa: {empresa_codigo}, Usuario ID: {usuario_id}")
    
    # Verificar autenticación para solicitudes AJAX usando sesiones de catastro
    if not empresa_codigo or not usuario_id:
        logger.warning(f"Usuario no autenticado en eliminar_area_rural - Empresa: {empresa_codigo}, Usuario: {usuario_id}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No autenticado. Por favor, inicie sesión en catastro.'}, status=401)
        return redirect('catastro:catastro_login')
    
    if request.method != 'POST':
        logger.warning(f"Método no permitido en eliminar_area_rural: {request.method}")
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        area_id = request.POST.get('area_id', '').strip()
        
        if not area_id:
            return JsonResponse({'success': False, 'error': 'ID de área rural no proporcionado'}, status=400)
        
        try:
            area_id = int(area_id)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'ID de área rural inválido'}, status=400)
        
        # Obtener el registro y verificar que pertenezca a la empresa del usuario
        from .models import AreasRurales
        
        try:
            area_rural = AreasRurales.objects.get(id=area_id, empresa=empresa_codigo)
        except AreasRurales.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Registro no encontrado o no tiene permisos para eliminarlo'}, status=404)
        
        # Eliminar el registro
        cocata1 = area_rural.cocata1
        area_rural.delete()
        
        logger.info(f"AreasRurales eliminado (ID: {area_id}) para cocata1={cocata1}, empresa={empresa_codigo}")
        
        # Recalcular la sumatoria del campo monto filtrado por empresa y cocata1
        # y actualizar bvlbas1 en bdterreno
        suma_monto = Decimal('0.00')
        try:
            from .models import BDTerreno
            from django.db.models import Sum
            
            # Calcular la sumatoria de monto después de eliminar
            # Filtrar por empresa y cocata1
            suma_monto = AreasRurales.objects.filter(
                cocata1=cocata1,
                empresa=empresa_codigo
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            
            logger.info(f"Sumatoria calculada después de eliminar (cocata1={cocata1}, empresa={empresa_codigo}): {suma_monto}")
            
            # Obtener o crear el registro de BDTerreno
            try:
                terreno = BDTerreno.objects.get(cocata1=cocata1)
            except BDTerreno.DoesNotExist:
                terreno = BDTerreno(cocata1=cocata1, empresa=empresa_codigo)
                # Guardar primero el registro recién creado
                terreno.save()
                logger.info(f"BDTerreno creado para cocata1={cocata1}, empresa={empresa_codigo}")
            
            # Actualizar el campo bvlbas1 con la nueva sumatoria
            terreno.bvlbas1 = suma_monto
            terreno.save(update_fields=['bvlbas1'])
            
            logger.info(f"bvlbas1 actualizado en BDTerreno después de eliminar (cocata1={cocata1}): {suma_monto}")
            
            # Recalcular el valor de tierra en bdcata1 y el impuesto
            try:
                from .models import BDCata1
                
                # Calcular el valor del terreno rural desde BDTerreno
                valor_terreno_rural = terreno.calcular_valor_terreno() if hasattr(terreno, 'calcular_valor_terreno') else suma_monto
                
                registro_bdcata1 = BDCata1.objects.get(cocata1=cocata1, empresa=empresa_codigo)
                registro_bdcata1.bvl2tie = valor_terreno_rural
                
                # Recalcular el impuesto
                impuesto_calculado = calcular_impuesto_bdcata1(registro_bdcata1, empresa_codigo)
                registro_bdcata1.impuesto = impuesto_calculado
                
                registro_bdcata1.save(update_fields=['bvl2tie', 'impuesto'])
                logger.info(f'Campo bvl2tie e impuesto recalculados en bdcata1 después de eliminar área rural: impuesto={impuesto_calculado} para clave {cocata1}')
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro_bdcata1, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro_bdcata1, empresa_codigo)
            except BDCata1.DoesNotExist:
                logger.warning(f"No se encontró registro en bdcata1 para clave={cocata1}, empresa={empresa_codigo} al eliminar área rural")
            except Exception as e:
                logger.warning(f"Error al recalcular impuesto después de eliminar área rural: {str(e)}")
            
        except Exception as e:
            import traceback
            logger.error(f"Error al actualizar bvlbas1 en BDTerreno después de eliminar: {str(e)}\n{traceback.format_exc()}")
            # No interrumpir el flujo si hay error, solo loguear
        
        mensaje = 'Registro eliminado exitosamente'
        if suma_monto > 0:
            mensaje += f'. Sumatoria actualizada en bvlbas1: {suma_monto:.2f}'
        else:
            mensaje += '. Sumatoria actualizada en bvlbas1: 0.00 (no hay más registros)'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'suma_monto': str(suma_monto)
        })
        
    except Exception as e:
        import traceback
        logger.error(f"Error al eliminar AreasRurales: {str(e)}\n{traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': f'Error al eliminar: {str(e)}'
        }, status=500)

# ============================================================================
# GESTIÓN DE DETALLES ADICIONALES
# ============================================================================

@catastro_require_auth
def detalles_adicionales_list(request, clave=None):
    """
    Lista de detalles adicionales para una clave catastral específica
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    # Si se proporciona clave, filtrar por ella
    detalles = DetalleAdicionales.objects.filter(empresa=empresa_codigo)
    if clave:
        detalles = detalles.filter(clave=clave)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        detalles = detalles.filter(
            Q(clave__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(codigo__icontains=search)
        )
    
    detalles = detalles.order_by('-id')
    
    # Calcular sumatorias
    suma_area = detalles.aggregate(Sum('area'))['area__sum'] or Decimal('0.00')
    suma_porce = detalles.aggregate(Sum('porce'))['porce__sum'] or Decimal('0.00')
    suma_unit = detalles.aggregate(Sum('unit'))['unit__sum'] or Decimal('0.00')
    suma_total = detalles.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave and empresa_codigo:
        try:
            from .models import BDCata1
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave
            ).first()
        except Exception:
            pass
    
    context = {
        'titulo': 'Detalles Adicionales - Catastro',
        'detalles': detalles,
        'clave': clave,
        'empresa': empresa_codigo,
        'bien_inmueble': bien_inmueble,
        'search': search,
        'total_registros': detalles.count(),
        'suma_area': suma_area,
        'suma_porce': suma_porce,
        'suma_unit': suma_unit,
        'suma_total': suma_total,
    }
    
    return render(request, 'detalles_adicionales_list.html', context)

@catastro_require_auth
def detalles_adicionales_export_excel(request, clave=None):
    """
    Exportar detalles adicionales a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        if clave:
            return redirect('catastro:detalles_adicionales_list_clave', clave=clave)
        return redirect('catastro:detalles_adicionales_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener los detalles
    detalles = DetalleAdicionales.objects.filter(empresa=empresa_codigo)
    if clave:
        detalles = detalles.filter(clave=clave)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        detalles = detalles.filter(
            Q(clave__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(codigo__icontains=search)
        )
    
    detalles = detalles.order_by('-id')
    
    # Calcular sumatorias
    suma_total = detalles.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalles Adicionales"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Clave Catastral', 'Código', 'Descripción', 'Área', 'Porcentaje (%)', 
               'Valor Unitario', 'Total', 'Edif. No.', 'Piso', 'Código Edific.', 'Usuario', 'Fecha']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    for detalle in detalles:
        row = [
            detalle.id,
            detalle.clave or '',
            detalle.codigo or '',
            detalle.descripcion or '',
            float(detalle.area) if detalle.area else 0.00,
            float(detalle.porce) if detalle.porce else 0.00,
            float(detalle.unit) if detalle.unit else 0.00,
            float(detalle.total) if detalle.total else 0.00,
            int(detalle.edifino) if detalle.edifino else 0,
            int(detalle.piso) if detalle.piso else 0,
            detalle.codedi or '',
            detalle.usuario or '',
            detalle.fechasys.strftime('%d/%m/%Y %H:%M') if detalle.fechasys else ''
        ]
        ws.append(row)
    
    # Aplicar estilo a datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border_style
            if cell.column in [5, 6, 7, 8, 9, 10]:  # Columnas numéricas (Área, Porcentaje, Valor Unitario, Total, Edif. No., Piso)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if cell.column in [5, 6, 7, 8]:  # Columnas con decimales
                    cell.number_format = '#,##0.00'
                else:  # Columnas enteras (Edif. No., Piso)
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Agregar fila de totales
    total_row = [''] * 7 + [float(suma_total)]
    ws.append(total_row)
    total_cell = ws.cell(row=ws.max_row, column=8)
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = '#,##0.00'
    total_cell.border = border_style
    
    # Ajustar ancho de columnas
    column_widths = [8, 18, 10, 25, 12, 15, 15, 15, 15, 15, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'detalles_adicionales_{clave or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def detalles_adicionales_export_pdf(request, clave=None):
    """
    Exportar detalles adicionales a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        if clave:
            return redirect('catastro:detalles_adicionales_list_clave', clave=clave)
        return redirect('catastro:detalles_adicionales_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener los detalles
    detalles = DetalleAdicionales.objects.filter(empresa=empresa_codigo)
    if clave:
        detalles = detalles.filter(clave=clave)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        detalles = detalles.filter(
            Q(clave__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(codigo__icontains=search)
        )
    
    detalles = detalles.order_by('-id')
    
    # Calcular sumatorias
    suma_total = detalles.aggregate(Sum('total'))['total__sum'] or Decimal('0.00')
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = f"Detalles Adicionales - Catastro"
    if clave:
        title_text += f"<br/>Clave Catastral: {clave}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Preparar datos de la tabla
    data = []
    
    # Encabezados
    headers = ['ID', 'Clave', 'Código', 'Descripción', 'Área', 'Porc. (%)', 
               'Valor Unit.', 'Total', 'Edif. No.', 'Piso', 'Cód. Edific.', 'Usuario', 'Fecha']
    data.append(headers)
    
    # Datos
    for detalle in detalles:
        row = [
            str(detalle.id),
            detalle.clave or '-',
            detalle.codigo or '-',
            detalle.descripcion or '-',
            f"{detalle.area:.2f}" if detalle.area else "0.00",
            f"{detalle.porce:.2f}" if detalle.porce else "0.00",
            f"{detalle.unit:.2f}" if detalle.unit else "0.00",
            f"{detalle.total:.2f}" if detalle.total else "0.00",
            str(int(detalle.edifino)) if detalle.edifino else '0',
            str(int(detalle.piso)) if detalle.piso else '0',
            detalle.codedi or '-',
            detalle.usuario or '-',
            detalle.fechasys.strftime('%d/%m/%Y') if detalle.fechasys else '-'
        ]
        data.append(row)
    
    # Fila de totales
    total_row = [''] * 7 + [f"{suma_total:.2f}"] + [''] * 3
    data.append(total_row)
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('ALIGN', (4, 1), (10, -2), 'RIGHT'),  # Columnas numéricas (Área, Porcentaje, Valor Unitario, Total, Edif. No., Piso)
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        
        # Fila de totales
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (6, -1), 'RIGHT'),
        ('ALIGN', (7, -1), (7, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Información adicional
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        alignment=0  # Izquierda
    )
    info_text = f"<b>Total de registros:</b> {detalles.count()}<br/>"
    info_text += f"<b>Suma total:</b> L. {suma_total:.2f}<br/>"
    info_text += f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'detalles_adicionales_{clave or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response

@catastro_require_auth
def detalle_adicional_create(request, clave=None):
    """
    Crear nuevo detalle adicional
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    if request.method == 'POST':
        form = DetalleAdicionalesForm(request.POST, empresa=empresa_codigo, clave=clave)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.empresa = empresa_codigo
            if clave:
                detalle.clave = clave
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            
            # Intentar obtener el objeto User para la bitácora
            usuario_obj = None
            try:
                usuario_id = request.session.get('catastro_usuario_id')
                if usuario_id:
                    usuario_obj = User.objects.filter(id=usuario_id).first()
                elif current_user:
                    # Intentar buscar por username
                    usuario_obj = User.objects.filter(username=current_user).first()
            except Exception:
                pass
            
            # Asignar el objeto User a la instancia para que los signals lo puedan usar
            if usuario_obj:
                detalle._usuario_obj = usuario_obj
            
            detalle.save()
            messages.success(request, 'Detalle adicional creado exitosamente. Puede continuar agregando más detalles.')
            
            # Calcular sumatoria de detalles adicionales y actualizar bdcata1 (siempre, no solo cuando se solicita)
            if clave and empresa_codigo:
                try:
                    # Obtener el registro de bdcata1
                    bdcata1 = BDCata1.objects.filter(empresa=empresa_codigo, cocata1=clave).first()
                    if bdcata1:
                        # Calcular la sumatoria de todos los detalles adicionales (incluyendo el recién guardado)
                        detalles_todos = DetalleAdicionales.objects.filter(
                            empresa=empresa_codigo,
                            clave=clave
                        )
                        suma_total_detalles = detalles_todos.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
                        
                        # Actualizar el campo detalle en bdcata1 con la sumatoria
                        bdcata1.detalle = suma_total_detalles
                        logger.info(f'Campo detalle actualizado en bdcata1: {suma_total_detalles} para clave {clave}')
                        
                        # Calcular el impuesto usando la misma lógica del frontend
                        impuesto_calculado = calcular_impuesto_bdcata1(bdcata1, empresa_codigo)
                        
                        # Actualizar el campo impuesto en bdcata1
                        bdcata1.impuesto = impuesto_calculado
                        
                        # Guardar ambos campos en una sola operación
                        bdcata1.save(update_fields=['detalle', 'impuesto'])
                        logger.info(f'Impuesto calculado y actualizado en bdcata1: {impuesto_calculado} para clave {clave}')
                        
                        # Actualizar el impuesto en tasasmunicipales
                        actualizar_impuesto_tasas_municipales(bdcata1, impuesto_calculado, empresa_codigo)
                        
                        # Calcular tasas municipales (rubros que empiezan con T)
                        calcular_tasas_municipales_automatico(bdcata1, empresa_codigo)
                except Exception as e:
                    logger.error(f'Error al calcular impuesto después de guardar detalle adicional: {str(e)}', exc_info=True)
            
            # En lugar de redirigir, recargar el formulario limpio para seguir agregando detalles
            # Inicializar formulario con usuario y fecha actual, manteniendo empresa y clave
            form = DetalleAdicionalesForm(
                empresa=empresa_codigo, 
                clave=clave,
                usuario=current_user[:50] if current_user else '',
                fecha_sistema=timezone.now()
            )
            
            context = {
                'titulo': 'Nuevo Detalle Adicional - Catastro',
                'form': form,
                'clave': clave,
            }
            
            return render(request, 'detalle_adicional_form.html', context)
    else:
        # Inicializar formulario con usuario y fecha actual
        form = DetalleAdicionalesForm(
            empresa=empresa_codigo, 
            clave=clave,
            usuario=current_user[:50] if current_user else '',
            fecha_sistema=timezone.now()
        )
    
    context = {
        'titulo': 'Nuevo Detalle Adicional - Catastro',
        'form': form,
        'clave': clave,
    }
    
    return render(request, 'detalle_adicional_form.html', context)

@catastro_require_auth
def detalle_adicional_especial_create(request, empresa=None, cocata1=None, clave=None):
    """
    Crear nuevo detalle especial (solo descripción y total)
    Recibe parámetros empresa y cocata1 de la tabla bdcata1
    """
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros (cocata1 o clave)
    clave_catastral = cocata1 or clave or request.GET.get('cocata1', '').strip()
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    if request.method == 'POST':
        form = DetalleAdicionalesForm(request.POST, empresa=empresa_codigo, clave=clave_catastral)
        # Remover readonly del campo total para que se envíe en POST
        if 'readonly' in form.fields['total'].widget.attrs:
            del form.fields['total'].widget.attrs['readonly']
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.empresa = empresa_codigo
            if clave_catastral:
                detalle.clave = clave_catastral
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            
            # Para detalles especiales:
            # - codigo siempre debe ser "E"
            detalle.codigo = 'E'
            
            # - total se toma directamente de request.POST porque el campo puede estar readonly
            # y Django no lo incluye en cleaned_data cuando está readonly
            total_value = request.POST.get('total', '').strip()
            if total_value:
                try:
                    detalle.total = Decimal(str(total_value))
                except (ValueError, InvalidOperation):
                    detalle.total = Decimal('0.00')
            elif 'total' in form.cleaned_data and form.cleaned_data['total'] is not None:
                detalle.total = form.cleaned_data['total']
            else:
                detalle.total = Decimal('0.00')
            
            # Los demás campos se dejan en 0 o vacío
            if not detalle.area:
                detalle.area = Decimal('0.00')
            if not detalle.porce:
                detalle.porce = Decimal('0.00')
            if not detalle.unit:
                detalle.unit = Decimal('0.00')
            
            # Intentar obtener el objeto User para la bitácora
            usuario_obj = None
            try:
                usuario_id = request.session.get('catastro_usuario_id')
                if usuario_id:
                    usuario_obj = User.objects.filter(id=usuario_id).first()
                elif current_user:
                    # Intentar buscar por username
                    usuario_obj = User.objects.filter(username=current_user).first()
            except Exception:
                pass
            
            # Asignar el objeto User a la instancia para que los signals lo puedan usar
            if usuario_obj:
                detalle._usuario_obj = usuario_obj
            
            detalle.save()
            messages.success(request, 'Detalle especial creado exitosamente.')
            if clave_catastral:
                return redirect('catastro:detalles_adicionales_list_clave', clave=clave_catastral)
            return redirect('catastro:detalles_adicionales_list')
    else:
        # Inicializar formulario con usuario y fecha actual
        form = DetalleAdicionalesForm(
            empresa=empresa_codigo, 
            clave=clave_catastral,
            usuario=current_user[:50] if current_user else '',
            fecha_sistema=timezone.now()
        )
        # Establecer valores por defecto para campos no requeridos
        form.fields['area'].initial = Decimal('0.00')
        form.fields['porce'].initial = Decimal('0.00')
        form.fields['unit'].initial = Decimal('0.00')
        # Permitir editar total en detalles especiales (remover readonly)
        if 'readonly' in form.fields['total'].widget.attrs:
            del form.fields['total'].widget.attrs['readonly']
    
    context = {
        'titulo': 'Nuevo Detalle Especial - Catastro',
        'form': form,
        'clave': clave_catastral,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'detalle_especial_form.html', context)

@catastro_require_auth
def edificacion_especial_create(request, clave=None):
    """
    Crear nueva edificación especial con valor estimado
    Solo requiere descripción y totedi
    Guarda con edifino=0, piso=0, uso='E'
    """
    # Obtener empresa del login
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    usuario_nombre = request.session.get('catastro_usuario_nombre', '') or request.session.get('catastro_usuario_id', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    # Obtener clave catastral de parámetros o GET
    clave_catastral = clave or request.GET.get('clave', '').strip() or request.GET.get('cocata1', '').strip()
    
    if not clave_catastral:
        messages.error(request, 'No se proporcionó la clave catastral.')
        return redirect('catastro:catastro_menu_principal')
    
    # Verificar que existe el registro de bien inmueble
    try:
        registro = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    if request.method == 'POST':
        form = EdificacionEspecialForm(
            request.POST,
            empresa=empresa_codigo,
            clave=clave_catastral,
            usuario=usuario_nombre[:50] if usuario_nombre else '',
            fecha_sistema=timezone.now()
        )
        
        if form.is_valid():
            edificacion = form.save(commit=False)
            edificacion.empresa = empresa_codigo
            edificacion.clave = clave_catastral
            # Usar los valores ingresados manualmente por el usuario
            edificacion.edifino = Decimal(str(form.cleaned_data.get('edifino', 0)))
            edificacion.piso = Decimal(str(form.cleaned_data.get('piso', 0)))
            edificacion.uso = 'E'  # uso = 'E' (especial)
            edificacion.clase = form.cleaned_data.get('clase', '') or ''
            edificacion.calidad = form.cleaned_data.get('calidad', '') or ''
            edificacion.descripcion = form.cleaned_data.get('descripcion', '') or ''
            edificacion.totedi = form.cleaned_data.get('totedi', Decimal('0.00'))
            edificacion.area = Decimal('0.00')
            edificacion.costo = Decimal('0.00')
            edificacion.bueno = Decimal('0')
            edificacion.usuario = usuario_nombre[:50] if usuario_nombre else ''
            edificacion.fechasys = timezone.now()
            
            # Intentar obtener el objeto User para la bitácora
            usuario_obj = None
            try:
                usuario_id = request.session.get('catastro_usuario_id')
                if usuario_id:
                    usuario_obj = User.objects.filter(id=usuario_id).first()
                elif usuario_nombre:
                    usuario_obj = User.objects.filter(username=usuario_nombre).first()
            except Exception:
                pass
            
            # Asignar el objeto User a la instancia para que los signals lo puedan usar
            if usuario_obj:
                edificacion._usuario_obj = usuario_obj
            
            edificacion.save()
            
            # Calcular sumatoria de todas las edificaciones y actualizar mejoras en bdcata1
            try:
                # Calcular la sumatoria de todas las edificaciones (incluyendo la recién guardada)
                edificaciones_todas = Edificacion.objects.filter(
                    empresa=empresa_codigo,
                    clave=clave_catastral
                )
                suma_total_edificaciones = edificaciones_todas.aggregate(total=Sum('totedi'))['total'] or Decimal('0.00')
                
                # Obtener el registro de bdcata1
                registro_bdcata1 = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
                
                # Actualizar el campo mejoras con la sumatoria total
                registro_bdcata1.mejoras = suma_total_edificaciones.quantize(Decimal('0.01'))
                logger.info(f'Campo mejoras actualizado en bdcata1: {suma_total_edificaciones} para clave {clave_catastral}')
                
                # Recalcular el impuesto usando la misma lógica del frontend
                impuesto_calculado = calcular_impuesto_bdcata1(registro_bdcata1, empresa_codigo)
                
                # Actualizar el campo impuesto en bdcata1
                registro_bdcata1.impuesto = impuesto_calculado
                
                # Guardar ambos campos en una sola operación
                registro_bdcata1.save(update_fields=['mejoras', 'impuesto'])
                logger.info(f'Impuesto calculado y actualizado en bdcata1 después de guardar edificación especial: {impuesto_calculado} para clave {clave_catastral}')
                
                # Actualizar el impuesto en tasasmunicipales
                actualizar_impuesto_tasas_municipales(registro_bdcata1, impuesto_calculado, empresa_codigo)
                
                # Calcular tasas municipales (rubros que empiezan con T)
                calcular_tasas_municipales_automatico(registro_bdcata1, empresa_codigo)
            except BDCata1.DoesNotExist:
                logger.error(f'No se encontró registro bdcata1 para clave {clave_catastral} al actualizar mejoras')
            except Exception as e:
                logger.error(f'Error al actualizar campo mejoras e impuesto en bdcata1: {str(e)}', exc_info=True)
            
            descripcion_texto = form.cleaned_data.get('descripcion', '') or 'Sin descripción'
            messages.success(request, f'Edificación especial creada exitosamente. Descripción: {descripcion_texto}')
            return redirect('catastro:edificaciones_form', clave=clave_catastral)
    else:
        form = EdificacionEspecialForm(
            empresa=empresa_codigo,
            clave=clave_catastral,
            usuario=usuario_nombre[:50] if usuario_nombre else '',
            fecha_sistema=timezone.now()
        )
    
    # Obtener información del municipio para mostrar
    municipio_descripcion = ''
    try:
        municipio = Municipio.objects.filter(codigo=empresa_codigo).first()
        if municipio:
            municipio_descripcion = municipio.descripcion
    except:
        pass
    
    context = {
        'titulo': 'Nueva Edificación Especial - Catastro',
        'form': form,
        'clave': clave_catastral,
        'empresa': empresa_codigo,
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': municipio_descripcion,
    }
    
    return render(request, 'edificacion_especial_form.html', context)

@catastro_require_auth
def comentario_catastro_create(request, clave=None):
    """
    Crear nuevo comentario para una clave catastral
    """
    # Obtener empresa del login
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    usuario_nombre = request.session.get('catastro_usuario_nombre', '') or request.session.get('catastro_usuario_id', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    # Obtener clave catastral de parámetros o GET
    clave_catastral = clave or request.GET.get('clave', '').strip() or request.GET.get('cocata1', '').strip()
    
    if not clave_catastral:
        messages.error(request, 'No se proporcionó la clave catastral.')
        return redirect('catastro:catastro_menu_principal')
    
    # Verificar que existe el registro de bien inmueble
    try:
        registro = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    if request.method == 'POST':
        form = ComentariosCatastroForm(
            request.POST,
            empresa=empresa_codigo,
            clave=clave_catastral,
            usuario=usuario_nombre[:50] if usuario_nombre else '',
            fecha_sistema=timezone.now()
        )
        
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.empresa = empresa_codigo
            comentario.clave = clave_catastral
            comentario.usuario = usuario_nombre[:50] if usuario_nombre else ''
            comentario.fecha = timezone.now()
            comentario.save()
            messages.success(request, 'Comentario agregado exitosamente.')
            return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}&comentario_agregado=1")
    else:
        form = ComentariosCatastroForm(
            empresa=empresa_codigo,
            clave=clave_catastral,
            usuario=usuario_nombre[:50] if usuario_nombre else '',
            fecha_sistema=timezone.now()
        )
    
    # Obtener información del municipio para mostrar
    municipio_descripcion = ''
    try:
        municipio = Municipio.objects.filter(codigo=empresa_codigo).first()
        if municipio:
            municipio_descripcion = municipio.descripcion
    except:
        pass
    
    context = {
        'titulo': 'Agregar Comentario - Catastro',
        'form': form,
        'clave': clave_catastral,
        'empresa': empresa_codigo,
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': municipio_descripcion,
    }
    
    return render(request, 'comentario_catastro_form.html', context)

@catastro_require_auth
def api_comentarios_catastro(request):
    """
    API para obtener comentarios de una clave catastral
    Retorna JSON con lista de comentarios ordenados por fecha descendente
    """
    empresa_codigo = request.session.get('catastro_empresa') or request.session.get('empresa') or request.session.get('catastro_municipio_codigo')
    clave_catastral = request.GET.get('clave', '').strip()
    
    if not clave_catastral or not empresa_codigo:
        return JsonResponse({'error': 'Clave catastral y empresa son requeridos'}, status=400)
    
    comentarios = ComentariosCatastro.objects.filter(
        clave=clave_catastral,
        empresa=empresa_codigo
    ).order_by('-fecha', '-id')
    
    comentarios_list = []
    for comentario in comentarios:
        comentarios_list.append({
            'id': comentario.id,
            'comentario': comentario.comentario or '',
            'usuario': comentario.usuario or '',
            'fecha': comentario.fecha.strftime('%Y-%m-%d %H:%M:%S') if comentario.fecha else '',
        })
    
    return JsonResponse({
        'comentarios': comentarios_list,
        'total': len(comentarios_list)
    })

@catastro_require_auth
def detalle_adicional_update(request, pk):
    """
    Actualizar detalle adicional
    Valida si codigo = "E" para cargar formulario especial o normal
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    detalle = get_object_or_404(DetalleAdicionales, pk=pk, empresa=empresa_codigo)
    
    # Validar si codigo = "E" para usar formulario especial
    if detalle.codigo == 'E':
        # Redirigir a la vista de edición de detalle especial
        return detalle_adicional_especial_update(request, pk)
    
    # Si no es "E", usar el formulario normal
    if request.method == 'POST':
        form = DetalleAdicionalesForm(request.POST, instance=detalle, empresa=empresa_codigo, clave=detalle.clave)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            
            # Intentar obtener el objeto User para la bitácora
            usuario_obj = None
            try:
                usuario_id = request.session.get('catastro_usuario_id')
                if usuario_id:
                    usuario_obj = User.objects.filter(id=usuario_id).first()
                elif current_user:
                    # Intentar buscar por username
                    usuario_obj = User.objects.filter(username=current_user).first()
            except Exception:
                pass
            
            # Asignar el objeto User a la instancia para que los signals lo puedan usar
            if usuario_obj:
                detalle._usuario_obj = usuario_obj
            
            detalle.save()
            messages.success(request, 'Detalle adicional actualizado exitosamente.')
            
            # Calcular sumatoria de detalles adicionales y actualizar bdcata1
            clave = detalle.clave
            if clave and empresa_codigo:
                try:
                    # Obtener el registro de bdcata1 usando la relación: bdcata1.empresa = detallesadicionales.empresa y bdcata1.cocata1 = detallesadicionales.clave
                    bdcata1 = BDCata1.objects.filter(empresa=empresa_codigo, cocata1=clave).first()
                    if bdcata1:
                        # Calcular la sumatoria de todos los detalles adicionales (incluyendo el recién actualizado)
                        detalles_todos = DetalleAdicionales.objects.filter(
                            empresa=empresa_codigo,
                            clave=clave
                        )
                        suma_total_detalles = detalles_todos.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
                        
                        # Actualizar el campo detalle en bdcata1 con la sumatoria
                        bdcata1.detalle = suma_total_detalles
                        logger.info(f'Campo detalle actualizado en bdcata1 después de editar: {suma_total_detalles} para clave {clave}')
                        
                        # Recalcular el impuesto usando la misma lógica del frontend
                        impuesto_calculado = calcular_impuesto_bdcata1(bdcata1, empresa_codigo)
                        
                        # Actualizar el campo impuesto en bdcata1
                        bdcata1.impuesto = impuesto_calculado
                        
                        # Guardar ambos campos en una sola operación
                        bdcata1.save(update_fields=['detalle', 'impuesto'])
                        logger.info(f'Impuesto recalculado y actualizado en bdcata1 después de editar detalle: {impuesto_calculado} para clave {clave}')
                        
                        # Actualizar el impuesto en tasasmunicipales
                        actualizar_impuesto_tasas_municipales(bdcata1, impuesto_calculado, empresa_codigo)
                        
                        # Calcular tasas municipales (rubros que empiezan con T)
                        calcular_tasas_municipales_automatico(bdcata1, empresa_codigo)
                except Exception as e:
                    logger.error(f'Error al recalcular impuesto después de editar detalle adicional: {str(e)}', exc_info=True)
            
            return redirect('catastro:detalles_adicionales_list_clave', clave=detalle.clave)
    else:
        form = DetalleAdicionalesForm(instance=detalle, empresa=empresa_codigo, clave=detalle.clave)
    
    context = {
        'titulo': f'Editar Detalle Adicional - Catastro',
        'form': form,
        'detalle': detalle,
        'clave': detalle.clave,
    }
    
    return render(request, 'detalle_adicional_form.html', context)

@catastro_require_auth
def detalle_adicional_especial_update(request, pk):
    """
    Actualizar detalle especial (codigo = "E")
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    detalle = get_object_or_404(DetalleAdicionales, pk=pk, empresa=empresa_codigo)
    
    if request.method == 'POST':
        form = DetalleAdicionalesForm(request.POST, instance=detalle, empresa=empresa_codigo, clave=detalle.clave)
        # Remover readonly del campo total para que se envíe en POST
        if 'readonly' in form.fields['total'].widget.attrs:
            del form.fields['total'].widget.attrs['readonly']
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            
            # Para detalles especiales, asegurar que codigo = "E"
            detalle.codigo = 'E'
            
            # - total se toma directamente de request.POST porque el campo puede estar readonly
            # y Django no lo incluye en cleaned_data cuando está readonly
            total_value = request.POST.get('total', '').strip()
            if total_value:
                try:
                    detalle.total = Decimal(str(total_value))
                except (ValueError, InvalidOperation):
                    detalle.total = Decimal('0.00')
            elif 'total' in form.cleaned_data and form.cleaned_data['total'] is not None:
                detalle.total = form.cleaned_data['total']
            # Si no hay valor, mantener el valor existente
            
            # Intentar obtener el objeto User para la bitácora
            usuario_obj = None
            try:
                usuario_id = request.session.get('catastro_usuario_id')
                if usuario_id:
                    usuario_obj = User.objects.filter(id=usuario_id).first()
                elif current_user:
                    # Intentar buscar por username
                    usuario_obj = User.objects.filter(username=current_user).first()
            except Exception:
                pass
            
            # Asignar el objeto User a la instancia para que los signals lo puedan usar
            if usuario_obj:
                detalle._usuario_obj = usuario_obj
            
            detalle.save()
            messages.success(request, 'Detalle especial actualizado exitosamente.')
            
            # Calcular sumatoria de detalles adicionales y actualizar bdcata1
            clave = detalle.clave
            if clave and empresa_codigo:
                try:
                    # Obtener el registro de bdcata1 usando la relación: bdcata1.empresa = detallesadicionales.empresa y bdcata1.cocata1 = detallesadicionales.clave
                    bdcata1 = BDCata1.objects.filter(empresa=empresa_codigo, cocata1=clave).first()
                    if bdcata1:
                        # Calcular la sumatoria de todos los detalles adicionales (incluyendo el recién actualizado)
                        detalles_todos = DetalleAdicionales.objects.filter(
                            empresa=empresa_codigo,
                            clave=clave
                        )
                        suma_total_detalles = detalles_todos.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
                        
                        # Actualizar el campo detalle en bdcata1 con la sumatoria
                        bdcata1.detalle = suma_total_detalles
                        logger.info(f'Campo detalle actualizado en bdcata1 después de editar detalle especial: {suma_total_detalles} para clave {clave}')
                        
                        # Recalcular el impuesto usando la misma lógica del frontend
                        impuesto_calculado = calcular_impuesto_bdcata1(bdcata1, empresa_codigo)
                        
                        # Actualizar el campo impuesto en bdcata1
                        bdcata1.impuesto = impuesto_calculado
                        
                        # Guardar ambos campos en una sola operación
                        bdcata1.save(update_fields=['detalle', 'impuesto'])
                        logger.info(f'Impuesto recalculado y actualizado en bdcata1 después de editar detalle especial: {impuesto_calculado} para clave {clave}')
                        
                        # Actualizar el impuesto en tasasmunicipales
                        actualizar_impuesto_tasas_municipales(bdcata1, impuesto_calculado, empresa_codigo)
                        
                        # Calcular tasas municipales (rubros que empiezan con T)
                        calcular_tasas_municipales_automatico(bdcata1, empresa_codigo)
                except Exception as e:
                    logger.error(f'Error al recalcular impuesto después de editar detalle especial: {str(e)}', exc_info=True)
            
            return redirect('catastro:detalles_adicionales_list_clave', clave=detalle.clave)
    else:
        form = DetalleAdicionalesForm(instance=detalle, empresa=empresa_codigo, clave=detalle.clave)
        # Permitir editar total en detalles especiales (remover readonly)
        if 'readonly' in form.fields['total'].widget.attrs:
            del form.fields['total'].widget.attrs['readonly']
    
    context = {
        'titulo': 'Editar Detalle Especial - Catastro',
        'form': form,
        'detalle': detalle,
        'clave': detalle.clave,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'detalle_especial_form.html', context)

@catastro_require_auth
def detalle_adicional_delete(request, pk):
    """
    Eliminar detalle adicional
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    detalle = get_object_or_404(DetalleAdicionales, pk=pk, empresa=empresa_codigo)
    clave = detalle.clave
    
    if request.method == 'POST':
        detalle.delete()
        messages.success(request, 'Detalle adicional eliminado exitosamente.')
        
        # Calcular sumatoria de detalles adicionales restantes y actualizar bdcata1
        if clave and empresa_codigo:
            try:
                # Obtener el registro de bdcata1
                bdcata1 = BDCata1.objects.filter(empresa=empresa_codigo, cocata1=clave).first()
                if bdcata1:
                    # Calcular la sumatoria de todos los detalles adicionales restantes (después de eliminar)
                    detalles_restantes = DetalleAdicionales.objects.filter(
                        empresa=empresa_codigo,
                        clave=clave
                    )
                    suma_total_detalles = detalles_restantes.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
                    
                    # Actualizar el campo detalle en bdcata1 con la sumatoria
                    bdcata1.detalle = suma_total_detalles
                    logger.info(f'Campo detalle actualizado en bdcata1: {suma_total_detalles} para clave {clave}')
                    
                    # Recalcular el impuesto usando la misma lógica del frontend
                    impuesto_calculado = calcular_impuesto_bdcata1(bdcata1, empresa_codigo)
                    
                    # Actualizar el campo impuesto en bdcata1
                    bdcata1.impuesto = impuesto_calculado
                    
                    # Guardar ambos campos en una sola operación
                    bdcata1.save(update_fields=['detalle', 'impuesto'])
                    logger.info(f'Impuesto recalculado y actualizado en bdcata1 después de eliminar detalle: {impuesto_calculado} para clave {clave}')
                    
                    # Actualizar el impuesto en tasasmunicipales
                    actualizar_impuesto_tasas_municipales(bdcata1, impuesto_calculado, empresa_codigo)
                    
                    # Calcular tasas municipales (rubros que empiezan con T)
                    calcular_tasas_municipales_automatico(bdcata1, empresa_codigo)
            except Exception as e:
                logger.error(f'Error al recalcular impuesto después de eliminar detalle adicional: {str(e)}', exc_info=True)
        
        return redirect('catastro:detalles_adicionales_list_clave', clave=clave)
    
    context = {
        'titulo': f'Eliminar Detalle Adicional - Catastro',
        'detalle': detalle,
        'clave': clave,
    }
    
    return render(request, 'detalle_adicional_confirm_delete.html', context)

# ============================================================================
# GESTIÓN DE TIPOS DE DETALLE (CATÁLOGO)
# ============================================================================

@catastro_require_auth
def tipos_detalle_list(request):
    """
    Lista de tipos de detalle (catálogo)
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    tipos = TipoDetalle.objects.all()
    if empresa_codigo:
        tipos = tipos.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        tipos = tipos.filter(
            Q(codigo__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    tipos = tipos.order_by('codigo')
    
    context = {
        'titulo': 'Tipos de Detalle - Catastro',
        'tipos': tipos,
        'search': search,
        'total_registros': tipos.count(),
    }
    
    return render(request, 'tipos_detalle_list.html', context)

@catastro_require_auth
def tipo_detalle_create(request):
    """
    Crear nuevo tipo de detalle
    Si ya existe un registro con la misma empresa y código, muestra los datos existentes
    para permitir modificarlos
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        costo = request.POST.get('costo', '0.000').strip()
        # Flag para indicar si se está editando un registro existente
        editar_existente = request.POST.get('editar_existente', '') == '1'
        
        if codigo and descripcion:
            try:
                from decimal import Decimal
                
                # Validar que el código no exceda 4 caracteres
                codigo_limite = codigo[:4]
                descripcion_limite = descripcion[:30]
                costo_decimal = Decimal(costo) if costo else Decimal('0.000')
                
                # Verificar si ya existe un registro con la misma empresa y código
                tipo_existente = TipoDetalle.objects.filter(
                    empresa=empresa_codigo if empresa_codigo else None,
                    codigo=codigo_limite
                ).first()
                
                if tipo_existente and not editar_existente:
                    # Si existe y no se está editando explícitamente, mostrar el formulario con los datos existentes
                    messages.warning(request, f'El código "{codigo_limite}" ya existe. Los datos existentes se muestran a continuación. Puede modificarlos y guardar.')
                    context = {
                        'titulo': 'Editar Tipo de Detalle Existente - Catastro',
                        'empresa': empresa_codigo,
                        'tipo': tipo_existente,  # Pasar el tipo existente para mostrar sus datos
                    }
                    return render(request, 'tipo_detalle_form.html', context)
                elif tipo_existente and editar_existente:
                    # Si existe y se está editando explícitamente, actualizar
                    tipo_existente.descripcion = descripcion_limite
                    tipo_existente.costo = costo_decimal
                    tipo_existente.save()
                    messages.success(request, f'Tipo de detalle con código "{codigo_limite}" actualizado exitosamente.')
                    return redirect('catastro:tipos_detalle_list')
                else:
                    # Si no existe, crear uno nuevo
                    TipoDetalle.objects.create(
                        empresa=empresa_codigo if empresa_codigo else None,
                        codigo=codigo_limite,
                        descripcion=descripcion_limite,
                        costo=costo_decimal
                    )
                    messages.success(request, f'Tipo de detalle con código "{codigo_limite}" creado exitosamente.')
                    return redirect('catastro:tipos_detalle_list')
            except Exception as e:
                messages.error(request, f'Error al crear/actualizar tipo de detalle: {str(e)}')
        else:
            messages.error(request, 'Código y descripción son obligatorios.')
    
    # GET: Verificar si se está intentando crear con un código que ya existe
    codigo_buscar = request.GET.get('codigo', '').strip()
    tipo_existente = None
    
    if codigo_buscar:
        tipo_existente = TipoDetalle.objects.filter(
            empresa=empresa_codigo if empresa_codigo else None,
            codigo=codigo_buscar[:4]
        ).first()
        
        if tipo_existente:
            # Si existe, mostrar los datos existentes para edición
            messages.info(request, f'El código "{codigo_buscar[:4]}" ya existe. Puede modificar los datos a continuación.')
    
    context = {
        'titulo': 'Nuevo Tipo de Detalle - Catastro',
        'empresa': empresa_codigo,
        'tipo': tipo_existente,  # Pasar el tipo existente si se encontró
    }
    
    return render(request, 'tipo_detalle_form.html', context)

@catastro_require_auth
def tipo_detalle_update(request, pk):
    """
    Actualizar tipo de detalle
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    tipo = get_object_or_404(TipoDetalle, pk=pk)
    
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        costo = request.POST.get('costo', '0.000').strip()
        
        if codigo and descripcion:
            try:
                from decimal import Decimal
                tipo.codigo = codigo[:4]
                tipo.descripcion = descripcion[:30]
                tipo.costo = Decimal(costo) if costo else Decimal('0.000')
                tipo.save()
                messages.success(request, 'Tipo de detalle actualizado exitosamente.')
                return redirect('catastro:tipos_detalle_list')
            except Exception as e:
                messages.error(request, f'Error al actualizar tipo de detalle: {str(e)}')
        else:
            messages.error(request, 'Código y descripción son obligatorios.')
    
    context = {
        'titulo': f'Editar Tipo de Detalle - Catastro',
        'tipo': tipo,
    }
    
    return render(request, 'tipo_detalle_form.html', context)

@catastro_require_auth
def tipo_detalle_delete(request, pk):
    """
    Eliminar tipo de detalle
    """
    tipo = get_object_or_404(TipoDetalle, pk=pk)
    
    if request.method == 'POST':
        tipo.delete()
        messages.success(request, 'Tipo de detalle eliminado exitosamente.')
        return redirect('catastro:tipos_detalle_list')
    
    context = {
        'titulo': f'Eliminar Tipo de Detalle - Catastro',
        'tipo': tipo,
    }
    
    return render(request, 'tipo_detalle_confirm_delete.html', context)

@catastro_require_auth
def api_buscar_tipo_detalle(request):
    """
    API endpoint para buscar un tipo de detalle por empresa, código o descripción.
    Búsqueda interactiva: campo empresa = Empresa y campo codigo = Código o descripción.
    Si existe en tipodetalle, devuelve descripcion y costo.
    Si hay múltiples coincidencias, devuelve una lista.
    """
    try:
        from .models import TipoDetalle
        from django.db.models import Q
    except ImportError:
        return JsonResponse({'encontrado': False, 'mensaje': 'Error al cargar el modelo'}, status=500)

    empresa = request.GET.get('empresa', '').strip()
    if not empresa:
        empresa = request.session.get('catastro_empresa', '') or request.session.get('empresa', '')

    termino_busqueda = request.GET.get('codigo', '').strip() # Usamos 'codigo' como término general

    if not termino_busqueda:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar un código o descripción para buscar'
        })

    try:
        # Primero, intentar una coincidencia exacta por código
        tipo_detalle_exacto = TipoDetalle.objects.filter(codigo=termino_busqueda)
        if empresa:
            tipo_detalle_exacto = tipo_detalle_exacto.filter(empresa=empresa)
        tipo_detalle_exacto = tipo_detalle_exacto.first()

        if tipo_detalle_exacto:
            return JsonResponse({
                'encontrado': True,
                'unico': True,
                'id': tipo_detalle_exacto.id,
                'codigo': tipo_detalle_exacto.codigo,
                'descripcion': tipo_detalle_exacto.descripcion or '',
                'costo': str(tipo_detalle_exacto.costo),
            })
        else:
            # Si no hay coincidencia exacta por código, buscar por código o descripción (parcial)
            # Limitar resultados para evitar sobrecarga
            resultados_query = TipoDetalle.objects.filter(
                Q(codigo__icontains=termino_busqueda) |
                Q(descripcion__icontains=termino_busqueda)
            )
            if empresa:
                resultados_query = resultados_query.filter(empresa=empresa)

            resultados = resultados_query.order_by('codigo', 'descripcion')[:20] # Limitar a 20 resultados

            if resultados.exists():
                lista_resultados = []
                for tipo_detalle in resultados:
                    lista_resultados.append({
                        'id': tipo_detalle.id,
                        'codigo': tipo_detalle.codigo,
                        'descripcion': tipo_detalle.descripcion or '',
                        'costo': str(tipo_detalle.costo),
                    })
                return JsonResponse({
                    'encontrado': True,
                    'unico': False,
                    'resultados': lista_resultados,
                    'total': len(lista_resultados),
                    'mensaje': f'Se encontraron {len(lista_resultados)} tipos de detalle.'
                })
            else:
                mensaje = f'No se encontró un tipo de detalle con código o descripción "{termino_busqueda}"'
                if empresa:
                    mensaje += f' para la empresa "{empresa}"'
                return JsonResponse({
                    'encontrado': False,
                    'mensaje': mensaje
                })

    except Exception as e:
        logger.error(f"Error al buscar tipo de detalle: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'mensaje': f'Error al buscar: {str(e)}'
        }, status=500)

@catastro_require_auth
@csrf_exempt
def api_listar_edificaciones(request):
    """
    API endpoint para listar edificaciones filtradas por empresa y clave
    Devuelve lista de edificaciones con edifino, piso, uso, clase, calidad
    """
    try:
        from .models import Edificacion
    except ImportError:
        return JsonResponse({'exito': False, 'mensaje': 'Error al cargar el modelo'}, status=500)
    
    empresa = request.GET.get('empresa', '').strip()
    if not empresa:
        empresa = request.session.get('catastro_empresa', '') or request.session.get('empresa', '')
    
    clave = request.GET.get('clave', '').strip()
    
    if not empresa or not clave:
        return JsonResponse({
            'exito': False,
            'mensaje': 'Debe proporcionar empresa y clave'
        })
    
    try:
        edificaciones = Edificacion.objects.filter(empresa=empresa, clave=clave).order_by('edifino', 'piso')
        
        lista_edificaciones = []
        for edificacion in edificaciones:
            # Concatenar uso, clase y calidad para mostrar
            uso_clase_calidad = f"{edificacion.uso or ''}{edificacion.clase or ''}{edificacion.calidad or ''}"
            texto_display = f"Edif. {edificacion.edifino}"
            if edificacion.piso and edificacion.piso != 0:
                texto_display += f" - Piso: {edificacion.piso}"
            texto_display += f" - {uso_clase_calidad}"
            
            lista_edificaciones.append({
                'id': edificacion.id,
                'edifino': str(edificacion.edifino),
                'piso': str(edificacion.piso) if edificacion.piso and edificacion.piso != 0 else '',
                'uso': edificacion.uso or '',
                'clase': edificacion.clase or '',
                'calidad': edificacion.calidad or '',
                'uso_clase_calidad': uso_clase_calidad,
                'texto_display': texto_display,
            })
        
        return JsonResponse({
            'exito': True,
            'edificaciones': lista_edificaciones,
            'total': len(lista_edificaciones)
        })
        
    except Exception as e:
        logger.error(f"Error al listar edificaciones: {str(e)}", exc_info=True)
        return JsonResponse({
            'exito': False,
            'mensaje': f'Error al listar edificaciones: {str(e)}'
        }, status=500)

# ============================================================================
# GESTIÓN DE BARRIOS / ALDEAS
# ============================================================================

@catastro_require_auth
def barrios_list(request):
    """
    Lista de barrios / aldeas
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    barrios = Barrios.objects.all()
    if empresa_codigo:
        barrios = barrios.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        barrios = barrios.filter(
            Q(codbarrio__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(depto__icontains=search) |
            Q(codmuni__icontains=search)
        )
    
    barrios = barrios.order_by('depto', 'codmuni', 'codbarrio')
    
    context = {
        'titulo': 'Barrios / Aldeas - Catastro',
        'barrios': barrios,
        'search': search,
        'total_registros': barrios.count(),
    }
    
    return render(request, 'barrios_list.html', context)

@catastro_require_auth
def barrio_create(request):
    """
    Crear nuevo barrio / aldea
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        # Verificar si se está intentando actualizar (viene con barrio_id)
        barrio_id = request.POST.get('barrio_id', '').strip()
        
        if barrio_id:
            # Si hay un ID, redirigir a la vista de actualización
            try:
                barrio = Barrios.objects.get(pk=barrio_id, empresa=empresa_codigo)
                return redirect('catastro:barrio_update', pk=barrio_id)
            except Barrios.DoesNotExist:
                pass
        
        form = BarriosForm(request.POST, empresa=empresa_codigo)
        if form.is_valid():
            try:
                codbarrio = form.cleaned_data.get('codbarrio', '').strip()
                
                # Verificar si ya existe un barrio con esta empresa y código
                barrio_existente = Barrios.objects.filter(empresa=empresa_codigo, codbarrio=codbarrio).first()
                
                if barrio_existente:
                    # Si existe, redirigir a edición
                    messages.info(request, f'El barrio con código {codbarrio} ya existe. Redirigiendo a edición.')
                    return redirect('catastro:barrio_update', pk=barrio_existente.id)
                
                barrio = form.save(commit=False)
                barrio.empresa = empresa_codigo
                
                # Calcular depto y codmuni desde empresa
                # depto = primeros dos dígitos de empresa
                # codmuni = últimos dos dígitos de empresa
                if empresa_codigo and len(empresa_codigo) >= 4:
                    barrio.depto = empresa_codigo[:2]  # Primeros 2 dígitos
                    barrio.codmuni = empresa_codigo[2:4]  # Últimos 2 dígitos
                elif empresa_codigo and len(empresa_codigo) == 2:
                    # Si solo tiene 2 dígitos, usar ambos para depto y codmuni
                    barrio.depto = empresa_codigo
                    barrio.codmuni = empresa_codigo
                else:
                    barrio.depto = ''
                    barrio.codmuni = ''
                
                barrio.save()
                messages.success(request, 'Barrio / Aldea creado exitosamente.')
                return redirect('catastro:barrios_list')
            except Exception as e:
                messages.error(request, f'Error al crear barrio: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = BarriosForm(empresa=empresa_codigo)
    
    context = {
        'titulo': 'Nuevo Barrio / Aldea - Catastro',
        'form': form,
    }
    
    return render(request, 'barrio_form.html', context)

@catastro_require_auth
def barrio_update(request, pk):
    """
    Actualizar barrio / aldea
    """
    barrio = get_object_or_404(Barrios, pk=pk)
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        form = BarriosForm(request.POST, instance=barrio, empresa=empresa_codigo)
        if form.is_valid():
            try:
                barrio = form.save(commit=False)
                barrio.empresa = empresa_codigo
                
                # Calcular depto y codmuni desde empresa
                # depto = primeros dos dígitos de empresa
                # codmuni = últimos dos dígitos de empresa
                if empresa_codigo and len(empresa_codigo) >= 4:
                    barrio.depto = empresa_codigo[:2]  # Primeros 2 dígitos
                    barrio.codmuni = empresa_codigo[2:4]  # Últimos 2 dígitos
                elif empresa_codigo and len(empresa_codigo) == 2:
                    # Si solo tiene 2 dígitos, usar ambos para depto y codmuni
                    barrio.depto = empresa_codigo
                    barrio.codmuni = empresa_codigo
                else:
                    barrio.depto = ''
                    barrio.codmuni = ''
                
                barrio.save()
                messages.success(request, 'Barrio / Aldea actualizado exitosamente.')
                return redirect('catastro:barrios_list')
            except Exception as e:
                messages.error(request, f'Error al actualizar barrio: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = BarriosForm(instance=barrio, empresa=empresa_codigo)
    
    context = {
        'titulo': f'Editar Barrio / Aldea - Catastro',
        'form': form,
        'barrio': barrio,
    }
    
    return render(request, 'barrio_form.html', context)

@catastro_require_auth
def barrio_delete(request, pk):
    """
    Eliminar barrio / aldea
    """
    barrio = get_object_or_404(Barrios, pk=pk)
    
    if request.method == 'POST':
        try:
            barrio.delete()
            messages.success(request, 'Barrio / Aldea eliminado exitosamente.')
            return redirect('catastro:barrios_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar barrio: {str(e)}')
            return redirect('catastro:barrios_list')
    
    context = {
        'titulo': 'Eliminar Barrio / Aldea - Catastro',
        'barrio': barrio,
    }
    
    return render(request, 'barrio_confirm_delete.html', context)

@catastro_require_auth
def barrios_export_excel(request):
    """
    Exportar barrios a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:barrios_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener los barrios
    barrios = Barrios.objects.all()
    if empresa_codigo:
        barrios = barrios.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        barrios = barrios.filter(
            Q(codbarrio__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(depto__icontains=search) |
            Q(codmuni__icontains=search)
        )
    
    barrios = barrios.order_by('depto', 'codmuni', 'codbarrio')
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barrios"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados (solo código y descripción)
    headers = ['Código Barrio', 'Descripción']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos (solo código y descripción)
    for barrio in barrios:
        row = [
            barrio.codbarrio or '',
            barrio.descripcion or ''
        ]
        ws.append(row)
    
    # Aplicar estilo a datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar ancho de columnas
    column_widths = [20, 40]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'barrios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def barrios_export_pdf(request):
    """
    Exportar barrios a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:barrios_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener los barrios
    barrios = Barrios.objects.all()
    if empresa_codigo:
        barrios = barrios.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        barrios = barrios.filter(
            Q(codbarrio__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(depto__icontains=search) |
            Q(codmuni__icontains=search)
        )
    
    barrios = barrios.order_by('depto', 'codmuni', 'codbarrio')
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = "Barrios / Aldeas - Catastro"
    if empresa_codigo:
        title_text += f"<br/>Empresa: {empresa_codigo}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Preparar datos de la tabla
    data = []
    
    # Encabezados (solo código y descripción)
    headers = ['Código Barrio', 'Descripción']
    data.append(headers)
    
    # Datos (solo código y descripción)
    for barrio in barrios:
        row = [
            barrio.codbarrio or '-',
            barrio.descripcion or '-'
        ]
        data.append(row)
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Información adicional
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        alignment=0  # Izquierda
    )
    info_text = f"<b>Total de registros:</b> {barrios.count()}<br/>"
    info_text += f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'barrios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


@catastro_require_auth
def usos_predio_list(request):
    """
    Listado de usos de predio (tabla usos).
    Desde cada fila se abre el catálogo de sub usos en Configuración general filtrado por ese uso.
    """
    usos_qs = Usos.objects.all()
    search = (request.GET.get('search') or '').strip()
    if search:
        usos_qs = usos_qs.filter(Q(uso__icontains=search) | Q(desuso__icontains=search))
    usos_qs = usos_qs.order_by('uso')
    context = {
        'titulo': 'Usos del predio — Catastro',
        'usos': usos_qs,
        'search': search,
        'total_registros': usos_qs.count(),
    }
    return render(request, 'usos_predio_list.html', context)


@catastro_require_auth
def usos_predio_create(request):
    """Alta de uso de predio (tabla usos)."""
    if request.method == 'POST':
        form = UsosPredioForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Uso guardado correctamente.')
                return redirect('catastro:usos_predio_list')
            except IntegrityError:
                messages.error(request, 'Ya existe un uso con ese código.')
        else:
            messages.error(request, 'Revise los datos del formulario.')
    else:
        form = UsosPredioForm()
    return render(request, 'usos_predio_form.html', {
        'titulo': 'Nuevo uso del predio — Catastro',
        'form': form,
        'modo': 'create',
    })


@catastro_require_auth
def usos_predio_update(request, pk):
    """Edición de uso de predio (no se modifica el código de uso)."""
    uso_obj = get_object_or_404(Usos, pk=pk)
    if request.method == 'POST':
        posted_uso = (request.POST.get('uso') or '').strip()
        if posted_uso != uso_obj.uso:
            messages.error(request, 'No está permitido cambiar el código de uso.')
            return redirect('catastro:usos_predio_update', pk=pk)
        form = UsosPredioForm(request.POST, instance=uso_obj, uso_readonly=True)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Uso actualizado correctamente.')
                return redirect('catastro:usos_predio_list')
            except IntegrityError:
                messages.error(request, 'No se pudo guardar el registro.')
        else:
            messages.error(request, 'Revise los datos del formulario.')
    else:
        form = UsosPredioForm(instance=uso_obj, uso_readonly=True)
    return render(request, 'usos_predio_form.html', {
        'titulo': 'Editar uso del predio — Catastro',
        'form': form,
        'uso_obj': uso_obj,
        'modo': 'update',
    })


@catastro_require_auth
def usos_predio_delete(request, pk):
    """Baja de uso si no hay predios ni sub usos que lo referencien."""
    uso_obj = get_object_or_404(Usos, pk=pk)
    sub_count = Subuso.objects.filter(uso=uso_obj.uso).count()
    pred_count = BDCata1.objects.filter(uso=uso_obj.uso).count()

    if request.method == 'POST':
        if sub_count > 0 or pred_count > 0:
            messages.error(
                request,
                'No se puede eliminar: hay sub usos o predios asociados a este código de uso.',
            )
            return redirect('catastro:usos_predio_list')
        try:
            uso_obj.delete()
            messages.success(request, 'Uso eliminado correctamente.')
        except IntegrityError as e:
            messages.error(request, f'No se puede eliminar el uso: {e}')
        return redirect('catastro:usos_predio_list')

    return render(request, 'usos_predio_confirm_delete.html', {
        'titulo': 'Eliminar uso del predio — Catastro',
        'uso_obj': uso_obj,
        'sub_count': sub_count,
        'pred_count': pred_count,
        'bloqueado': sub_count > 0 or pred_count > 0,
    })


@catastro_require_auth
def topografia_list(request):
    """
    Lista de topografías del predio
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    topografias = Topografia.objects.all()
    if empresa_codigo:
        topografias = topografias.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        topografias = topografias.filter(
            Q(cotopo__icontains=search) |
            Q(descritopo__icontains=search)
        )
    
    topografias = topografias.order_by('cotopo')
    
    context = {
        'titulo': 'Topografía del Predio - Catastro',
        'topografias': topografias,
        'search': search,
        'total_registros': topografias.count(),
    }
    
    return render(request, 'topografia_list.html', context)

@catastro_require_auth
def topografia_create(request):
    """
    Crear nueva topografía
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        # Obtener cotopo del POST antes de validar el formulario
        cotopo_post = request.POST.get('cotopo', '').strip()
        
        # Verificar si ya existe una topografía con esta empresa y código ANTES de validar
        # La restricción única es sobre (empresa, cotopo) juntos
        topografia_existente = None
        if empresa_codigo and cotopo_post:
            topografia_existente = Topografia.objects.filter(
                empresa=empresa_codigo,
                cotopo=cotopo_post
            ).first()
        
        # Si existe un registro, crear el formulario con la instancia existente (modo actualización)
        if topografia_existente:
            form = TopografiaForm(request.POST, instance=topografia_existente, empresa=empresa_codigo)
        else:
            # Si no existe, crear formulario normal (modo creación)
            form = TopografiaForm(request.POST, empresa=empresa_codigo)
        
        if form.is_valid():
            try:
                # Guardar usando el método del formulario (funciona tanto para crear como actualizar)
                topografia = form.save(commit=False)
                topografia.empresa = empresa_codigo
                topografia.save()
                
                if topografia_existente:
                    logger.info(f"Topografía actualizada: ID={topografia.id}, empresa={empresa_codigo}, cotopo={cotopo_post}, factopo={topografia.factopo}, descritopo={topografia.descritopo}")
                    messages.success(request, f'Topografía con código {cotopo_post} actualizada exitosamente.')
                else:
                    logger.info(f"Topografía creada: ID={topografia.id}, empresa={empresa_codigo}, cotopo={cotopo_post}, factopo={topografia.factopo}, descritopo={topografia.descritopo}")
                    messages.success(request, 'Topografía creada exitosamente.')
                
                return redirect('catastro:topografia_list')
            except Exception as e:
                logger.error(f"Error al crear/actualizar topografía: {str(e)}", exc_info=True)
                messages.error(request, f'Error al crear/actualizar topografía: {str(e)}')
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TopografiaForm(empresa=empresa_codigo)
    
    context = {
        'titulo': 'Nueva Topografía - Catastro',
        'form': form,
        'empresa_codigo': empresa_codigo,
    }
    
    return render(request, 'topografia_form.html', context)

@catastro_require_auth
def topografia_update(request, pk):
    """
    Actualizar topografía
    """
    topografia = get_object_or_404(Topografia, pk=pk)
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        form = TopografiaForm(request.POST, instance=topografia, empresa=empresa_codigo)
        if form.is_valid():
            try:
                topografia = form.save(commit=False)
                topografia.empresa = empresa_codigo
                topografia.save()
                messages.success(request, 'Topografía actualizada exitosamente.')
                return redirect('catastro:topografia_list')
            except Exception as e:
                messages.error(request, f'Error al actualizar topografía: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = TopografiaForm(instance=topografia, empresa=empresa_codigo)
    
    context = {
        'titulo': 'Editar Topografía - Catastro',
        'form': form,
        'topografia': topografia,
        'empresa_codigo': empresa_codigo,
    }
    
    return render(request, 'topografia_form.html', context)

@catastro_require_auth
def topografia_delete(request, pk):
    """
    Eliminar topografía
    """
    topografia = get_object_or_404(Topografia, pk=pk)
    
    if request.method == 'POST':
        try:
            topografia.delete()
            messages.success(request, 'Topografía eliminada exitosamente.')
            return redirect('catastro:topografia_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar topografía: {str(e)}')
            return redirect('catastro:topografia_list')
    
    context = {
        'titulo': 'Eliminar Topografía - Catastro',
        'topografia': topografia,
    }
    
    return render(request, 'topografia_confirm_delete.html', context)

@catastro_require_auth
def topografia_export_excel(request):
    """
    Exportar topografías a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:topografia_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener las topografías
    topografias = Topografia.objects.all()
    if empresa_codigo:
        topografias = topografias.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        topografias = topografias.filter(
            Q(cotopo__icontains=search) |
            Q(descritopo__icontains=search)
        )
    
    topografias = topografias.order_by('cotopo')
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Topografías"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Código', 'Descripción', 'Factor Topografía']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    for topografia in topografias:
        row = [
            topografia.cotopo or '',
            topografia.descritopo or '',
            float(topografia.factopo) if topografia.factopo else 0.00
        ]
        ws.append(row)
    
    # Aplicar estilo a datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border_style
            if cell.column == 3:  # Columna de factor (numérica)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar ancho de columnas
    column_widths = [15, 40, 20]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'topografias_{empresa_codigo or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def topografia_export_pdf(request):
    """
    Exportar topografías a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from io import BytesIO
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:topografia_list')
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener las topografías
    topografias = Topografia.objects.all()
    if empresa_codigo:
        topografias = topografias.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        topografias = topografias.filter(
            Q(cotopo__icontains=search) |
            Q(descritopo__icontains=search)
        )
    
    topografias = topografias.order_by('cotopo')
    
    # Crear buffer para el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = "Topografías del Predio - Catastro"
    if empresa_codigo:
        title_text += f"<br/>Empresa: {empresa_codigo}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Preparar datos de la tabla
    data = []
    
    # Encabezados
    headers = ['Código', 'Descripción', 'Factor Topografía']
    data.append(headers)
    
    # Datos
    for topografia in topografias:
        row = [
            topografia.cotopo or '-',
            topografia.descritopo or '-',
            f"{float(topografia.factopo):.2f}" if topografia.factopo else '0.00'
        ]
        data.append(row)
    
    # Crear tabla
    table = Table(data)
    
    # Estilo de la tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Factor alineado a la derecha
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Información adicional
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        alignment=0  # Izquierda
    )
    info_text = f"<b>Total de registros:</b> {topografias.count()}<br/>"
    info_text += f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'topografias_{empresa_codigo or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response

@catastro_require_auth
@csrf_exempt
@require_http_methods(["GET"])
def api_buscar_topografia(request):
    """
    API endpoint para buscar topografía por empresa y código
    """
    try:
        from .models import Topografia
    except ImportError:
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    # Obtener empresa de la sesión
    empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '') or request.session.get('empresa', '')
    cotopo = request.GET.get('cotopo', '').strip()
    
    if not empresa:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'No se encontró información de empresa en la sesión'
        })
    
    if not cotopo:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar el código de topografía'
        })
    
    try:
        topografia_obj = Topografia.objects.filter(
            empresa=empresa,
            cotopo=cotopo
        ).first()
        
        if topografia_obj:
            return JsonResponse({
                'encontrado': True,
                'descritopo': topografia_obj.descritopo or '',
                'factopo': str(topografia_obj.factopo) if topografia_obj.factopo else '0.00',
                'id': topografia_obj.id,
                'mensaje': 'Topografía encontrada'
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': f'No se encontró topografía para empresa={empresa}, código={cotopo}'
            })
            
    except Exception as e:
        logger.error(f"Error al buscar topografía: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'error': f'Error al buscar topografía: {str(e)}'
        }, status=500)

@catastro_require_auth
@csrf_exempt
@require_http_methods(["GET"])
def api_buscar_barrio(request):
    """
    API endpoint para buscar un barrio por empresa y código de barrio
    """
    try:
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        codbarrio = request.GET.get('codbarrio', '').strip()

        if not empresa or not codbarrio:
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'Debe proporcionar empresa y código de barrio.'
            })

        barrio = Barrios.objects.filter(empresa=empresa, codbarrio=codbarrio).first()

        if barrio:
            return JsonResponse({
                'encontrado': True,
                'data': {
                    'id': barrio.id,
                    'empresa': barrio.empresa or '',
                    'codbarrio': barrio.codbarrio or '',
                    'descripcion': barrio.descripcion or '',
                    'tipica': str(barrio.tipica) if barrio.tipica else '0.00',
                }
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'Barrio no encontrado.'
            })

    except Exception as e:
        logger.error(f"Error en api_buscar_barrio: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'mensaje': f'Error interno del servidor: {str(e)}'
        }, status=500)

# ============================================================================
# GESTIÓN DE TIPOS DE MATERIAL
# ============================================================================

@catastro_require_auth
def tipo_material_list(request):
    """
    Lista de tipos de material
    """
    tipos_material = TipoMaterial.objects.all()
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        tipos_material = tipos_material.filter(
            Q(No__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    tipos_material = tipos_material.order_by('No')
    
    context = {
        'titulo': 'Tipos de Material - Catastro',
        'tipos_material': tipos_material,
        'search': search,
        'total_registros': tipos_material.count(),
    }
    
    return render(request, 'tipo_material_list.html', context)

@catastro_require_auth
def tipo_material_create(request):
    """
    Crear nuevo tipo de material
    """
    if request.method == 'POST':
        form = TipoMaterialForm(request.POST)
        if form.is_valid():
            try:
                tipo_material = form.save()
                messages.success(request, 'Tipo de material creado exitosamente.')
                return redirect('catastro:tipo_material_list')
            except Exception as e:
                logger.error(f"Error al crear tipo de material: {str(e)}", exc_info=True)
                messages.error(request, f'Error al crear tipo de material: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TipoMaterialForm()
    
    context = {
        'titulo': 'Nuevo Tipo de Material - Catastro',
        'form': form,
    }
    
    return render(request, 'tipo_material_form.html', context)

@catastro_require_auth
def tipo_material_update(request, pk):
    """
    Actualizar tipo de material
    """
    tipo_material = get_object_or_404(TipoMaterial, pk=pk)
    
    if request.method == 'POST':
        form = TipoMaterialForm(request.POST, instance=tipo_material)
        if form.is_valid():
            try:
                tipo_material = form.save()
                messages.success(request, 'Tipo de material actualizado exitosamente.')
                return redirect('catastro:tipo_material_list')
            except Exception as e:
                messages.error(request, f'Error al actualizar tipo de material: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = TipoMaterialForm(instance=tipo_material)
    
    context = {
        'titulo': 'Editar Tipo de Material - Catastro',
        'form': form,
        'tipo_material': tipo_material,
    }
    
    return render(request, 'tipo_material_form.html', context)

@catastro_require_auth
def tipo_material_delete(request, pk):
    """
    Eliminar tipo de material
    """
    tipo_material = get_object_or_404(TipoMaterial, pk=pk)
    
    if request.method == 'POST':
        try:
            tipo_material.delete()
            messages.success(request, 'Tipo de material eliminado exitosamente.')
            return redirect('catastro:tipo_material_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar tipo de material: {str(e)}')
            return redirect('catastro:tipo_material_list')
    
    context = {
        'titulo': 'Eliminar Tipo de Material - Catastro',
        'tipo_material': tipo_material,
    }
    
    return render(request, 'tipo_material_confirm_delete.html', context)

# ============================================================================
# GESTIÓN DE USOS DE EDIFICACIÓN
# ============================================================================

@catastro_require_auth
def usos_edifica_list(request):
    """
    Lista de usos de edificación (catálogo)
    """
    usos = UsoEdifica.objects.all()
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        usos = usos.filter(
            Q(codigo__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    usos = usos.order_by('codigo')
    
    context = {
        'titulo': 'Usos de Edificación - Catastro',
        'usos': usos,
        'search': search,
        'total_registros': usos.count(),
    }
    
    return render(request, 'usos_edifica_list.html', context)

@catastro_require_auth
def uso_edifica_create(request):
    """
    Crear nuevo uso de edificación
    Si ya existe un registro con el mismo código, muestra los datos existentes
    para permitir modificarlos
    """
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        # Flag para indicar si se está editando un registro existente
        editar_existente = request.POST.get('editar_existente', '') == '1'
        
        if codigo and descripcion:
            try:
                # Validar que el código no exceda 3 caracteres
                codigo_limite = codigo[:3]
                descripcion_limite = descripcion[:50]
                
                # Verificar si ya existe un registro con el mismo código
                uso_existente = UsoEdifica.objects.filter(
                    codigo=codigo_limite
                ).first()
                
                if uso_existente and not editar_existente:
                    # Si existe y no se está editando explícitamente, mostrar el formulario con los datos existentes
                    messages.warning(request, f'El código "{codigo_limite}" ya existe. Los datos existentes se muestran a continuación. Puede modificarlos y guardar.')
                    context = {
                        'titulo': 'Editar Uso de Edificación Existente - Catastro',
                        'uso': uso_existente,
                    }
                    return render(request, 'uso_edifica_form.html', context)
                elif uso_existente and editar_existente:
                    # Si existe y se está editando explícitamente, actualizar
                    uso_existente.descripcion = descripcion_limite
                    uso_existente.save()
                    messages.success(request, f'Uso de edificación con código "{codigo_limite}" actualizado exitosamente.')
                    return redirect('catastro:usos_edifica_list')
                else:
                    # Si no existe, crear uno nuevo
                    UsoEdifica.objects.create(
                        codigo=codigo_limite,
                        descripcion=descripcion_limite
                    )
                    messages.success(request, f'Uso de edificación con código "{codigo_limite}" creado exitosamente.')
                    return redirect('catastro:usos_edifica_list')
            except Exception as e:
                messages.error(request, f'Error al crear/actualizar uso de edificación: {str(e)}')
        else:
            messages.error(request, 'Código y descripción son obligatorios.')
    
    # GET: Verificar si se está intentando crear con un código que ya existe
    codigo_buscar = request.GET.get('codigo', '').strip()
    uso_existente = None
    
    if codigo_buscar:
        uso_existente = UsoEdifica.objects.filter(
            codigo=codigo_buscar[:3]
        ).first()
        
        if uso_existente:
            # Si existe, mostrar los datos existentes para edición
            messages.info(request, f'El código "{codigo_buscar[:3]}" ya existe. Puede modificar los datos a continuación.')
    
    context = {
        'titulo': 'Nuevo Uso de Edificación - Catastro',
        'uso': uso_existente,
    }
    
    return render(request, 'uso_edifica_form.html', context)

@catastro_require_auth
def uso_edifica_update(request, pk):
    """
    Actualizar uso de edificación
    """
    uso = get_object_or_404(UsoEdifica, pk=pk)
    
    if request.method == 'POST':
        codigo = request.POST.get('codigo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        if codigo and descripcion:
            try:
                uso.codigo = codigo[:3]
                uso.descripcion = descripcion[:50]
                uso.save()
                messages.success(request, 'Uso de edificación actualizado exitosamente.')
                return redirect('catastro:usos_edifica_list')
            except Exception as e:
                messages.error(request, f'Error al actualizar uso de edificación: {str(e)}')
        else:
            messages.error(request, 'Código y descripción son obligatorios.')
    
    context = {
        'titulo': f'Editar Uso de Edificación - Catastro',
        'uso': uso,
    }
    
    return render(request, 'uso_edifica_form.html', context)

@catastro_require_auth
def uso_edifica_delete(request, pk):
    """
    Eliminar uso de edificación
    """
    uso = get_object_or_404(UsoEdifica, pk=pk)
    
    if request.method == 'POST':
        uso.delete()
        messages.success(request, 'Uso de edificación eliminado exitosamente.')
        return redirect('catastro:usos_edifica_list')
    
    context = {
        'titulo': f'Eliminar Uso de Edificación - Catastro',
        'uso': uso,
    }
    
    return render(request, 'uso_edifica_confirm_delete.html', context)

@catastro_require_auth
def api_buscar_uso_edifica(request):
    """
    API endpoint para buscar un uso de edificación por código
    Búsqueda interactiva: campo codigo = Código
    Si existe en usoedifica, devuelve descripcion
    """
    try:
        from .models import UsoEdifica
    except ImportError:
        return JsonResponse({'encontrado': False, 'mensaje': 'Error al cargar el modelo'}, status=500)
    
    codigo = request.GET.get('codigo', '').strip()
    
    if not codigo:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar un código'
        })
    
    try:
        # Buscar uso de edificación por código
        uso_edifica = UsoEdifica.objects.filter(codigo=codigo[:3]).first()
        
        if uso_edifica:
            # Si existe, devolver descripcion e ID
            return JsonResponse({
                'encontrado': True,
                'id': uso_edifica.id,
                'codigo': uso_edifica.codigo,
                'descripcion': uso_edifica.descripcion or '',
            })
        else:
            # No se encontró el uso de edificación
            return JsonResponse({
                'encontrado': False,
                'mensaje': f'No se encontró un uso de edificación con código "{codigo}"'
            })
            
    except Exception as e:
        logger.error(f"Error al buscar uso de edificación: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'mensaje': f'Error al buscar: {str(e)}'
        }, status=500)

@catastro_require_auth
def confi_tipologia_list(request):
    """
    Lista de configuraciones de tipología
    """
    tipologias = ConfiTipologia.objects.all()
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        tipologias = tipologias.filter(
            Q(uso__icontains=search) |
            Q(clase__icontains=search) |
            Q(tipo__icontains=search) |
            Q(categoria__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    # Ordenar numéricamente por tipo (para que 10 venga después de 9)
    # Coalesce maneja valores NULL convirtiéndolos a '0' antes de convertir a entero
    tipologias = tipologias.annotate(
        tipo_numeric=Cast(Coalesce('tipo', Value('0')), IntegerField())
    ).order_by('uso', 'clase', 'tipo_numeric')
    
    context = {
        'titulo': 'Configuraciones de Tipología - Catastro',
        'tipologias': tipologias,
        'search': search,
        'total_registros': tipologias.count(),
    }
    
    return render(request, 'confi_tipologia_list.html', context)

@catastro_require_auth
def confi_tipologia_create(request):
    """
    Crear nueva configuración de tipología
    Si ya existe un registro con los mismos uso, clase, tipo y categoria, lo actualiza
    """
    if request.method == 'POST':
        # Obtener valores del POST antes de crear el formulario
        uso = request.POST.get('uso', '').strip()
        clase = request.POST.get('clase', '').strip()
        tipo = request.POST.get('tipo', '').strip() or None
        categoria = request.POST.get('categoria', '').strip() or None
        
        # Buscar si ya existe un registro con los mismos parámetros ANTES de validar el formulario
        tipologia_existente = None
        if uso and clase and tipo is not None:
            filtros = {
                'uso': uso,
                'clase': clase,
                'tipo': tipo
            }
            if categoria:
                filtros['categoria'] = categoria
            else:
                filtros['categoria__isnull'] = True
            
            tipologia_existente = ConfiTipologia.objects.filter(**filtros).first()
        
        # Si existe un registro, crear el formulario con la instancia para que sea una actualización
        if tipologia_existente:
            form = ConfiTipologiaForm(request.POST, instance=tipologia_existente)
        else:
            form = ConfiTipologiaForm(request.POST)
        
        if form.is_valid():
            tipologia = form.save()
            
            if tipologia_existente:
                messages.success(request, 'Configuración de tipología actualizada exitosamente.')
            else:
                messages.success(request, 'Configuración de tipología creada exitosamente.')
            
            return redirect('catastro:confi_tipologia_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ConfiTipologiaForm()
    
    context = {
        'titulo': 'Nueva Configuración de Tipología - Catastro',
        'form': form,
    }
    
    return render(request, 'confi_tipologia_form.html', context)

@catastro_require_auth
def confi_tipologia_update(request, pk):
    """
    Actualizar configuración de tipología
    """
    tipologia = get_object_or_404(ConfiTipologia, pk=pk)
    
    if request.method == 'POST':
        form = ConfiTipologiaForm(request.POST, instance=tipologia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración de tipología actualizada exitosamente.')
            return redirect('catastro:confi_tipologia_list')
    else:
        form = ConfiTipologiaForm(instance=tipologia)
    
    context = {
        'titulo': f'Editar Configuración de Tipología - Catastro',
        'form': form,
        'tipologia': tipologia,
    }
    
    return render(request, 'confi_tipologia_form.html', context)

@catastro_require_auth
def confi_tipologia_create_clasificacion(request, uso, clase):
    """
    Crear nueva configuración de tipología desde la página de clasificación de pesos
    Si ya existe un registro con los mismos uso, clase, tipo y categoria, lo actualiza
    """
    if request.method == 'POST':
        # Obtener tipo y categoria del POST antes de crear el formulario
        tipo = request.POST.get('tipo', '').strip()
        categoria = request.POST.get('categoria', '').strip() or None
        
        # Buscar si ya existe un registro con los mismos parámetros ANTES de validar el formulario
        tipologia_existente = None
        if tipo:
            filtros = {
                'uso': uso,
                'clase': clase,
                'tipo': tipo
            }
            if categoria:
                filtros['categoria'] = categoria
            else:
                filtros['categoria__isnull'] = True
            
            tipologia_existente = ConfiTipologia.objects.filter(**filtros).first()
        
        # Si existe un registro, crear el formulario con la instancia para que sea una actualización
        if tipologia_existente:
            form = ConfiTipologiaForm(request.POST, instance=tipologia_existente)
        else:
            form = ConfiTipologiaForm(request.POST)
        
        if form.is_valid():
            tipologia = form.save(commit=False)
            # Asegurar que uso y clase sean los correctos
            tipologia.uso = uso
            tipologia.clase = clase
            tipologia.save()
            
            if tipologia_existente:
                messages.success(request, 'Registro actualizado exitosamente.')
            else:
                messages.success(request, 'Registro creado exitosamente.')
            
            return redirect('catastro:costos_clasificacion_pesos', uso=uso, clase=clase)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        # Pre-llenar uso y clase
        form = ConfiTipologiaForm(initial={'uso': uso, 'clase': clase})
        # Hacer uso y clase readonly
        form.fields['uso'].widget.attrs['readonly'] = True
        form.fields['clase'].widget.attrs['readonly'] = True
    
    # Obtener tipos de material para el contexto
    tipos_material = TipoMaterial.objects.all().order_by('No')
    
    context = {
        'titulo': 'Nueva Clasificación de Pesos - Catastro',
        'form': form,
        'uso': uso,
        'clase': clase,
        'tipos_material': tipos_material,
        'volver_url': 'catastro:costos_clasificacion_pesos',
        'volver_params': {'uso': uso, 'clase': clase},
    }
    
    return render(request, 'confi_tipologia_form_clasificacion.html', context)

@catastro_require_auth
def confi_tipologia_update_clasificacion(request, pk, uso, clase):
    """
    Actualizar configuración de tipología desde la página de clasificación de pesos
    """
    tipologia = get_object_or_404(ConfiTipologia, pk=pk, uso=uso, clase=clase)
    
    if request.method == 'POST':
        form = ConfiTipologiaForm(request.POST, instance=tipologia)
        if form.is_valid():
            tipologia = form.save(commit=False)
            # Asegurar que uso y clase sean los correctos
            tipologia.uso = uso
            tipologia.clase = clase
            tipologia.save()
            messages.success(request, 'Registro actualizado exitosamente.')
            return redirect('catastro:costos_clasificacion_pesos', uso=uso, clase=clase)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ConfiTipologiaForm(instance=tipologia)
        # Hacer uso y clase readonly
        form.fields['uso'].widget.attrs['readonly'] = True
        form.fields['clase'].widget.attrs['readonly'] = True
    
    # Obtener tipos de material para el contexto
    tipos_material = TipoMaterial.objects.all().order_by('No')
    
    context = {
        'titulo': 'Editar Clasificación de Pesos - Catastro',
        'form': form,
        'tipologia': tipologia,
        'uso': uso,
        'clase': clase,
        'tipos_material': tipos_material,
        'volver_url': 'catastro:costos_clasificacion_pesos',
        'volver_params': {'uso': uso, 'clase': clase},
    }
    
    return render(request, 'confi_tipologia_form_clasificacion.html', context)

@catastro_require_auth
def confi_tipologia_delete(request, pk):
    """
    Eliminar configuración de tipología
    """
    tipologia = get_object_or_404(ConfiTipologia, pk=pk)
    
    # Obtener parámetros de redirección si vienen desde clasificación de pesos
    uso = request.GET.get('uso', '')
    clase = request.GET.get('clase', '')
    
    if request.method == 'POST':
        uso_post = request.POST.get('uso', '')
        clase_post = request.POST.get('clase', '')
        tipologia.delete()
        messages.success(request, 'Configuración de tipología eliminada exitosamente.')
        
        # Si viene desde clasificación de pesos, redirigir ahí
        if uso_post and clase_post:
            return redirect('catastro:costos_clasificacion_pesos', uso=uso_post, clase=clase_post)
        return redirect('catastro:confi_tipologia_list')
    
    context = {
        'titulo': f'Eliminar Configuración de Tipología - Catastro',
        'tipologia': tipologia,
        'uso': uso,
        'clase': clase,
    }
    
    return render(request, 'confi_tipologia_confirm_delete.html', context)

@catastro_require_auth
def api_calcular_calidad(request):
    """
    API endpoint para calcular la calidad de una edificación basándose en parámetros.
    Utiliza la tabla costos para determinar la calidad según uso, clase y pesos totales.
    Busca en costos donde: uso, clase y pesos_total está entre rango1 y rango2.
    """
    try:
        from decimal import Decimal
        
        empresa_codigo_api = (request.session.get('catastro_empresa') or request.session.get('empresa') or '').strip()
        empresas_buscar = codigos_empresa_equivalentes(empresa_codigo_api) if empresa_codigo_api else []
        
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        pesos_total = request.GET.get('pesos_total', '').strip()
        
        if not empresa_codigo_api:
            return JsonResponse({
                'calidad': None,
                'mensaje': 'No hay código de municipio en sesión (empresa).'
            })
        
        if not uso or not clase:
            return JsonResponse({
                'calidad': None,
                'mensaje': 'Uso y Clase son requeridos para calcular la calidad.'
            })
        
        if not pesos_total:
            return JsonResponse({
                'calidad': None,
                'mensaje': 'Pesos totales es requerido para calcular la calidad.'
            })
        
        try:
            pesos_total_decimal = Decimal(str(pesos_total))
        except (ValueError, InvalidOperation):
            return JsonResponse({
                'calidad': None,
                'mensaje': f'Pesos totales inválido: {pesos_total}'
            })
        
        # Buscar en costos del municipio en sesión donde Pesos Totales está entre rango1 y rango2
        qs_calidad = Costos.objects.filter(
            empresa__in=empresas_buscar,
            uso=uso,
            clase=clase,
            rango1__lte=pesos_total_decimal,
            rango2__gte=pesos_total_decimal
        )
        costo = qs_calidad.first()
        
        if costo:
            logger.info(f"Calidad encontrada en costos: {costo.calidad} para uso={uso}, clase={clase}, pesos={pesos_total_decimal} (rango: {costo.rango1}-{costo.rango2})")
            return JsonResponse({
                'calidad': costo.calidad,
                'costo': str(costo.costo),
                'rango1': str(costo.rango1),
                'rango2': str(costo.rango2),
                'mensaje': f'Calidad encontrada: {costo.calidad}'
            })
        else:
            logger.warning(f"No se encontró calidad en costos para uso={uso}, clase={clase}, pesos={pesos_total_decimal}")
            return JsonResponse({
                'calidad': None,
                'mensaje': f'No se encontró calidad en costos para uso={uso}, clase={clase}, pesos={pesos_total_decimal}. Verifique que existan registros en la tabla costos con estos parámetros y rangos.'
            })
            
    except Exception as e:
        logger.error(f"Error al calcular calidad: {str(e)}", exc_info=True)
        return JsonResponse({
            'calidad': None,
            'mensaje': f'Error al calcular la calidad: {str(e)}'
        }, status=500)

@catastro_require_auth
def especificaciones_calcular_calidad(request):
    """
    Formulario para calcular calidad basado en especificaciones
    Recibe parámetros: edifino, piso, uso, clase
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    # Obtener parámetros de la URL
    edifino = request.GET.get('edifino', '').strip()
    piso = request.GET.get('piso', '').strip()
    uso = request.GET.get('uso', '').strip()
    clase = request.GET.get('clase', '').strip()
    clave = request.GET.get('clave', '').strip()
    
    if request.method == 'POST':
        form = EspecificacionesForm(
            request.POST,
            clave=clave,
            usuario=current_user[:50] if current_user else '',
            fecha_sistema=timezone.now()
        )
        if form.is_valid():
            # Buscar si ya existe un registro con la misma combinación de clave, piso, edifino, uso, clase
            especificacion_existente = None
            if clave and edifino and uso and clase:
                especificacion_existente = Especificaciones.objects.filter(
                    clave=clave,
                    edifino=Decimal(edifino) if edifino else 0,
                    piso=piso if piso else None,
                    uso=uso,
                    clase=clase
                ).first()
            
            if especificacion_existente:
                # Actualizar registro existente - usar el formulario con la instancia existente
                especificacion = form.save(commit=False)
                especificacion.id = especificacion_existente.id
                especificacion.pk = especificacion_existente.pk
                logger.info(f"Actualizando especificación existente ID={especificacion_existente.id} para clave={clave}, edifino={edifino}, piso={piso}, uso={uso}, clase={clase}")
            else:
                # Crear nuevo registro
                especificacion = form.save(commit=False)
                logger.info(f"Creando nueva especificación para clave={clave}, edifino={edifino}, piso={piso}, uso={uso}, clase={clase}")
            
            # Establecer valores (asegurar que todos los campos estén correctos)
            if clave:
                especificacion.clave = clave
            if edifino:
                especificacion.edifino = Decimal(edifino)
            if piso:
                especificacion.piso = piso
            else:
                especificacion.piso = None
            if uso:
                especificacion.uso = uso
            if clase:
                especificacion.clase = clase
            especificacion.usuario = current_user[:50] if current_user else ''
            especificacion.fechasys = timezone.now()
            
            # Calcular suma de pesos
            pesos_total = (
                (especificacion.pesofun or Decimal('0')) +
                (especificacion.pesopiso or Decimal('0')) +
                (especificacion.pesoparext or Decimal('0')) +
                (especificacion.pesotecho or Decimal('0')) +
                (especificacion.pesoparint or Decimal('0')) +
                (especificacion.pesocielo or Decimal('0')) +
                (especificacion.pesocarpini or Decimal('0')) +
                (especificacion.pesoelectri or Decimal('0')) +
                (especificacion.pesoplome or Decimal('0')) +
                (especificacion.pesotros or Decimal('0'))
            )
            # Guardar el total sin redondeo
            especificacion.pesos = pesos_total
            
            # Buscar calidad en la tabla costos según empresa, uso, clase y rango de pesos
            # La calidad DEBE obtenerse de la tabla costos, NO usar lógica por defecto
            try:
                # Obtener empresa de la sesión
                empresa_codigo = request.session.get('catastro_empresa', '')
                
                if not empresa_codigo:
                    messages.error(request, 'No se encontró la empresa en la sesión. No se puede calcular la calidad.')
                    return render(request, 'especificaciones_calcular_calidad.html', {
                        'titulo': 'Calcular Calidad - Especificaciones',
                        'form': form,
                        'edifino': edifino,
                        'piso': piso,
                        'uso': uso,
                        'clase': clase,
                        'clave': clave,
                    })
                
                if not uso or not clase:
                    messages.error(request, 'Uso y Clase son requeridos para calcular la calidad.')
                    return render(request, 'especificaciones_calcular_calidad.html', {
                        'titulo': 'Calcular Calidad - Especificaciones',
                        'form': form,
                        'edifino': edifino,
                        'piso': piso,
                        'uso': uso,
                        'clase': clase,
                        'clave': clave,
                    })
                
                emp_buscar = codigos_empresa_equivalentes(empresa_codigo) if empresa_codigo else []
                costo = Costos.objects.filter(
                    empresa__in=emp_buscar,
                    uso=uso,
                    clase=clase,
                    rango1__lte=pesos_total,
                    rango2__gte=pesos_total
                ).first() if emp_buscar else None
                
                if costo:
                    especificacion.calidad = costo.calidad
                    logger.info(f"Calidad encontrada en costos: {costo.calidad} para uso={uso}, clase={clase}, pesos={pesos_total} (rango: {costo.rango1}-{costo.rango2})")
                else:
                    # NO usar lógica por defecto - mostrar error
                    error_msg = f'No se encontró calidad en costos para uso={uso}, clase={clase}, pesos={pesos_total}. Verifique que existan registros en la tabla costos con estos parámetros y rangos.'
                    logger.error(error_msg)
                    messages.error(request, error_msg)
                    return render(request, 'especificaciones_calcular_calidad.html', {
                        'titulo': 'Calcular Calidad - Especificaciones',
                        'form': form,
                        'edifino': edifino,
                        'piso': piso,
                        'uso': uso,
                        'clase': clase,
                        'clave': clave,
                    })
            except Exception as e:
                error_msg = f'Error al buscar calidad en costos: {str(e)}'
                logger.error(error_msg, exc_info=True)
                messages.error(request, error_msg)
                return render(request, 'especificaciones_calcular_calidad.html', {
                    'titulo': 'Calcular Calidad - Especificaciones',
                    'form': form,
                    'edifino': edifino,
                    'piso': piso,
                    'uso': uso,
                    'clase': clase,
                    'clave': clave,
                })
            
            especificacion.save()
            
            # Actualizar el campo calidad en la tabla edificacion
            try:
                # Construir filtro de búsqueda para la edificación
                filtro_edificacion = {
                    'clave': clave,
                    'edifino': Decimal(edifino) if edifino else 0
                }
                
                # Si hay empresa en la sesión, incluirla en el filtro
                if empresa_codigo:
                    filtro_edificacion['empresa'] = empresa_codigo
                
                # Buscar la edificación
                edificacion = Edificacion.objects.filter(**filtro_edificacion)
                
                # Si hay piso, filtrar por piso, si no, buscar donde piso es None
                if piso and piso.strip():
                    try:
                        piso_decimal = Decimal(piso)
                        edificacion = edificacion.filter(piso=piso_decimal)
                    except (ValueError, InvalidOperation):
                        # Si piso no es un número válido, buscar donde piso es None
                        edificacion = edificacion.filter(piso__isnull=True)
                else:
                    # Si no hay piso, buscar donde piso es None
                    edificacion = edificacion.filter(piso__isnull=True)
                
                edificacion = edificacion.first()
                
                if edificacion:
                    # Truncar calidad a 2 caracteres si es necesario (el campo en Edificacion es max_length=2)
                    calidad_para_edificacion = str(especificacion.calidad)[:2] if especificacion.calidad else None
                    edificacion.calidad = calidad_para_edificacion
                    edificacion.save()
                    logger.info(f"Calidad actualizada en edificación ID={edificacion.id}: {calidad_para_edificacion} (calidad original: {especificacion.calidad})")
                    messages.success(request, f'Calidad calculada y guardada en edificación: {calidad_para_edificacion} (Pesos: {pesos_total})')
                else:
                    logger.warning(f"No se encontró edificación para actualizar: clave={clave}, edifino={edifino}, piso={piso}")
                    messages.warning(request, f'Calidad calculada: {especificacion.calidad} (Pesos: {pesos_total}), pero no se encontró la edificación para actualizar. Verifique que la edificación exista.')
            except Exception as e:
                logger.error(f"Error al actualizar calidad en edificación: {str(e)}", exc_info=True)
                messages.warning(request, f'Calidad calculada: {especificacion.calidad} (Pesos: {pesos_total}), pero hubo un error al actualizar la edificación: {str(e)}')
            
            # Redirigir de vuelta al formulario de edificaciones con la calidad calculada
            redirect_url = reverse('catastro:edificaciones_form', kwargs={'clave': clave})
            if edifino:
                redirect_url += f'?edifino={edifino}'
            if piso:
                redirect_url += f'&piso={piso}'
            redirect_url += f'&calidad={especificacion.calidad}'
            return redirect(redirect_url)
    else:
        # Buscar si ya existe una especificación para estos parámetros
        especificacion_existente = None
        if clave and edifino:
            especificacion_existente = Especificaciones.objects.filter(
                clave=clave,
                edifino=Decimal(edifino) if edifino else 0,
                piso=piso if piso else None
            ).first()
        
        if especificacion_existente:
            form = EspecificacionesForm(
                instance=especificacion_existente,
                clave=clave,
                usuario=current_user[:50] if current_user else '',
                fecha_sistema=timezone.now()
            )
        else:
            form = EspecificacionesForm(
                clave=clave,
                usuario=current_user[:50] if current_user else '',
                fecha_sistema=timezone.now()
            )
            # Establecer valores iniciales
            if edifino:
                form.fields['edifino'].initial = edifino
            if piso:
                form.fields['piso'].initial = piso
            if uso:
                form.fields['uso'].initial = uso
            if clase:
                form.fields['clase'].initial = clase
    
    context = {
        'titulo': 'Calcular Calidad - Especificaciones',
        'form': form,
        'edifino': edifino,
        'piso': piso,
        'uso': uso,
        'clase': clase,
        'clave': clave,
    }
    
    return render(request, 'especificaciones_calcular_calidad.html', context)

@catastro_require_auth
def especificaciones_list(request):
    """
    Lista de especificaciones
    """
    especificaciones = Especificaciones.objects.all()
    
    # Filtros
    clave = request.GET.get('clave', '').strip()
    if clave:
        especificaciones = especificaciones.filter(clave=clave)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        especificaciones = especificaciones.filter(
            Q(clave__icontains=search) |
            Q(calidad__icontains=search)
        )
    
    especificaciones = especificaciones.order_by('-fechasys')
    
    context = {
        'titulo': 'Especificaciones - Catastro',
        'especificaciones': especificaciones,
        'search': search,
        'clave': clave,
        'total_registros': especificaciones.count(),
    }
    
    return render(request, 'especificaciones_list.html', context)

@catastro_require_auth
def especificaciones_update(request, pk):
    """
    Actualizar especificación
    """
    especificacion = get_object_or_404(Especificaciones, pk=pk)
    current_user = request.session.get('catastro_usuario_nombre', '')
    
    if request.method == 'POST':
        form = EspecificacionesForm(
            request.POST,
            instance=especificacion,
            clave=especificacion.clave,
            usuario=current_user[:50] if current_user else '',
            fecha_sistema=timezone.now()
        )
        if form.is_valid():
            especificacion = form.save(commit=False)
            especificacion.usuario = current_user[:50] if current_user else ''
            especificacion.fechasys = timezone.now()
            
            # Recalcular pesos y calidad
            pesos_total = (
                (especificacion.pesofun or Decimal('0')) +
                (especificacion.pesopiso or Decimal('0')) +
                (especificacion.pesoparext or Decimal('0')) +
                (especificacion.pesotecho or Decimal('0')) +
                (especificacion.pesoparint or Decimal('0')) +
                (especificacion.pesocielo or Decimal('0')) +
                (especificacion.pesocarpini or Decimal('0')) +
                (especificacion.pesoelectri or Decimal('0')) +
                (especificacion.pesoplome or Decimal('0')) +
                (especificacion.pesotros or Decimal('0'))
            )
            # Guardar el total sin redondeo
            especificacion.pesos = pesos_total
            
            # Buscar calidad en la tabla costos según uso, clase y rango de pesos
            # La calidad DEBE obtenerse de la tabla costos, NO usar lógica por defecto
            try:
                if not especificacion.uso or not especificacion.clase:
                    messages.error(request, 'Uso y Clase son requeridos para calcular la calidad.')
                    return redirect('catastro:especificaciones_list')
                
                empresa_ed = (request.session.get('catastro_empresa') or request.session.get('empresa') or '').strip()
                emp_ed = codigos_empresa_equivalentes(empresa_ed) if empresa_ed else []
                costo = Costos.objects.filter(
                    empresa__in=emp_ed,
                    uso=especificacion.uso,
                    clase=especificacion.clase,
                    rango1__lte=pesos_total,
                    rango2__gte=pesos_total
                ).first() if emp_ed else None
                
                if costo:
                    especificacion.calidad = costo.calidad
                    logger.info(f"Calidad encontrada en costos: {costo.calidad} para uso={especificacion.uso}, clase={especificacion.clase}, pesos={pesos_total} (rango: {costo.rango1}-{costo.rango2})")
                else:
                    # NO usar lógica por defecto - mostrar error
                    error_msg = f'No se encontró calidad en costos para uso={especificacion.uso}, clase={especificacion.clase}, pesos={pesos_total}. Verifique que existan registros en la tabla costos con estos parámetros y rangos.'
                    logger.error(error_msg)
                    messages.error(request, error_msg)
                    return redirect('catastro:especificaciones_list')
            except Exception as e:
                error_msg = f'Error al buscar calidad en costos: {str(e)}'
                logger.error(error_msg, exc_info=True)
                messages.error(request, error_msg)
                return redirect('catastro:especificaciones_list')
            
            especificacion.save()
            messages.success(request, 'Especificación actualizada exitosamente.')
            return redirect('catastro:especificaciones_list')
    else:
        form = EspecificacionesForm(
            instance=especificacion,
            clave=especificacion.clave,
            usuario=current_user[:50] if current_user else '',
            fecha_sistema=timezone.now()
        )
    
    context = {
        'titulo': f'Editar Especificación - Catastro',
        'form': form,
        'especificacion': especificacion,
    }
    
    return render(request, 'especificaciones_form.html', context)

@catastro_require_auth
def especificaciones_delete(request, pk):
    """
    Eliminar especificación
    """
    especificacion = get_object_or_404(Especificaciones, pk=pk)
    
    if request.method == 'POST':
        especificacion.delete()
        messages.success(request, 'Especificación eliminada exitosamente.')
        return redirect('catastro:especificaciones_list')
    
    context = {
        'titulo': f'Eliminar Especificación - Catastro',
        'especificacion': especificacion,
    }
    
    return render(request, 'especificaciones_confirm_delete.html', context)

@catastro_require_auth
def api_listar_fundiciones(request):
    """
    API endpoint para listar fundiciones filtradas por uso, clase y tipo='1'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar fundiciones.'
            })
        
        # Buscar en confi_tipologia donde uso=Uso, clase=Clase y tipo='1'
        # SELECT categoria, descripcion FROM confi_tipologia 
        # WHERE uso = uso AND clase = clase AND tipo = '1' 
        # ORDER BY categoria
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='1'
        ).order_by('categoria', 'descripcion')
        
        logger.info(f"Buscando fundiciones: uso={uso}, clase={clase}, tipo=1. Encontrados: {items.count()}")
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codfun, codpiso, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codfun, codpiso, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '1'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar fundiciones: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar fundiciones: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_pisos(request):
    """
    API endpoint para listar pisos filtrados por uso, clase y tipo='2'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar pisos.'
            })
        
        # Buscar en confi_tipologia donde tipo='2'
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='2'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar pisos: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar pisos: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_paredes_exteriores(request):
    """
    API endpoint para listar paredes exteriores filtradas por uso, clase y tipo='3'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar paredes exteriores.'
            })
        
        # Buscar en confi_tipologia donde tipo='3'
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='3'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar paredes exteriores: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar paredes exteriores: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_techos(request):
    """
    API endpoint para listar techos filtrados por uso, clase y tipo='4'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar techos.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='4'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar techos: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar techos: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_paredes_interiores(request):
    """
    API endpoint para listar paredes interiores filtradas por uso, clase y tipo='5'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar paredes interiores.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='5'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar paredes interiores: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar paredes interiores: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_cielo_falso(request):
    """
    API endpoint para listar cielos falsos filtrados por uso, clase y tipo='6'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar cielos falsos.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='6'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar cielos falsos: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar cielos falsos: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_carpinteria(request):
    """
    API endpoint para listar carpintería filtrada por uso, clase y tipo='7'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar carpintería.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='7'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar carpintería: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar carpintería: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_electricidad(request):
    """
    API endpoint para listar electricidad filtrada por uso, clase y tipo='8'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar electricidad.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='8'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar electricidad: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar electricidad: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_plomeria(request):
    """
    API endpoint para listar plomería filtrada por uso, clase y tipo='9'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar plomería.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='9'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar plomería: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar plomería: {str(e)}'
        }, status=500)

@catastro_require_auth
def api_listar_otros_detalles(request):
    """
    API endpoint para listar otros detalles filtrados por uso, clase y tipo='10'
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        
        if not uso or not clase:
            return JsonResponse({
                'items': [],
                'mensaje': 'Uso y Clase son requeridos para filtrar otros detalles.'
            })
        
        items = ConfiTipologia.objects.filter(
            uso=uso,
            clase=clase,
            tipo='10'
        ).order_by('descripcion')
        
        items_list = []
        for item in items:
            # El código será la categoría (que se guardará en codpiso, codparext, etc.)
            categoria = item.categoria or ''
            descripcion = item.descripcion or ''
            items_list.append({
                'id': item.id,
                'codigo': categoria,  # El código es la categoría (se guarda en codpiso, codparext, etc.)
                'categoria': categoria,
                'descripcion': descripcion,
                'peso': str(item.peso),
                'uso': item.uso,
                'clase': item.clase,
                'tipo': item.tipo or '2'
            })
        
        return JsonResponse({
            'items': items_list,
            'total': len(items_list)
        })
        
    except Exception as e:
        logger.error(f"Error al listar otros detalles: {str(e)}", exc_info=True)
        return JsonResponse({
            'items': [],
            'mensaje': f'Error al listar otros detalles: {str(e)}'
        }, status=500)

@csrf_exempt
@catastro_require_auth
def api_buscar_confi_tipologia(request):
    """
    API endpoint para buscar una configuración de tipología por uso, clase, tipo y categoria
    """
    try:
        uso = request.GET.get('uso', '').strip()
        clase = request.GET.get('clase', '').strip()
        tipo = request.GET.get('tipo', '').strip()
        categoria = request.GET.get('categoria', '').strip()
        
        if not uso or not clase or not tipo:
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'Uso, Clase y Clasificación son requeridos para buscar.'
            })
        
        # Buscar registro con uso, clase, tipo y categoria (si se proporciona)
        filtros = {
            'uso': uso,
            'clase': clase,
            'tipo': tipo
        }
        
        if categoria:
            filtros['categoria'] = categoria
        
        tipologia = ConfiTipologia.objects.filter(**filtros).first()
        
        if tipologia:
            return JsonResponse({
                'encontrado': True,
                'descripcion': tipologia.descripcion or '',
                'peso': str(tipologia.peso) if tipologia.peso else '0',
                'id': tipologia.id
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'No se encontró registro con los parámetros proporcionados.'
            })
        
    except Exception as e:
        logger.error(f"Error al buscar configuración de tipología: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'mensaje': f'Error al buscar configuración de tipología: {str(e)}'
        }, status=500)

@catastro_require_auth
def det_especificacion_list(request):
    """
    Lista de detalles de especificaciones
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    detalles = DetEspecificacion.objects.all()
    if empresa_codigo:
        detalles = detalles.filter(empresa=empresa_codigo)
    
    # Filtros
    clave = request.GET.get('clave', '').strip()
    if clave:
        detalles = detalles.filter(clave=clave)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        detalles = detalles.filter(
            Q(clave__icontains=search) |
            Q(edifino__icontains=search) |
            Q(piso__icontains=search)
        )
    
    detalles = detalles.order_by('-fechasys')
    
    context = {
        'titulo': 'Detalles de Especificaciones - Catastro',
        'detalles': detalles,
        'search': search,
        'clave': clave,
        'total_registros': detalles.count(),
    }
    
    return render(request, 'det_especificacion_list.html', context)

@catastro_require_auth
def det_especificacion_create(request):
    """
    Crear nuevo detalle de especificación
    """
    empresa_codigo = request.session.get('catastro_empresa', '')
    # Obtener usuario de la sesión o del request.user como respaldo
    current_user = request.session.get('catastro_usuario_nombre', '')
    if not current_user and request.user.is_authenticated:
        current_user = request.user.username or request.user.get_full_name() or str(request.user)
    
    if request.method == 'POST':
        form = DetEspecificacionForm(request.POST)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.empresa = empresa_codigo
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            detalle.save()
            messages.success(request, 'Detalle de especificación creado exitosamente.')
            return redirect('catastro:det_especificacion_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = DetEspecificacionForm(initial={'empresa': empresa_codigo})
    
    context = {
        'titulo': 'Nuevo Detalle de Especificación - Catastro',
        'form': form,
    }
    
    return render(request, 'det_especificacion_form.html', context)

@catastro_require_auth
def det_especificacion_create_from_edificacion(request):
    """
    Crear nuevo detalle de especificación desde el formulario de edificaciones
    Recibe parámetros: empresa, clave, edifino, piso
    """
    empresa_codigo = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
    clave = request.GET.get('clave', '').strip()
    edifino = request.GET.get('edifino', '').strip()
    piso = request.GET.get('piso', '').strip() or '0'
    
    # Obtener usuario de la sesión o del request.user como respaldo
    current_user = request.session.get('catastro_usuario_nombre', '')
    if not current_user and request.user.is_authenticated:
        current_user = request.user.username or request.user.get_full_name() or str(request.user)
    
    if not clave or not edifino:
        messages.error(request, 'Clave catastral y No. Edificación son requeridos.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Buscar si ya existe un registro con estos parámetros
    try:
        edifino_decimal = Decimal(edifino)
        piso_decimal = Decimal(piso) if piso else Decimal('0')
    except (ValueError, InvalidOperation):
        messages.error(request, 'No. Edificación y Piso deben ser valores numéricos válidos.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    detalle_existente = DetEspecificacion.objects.filter(
        empresa=empresa_codigo,
        clave=clave,
        edifino=edifino_decimal,
        piso=piso_decimal
    ).first()
    
    if request.method == 'POST':
        if detalle_existente:
            form = DetEspecificacionForm(request.POST, instance=detalle_existente)
        else:
            form = DetEspecificacionForm(request.POST)
        
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.empresa = empresa_codigo
            detalle.clave = clave
            detalle.edifino = edifino_decimal
            detalle.piso = piso_decimal
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            detalle.save()
            
            if detalle_existente:
                messages.success(request, 'Detalle de especificación actualizado exitosamente.')
            else:
                messages.success(request, 'Detalle de especificación creado exitosamente.')
            
            # Redirigir de vuelta al formulario de edificaciones
            return redirect(f"{reverse('catastro:edificaciones_form', args=[clave])}?edifino={edifino}&piso={piso}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        if detalle_existente:
            form = DetEspecificacionForm(instance=detalle_existente)
        else:
            form = DetEspecificacionForm(initial={
                'empresa': empresa_codigo,
                'clave': clave,
                'edifino': edifino_decimal,
                'piso': piso_decimal
            })
        
        # Hacer readonly los campos clave, edifino y piso
        form.fields['clave'].widget.attrs['readonly'] = True
        form.fields['edifino'].widget.attrs['readonly'] = True
        form.fields['piso'].widget.attrs['readonly'] = True
    
    context = {
        'titulo': f'Especificaciones de la Edificación - Catastro',
        'form': form,
        'clave': clave,
        'edifino': edifino,
        'piso': piso,
        'volver_url': 'catastro:edificaciones_form',
        'volver_params': {'clave': clave, 'edifino': edifino, 'piso': piso},
    }
    
    return render(request, 'det_especificacion_form.html', context)

@catastro_require_auth
def det_especificacion_update(request, pk):
    """
    Actualizar detalle de especificación
    """
    detalle = get_object_or_404(DetEspecificacion, pk=pk)
    # Obtener usuario de la sesión o del request.user como respaldo
    current_user = request.session.get('catastro_usuario_nombre', '')
    if not current_user and request.user.is_authenticated:
        current_user = request.user.username or request.user.get_full_name() or str(request.user)
    
    if request.method == 'POST':
        form = DetEspecificacionForm(request.POST, instance=detalle)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.usuario = current_user[:50] if current_user else ''
            detalle.fechasys = timezone.now()
            detalle.save()
            messages.success(request, 'Detalle de especificación actualizado exitosamente.')
            return redirect('catastro:det_especificacion_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = DetEspecificacionForm(instance=detalle)
    
    context = {
        'titulo': f'Editar Detalle de Especificación - Catastro',
        'form': form,
        'detalle': detalle,
    }
    
    return render(request, 'det_especificacion_form.html', context)

@catastro_require_auth
def det_especificacion_delete(request, pk):
    """
    Eliminar detalle de especificación
    """
    detalle = get_object_or_404(DetEspecificacion, pk=pk)
    
    if request.method == 'POST':
        try:
            detalle.delete()
            messages.success(request, 'Detalle de especificación eliminado exitosamente.')
            return redirect('catastro:det_especificacion_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar detalle de especificación: {str(e)}')
            return redirect('catastro:det_especificacion_list')
    
    context = {
        'titulo': f'Eliminar Detalle de Especificación - Catastro',
        'detalle': detalle,
    }
    
    return render(request, 'det_especificacion_confirm_delete.html', context)

# ============================================================================
# GESTIÓN DE VALOR ARBOL (CLASE Y VARIEDAD DE CULTIVO)
# ============================================================================

@catastro_require_auth
def valor_arbol_list(request):
    """
    Lista de valores de árbol (clases y variedades de cultivo)
    """
    from .models import ValorArbol
    from .forms import ValorArbolForm
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    valor_arboles = ValorArbol.objects.all()
    if empresa_codigo:
        valor_arboles = valor_arboles.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        valor_arboles = valor_arboles.filter(
            Q(codigo__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    valor_arboles = valor_arboles.order_by('codigo')
    
    context = {
        'titulo': 'Clase y Variedad de Cultivo - Catastro',
        'valor_arboles': valor_arboles,
        'search': search,
        'total_registros': valor_arboles.count(),
    }
    
    return render(request, 'valor_arbol_list.html', context)

@catastro_require_auth
def valor_arbol_create(request):
    """
    Crear nuevo valor de árbol
    Si ya existe un registro con la misma empresa y código, redirige a la edición
    """
    from .models import ValorArbol
    from .forms import ValorArbolForm
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        form = ValorArbolForm(request.POST, empresa=empresa_codigo)
        
        if form.is_valid():
            codigo = form.cleaned_data.get('codigo', '').strip()
            
            # Verificar si ya existe un registro con la misma empresa y código
            valor_arbol_existente = None
            if codigo and empresa_codigo:
                valor_arbol_existente = ValorArbol.objects.filter(
                    empresa=empresa_codigo,
                    codigo=codigo
                ).first()
            
            if valor_arbol_existente:
                # Si existe, actualizar el registro existente directamente sin redirigir
                valor_arbol_existente.descripcion = form.cleaned_data.get('descripcion', '')
                valor_arbol_existente.valor = form.cleaned_data.get('valor', 0.00)
                valor_arbol_existente.save()
                messages.success(request, 'Valor Árbol actualizado exitosamente.')
                return redirect('catastro:valor_arbol_list')
            else:
                # Si no existe, crear el nuevo registro
                try:
                    valor_arbol = form.save(commit=False)
                    valor_arbol.empresa = empresa_codigo
                    valor_arbol.save()
                    messages.success(request, 'Clase y variedad de cultivo creada exitosamente.')
                    return redirect('catastro:valor_arbol_list')
                except Exception as e:
                    # Si hay un error de integridad (duplicado), actualizar el existente
                    error_str = str(e).lower()
                    if 'unique' in error_str or 'duplicate' in error_str or 'ya existe' in error_str:
                        if codigo and empresa_codigo:
                            valor_arbol_existente = ValorArbol.objects.filter(
                                empresa=empresa_codigo,
                                codigo=codigo
                            ).first()
                            
                            if valor_arbol_existente:
                                # Actualizar el registro existente directamente
                                valor_arbol_existente.descripcion = form.cleaned_data.get('descripcion', '')
                                valor_arbol_existente.valor = form.cleaned_data.get('valor', 0.00)
                                valor_arbol_existente.save()
                                messages.success(request, 'Valor Árbol actualizado exitosamente.')
                                return redirect('catastro:valor_arbol_list')
                    
                    # Si es otro tipo de error, mostrarlo
                    messages.error(request, f'Error al guardar: {str(e)}')
        else:
            # Si hay errores de validación, verificar si es por duplicado y permitir actualizar
            codigo = request.POST.get('codigo', '').strip()
            if codigo and empresa_codigo:
                valor_arbol_existente = ValorArbol.objects.filter(
                    empresa=empresa_codigo,
                    codigo=codigo
                ).first()
                
                if valor_arbol_existente:
                    # Si existe y hay datos válidos, intentar actualizar directamente
                    descripcion = request.POST.get('descripcion', '').strip()
                    valor_str = request.POST.get('valor', '').strip()
                    
                    try:
                        from decimal import Decimal
                        valor = Decimal(valor_str) if valor_str else Decimal('0.00')
                        
                        # Actualizar el registro existente
                        valor_arbol_existente.descripcion = descripcion
                        valor_arbol_existente.valor = valor
                        valor_arbol_existente.save()
                        messages.success(request, 'Valor Árbol actualizado exitosamente.')
                        return redirect('catastro:valor_arbol_list')
                    except (ValueError, InvalidOperation):
                        # Si hay error en la conversión, mostrar errores de validación
                        pass
            
            # Mostrar otros errores de validación
            for field, errors in form.errors.items():
                for error in errors:
                    error_str = str(error).lower()
                    if 'ya existe' not in error_str and 'already exists' not in error_str and 'unique' not in error_str:
                        messages.error(request, f'{field}: {error}')
    else:
        form = ValorArbolForm(empresa=empresa_codigo)
    
    context = {
        'titulo': 'Nueva Clase y Variedad de Cultivo - Catastro',
        'form': form,
        'valor_arbol': None,
    }
    
    return render(request, 'valor_arbol_form.html', context)

@catastro_require_auth
def valor_arbol_update(request, pk):
    """
    Actualizar valor de árbol
    """
    from .models import ValorArbol
    from .forms import ValorArbolForm
    
    valor_arbol = get_object_or_404(ValorArbol, pk=pk)
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        form = ValorArbolForm(request.POST, instance=valor_arbol, empresa=empresa_codigo)
        if form.is_valid():
            valor_arbol = form.save(commit=False)
            valor_arbol.empresa = empresa_codigo
            valor_arbol.save()
            messages.success(request, 'Clase y variedad de cultivo actualizada exitosamente.')
            return redirect('catastro:valor_arbol_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ValorArbolForm(instance=valor_arbol, empresa=empresa_codigo)
    
    context = {
        'titulo': f'Editar Clase y Variedad de Cultivo - Catastro',
        'form': form,
        'valor_arbol': valor_arbol,
    }
    
    return render(request, 'valor_arbol_form.html', context)

@csrf_exempt
@catastro_require_auth
def buscar_valor_arbol_ajax(request):
    """
    API endpoint para búsqueda de valor de árbol por empresa y código
    """
    from .models import ValorArbol
    
    try:
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        codigo = request.GET.get('codigo', '').strip()
        
        if not codigo:
            return JsonResponse({
                'existe': False,
                'mensaje': 'Debe proporcionar un código'
            })
        
        # Buscar el valor de árbol
        if empresa:
            valor_arbol = ValorArbol.objects.filter(
                empresa=empresa,
                codigo=codigo
            ).first()
        else:
            valor_arbol = ValorArbol.objects.filter(
                codigo=codigo
            ).first()
        
        if valor_arbol:
            # Generar URL de edición
            from django.urls import reverse
            url_editar = reverse('catastro:valor_arbol_update', kwargs={'pk': valor_arbol.id})
            
            return JsonResponse({
                'existe': True,
                'id': valor_arbol.id,
                'empresa': valor_arbol.empresa or '',
                'codigo': valor_arbol.codigo or '',
                'descripcion': valor_arbol.descripcion or '',
                'valor': str(valor_arbol.valor) if valor_arbol.valor else '0.00',
                'url_editar': url_editar,
                'mensaje': f'Valor Árbol con este Empresa y Código ya existe.'
            })
        else:
            return JsonResponse({
                'existe': False,
                'mensaje': f'No se encontró valor de árbol con código {codigo}'
            })
            
    except Exception as e:
        logger.error(f"Error en búsqueda de valor de árbol: {str(e)}", exc_info=True)
        return JsonResponse({
            'existe': False,
            'error': str(e),
            'mensaje': 'Error al buscar el valor de árbol'
        }, status=500)

@catastro_require_auth
def valor_arbol_delete(request, pk):
    """
    Eliminar valor de árbol
    """
    from .models import ValorArbol
    
    valor_arbol = get_object_or_404(ValorArbol, pk=pk)
    
    if request.method == 'POST':
        try:
            valor_arbol.delete()
            messages.success(request, 'Clase y variedad de cultivo eliminada exitosamente.')
            return redirect('catastro:valor_arbol_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar clase y variedad de cultivo: {str(e)}')
            return redirect('catastro:valor_arbol_list')
    
    context = {
        'titulo': f'Eliminar Clase y Variedad de Cultivo - Catastro',
        'valor_arbol': valor_arbol,
    }
    
    return render(request, 'valor_arbol_confirm_delete.html', context)

@catastro_require_auth
def valor_arbol_export_excel(request):
    """
    Exportar valores de árbol a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        return redirect('catastro:valor_arbol_list')
    
    from .models import ValorArbol
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener todos los valores de árbol
    valor_arboles = ValorArbol.objects.all()
    if empresa_codigo:
        valor_arboles = valor_arboles.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        valor_arboles = valor_arboles.filter(
            Q(codigo__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    valor_arboles = valor_arboles.order_by('codigo')
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clase y Variedad de Cultivo"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.append(['CLASE Y VARIEDAD DE CULTIVO'])
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Encabezados
    headers = ['ID', 'Código', 'Descripción', 'Valor']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    row_num = 4
    suma_total = Decimal('0.00')
    
    for valor_arbol in valor_arboles:
        ws.append([
            valor_arbol.id,
            valor_arbol.codigo or '',
            valor_arbol.descripcion or '',
            float(valor_arbol.valor) if valor_arbol.valor else 0.00
        ])
        
        # Aplicar bordes
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border_style
            if col == 4:  # Columna de valor
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
        
        suma_total += valor_arbol.valor or Decimal('0.00')
        row_num += 1
    
    # Fila de totales
    if row_num > 4:
        ws.append([])
        row_num += 1
        ws.append(['', '', 'TOTAL:', float(suma_total)])
        total_row = row_num
        for col in range(1, 5):
            cell = ws.cell(row=total_row, column=col)
            cell.border = border_style
            if col == 3:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 4:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'clase_variedad_cultivo_{empresa_codigo or "all"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def valor_arbol_export_pdf(request):
    """
    Exportar valores de árbol a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        return redirect('catastro:valor_arbol_list')
    
    from .models import ValorArbol
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener todos los valores de árbol
    valor_arboles = ValorArbol.objects.all()
    if empresa_codigo:
        valor_arboles = valor_arboles.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        valor_arboles = valor_arboles.filter(
            Q(codigo__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    valor_arboles = valor_arboles.order_by('codigo')
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    title_text = "CLASE Y VARIEDAD DE CULTIVO"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Preparar datos para la tabla
    data = [['ID', 'Código', 'Descripción', 'Valor']]
    suma_total = Decimal('0.00')
    
    for valor_arbol in valor_arboles:
        data.append([
            str(valor_arbol.id),
            valor_arbol.codigo or '',
            valor_arbol.descripcion or '',
            f"{valor_arbol.valor:.2f}" if valor_arbol.valor else '0.00'
        ])
        suma_total += valor_arbol.valor or Decimal('0.00')
    
    # Agregar fila de totales
    if len(data) > 1:
        data.append(['', '', 'TOTAL:', f"{suma_total:.2f}"])
    
    # Crear tabla
    table = Table(data, colWidths=[0.8*inch, 1.2*inch, 4*inch, 1*inch])
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('ALIGN', (3, 1), (3, -2), 'RIGHT'),  # Alinear valores a la derecha
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        # Fila de totales
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F4F8')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Crear respuesta HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f'clase_variedad_cultivo_{empresa_codigo or "all"}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# ============================================================================
# GESTIÓN DE FACTOR CULTIVO
# ============================================================================

@catastro_require_auth
def factor_cultivo(request, empresa, codigo):
    """
    Lista de factores de cultivo para un código específico
    """
    from .models import FactorCultivo, ValorArbol
    from .forms import FactorCultivoForm
    
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    
    # Obtener información del cultivo si existe
    valor_arbol = None
    try:
        valor_arbol = ValorArbol.objects.filter(
            empresa=empresa_codigo,
            codigo=codigo
        ).first()
    except Exception:
        pass
    
    # Obtener factores de cultivo
    factores = FactorCultivo.objects.filter(
        empresa=empresa_codigo,
        codigo=codigo
    )
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        try:
            search_num = int(search)
            factores = factores.filter(
                Q(rango1=search_num) |
                Q(rango2=search_num) |
                Q(factor__icontains=search)
            )
        except ValueError:
            factores = factores.filter(
                Q(factor__icontains=search)
            )
    
    factores = factores.order_by('rango1', 'rango2')
    
    context = {
        'titulo': f'Factores de Cultivo - Código: {codigo}',
        'factores': factores,
        'valor_arbol': valor_arbol,
        'empresa': empresa_codigo,
        'codigo': codigo,
        'search': search,
        'total_registros': factores.count(),
    }
    
    return render(request, 'factor_cultivo.html', context)

@catastro_require_auth
def factor_cultivo_create(request, empresa, codigo):
    """
    Crear nuevo factor de cultivo
    Si ya existe un registro con la misma empresa, código y rango1, actualiza el registro existente.
    """
    from .models import FactorCultivo, ValorArbol
    from .forms import FactorCultivoForm
    
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    
    # Obtener información del cultivo si existe
    valor_arbol = None
    try:
        valor_arbol = ValorArbol.objects.filter(
            empresa=empresa_codigo,
            codigo=codigo
        ).first()
    except Exception:
        pass
    
    if request.method == 'POST':
        form = FactorCultivoForm(request.POST, empresa=empresa_codigo, codigo_cultivo=codigo)
        if form.is_valid():
            rango1 = form.cleaned_data.get('rango1')
            
            # Verificar si ya existe un registro con la misma empresa, código y rango1
            if rango1 is not None:
                factor_existente = FactorCultivo.objects.filter(
                    empresa=empresa_codigo,
                    codigo=codigo,
                    rango1=rango1
                ).first()
                
                if factor_existente:
                    # Si existe, actualizar el registro existente directamente
                    factor_existente.rango2 = form.cleaned_data.get('rango2')
                    factor_existente.factor = form.cleaned_data.get('factor')
                    factor_existente.save()
                    messages.success(request, 'Factor de cultivo actualizado exitosamente.')
                    return redirect('catastro:factor_cultivo', empresa=empresa_codigo, codigo=codigo)
            
            # Si no existe, crear el nuevo registro
            try:
                factor = form.save(commit=False)
                factor.empresa = empresa_codigo
                factor.codigo = codigo
                factor.save()
                messages.success(request, 'Factor de cultivo creado exitosamente.')
                return redirect('catastro:factor_cultivo', empresa=empresa_codigo, codigo=codigo)
            except Exception as e:
                # Fallback para errores de integridad si la validación previa falla por alguna razón
                error_str = str(e).lower()
                if 'unique' in error_str or 'duplicate' in error_str:
                    messages.error(request, 'Error: Ya existe un registro con esta Empresa, Código y Rango 1.')
                else:
                    messages.error(request, f'Error al guardar el factor de cultivo: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FactorCultivoForm(empresa=empresa_codigo, codigo_cultivo=codigo)
    
    context = {
        'titulo': f'Nuevo Factor de Cultivo - Código: {codigo}',
        'form': form,
        'factor': None,
        'valor_arbol': valor_arbol,
        'empresa': empresa_codigo,
        'codigo': codigo,
    }
    
    return render(request, 'factor_cultivo_form.html', context)

@catastro_require_auth
def factor_cultivo_update(request, empresa, codigo, pk):
    """
    Actualizar factor de cultivo
    """
    from .models import FactorCultivo, ValorArbol
    from .forms import FactorCultivoForm
    
    factor = get_object_or_404(FactorCultivo, pk=pk)
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    
    # Obtener información del cultivo si existe
    valor_arbol = None
    try:
        valor_arbol = ValorArbol.objects.filter(
            empresa=empresa_codigo,
            codigo=codigo
        ).first()
    except Exception:
        pass
    
    if request.method == 'POST':
        form = FactorCultivoForm(request.POST, instance=factor, empresa=empresa_codigo, codigo_cultivo=codigo)
        if form.is_valid():
            factor = form.save(commit=False)
            factor.empresa = empresa_codigo
            factor.codigo = codigo
            factor.save()
            messages.success(request, 'Factor de cultivo actualizado exitosamente.')
            return redirect('catastro:factor_cultivo', empresa=empresa_codigo, codigo=codigo)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FactorCultivoForm(instance=factor, empresa=empresa_codigo, codigo_cultivo=codigo)
    
    context = {
        'titulo': f'Editar Factor de Cultivo - Código: {codigo}',
        'form': form,
        'factor': factor,
        'valor_arbol': valor_arbol,
        'empresa': empresa_codigo,
        'codigo': codigo,
    }
    
    return render(request, 'factor_cultivo_form.html', context)

@catastro_require_auth
def factor_cultivo_delete(request, empresa, codigo, pk):
    """
    Eliminar factor de cultivo
    """
    from .models import FactorCultivo
    
    factor = get_object_or_404(FactorCultivo, pk=pk)
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        try:
            factor.delete()
            messages.success(request, 'Factor de cultivo eliminado exitosamente.')
            return redirect('catastro:factor_cultivo', empresa=empresa_codigo, codigo=codigo)
        except Exception as e:
            messages.error(request, f'Error al eliminar factor de cultivo: {str(e)}')
            return redirect('catastro:factor_cultivo', empresa=empresa_codigo, codigo=codigo)
    
    context = {
        'titulo': f'Eliminar Factor de Cultivo - Código: {codigo}',
        'factor': factor,
        'empresa': empresa_codigo,
        'codigo': codigo,
    }
    
    return render(request, 'factor_cultivo_confirm_delete.html', context)

@csrf_exempt
@catastro_require_auth
def buscar_factor_cultivo_ajax(request):
    """
    API endpoint para búsqueda de factor de cultivo por empresa, código y rango1
    """
    from .models import FactorCultivo
    
    try:
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        codigo = request.GET.get('codigo', '').strip()
        rango1 = request.GET.get('rango1', '').strip()
        
        if not empresa or not codigo or not rango1:
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'Debe proporcionar empresa, código y rango1'
            })
        
        try:
            rango1_decimal = Decimal(rango1)
        except (ValueError, InvalidOperation):
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'Rango1 debe ser un número válido'
            })
        
        # Buscar el factor de cultivo
        factor = FactorCultivo.objects.filter(
            empresa=empresa,
            codigo=codigo,
            rango1=rango1_decimal
        ).first()
        
        if factor:
            return JsonResponse({
                'existe': True,
                'encontrado': True,
                'id': factor.id,
                'empresa': factor.empresa or '',
                'codigo': factor.codigo or '',
                'rango1': str(int(factor.rango1)) if factor.rango1 else '0',
                'rango2': str(int(factor.rango2)) if factor.rango2 else '0',
                'factor': str(factor.factor) if factor.factor else '0.000',
                'mensaje': f'Factor encontrado: Rango {factor.rango1}-{factor.rango2}'
            })
        else:
            return JsonResponse({
                'existe': False,
                'encontrado': False,
                'mensaje': f'No se encontró factor para código {codigo} y rango1 {rango1}'
            })
            
    except Exception as e:
        logger.error(f"Error en búsqueda de factor de cultivo: {str(e)}", exc_info=True)
        return JsonResponse({
            'encontrado': False,
            'error': str(e),
            'mensaje': 'Error al buscar el factor de cultivo'
        }, status=500)

@catastro_require_auth
def buscar_tierra_rural_ajax(request):
    """
    API endpoint para búsqueda de valor de tierra rural por empresa y código
    """
    from .models import FactoresRiego
    
    try:
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        codigo = request.GET.get('codigo', '').strip()
        
        if not codigo:
            return JsonResponse({
                'existe': False,
                'mensaje': 'Debe proporcionar un código'
            })
        
        if not empresa:
            return JsonResponse({
                'existe': False,
                'mensaje': 'No se ha seleccionado un municipio'
            })
        
        # Buscar el factor de riego por empresa y código (único por empresa+código)
        factor_riego = FactoresRiego.objects.filter(
            empresa=empresa,
            codigo=codigo
        ).first()
        
        if factor_riego:
            return JsonResponse({
                'existe': True,
                'id': factor_riego.id,
                'codigo': factor_riego.codigo or '',
                'descripcion': factor_riego.descripcion or '',
                'valor': str(factor_riego.valor) if factor_riego.valor else '0.00',
                'empresa': factor_riego.empresa or '',
                'mensaje': 'Registro encontrado'
            })
        else:
            return JsonResponse({
                'existe': False,
                'mensaje': 'No se encontró un registro con este código para esta empresa'
            })
    except Exception as e:
        return JsonResponse({
            'existe': False,
            'mensaje': f'Error al buscar el valor de tierra rural: {str(e)}'
        }, status=500)

# ============================================================================
# VALORES DE TIERRA RURAL (FactoresRiego)
# ============================================================================

@catastro_require_auth
def tierra_rural_list(request):
    """
    Lista de valores de tierra rural
    """
    from .models import FactoresRiego
    from .forms import TierraRuralForm
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    # Obtener todos los valores de tierra rural
    tierra_rural = FactoresRiego.objects.all()
    if empresa_codigo:
        tierra_rural = tierra_rural.filter(empresa=empresa_codigo)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        try:
            search_num = float(search)
            tierra_rural = tierra_rural.filter(
                Q(codigo__icontains=search) |
                Q(descripcion__icontains=search) |
                Q(valor=search_num)
            )
        except ValueError:
            tierra_rural = tierra_rural.filter(
                Q(codigo__icontains=search) |
                Q(descripcion__icontains=search)
            )
    
    tierra_rural = tierra_rural.order_by('codigo')
    
    context = {
        'titulo': 'Valores de Tierra Rural - Catastro',
        'tierra_rural': tierra_rural,
        'search': search,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'tierra_rural_list.html', context)

@catastro_require_auth
def tierra_rural_create(request):
    """
    Crear nuevo valor de tierra rural
    Si ya existe un registro con la misma empresa y código, actualiza el registro existente directamente.
    """
    from .models import FactoresRiego
    from .forms import TierraRuralForm
    
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        form = TierraRuralForm(request.POST, empresa=empresa_codigo)
        
        # Obtener código del POST para verificar existencia antes de validar
        codigo = request.POST.get('codigo', '').strip()
        
        # Verificar si ya existe un registro con la misma empresa y código
        tierra_rural_existente = None
        if codigo and empresa_codigo:
            tierra_rural_existente = FactoresRiego.objects.filter(
                empresa=empresa_codigo,
                codigo=codigo
            ).first()
        
        if form.is_valid():
            if tierra_rural_existente:
                # Si existe, actualizar el registro existente directamente
                tierra_rural_existente.descripcion = form.cleaned_data.get('descripcion', '')
                tierra_rural_existente.valor = form.cleaned_data.get('valor', 0.00)
                tierra_rural_existente.save()
                messages.success(request, 'Valor de tierra rural actualizado exitosamente.')
                return redirect('catastro:tierra_rural_list')
            else:
                # Si no existe, crear el nuevo registro
                try:
                    tierra_rural = form.save(commit=False)
                    tierra_rural.empresa = empresa_codigo
                    tierra_rural.save()
                    messages.success(request, 'Valor de tierra rural creado exitosamente.')
                    return redirect('catastro:tierra_rural_list')
                except Exception as e:
                    # Si hay un error de integridad (duplicado), actualizar el existente
                    error_str = str(e).lower()
                    if 'unique' in error_str or 'duplicate' in error_str or 'ya existe' in error_str:
                        if codigo and empresa_codigo:
                            tierra_rural_existente = FactoresRiego.objects.filter(
                                empresa=empresa_codigo,
                                codigo=codigo
                            ).first()
                            
                            if tierra_rural_existente:
                                # Actualizar el registro existente directamente
                                try:
                                    descripcion = request.POST.get('descripcion', '').strip()
                                    valor_str = request.POST.get('valor', '0.00').strip()
                                    try:
                                        valor = Decimal(valor_str) if valor_str else Decimal('0.00')
                                    except:
                                        valor = Decimal('0.00')
                                    
                                    tierra_rural_existente.descripcion = descripcion
                                    tierra_rural_existente.valor = valor
                                    tierra_rural_existente.save()
                                    messages.success(request, 'Valor de tierra rural actualizado exitosamente.')
                                    return redirect('catastro:tierra_rural_list')
                                except Exception as e2:
                                    messages.error(request, f'Error al actualizar: {str(e2)}')
                    
                    # Si es otro tipo de error, mostrarlo
                    messages.error(request, f'Error al guardar: {str(e)}')
        else:
            # Si hay errores de validación, verificar si es por duplicado y permitir actualizar
            if tierra_rural_existente:
                # Si existe un registro, intentar actualizar directamente con los datos del POST
                try:
                    descripcion = request.POST.get('descripcion', '').strip()
                    valor_str = request.POST.get('valor', '0.00').strip()
                    try:
                        valor = Decimal(valor_str) if valor_str else Decimal('0.00')
                    except:
                        valor = Decimal('0.00')
                    
                    # Validar que los datos sean válidos
                    if descripcion and len(descripcion) <= 45:
                        tierra_rural_existente.descripcion = descripcion
                        tierra_rural_existente.valor = valor
                        tierra_rural_existente.save()
                        messages.success(request, 'Valor de tierra rural actualizado exitosamente.')
                        return redirect('catastro:tierra_rural_list')
                    else:
                        # Si hay errores de validación en los datos, mostrar los errores del formulario
                        for field, errors in form.errors.items():
                            for error in errors:
                                messages.error(request, f'{field}: {error}')
                except Exception as e:
                    messages.error(request, f'Error al actualizar: {str(e)}')
            else:
                # Si no existe y hay errores, mostrar los errores del formulario
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
    else:
        form = TierraRuralForm(empresa=empresa_codigo)
    
    context = {
        'titulo': 'Nuevo Valor de Tierra Rural - Catastro',
        'form': form,
        'tierra_rural': None,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'tierra_rural_form.html', context)

@catastro_require_auth
def tierra_rural_update(request, pk):
    """
    Actualizar valor de tierra rural
    """
    from .models import FactoresRiego
    from .forms import TierraRuralForm
    
    tierra_rural = get_object_or_404(FactoresRiego, pk=pk)
    empresa_codigo = request.session.get('catastro_empresa', '')
    
    if request.method == 'POST':
        form = TierraRuralForm(request.POST, instance=tierra_rural, empresa=empresa_codigo)
        if form.is_valid():
            tierra_rural = form.save(commit=False)
            tierra_rural.empresa = empresa_codigo
            tierra_rural.save()
            messages.success(request, 'Valor de tierra rural actualizado exitosamente.')
            return redirect('catastro:tierra_rural_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TierraRuralForm(instance=tierra_rural, empresa=empresa_codigo)
    
    context = {
        'titulo': 'Editar Valor de Tierra Rural - Catastro',
        'form': form,
        'tierra_rural': tierra_rural,
        'empresa': empresa_codigo,
    }
    
    return render(request, 'tierra_rural_form.html', context)

@catastro_require_auth
def tierra_rural_delete(request, pk):
    """
    Eliminar valor de tierra rural
    """
    from .models import FactoresRiego
    
    tierra_rural = get_object_or_404(FactoresRiego, pk=pk)
    
    if request.method == 'POST':
        try:
            tierra_rural.delete()
            messages.success(request, 'Valor de tierra rural eliminado exitosamente.')
            return redirect('catastro:tierra_rural_list')
        except Exception as e:
            messages.error(request, f'Error al eliminar valor de tierra rural: {str(e)}')
            return redirect('catastro:tierra_rural_list')
    
    context = {
        'titulo': f'Eliminar Valor de Tierra Rural - Catastro',
        'tierra_rural': tierra_rural,
    }
    
    return render(request, 'tierra_rural_confirm_delete.html', context)

# ============================================================================
# CULTIVO PERMANENTE
# ============================================================================

@catastro_require_auth
def cultivo_permanente_list(request, empresa=None, cocata1=None):
    """
    Lista de cultivos permanentes
    Recibe parámetros empresa y cocata1 de la tabla bdcata1
    """
    from .models import CultivoPermanente, BDCata1
    from .forms import CultivoPermanenteForm
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Obtener cultivos permanentes filtrados por empresa y clave
    cultivos = CultivoPermanente.objects.all()
    if empresa_codigo:
        cultivos = cultivos.filter(empresa=empresa_codigo)
    if clave_catastral:
        cultivos = cultivos.filter(clave=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        try:
            search_num = float(search)
            cultivos = cultivos.filter(
                Q(clave__icontains=search) |
                Q(clase__icontains=search) |
                Q(estado__icontains=search) |
                Q(arbol=search_num) |
                Q(edad=search_num) |
                Q(factor=search_num) |
                Q(valor=search_num)
            )
        except ValueError:
            cultivos = cultivos.filter(
                Q(clave__icontains=search) |
                Q(clase__icontains=search) |
                Q(estado__icontains=search)
            )
    
    cultivos = cultivos.order_by('clave', 'clase', 'edad')
    
    # Calcular totales
    total_arboles = sum(c.arbol for c in cultivos if c.arbol) or 0
    total_valor = sum(c.valor for c in cultivos if c.valor) or Decimal('0.00')
    
    context = {
        'titulo': f'Cultivos Permanentes - Clave: {clave_catastral}' if clave_catastral else 'Cultivos Permanentes - Catastro',
        'cultivos': cultivos,
        'bien_inmueble': bien_inmueble,
        'search': search,
        'empresa': empresa_codigo,
        'clave': clave_catastral,
        'total_arboles': total_arboles,
        'total_valor': total_valor,
    }
    
    return render(request, 'cultivo_permanente_list.html', context)

@catastro_require_auth
def cultivo_permanente_export_excel(request, empresa=None, cocata1=None):
    """
    Exportar cultivos permanentes a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        if empresa and cocata1:
            return redirect('catastro:cultivo_permanente_list', empresa=empresa, cocata1=cocata1)
        return redirect('catastro:cultivo_permanente_list')
    
    from .models import CultivoPermanente, BDCata1, ValorArbol
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener cultivos permanentes filtrados por empresa y clave
    cultivos = CultivoPermanente.objects.all()
    if empresa_codigo:
        cultivos = cultivos.filter(empresa=empresa_codigo)
    if clave_catastral:
        cultivos = cultivos.filter(clave=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        try:
            search_num = float(search)
            cultivos = cultivos.filter(
                Q(clave__icontains=search) |
                Q(clase__icontains=search) |
                Q(estado__icontains=search) |
                Q(arbol=search_num) |
                Q(edad=search_num) |
                Q(factor=search_num) |
                Q(valor=search_num)
            )
        except ValueError:
            cultivos = cultivos.filter(
                Q(clave__icontains=search) |
                Q(clase__icontains=search) |
                Q(estado__icontains=search)
            )
    
    cultivos = cultivos.order_by('clave', 'clase', 'edad')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cultivos Permanentes"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    titulo_texto = f"CULTIVOS PERMANENTES"
    if bien_inmueble:
        titulo_texto += f" - Clave: {clave_catastral}"
    ws.append([titulo_texto])
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Información del bien inmueble si existe
    if bien_inmueble:
        ws.append(['INFORMACIÓN DEL BIEN INMUEBLE'])
        ws.merge_cells('A3:I3')
        info_cell = ws['A3']
        info_cell.font = Font(bold=True, size=12)
        ws.append(['Clave Catastral:', bien_inmueble.cocata1 or '', '', 
                   'Propietario:', f"{bien_inmueble.nombres or ''} {bien_inmueble.apellidos or ''}".strip()])
        ws.append(['Identidad:', bien_inmueble.identidad or '', '', 
                   'Ubicación:', bien_inmueble.ubicacion or ''])
        ws.append([])
        row_start = 7
    else:
        row_start = 3
    
    # Encabezados
    headers = ['ID', 'Clave Catastral', 'Clase', 'Número Árboles', 'Estado Fitosanitario', 'Edad', 'Factor de Modificación', 'Valor']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for cell in ws[row_start]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
    
    # Agregar datos
    row_num = row_start + 1
    total_arboles = 0
    total_valor = Decimal('0.00')
    
    # Obtener descripciones de clases desde ValorArbol
    clases_dict = {}
    if empresa_codigo:
        valor_arboles = ValorArbol.objects.filter(empresa=empresa_codigo)
        clases_dict = {va.codigo: va.descripcion for va in valor_arboles}
    
    for cultivo in cultivos:
        # Obtener descripción de la clase
        clase_descripcion = clases_dict.get(cultivo.clase, cultivo.clase or '')
        
        # Estado fitosanitario
        estado_texto = ''
        if cultivo.estado == '1':
            estado_texto = '1 - Bueno'
        elif cultivo.estado == '2':
            estado_texto = '2 - Regular'
        elif cultivo.estado == '3':
            estado_texto = '3 - Malo'
        else:
            estado_texto = cultivo.estado or ''
        
        ws.append([
            cultivo.id,
            cultivo.clave or '',
            clase_descripcion,
            float(cultivo.arbol) if cultivo.arbol else 0,
            estado_texto,
            float(cultivo.edad) if cultivo.edad else 0,
            float(cultivo.factor) if cultivo.factor else 0.00,
            float(cultivo.valor) if cultivo.valor else 0.00
        ])
        
        # Aplicar bordes
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border_style
            if col in [4, 6, 7, 8]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if col == 7:  # Factor
                    cell.number_format = '#,##0.00'
                elif col == 8:  # Valor
                    cell.number_format = '#,##0.00'
        
        total_arboles += cultivo.arbol or 0
        total_valor += cultivo.valor or Decimal('0.00')
        row_num += 1
    
    # Fila de totales
    if row_num > row_start + 1:
        ws.append([])
        row_num += 1
        ws.append(['', '', 'TOTALES:', total_arboles, '', '', '', float(total_valor)])
        total_row = row_num
        for col in range(1, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.border = border_style
            if col == 3:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 4:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
            elif col == 8:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'cultivos_permanentes_{clave_catastral or empresa_codigo or "all"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def cultivo_permanente_export_pdf(request, empresa=None, cocata1=None):
    """
    Exportar cultivos permanentes a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        if empresa and cocata1:
            return redirect('catastro:cultivo_permanente_list', empresa=empresa, cocata1=cocata1)
        return redirect('catastro:cultivo_permanente_list')
    
    from .models import CultivoPermanente, BDCata1, ValorArbol
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener cultivos permanentes filtrados por empresa y clave
    cultivos = CultivoPermanente.objects.all()
    if empresa_codigo:
        cultivos = cultivos.filter(empresa=empresa_codigo)
    if clave_catastral:
        cultivos = cultivos.filter(clave=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        try:
            search_num = float(search)
            cultivos = cultivos.filter(
                Q(clave__icontains=search) |
                Q(clase__icontains=search) |
                Q(estado__icontains=search) |
                Q(arbol=search_num) |
                Q(edad=search_num) |
                Q(factor=search_num) |
                Q(valor=search_num)
            )
        except ValueError:
            cultivos = cultivos.filter(
                Q(clave__icontains=search) |
                Q(clase__icontains=search) |
                Q(estado__icontains=search)
            )
    
    cultivos = cultivos.order_by('clave', 'clase', 'edad')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    titulo_texto = "CULTIVOS PERMANENTES"
    if bien_inmueble:
        titulo_texto += f"<br/>Clave: {clave_catastral}"
    title = Paragraph(titulo_texto, title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del bien inmueble si existe
    if bien_inmueble:
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=10
        )
        info_text = f"<b>Clave Catastral:</b> {bien_inmueble.cocata1 or ''}<br/>"
        info_text += f"<b>Propietario:</b> {bien_inmueble.nombres or ''} {bien_inmueble.apellidos or ''}<br/>"
        info_text += f"<b>Identidad:</b> {bien_inmueble.identidad or '-'}<br/>"
        info_text += f"<b>Ubicación:</b> {bien_inmueble.ubicacion or '-'}"
        info_para = Paragraph(info_text, info_style)
        elements.append(info_para)
        elements.append(Spacer(1, 0.2*inch))
    
    # Preparar datos para la tabla
    data = [['ID', 'Clave', 'Clase', 'N° Árboles', 'Estado', 'Edad', 'Factor', 'Valor']]
    total_arboles = 0
    total_valor = Decimal('0.00')
    
    # Obtener descripciones de clases desde ValorArbol
    clases_dict = {}
    if empresa_codigo:
        valor_arboles = ValorArbol.objects.filter(empresa=empresa_codigo)
        clases_dict = {va.codigo: va.descripcion for va in valor_arboles}
    
    for cultivo in cultivos:
        # Obtener descripción de la clase
        clase_descripcion = clases_dict.get(cultivo.clase, cultivo.clase or '')
        
        # Estado fitosanitario
        estado_texto = ''
        if cultivo.estado == '1':
            estado_texto = '1 - Bueno'
        elif cultivo.estado == '2':
            estado_texto = '2 - Regular'
        elif cultivo.estado == '3':
            estado_texto = '3 - Malo'
        else:
            estado_texto = cultivo.estado or ''
        
        data.append([
            str(cultivo.id),
            cultivo.clave or '',
            clase_descripcion[:30] if len(clase_descripcion) > 30 else clase_descripcion,  # Truncar si es muy largo
            str(cultivo.arbol) if cultivo.arbol else '0',
            estado_texto,
            str(cultivo.edad) if cultivo.edad else '0',
            f"{cultivo.factor:.2f}" if cultivo.factor else '0.00',
            f"{cultivo.valor:.2f}" if cultivo.valor else '0.00'
        ])
        
        total_arboles += cultivo.arbol or 0
        total_valor += cultivo.valor or Decimal('0.00')
    
    # Agregar fila de totales
    if len(data) > 1:
        data.append(['', '', 'TOTALES:', str(total_arboles), '', '', '', f"{total_valor:.2f}"])
    
    # Crear tabla
    table = Table(data, colWidths=[0.5*inch, 1.2*inch, 2*inch, 0.8*inch, 1*inch, 0.6*inch, 0.8*inch, 1*inch])
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('ALIGN', (3, 1), (3, -2), 'RIGHT'),  # N° Árboles
        ('ALIGN', (5, 1), (5, -2), 'RIGHT'),  # Edad
        ('ALIGN', (6, 1), (6, -2), 'RIGHT'),  # Factor
        ('ALIGN', (7, 1), (7, -2), 'RIGHT'),  # Valor
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        # Fila de totales
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F4F8')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, -1), (-1, -1), 'LEFT'),
        ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
        ('ALIGN', (7, -1), (7, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y crear respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f'cultivos_permanentes_{clave_catastral or empresa_codigo or "all"}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def actualizar_cultivo_bdcata1(empresa, clave):
    """
    Función auxiliar para calcular la sumatoria de valores de cultivopermanente
    y actualizar el campo cultivo en BDCata1, además de recalcular el impuesto
    """
    from .models import CultivoPermanente, BDCata1
    from django.db.models import Sum
    from decimal import Decimal
    
    try:
        # Calcular la sumatoria de todos los valores de cultivopermanente
        # para la misma empresa y clave
        sumatoria = CultivoPermanente.objects.filter(
            empresa=empresa,
            clave=clave
        ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
        
        # Convertir a Decimal si es necesario
        if not isinstance(sumatoria, Decimal):
            sumatoria = Decimal(str(sumatoria))
        
        # Actualizar el campo cultivo en BDCata1
        bien_inmueble = BDCata1.objects.filter(
            empresa=empresa,
            cocata1=clave
        ).first()
        
        if bien_inmueble:
            bien_inmueble.cultivo = sumatoria
            
            # Recalcular el impuesto usando la misma lógica del frontend
            impuesto_calculado = calcular_impuesto_bdcata1(bien_inmueble, empresa)
            
            # Actualizar el campo impuesto en bdcata1
            bien_inmueble.impuesto = impuesto_calculado
            
            # Guardar ambos campos en una sola operación
            bien_inmueble.save(update_fields=['cultivo', 'impuesto'])
            logger.info(f'Campo cultivo e impuesto actualizados en BDCata1: empresa={empresa}, clave={clave}, cultivo={sumatoria}, impuesto={impuesto_calculado}')
            
            # Actualizar el impuesto en tasasmunicipales
            actualizar_impuesto_tasas_municipales(bien_inmueble, impuesto_calculado, empresa)
            
            # Calcular tasas municipales (rubros que empiezan con T)
            calcular_tasas_municipales_automatico(bien_inmueble, empresa)
        else:
            logger.warning(f'No se encontró BDCata1 para actualizar: empresa={empresa}, clave={clave}')
    except Exception as e:
        logger.error(f'Error al actualizar cultivo e impuesto en BDCata1: {str(e)}', exc_info=True)

@catastro_require_auth
def cultivo_permanente_create(request, empresa=None, clave=None):
    """
    Crear nuevo cultivo permanente
    """
    from .models import CultivoPermanente, BDCata1
    from .forms import CultivoPermanenteForm
    
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    clave_catastral = clave or request.GET.get('cocata1', '').strip()
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    if not empresa_codigo or not clave_catastral:
        messages.error(request, 'Debe proporcionar empresa y clave catastral.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    if request.method == 'POST':
        form = CultivoPermanenteForm(request.POST, empresa=empresa_codigo, clave=clave_catastral)
        if form.is_valid():
            cultivo = form.save(commit=False)
            cultivo.empresa = empresa_codigo
            cultivo.clave = clave_catastral
            cultivo.usuario = request.session.get('catastro_usuario_nombre', '') or ''
            cultivo.fechasys = timezone.now()
            cultivo.save()
            
            # Calcular la sumatoria de valores y actualizar BDCata1
            actualizar_cultivo_bdcata1(empresa_codigo, clave_catastral)
            
            messages.success(request, 'Cultivo permanente creado exitosamente.')
            return redirect('catastro:cultivo_permanente_list', empresa=empresa_codigo, cocata1=clave_catastral)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CultivoPermanenteForm(empresa=empresa_codigo, clave=clave_catastral)
    
    context = {
        'titulo': f'Nuevo Cultivo Permanente - Clave: {clave_catastral}',
        'form': form,
        'cultivo': None,
        'bien_inmueble': bien_inmueble,
        'empresa': empresa_codigo,
        'clave': clave_catastral,
    }
    
    return render(request, 'cultivo_permanente_form.html', context)

@catastro_require_auth
def cultivo_permanente_update(request, empresa=None, cocata1=None, pk=None):
    """
    Actualizar cultivo permanente
    """
    from .models import CultivoPermanente, BDCata1
    from .forms import CultivoPermanenteForm
    
    cultivo = get_object_or_404(CultivoPermanente, pk=pk)
    # Obtener empresa y clave directamente de la instancia (no cambiar estos valores)
    empresa_codigo = cultivo.empresa if cultivo.empresa else request.session.get('catastro_empresa', '')
    clave_catastral = cultivo.clave if cultivo.clave else ''
    
    # Validar que tenemos los valores necesarios
    if not empresa_codigo:
        messages.error(request, 'Error: No se pudo determinar la empresa del cultivo.')
        return redirect('catastro:cultivo_permanente_list', empresa=request.session.get('catastro_empresa', ''), cocata1=cultivo.clave or '')
    if not clave_catastral:
        messages.error(request, 'Error: No se pudo determinar la clave catastral del cultivo.')
        return redirect('catastro:cultivo_permanente_list', empresa=empresa_codigo, cocata1='')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    if request.method == 'POST':
        # Al editar, NO pasar empresa y clave como parámetros
        # Django usará automáticamente los valores de la instancia
        form = CultivoPermanenteForm(request.POST, instance=cultivo)
        if form.is_valid():
            cultivo = form.save(commit=False)
            # NO cambiar empresa y clave - el formulario ya preserva los valores de la instancia
            # Solo actualizar usuario y fechasys
            cultivo.usuario = request.session.get('catastro_usuario_nombre', '') or ''
            cultivo.fechasys = timezone.now()
            cultivo.save()
            
            # Usar los valores reales del cultivo guardado para actualizar BDCata1
            actualizar_cultivo_bdcata1(cultivo.empresa, cultivo.clave)
            
            messages.success(request, 'Cultivo permanente actualizado exitosamente.')
            # Usar los valores reales del cultivo guardado para el redirect
            empresa_redirect = empresa or cultivo.empresa
            cocata1_redirect = cocata1 or cultivo.clave
            return redirect('catastro:cultivo_permanente_list', empresa=empresa_redirect, cocata1=cocata1_redirect)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        # Al editar, NO pasar empresa y clave como parámetros para evitar conflictos
        # Django usará automáticamente los valores de la instancia
        form = CultivoPermanenteForm(instance=cultivo)
    
    context = {
        'titulo': f'Editar Cultivo Permanente - Clave: {clave_catastral}',
        'form': form,
        'cultivo': cultivo,  # Pasar la instancia para que el template use sus valores directamente
        'bien_inmueble': bien_inmueble,
        # Usar valores de la instancia para el contexto (no los parámetros)
        'empresa': cultivo.empresa if cultivo.empresa else empresa_codigo,
        'clave': cultivo.clave if cultivo.clave else clave_catastral,
    }
    
    return render(request, 'cultivo_permanente_form.html', context)

@catastro_require_auth
def cultivo_permanente_delete(request, empresa=None, cocata1=None, pk=None):
    """
    Eliminar cultivo permanente
    """
    from .models import CultivoPermanente
    
    cultivo = get_object_or_404(CultivoPermanente, pk=pk)
    empresa_codigo = empresa or cultivo.empresa or request.session.get('catastro_empresa', '')
    clave_catastral = cocata1 or cultivo.clave
    
    if request.method == 'POST':
        try:
            cultivo.delete()
            
            # Recalcular la sumatoria de valores y actualizar BDCata1
            actualizar_cultivo_bdcata1(empresa_codigo, clave_catastral)
            
            messages.success(request, 'Cultivo permanente eliminado exitosamente.')
            empresa_redirect = empresa or empresa_codigo
            cocata1_redirect = cocata1 or clave_catastral
            return redirect('catastro:cultivo_permanente_list', empresa=empresa_redirect, cocata1=cocata1_redirect)
        except Exception as e:
            messages.error(request, f'Error al eliminar cultivo permanente: {str(e)}')
            empresa_redirect = empresa or empresa_codigo
            cocata1_redirect = cocata1 or clave_catastral
            return redirect('catastro:cultivo_permanente_list', empresa=empresa_redirect, cocata1=cocata1_redirect)
    
    context = {
        'titulo': f'Eliminar Cultivo Permanente - Catastro',
        'cultivo': cultivo,
        'empresa': empresa_codigo,
        'clave': clave_catastral,
    }
    
    return render(request, 'cultivo_permanente_confirm_delete.html', context)

@catastro_require_auth
def buscar_clase_cultivo_ajax(request):
    """
    API endpoint para búsqueda de clase de cultivo (ValorArbol) por empresa y código o ID
    """
    from .models import ValorArbol
    
    try:
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        codigo = request.GET.get('codigo', '').strip()
        id_obj = request.GET.get('id', '').strip()
        
        if not codigo and not id_obj:
            return JsonResponse({
                'existe': False,
                'mensaje': 'Debe proporcionar un código o ID'
            })
        
        if not empresa:
            return JsonResponse({
                'existe': False,
                'mensaje': 'No se ha seleccionado un municipio'
            })
        
        # Buscar el valor de árbol por empresa y código o ID
        valor_arbol = None
        if id_obj:
            try:
                # Si se proporciona un ID, buscar por ID
                valor_arbol = ValorArbol.objects.filter(
                    id=int(id_obj),
                    empresa=empresa
                ).first()
            except (ValueError, TypeError):
                pass
        
        if not valor_arbol and codigo:
            # Buscar por código
            valor_arbol = ValorArbol.objects.filter(
                empresa=empresa,
                codigo=codigo
            ).first()
        
        if valor_arbol:
            return JsonResponse({
                'existe': True,
                'codigo': valor_arbol.codigo or '',
                'descripcion': valor_arbol.descripcion or '',
                'valor': str(valor_arbol.valor) if valor_arbol.valor else '0.00',
                'mensaje': 'Clase de cultivo encontrada'
            })
        else:
            return JsonResponse({
                'existe': False,
                'mensaje': 'No se encontró una clase de cultivo con este código o ID'
            })
    except Exception as e:
        return JsonResponse({
            'existe': False,
            'mensaje': f'Error al buscar la clase de cultivo: {str(e)}'
        }, status=500)

@csrf_exempt
def buscar_factor_cultivo_ajax(request):
    """
    API endpoint para búsqueda de factor de cultivo por empresa, código y edad
    Busca en factorcultivo donde edad >= rango1 AND edad <= rango2
    """
    from .models import FactorCultivo
    from decimal import Decimal, InvalidOperation
    
    try:
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        codigo = request.GET.get('codigo', '').strip()
        edad_str = request.GET.get('edad', '').strip()
        
        if not codigo:
            return JsonResponse({
                'existe': False,
                'mensaje': 'Debe proporcionar un código'
            })
        
        if not empresa:
            return JsonResponse({
                'existe': False,
                'mensaje': 'No se ha seleccionado un municipio'
            })
        
        if not edad_str:
            return JsonResponse({
                'existe': False,
                'mensaje': 'Debe proporcionar una edad'
            })
        
        try:
            edad = Decimal(edad_str)
        except (InvalidOperation, ValueError):
            return JsonResponse({
                'existe': False,
                'mensaje': 'La edad debe ser un número válido'
            })
        
        # Buscar el factor de cultivo por empresa, código y rango de edad
        # edad >= rango1 AND edad <= rango2
        factor_cultivo = FactorCultivo.objects.filter(
            empresa=empresa,
            codigo=codigo,
            rango1__lte=edad,  # edad >= rango1
            rango2__gte=edad   # edad <= rango2
        ).first()
        
        if factor_cultivo:
            return JsonResponse({
                'existe': True,
                'factor': str(factor_cultivo.factor) if factor_cultivo.factor else '0.000',
                'rango1': str(factor_cultivo.rango1) if factor_cultivo.rango1 else '0',
                'rango2': str(factor_cultivo.rango2) if factor_cultivo.rango2 else '0',
                'mensaje': f'Factor encontrado para edad {edad} en rango [{factor_cultivo.rango1}-{factor_cultivo.rango2}]'
            })
        else:
            return JsonResponse({
                'existe': False,
                'mensaje': f'No se encontró un factor para código {codigo} y edad {edad}'
            })
    except Exception as e:
        return JsonResponse({
            'existe': False,
            'mensaje': f'Error al buscar el factor de cultivo: {str(e)}'
        }, status=500)

@catastro_require_auth
def informacion_legal_form(request, empresa=None, cocata1=None):
    """
    Formulario único de información legal del predio
    Relación uno a uno: un registro por cada clave catastral
    """
    from .models import Legales, BDCata1
    from .forms import LegalesForm
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    usuario_nombre = request.session.get('catastro_usuario_nombre', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    if not clave_catastral:
        messages.error(request, 'Debe proporcionar un código catastral.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Verificar que existe el registro de bien inmueble
    try:
        registro = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    # Intentar obtener el registro legal existente (si existe)
    legal_existente = None
    try:
        legal_existente = Legales.objects.get(empresa=empresa_codigo, colegal=clave_catastral)
    except Legales.DoesNotExist:
        pass
    
    if request.method == 'POST':
        # Verificar si es una acción de eliminación
        accion = request.POST.get('action', '')
        if accion == 'eliminar' and legal_existente:
            try:
                legal_existente.delete()
                messages.success(request, 'Información legal eliminada correctamente.')
                return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
            except Exception as e:
                messages.error(request, f'Error al eliminar la información legal: {str(e)}')
        else:
            # Si existe, actualizar; si no, crear
            if legal_existente:
                form = LegalesForm(request.POST, instance=legal_existente, empresa=empresa_codigo, cocata1=clave_catastral)
                es_nuevo = False
            else:
                form = LegalesForm(request.POST, empresa=empresa_codigo, cocata1=clave_catastral)
                es_nuevo = True
            
            if form.is_valid():
                legal = form.save(commit=False)
                legal.empresa = empresa_codigo
                legal.colegal = clave_catastral
                legal.usuario = usuario_nombre[:50] if usuario_nombre else ''
                legal.fechasys = timezone.now()
                legal.save()
                
                if es_nuevo:
                    messages.success(request, 'Información legal creada correctamente.')
                else:
                    messages.success(request, 'Información legal actualizada correctamente.')
                
                # Recargar el registro legal actualizado para mostrarlo en el formulario
                try:
                    legal_existente = Legales.objects.get(empresa=empresa_codigo, colegal=clave_catastral)
                except Legales.DoesNotExist:
                    legal_existente = None
                
                # Recargar el formulario con los datos actualizados
                form = LegalesForm(instance=legal_existente, empresa=empresa_codigo, cocata1=clave_catastral)
                # Establecer fecha de inscripción si existe
                if legal_existente and legal_existente.inscripcion:
                    # Convertir DateField a formato YYYY-MM-DD para el input type="date"
                    if isinstance(legal_existente.inscripcion, str):
                        try:
                            from datetime import datetime
                            fecha_obj = datetime.strptime(legal_existente.inscripcion, '%d/%m/%Y').date()
                            form.fields['inscripcion'].initial = fecha_obj.strftime('%Y-%m-%d')
                        except (ValueError, TypeError):
                            try:
                                from datetime import datetime
                                fecha_obj = datetime.strptime(legal_existente.inscripcion, '%Y-%m-%d').date()
                                form.fields['inscripcion'].initial = fecha_obj.strftime('%Y-%m-%d')
                            except (ValueError, TypeError):
                                pass
                    else:
                        form.fields['inscripcion'].initial = legal_existente.inscripcion.strftime('%Y-%m-%d')
                # Establecer fecha sistema
                if legal_existente and legal_existente.fechasys:
                    form.fields['fechasys'].initial = legal_existente.fechasys.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Establecer usuario
                if usuario_nombre:
                    form.fields['usuario'].initial = usuario_nombre[:50]
                
                # NO redirigir, continuar con el renderizado del formulario más abajo
            else:
                messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        # Cargar formulario con datos existentes o vacío (método GET)
        if legal_existente:
            form = LegalesForm(instance=legal_existente, empresa=empresa_codigo, cocata1=clave_catastral)
            # Establecer fecha de inscripción si existe
            if legal_existente.inscripcion:
                # Convertir DateField a formato YYYY-MM-DD para el input type="date"
                # Asegurar que sea un objeto date, no string
                if isinstance(legal_existente.inscripcion, str):
                    # Si es string, intentar parsearlo (puede venir en formato DD/MM/YYYY)
                    try:
                        from datetime import datetime
                        fecha_obj = datetime.strptime(legal_existente.inscripcion, '%d/%m/%Y').date()
                        form.fields['inscripcion'].initial = fecha_obj.strftime('%Y-%m-%d')
                    except (ValueError, TypeError):
                        try:
                            from datetime import datetime
                            fecha_obj = datetime.strptime(legal_existente.inscripcion, '%Y-%m-%d').date()
                            form.fields['inscripcion'].initial = fecha_obj.strftime('%Y-%m-%d')
                        except (ValueError, TypeError):
                            pass
                else:
                    # Si es un objeto date, formatearlo directamente
                    form.fields['inscripcion'].initial = legal_existente.inscripcion.strftime('%Y-%m-%d')
            # Establecer fecha sistema
            if legal_existente.fechasys:
                form.fields['fechasys'].initial = legal_existente.fechasys.strftime('%Y-%m-%d %H:%M:%S')
            else:
                form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            form = LegalesForm(empresa=empresa_codigo, cocata1=clave_catastral)
            form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if usuario_nombre:
            form.fields['usuario'].initial = usuario_nombre[:50]
    
    context = {
        'titulo': 'Información Legal del Predio',
        'form': form,
        'legal': legal_existente,
        'cocata1': clave_catastral,
        'registro': registro,
        'empresa': empresa_codigo,
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
    }
    
    return render(request, 'informacion_legal_form.html', context)

@catastro_require_auth
def caracteristicas_form(request, empresa=None, cocata1=None):
    """
    Formulario único de características del predio
    Relación uno a uno: un registro por cada clave catastral
    """
    from .models import Caracteristicas, BDCata1
    from .forms import CaracteristicasForm
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    usuario_nombre = request.session.get('catastro_usuario_nombre', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    if not clave_catastral:
        messages.error(request, 'Debe proporcionar un código catastral.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Verificar que existe el registro de bien inmueble
    try:
        registro = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    # Intentar obtener el registro de características existente (si existe)
    caracteristicas_existente = None
    try:
        caracteristicas_existente = Caracteristicas.objects.get(empresa=empresa_codigo, cocata1=clave_catastral)
    except Caracteristicas.DoesNotExist:
        pass
    
    if request.method == 'POST':
        # Verificar si es una acción de eliminación
        accion = request.POST.get('action', '')
        if accion == 'eliminar' and caracteristicas_existente:
            try:
                caracteristicas_existente.delete()
                messages.success(request, 'Características eliminadas correctamente.')
                return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
            except Exception as e:
                messages.error(request, f'Error al eliminar las características: {str(e)}')
        else:
            # Si existe, actualizar; si no, crear
            if caracteristicas_existente:
                form = CaracteristicasForm(request.POST, instance=caracteristicas_existente, empresa=empresa_codigo, cocata1=clave_catastral)
                es_nuevo = False
            else:
                form = CaracteristicasForm(request.POST, empresa=empresa_codigo, cocata1=clave_catastral)
                es_nuevo = True
            
            if form.is_valid():
                caracteristicas = form.save(commit=False)
                caracteristicas.empresa = empresa_codigo
                caracteristicas.cocata1 = clave_catastral
                caracteristicas.usuario = usuario_nombre[:50] if usuario_nombre else ''
                caracteristicas.fechasys = timezone.now()
                caracteristicas.save()
                
                if es_nuevo:
                    messages.success(request, 'Características creadas correctamente.')
                else:
                    messages.success(request, 'Características actualizadas correctamente.')
                
                # Recargar el registro de características actualizado para mostrarlo en el formulario
                try:
                    caracteristicas_existente = Caracteristicas.objects.get(empresa=empresa_codigo, cocata1=clave_catastral)
                except Caracteristicas.DoesNotExist:
                    caracteristicas_existente = None
                
                # Recargar el formulario con los datos actualizados
                form = CaracteristicasForm(instance=caracteristicas_existente, empresa=empresa_codigo, cocata1=clave_catastral)
                # Establecer fecha sistema
                if caracteristicas_existente and caracteristicas_existente.fechasys:
                    form.fields['fechasys'].initial = caracteristicas_existente.fechasys.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Establecer usuario
                if usuario_nombre:
                    form.fields['usuario'].initial = usuario_nombre[:50]
                
                # NO redirigir, continuar con el renderizado del formulario más abajo
            else:
                messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        # Cargar formulario con datos existentes o vacío (método GET)
        if caracteristicas_existente:
            form = CaracteristicasForm(instance=caracteristicas_existente, empresa=empresa_codigo, cocata1=clave_catastral)
            # Establecer fecha sistema
            if caracteristicas_existente.fechasys:
                form.fields['fechasys'].initial = caracteristicas_existente.fechasys.strftime('%Y-%m-%d %H:%M:%S')
            else:
                form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            form = CaracteristicasForm(empresa=empresa_codigo, cocata1=clave_catastral)
            form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if usuario_nombre:
            form.fields['usuario'].initial = usuario_nombre[:50]
    
    # Asegurar que caracteristicas_existente esté actualizado para el contexto
    # (necesario después de guardar para mostrar el botón eliminar)
    if not caracteristicas_existente:
        try:
            caracteristicas_existente = Caracteristicas.objects.get(empresa=empresa_codigo, cocata1=clave_catastral)
        except Caracteristicas.DoesNotExist:
            caracteristicas_existente = None
    
    context = {
        'titulo': 'Características del Predio',
        'form': form,
        'caracteristicas': caracteristicas_existente,
        'cocata1': clave_catastral,
        'registro': registro,
        'empresa': empresa_codigo,
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
    }
    
    return render(request, 'caracteristicas_form.html', context)

@catastro_require_auth
def complemento_form(request, empresa=None, cocata1=None):
    """
    Formulario único de datos complementarios del predio
    Relación uno a uno: un registro por cada clave catastral
    """
    from .models import Complemento, BDCata1
    from .forms import ComplementoForm
    from django.utils import timezone
    from django.urls import reverse
    from django.contrib import messages
    
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    usuario_nombre = request.session.get('catastro_usuario_nombre', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    if not clave_catastral:
        messages.error(request, 'Debe proporcionar un código catastral.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    try:
        registro = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    complemento_existente = None
    try:
        complemento_existente = Complemento.objects.get(empresa=empresa_codigo, cocomple=clave_catastral)
    except Complemento.DoesNotExist:
        pass
    
    if request.method == 'POST':
        accion = request.POST.get('action', '')
        if accion == 'eliminar' and complemento_existente:
            try:
                complemento_existente.delete()
                messages.success(request, 'Datos complementarios eliminados correctamente.')
                return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
            except Exception as e:
                messages.error(request, f'Error al eliminar los datos complementarios: {str(e)}')
        else:
            if complemento_existente:
                form = ComplementoForm(request.POST, instance=complemento_existente, empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
                es_nuevo = False
            else:
                form = ComplementoForm(request.POST, empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
                es_nuevo = True
            
            if form.is_valid():
                complemento = form.save(commit=False)
                complemento.empresa = empresa_codigo
                complemento.cocomple = clave_catastral
                complemento.usuario = usuario_nombre[:50] if usuario_nombre else ''
                complemento.fechasys = timezone.now()
                complemento.save()
                
                if es_nuevo:
                    messages.success(request, 'Datos complementarios creados correctamente.')
                else:
                    messages.success(request, 'Datos complementarios actualizados correctamente.')
                
                try:
                    complemento_existente = Complemento.objects.get(empresa=empresa_codigo, cocomple=clave_catastral)
                except Complemento.DoesNotExist:
                    complemento_existente = None
                
                form = ComplementoForm(instance=complemento_existente, empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
                
                # Asegurar que las fechas se formateen correctamente para los inputs HTML5
                if complemento_existente:
                    from datetime import date
                    # Fecha Adquisición
                    if complemento_existente.fechaadqui:
                        if isinstance(complemento_existente.fechaadqui, date):
                            form.fields['fechaadqui'].initial = complemento_existente.fechaadqui.strftime('%Y-%m-%d')
                    # Fecha Calculo
                    if complemento_existente.Bfecal:
                        if isinstance(complemento_existente.Bfecal, date):
                            form.fields['Bfecal'].initial = complemento_existente.Bfecal.strftime('%Y-%m-%d')
                
                if complemento_existente and complemento_existente.fechasys:
                    form.fields['fechasys'].initial = complemento_existente.fechasys.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                
                if usuario_nombre:
                    form.fields['usuario'].initial = usuario_nombre[:50]
            else:
                messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        if complemento_existente:
            form = ComplementoForm(instance=complemento_existente, empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
            
            # Asegurar que las fechas se formateen correctamente para los inputs HTML5
            from datetime import date
            # Fecha Adquisición
            if complemento_existente.fechaadqui:
                if isinstance(complemento_existente.fechaadqui, date):
                    form.fields['fechaadqui'].initial = complemento_existente.fechaadqui.strftime('%Y-%m-%d')
            # Fecha Calculo
            if complemento_existente.Bfecal:
                if isinstance(complemento_existente.Bfecal, date):
                    form.fields['Bfecal'].initial = complemento_existente.Bfecal.strftime('%Y-%m-%d')
            
            if complemento_existente.fechasys:
                form.fields['fechasys'].initial = complemento_existente.fechasys.strftime('%Y-%m-%d %H:%M:%S')
            else:
                form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            form = ComplementoForm(empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
            form.fields['fechasys'].initial = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if usuario_nombre:
            form.fields['usuario'].initial = usuario_nombre[:50]
    
    context = {
        'titulo': 'Datos Complementarios',
        'form': form,
        'complemento': complemento_existente,
        'cocata1': clave_catastral,
        'registro': registro,
        'empresa': empresa_codigo,
        'usuario_nombre': usuario_nombre,
        'municipio_descripcion': request.session.get('catastro_municipio_descripcion', ''),
    }
    return render(request, 'complemento_form.html', context)

@catastro_require_auth
def colindantes_list(request, empresa=None, cocata1=None):
    """
    Lista de colindantes para una clave catastral específica
    Recibe parámetros empresa y cocata1 de la tabla bdcata1
    """
    from .models import Colindantes, BDCata1, Colindancias
    from .forms import ColindantesForm
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Obtener colindantes filtrados por empresa y clave
    colindantes = Colindantes.objects.all()
    if empresa_codigo:
        colindantes = colindantes.filter(empresa=empresa_codigo)
    if clave_catastral:
        colindantes = colindantes.filter(cocata1=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        colindantes = colindantes.filter(
            Q(cocata1__icontains=search) |
            Q(tipo__icontains=search) |
            Q(colindante__icontains=search) |
            Q(codcolinda__icontains=search)
        )
    
    colindantes = colindantes.order_by('tipo', 'id')
    
    # Obtener descripciones de códigos de colindancia
    codigos_colindancia = set()
    for colindante in colindantes:
        if colindante.codcolinda:
            codigos_colindancia.add(colindante.codcolinda)
    
    # Crear diccionario de código -> descripción
    descripciones_colindancia = {}
    if codigos_colindancia:
        colindancias_objs = Colindancias.objects.filter(codigo__in=codigos_colindancia)
        descripciones_colindancia = {str(c.codigo): c.descripcion for c in colindancias_objs}
    
    # Agrupar colindantes por tipo y agregar descripción del código
    colindantes_por_tipo = {}
    tipo_labels = {
        'N': 'Norte',
        'S': 'Sur',
        'E': 'Este',
        'O': 'Oeste',
    }
    
    for colindante in colindantes:
        tipo_key = colindante.tipo if colindante.tipo else 'Otro'
        tipo_label = tipo_labels.get(tipo_key, tipo_key)
        
        # Agregar descripción del código de colindancia
        if colindante.codcolinda:
            codigo_key = str(colindante.codcolinda).strip()
            colindante.descripcion_codcolinda = descripciones_colindancia.get(codigo_key, '')
        else:
            colindante.descripcion_codcolinda = ''
        
        if tipo_key not in colindantes_por_tipo:
            colindantes_por_tipo[tipo_key] = {
                'tipo': tipo_key,
                'tipo_label': tipo_label,
                'colindantes': [],
            }
        colindantes_por_tipo[tipo_key]['colindantes'].append(colindante)
    
    # Convertir a lista ordenada por tipo (N, S, E, O, Otro)
    orden_tipos = ['N', 'S', 'E', 'O', 'Otro']
    colindantes_por_tipo_lista = sorted(
        colindantes_por_tipo.items(),
        key=lambda x: orden_tipos.index(x[0]) if x[0] in orden_tipos else 999
    )
    
    context = {
        'titulo': f'Colindantes - Clave: {clave_catastral}' if clave_catastral else 'Colindantes - Catastro',
        'colindantes': colindantes,
        'colindantes_por_tipo': colindantes_por_tipo_lista,
        'bien_inmueble': bien_inmueble,
        'search': search,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
        'total_registros': colindantes.count(),
    }
    
    return render(request, 'colindantes_list.html', context)

@catastro_require_auth
def colindante_create(request, empresa=None, cocata1=None):
    """
    Crear nuevo colindante
    Recibe parámetros empresa y cocata1 de la tabla bdcata1
    """
    from .models import Colindantes, BDCata1
    from .forms import ColindantesForm
    
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    usuario_nombre = request.session.get('catastro_usuario_nombre', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    if not clave_catastral:
        messages.error(request, 'Debe proporcionar un código catastral.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    try:
        bien_inmueble = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    if request.method == 'POST':
        form = ColindantesForm(request.POST, empresa=empresa_codigo, cocata1=clave_catastral)
        if form.is_valid():
            colindante = form.save(commit=False)
            colindante.empresa = empresa_codigo
            colindante.cocata1 = clave_catastral
            colindante.usuario = usuario_nombre[:50] if usuario_nombre else ''
            colindante.fechasys = timezone.now()
            colindante.save()
            
            messages.success(request, 'Colindante creado exitosamente.')
            return redirect('catastro:colindantes_list', empresa=empresa_codigo, cocata1=clave_catastral)
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = ColindantesForm(empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
    
    context = {
        'titulo': f'Nuevo Colindante - Clave: {clave_catastral}',
        'form': form,
        'colindante': None,
        'bien_inmueble': bien_inmueble,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
        'usuario_nombre': usuario_nombre,
    }
    
    return render(request, 'colindante_form.html', context)

@catastro_require_auth
def colindante_update(request, empresa=None, cocata1=None, pk=None):
    """
    Actualizar colindante
    """
    from .models import Colindantes, BDCata1
    from .forms import ColindantesForm
    
    colindante = get_object_or_404(Colindantes, pk=pk)
    # Obtener empresa y clave directamente de la instancia
    empresa_codigo = colindante.empresa if colindante.empresa else request.session.get('catastro_empresa', '')
    clave_catastral = colindante.cocata1 if colindante.cocata1 else ''
    usuario_nombre = request.session.get('catastro_usuario_nombre', '')
    
    # Validar que tenemos los valores necesarios
    if not empresa_codigo:
        messages.error(request, 'Error: No se pudo determinar la empresa del colindante.')
        return redirect('catastro:colindantes_list', empresa=request.session.get('catastro_empresa', ''), cocata1=colindante.cocata1 or '')
    if not clave_catastral:
        messages.error(request, 'Error: No se pudo determinar la clave catastral del colindante.')
        return redirect('catastro:colindantes_list', empresa=empresa_codigo, cocata1='')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    try:
        bien_inmueble = BDCata1.objects.filter(
            empresa=empresa_codigo,
            cocata1=clave_catastral
        ).first()
    except Exception:
        pass
    
    if request.method == 'POST':
        form = ColindantesForm(request.POST, instance=colindante, empresa=empresa_codigo, cocata1=clave_catastral)
        if form.is_valid():
            colindante = form.save(commit=False)
            colindante.empresa = empresa_codigo
            colindante.cocata1 = clave_catastral
            colindante.usuario = usuario_nombre[:50] if usuario_nombre else ''
            colindante.fechasys = timezone.now()
            colindante.save()
            
            messages.success(request, 'Colindante actualizado exitosamente.')
            return redirect('catastro:colindantes_list', empresa=empresa_codigo, cocata1=clave_catastral)
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = ColindantesForm(instance=colindante, empresa=empresa_codigo, cocata1=clave_catastral, initial={'usuario': usuario_nombre})
    
    context = {
        'titulo': f'Editar Colindante - Clave: {clave_catastral}',
        'form': form,
        'colindante': colindante,
        'bien_inmueble': bien_inmueble,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
        'usuario_nombre': usuario_nombre,
    }
    
    return render(request, 'colindante_form.html', context)

@catastro_require_auth
def colindante_delete(request, empresa=None, cocata1=None, pk=None):
    """
    Eliminar colindante
    """
    from .models import Colindantes
    
    colindante = get_object_or_404(Colindantes, pk=pk)
    empresa_codigo = colindante.empresa if colindante.empresa else request.session.get('catastro_empresa', '')
    clave_catastral = colindante.cocata1 if colindante.cocata1 else ''
    
    if request.method == 'POST':
        colindante.delete()
        messages.success(request, 'Colindante eliminado exitosamente.')
        return redirect('catastro:colindantes_list', empresa=empresa_codigo, cocata1=clave_catastral)
    
    context = {
        'titulo': f'Eliminar Colindante - Catastro',
        'colindante': colindante,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
    }
    
    return render(request, 'colindante_confirm_delete.html', context)

@catastro_require_auth
def colindantes_export_excel(request, empresa=None, cocata1=None):
    """
    Exportar colindantes a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. Instálela con: pip install openpyxl')
        if empresa and cocata1:
            return redirect('catastro:colindantes_list', empresa=empresa, cocata1=cocata1)
        return redirect('catastro:colindantes_list')
    
    from .models import Colindantes, BDCata1, Colindancias
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener colindantes filtrados por empresa y clave
    colindantes = Colindantes.objects.all()
    if empresa_codigo:
        colindantes = colindantes.filter(empresa=empresa_codigo)
    if clave_catastral:
        colindantes = colindantes.filter(cocata1=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        colindantes = colindantes.filter(
            Q(cocata1__icontains=search) |
            Q(tipo__icontains=search) |
            Q(colindante__icontains=search) |
            Q(codcolinda__icontains=search)
        )
    
    colindantes = colindantes.order_by('tipo', 'id')
    
    # Obtener descripciones de códigos de colindancia
    codigos_colindancia = set()
    for colindante in colindantes:
        if colindante.codcolinda:
            codigos_colindancia.add(colindante.codcolinda)
    
    # Crear diccionario de código -> descripción
    descripciones_colindancia = {}
    if codigos_colindancia:
        colindancias_objs = Colindancias.objects.filter(codigo__in=codigos_colindancia)
        descripciones_colindancia = {str(c.codigo): c.descripcion for c in colindancias_objs}
    
    # Agrupar colindantes por tipo y agregar descripción del código
    colindantes_por_tipo = {}
    tipo_labels = {
        'N': 'Norte',
        'S': 'Sur',
        'E': 'Este',
        'O': 'Oeste',
    }
    
    for colindante in colindantes:
        tipo_key = colindante.tipo if colindante.tipo else 'Otro'
        tipo_label = tipo_labels.get(tipo_key, tipo_key)
        
        # Agregar descripción del código de colindancia
        if colindante.codcolinda:
            codigo_key = str(colindante.codcolinda).strip()
            colindante.descripcion_codcolinda = descripciones_colindancia.get(codigo_key, '')
        else:
            colindante.descripcion_codcolinda = ''
        
        if tipo_key not in colindantes_por_tipo:
            colindantes_por_tipo[tipo_key] = {
                'tipo': tipo_key,
                'tipo_label': tipo_label,
                'colindantes': [],
            }
        colindantes_por_tipo[tipo_key]['colindantes'].append(colindante)
    
    # Convertir a lista ordenada por tipo (N, S, E, O, Otro)
    orden_tipos = ['N', 'S', 'E', 'O', 'Otro']
    colindantes_por_tipo_lista = sorted(
        colindantes_por_tipo.items(),
        key=lambda x: orden_tipos.index(x[0]) if x[0] in orden_tipos else 999
    )
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Crear libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colindantes"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    tipo_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    tipo_font = Font(bold=True, size=12)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    titulo_texto = f"COLINDANTES"
    if bien_inmueble:
        titulo_texto += f" - Clave: {clave_catastral}"
    ws.append([titulo_texto])
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.append([])
    
    # Información del bien inmueble si existe
    if bien_inmueble:
        ws.append(['INFORMACIÓN DEL BIEN INMUEBLE'])
        ws.merge_cells('A3:G3')
        info_cell = ws['A3']
        info_cell.font = Font(bold=True, size=12)
        ws.append(['Clave Catastral:', clave_catastral, '', 
                   'Propietario:', f"{bien_inmueble.nombres or ''} {bien_inmueble.apellidos or ''}".strip()])
        ws.append(['Identidad:', bien_inmueble.identidad or '', '', 
                   'Ubicación:', bien_inmueble.ubicacion or ''])
        ws.append([])
        row_start = 7
    else:
        row_start = 3
    
    # Agregar colindantes agrupados por tipo
    current_row = row_start
    
    for tipo_key, datos_tipo in colindantes_por_tipo_lista:
        # Encabezado de tipo
        tipo_header = f"Colindantes al {datos_tipo['tipo_label']} ({tipo_key})"
        ws.append([tipo_header])
        ws.merge_cells(f'A{current_row}:G{current_row}')
        tipo_cell = ws[f'A{current_row}']
        tipo_cell.fill = tipo_fill
        tipo_cell.font = tipo_font
        tipo_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Encabezados de tabla
        headers = ['ID', 'Clave Catastral', 'Colindante', 'Código Colindante', 'Usuario', 'Fecha Sistema']
        ws.append(headers)
        
        # Aplicar estilo a encabezados
        for cell in ws[current_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        current_row += 1
        
        # Agregar datos de colindantes
        for colindante in datos_tipo['colindantes']:
            # Formatear código colindante con descripción
            codigo_colindante = colindante.codcolinda or ''
            if codigo_colindante and hasattr(colindante, 'descripcion_codcolinda') and colindante.descripcion_codcolinda:
                codigo_colindante = f"{codigo_colindante} - {colindante.descripcion_codcolinda}"
            
            row = [
                colindante.id,
                colindante.cocata1 or '',
                colindante.colindante or '',
                codigo_colindante,
                colindante.usuario or '',
                colindante.fechasys.strftime('%d/%m/%Y %H:%M') if colindante.fechasys else ''
            ]
            ws.append(row)
            
            # Aplicar estilo a datos
            for cell in ws[current_row]:
                cell.border = border_style
                cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        # Agregar espacio entre grupos
        ws.append([])
        current_row += 1
    
    # Ajustar ancho de columnas
    column_widths = [8, 20, 40, 20, 30, 20]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'colindantes_{clave_catastral or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@catastro_require_auth
def colindantes_export_pdf(request, empresa=None, cocata1=None):
    """
    Exportar colindantes a PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. Instálela con: pip install reportlab')
        if empresa and cocata1:
            return redirect('catastro:colindantes_list', empresa=empresa, cocata1=cocata1)
        return redirect('catastro:colindantes_list')
    
    from .models import Colindantes, BDCata1, Colindancias
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener colindantes filtrados por empresa y clave
    colindantes = Colindantes.objects.all()
    if empresa_codigo:
        colindantes = colindantes.filter(empresa=empresa_codigo)
    if clave_catastral:
        colindantes = colindantes.filter(cocata1=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        colindantes = colindantes.filter(
            Q(cocata1__icontains=search) |
            Q(tipo__icontains=search) |
            Q(colindante__icontains=search) |
            Q(codcolinda__icontains=search)
        )
    
    colindantes = colindantes.order_by('tipo', 'id')
    
    # Obtener descripciones de códigos de colindancia
    codigos_colindancia = set()
    for colindante in colindantes:
        if colindante.codcolinda:
            codigos_colindancia.add(colindante.codcolinda)
    
    # Crear diccionario de código -> descripción
    descripciones_colindancia = {}
    if codigos_colindancia:
        colindancias_objs = Colindancias.objects.filter(codigo__in=codigos_colindancia)
        descripciones_colindancia = {str(c.codigo): c.descripcion for c in colindancias_objs}
    
    # Agrupar colindantes por tipo y agregar descripción del código
    colindantes_por_tipo = {}
    tipo_labels = {
        'N': 'Norte',
        'S': 'Sur',
        'E': 'Este',
        'O': 'Oeste',
    }
    
    for colindante in colindantes:
        tipo_key = colindante.tipo if colindante.tipo else 'Otro'
        tipo_label = tipo_labels.get(tipo_key, tipo_key)
        
        # Agregar descripción del código de colindancia
        if colindante.codcolinda:
            codigo_key = str(colindante.codcolinda).strip()
            colindante.descripcion_codcolinda = descripciones_colindancia.get(codigo_key, '')
        else:
            colindante.descripcion_codcolinda = ''
        
        if tipo_key not in colindantes_por_tipo:
            colindantes_por_tipo[tipo_key] = {
                'tipo': tipo_key,
                'tipo_label': tipo_label,
                'colindantes': [],
            }
        colindantes_por_tipo[tipo_key]['colindantes'].append(colindante)
    
    # Convertir a lista ordenada por tipo (N, S, E, O, Otro)
    orden_tipos = ['N', 'S', 'E', 'O', 'Otro']
    colindantes_por_tipo_lista = sorted(
        colindantes_por_tipo.items(),
        key=lambda x: orden_tipos.index(x[0]) if x[0] in orden_tipos else 999
    )
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Centrado
        spaceAfter=20
    )
    
    titulo_texto = "COLINDANTES"
    if bien_inmueble:
        titulo_texto += f" - Clave: {clave_catastral}"
    elements.append(Paragraph(titulo_texto, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del bien inmueble si existe
    if bien_inmueble:
        info_data = [
            ['INFORMACIÓN DEL BIEN INMUEBLE', ''],
            ['Clave Catastral:', clave_catastral],
            ['Propietario:', f"{bien_inmueble.nombres or ''} {bien_inmueble.apellidos or ''}".strip()],
            ['Identidad:', bien_inmueble.identidad or ''],
            ['Ubicación:', bien_inmueble.ubicacion or '']
        ]
        info_table = Table(info_data, colWidths=[2*inch, 5*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D3D3D3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Agregar colindantes agrupados por tipo
    for tipo_key, datos_tipo in colindantes_por_tipo_lista:
        # Encabezado de tipo
        tipo_style = ParagraphStyle(
            'TipoStyle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            alignment=0,  # Izquierda
            spaceAfter=10
        )
        tipo_texto = f"Colindantes al {datos_tipo['tipo_label']} ({tipo_key})"
        elements.append(Paragraph(tipo_texto, tipo_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Encabezados de tabla
        headers = ['ID', 'Clave Catastral', 'Colindante', 'Código Colindante']
        data = [headers]
        
        # Agregar datos
        for colindante in datos_tipo['colindantes']:
            # Formatear código colindante con descripción
            codigo_colindante = colindante.codcolinda or ''
            if codigo_colindante and hasattr(colindante, 'descripcion_codcolinda') and colindante.descripcion_codcolinda:
                codigo_colindante = f"{codigo_colindante} - {colindante.descripcion_codcolinda}"
            
            row = [
                str(colindante.id),
                colindante.cocata1 or '',
                colindante.colindante or '',
                codigo_colindante,
            ]
            data.append(row)
        
        # Crear tabla con anchos ajustados para acomodar descripción del código
        # ID, Clave Catastral, Colindante, Código Colindante (con descripción)
        table = Table(data, colWidths=[0.8*inch, 2.5*inch, 3.5*inch, 4.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('WORDWRAP', (0, 0), (-1, -1), True),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener valor del buffer
    buffer.seek(0)
    
    # Crear respuesta HTTP
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f'colindantes_{clave_catastral or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@catastro_require_auth
def copropietarios_list(request, empresa=None, cocata1=None):
    """
    Lista de copropietarios para una clave catastral específica
    Recibe parámetros empresa y cocata1 de la tabla bdcata1
    """
    from .models import Copropietarios, BDCata1
    from .forms import CopropietariosForm
    
    # Obtener empresa de parámetros o sesión
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    # Obtener clave catastral de parámetros o GET
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    if clave_catastral and empresa_codigo:
        try:
            bien_inmueble = BDCata1.objects.filter(
                empresa=empresa_codigo,
                cocata1=clave_catastral
            ).first()
        except Exception:
            pass
    
    # Obtener copropietarios filtrados por empresa y clave
    copropietarios = Copropietarios.objects.all()
    if empresa_codigo:
        copropietarios = copropietarios.filter(empresa=empresa_codigo)
    if clave_catastral:
        copropietarios = copropietarios.filter(cocata1=clave_catastral)
    
    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        copropietarios = copropietarios.filter(
            Q(cocata1__icontains=search) |
            Q(nombre__icontains=search) |
            Q(identidad__icontains=search)
        )
    
    copropietarios = copropietarios.order_by('id')
    
    # Calcular suma de porcentajes
    suma_porcentajes = copropietarios.aggregate(Sum('porcentaje'))['porcentaje__sum'] or Decimal('0.00')
    
    context = {
        'titulo': 'Gestión de Copropietarios - Catastro',
        'copropietarios': copropietarios,
        'bien_inmueble': bien_inmueble,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
        'suma_porcentajes': suma_porcentajes,
        'search': search,
    }
    
    return render(request, 'copropietarios_list.html', context)

@catastro_require_auth
def copropietario_create(request, empresa=None, cocata1=None):
    """
    Crear nuevo copropietario
    Recibe parámetros empresa y cocata1 de la tabla bdcata1
    """
    from .models import Copropietarios, BDCata1
    from .forms import CopropietariosForm
    
    empresa_codigo = empresa or request.session.get('catastro_empresa', '')
    clave_catastral = cocata1 or request.GET.get('cocata1', '').strip()
    usuario_nombre = request.session.get('catastro_usuario_nombre', '')
    
    if not empresa_codigo:
        messages.error(request, 'No se encontró información de empresa en la sesión.')
        return redirect('catastro:catastro_login')
    
    if not clave_catastral:
        messages.error(request, 'Debe proporcionar un código catastral.')
        return redirect('catastro:bienes_inmuebles_registrar')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    try:
        bien_inmueble = BDCata1.objects.get(cocata1=clave_catastral, empresa=empresa_codigo)
    except BDCata1.DoesNotExist:
        messages.error(request, f'No se encontró un registro con la clave catastral {clave_catastral}.')
        return redirect(f"{reverse('catastro:bienes_inmuebles_registrar')}?cocata1={clave_catastral}")
    
    if request.method == 'POST':
        form = CopropietariosForm(request.POST, empresa=empresa_codigo, cocata1=clave_catastral)
        if form.is_valid():
            try:
                copropietario = form.save(commit=False)
                copropietario.empresa = empresa_codigo
                copropietario.cocata1 = clave_catastral
                copropietario.save()
                messages.success(request, 'Copropietario creado exitosamente.')
                return redirect('catastro:copropietarios_list', empresa=empresa_codigo, cocata1=clave_catastral)
            except Exception as e:
                if 'copropietarios_idx1' in str(e) or 'UNIQUE' in str(e).upper():
                    # Si ya existe, actualizar el registro existente
                    try:
                        copropietario_existente = Copropietarios.objects.get(
                            empresa=empresa_codigo,
                            cocata1=clave_catastral,
                            identidad=form.cleaned_data.get('identidad', '')
                        )
                        # Actualizar los campos
                        copropietario_existente.nombre = form.cleaned_data.get('nombre', '')
                        copropietario_existente.porcentaje = form.cleaned_data.get('porcentaje')
                        copropietario_existente.save()
                        messages.success(request, 'Copropietario actualizado exitosamente.')
                        return redirect('catastro:copropietarios_list', empresa=empresa_codigo, cocata1=clave_catastral)
                    except Copropietarios.DoesNotExist:
                        messages.error(request, 'Ya existe un copropietario con esta empresa, clave catastral e identidad.')
                else:
                    messages.error(request, f'Error al crear el copropietario: {str(e)}')
    else:
        form = CopropietariosForm(empresa=empresa_codigo, cocata1=clave_catastral)
    
    context = {
        'titulo': 'Nuevo Copropietario - Catastro',
        'form': form,
        'bien_inmueble': bien_inmueble,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
    }
    
    return render(request, 'copropietario_form.html', context)

@catastro_require_auth
def copropietario_update(request, empresa=None, cocata1=None, pk=None):
    """
    Actualizar copropietario
    """
    from .models import Copropietarios, BDCata1
    from .forms import CopropietariosForm
    
    copropietario = get_object_or_404(Copropietarios, pk=pk)
    # Obtener empresa y clave directamente de la instancia
    empresa_codigo = copropietario.empresa if copropietario.empresa else request.session.get('catastro_empresa', '')
    clave_catastral = copropietario.cocata1 if copropietario.cocata1 else ''
    
    # Validar que tenemos los valores necesarios
    if not empresa_codigo:
        messages.error(request, 'Error: No se pudo determinar la empresa del copropietario.')
        return redirect('catastro:copropietarios_list', empresa=request.session.get('catastro_empresa', ''), cocata1=copropietario.cocata1 or '')
    if not clave_catastral:
        messages.error(request, 'Error: No se pudo determinar la clave catastral del copropietario.')
        return redirect('catastro:copropietarios_list', empresa=empresa_codigo, cocata1='')
    
    # Obtener información del bien inmueble si existe
    bien_inmueble = None
    try:
        bien_inmueble = BDCata1.objects.filter(
            empresa=empresa_codigo,
            cocata1=clave_catastral
        ).first()
    except Exception:
        pass
    
    if request.method == 'POST':
        form = CopropietariosForm(request.POST, instance=copropietario, empresa=empresa_codigo, cocata1=clave_catastral)
        if form.is_valid():
            try:
                copropietario = form.save(commit=False)
                copropietario.save()
                messages.success(request, 'Copropietario actualizado exitosamente.')
                return redirect('catastro:copropietarios_list', empresa=empresa_codigo, cocata1=clave_catastral)
            except Exception as e:
                if 'copropietarios_idx1' in str(e) or 'UNIQUE' in str(e).upper():
                    messages.error(request, 'Ya existe un copropietario con esta empresa, clave catastral e identidad.')
                else:
                    messages.error(request, f'Error al actualizar el copropietario: {str(e)}')
    else:
        form = CopropietariosForm(instance=copropietario, empresa=empresa_codigo, cocata1=clave_catastral)
    
    context = {
        'titulo': 'Editar Copropietario - Catastro',
        'form': form,
        'copropietario': copropietario,
        'bien_inmueble': bien_inmueble,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
    }
    
    return render(request, 'copropietario_form.html', context)

@catastro_require_auth
def copropietario_delete(request, empresa=None, cocata1=None, pk=None):
    """
    Eliminar copropietario
    """
    from .models import Copropietarios
    
    copropietario = get_object_or_404(Copropietarios, pk=pk)
    empresa_codigo = copropietario.empresa if copropietario.empresa else request.session.get('catastro_empresa', '')
    clave_catastral = copropietario.cocata1 if copropietario.cocata1 else ''
    
    if request.method == 'POST':
        copropietario.delete()
        messages.success(request, 'Copropietario eliminado exitosamente.')
        return redirect('catastro:copropietarios_list', empresa=empresa_codigo, cocata1=clave_catastral)
    
    context = {
        'titulo': f'Eliminar Copropietario - Catastro',
        'copropietario': copropietario,
        'empresa': empresa_codigo,
        'cocata1': clave_catastral,
    }
    
    return render(request, 'copropietario_confirm_delete.html', context)

@csrf_exempt
@csrf_exempt
@catastro_require_auth
def buscar_identidad_copropietario_ajax(request):
    """
    API endpoint para buscar identificación por número de identidad (DNI) para copropietarios
    Busca en la tabla identificacion y devuelve el nombre concatenado
    """
    try:
        from .models import Identificacion
    except ImportError as e:
        logger.error(f"Error al importar Identificacion: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error al cargar el modelo'}, status=500)
    
    identidad = request.GET.get('identidad', '').strip()
    
    if not identidad:
        return JsonResponse({
            'encontrado': False,
            'mensaje': 'Debe proporcionar un número de identidad'
        })
    
    try:
        # Buscar identificación por número de identidad
        identificacion = Identificacion.objects.filter(identidad=identidad).first()
        
        if identificacion:
            # Concatenar nombres y apellidos
            nombre_completo = f"{identificacion.nombres or ''} {identificacion.apellidos or ''}".strip()
            
            return JsonResponse({
                'encontrado': True,
                'identidad': identificacion.identidad,
                'nombres': identificacion.nombres or '',
                'apellidos': identificacion.apellidos or '',
                'nombre_completo': nombre_completo,
                'fechanac': identificacion.fechanac.strftime('%Y-%m-%d') if identificacion.fechanac else None,
            })
        else:
            return JsonResponse({
                'encontrado': False,
                'mensaje': 'No se encontró una identificación con ese número'
            })
    except Exception as e:
        logger.error(f"Error al buscar identificación: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': f'Error al buscar identificación: {str(e)}'
        }, status=500)


# =============================================================================
# VISTAS AJAX PARA MISCELÁNEOS - CATASTRO
# =============================================================================

@csrf_exempt
@catastro_require_auth
def buscar_concepto_miscelaneos_ajax(request):
    """Vista AJAX para buscar conceptos de cobro en misceláneos - Catastro"""
    if request.method != 'POST':
        return JsonResponse({
            'exito': False,
            'mensaje': 'Método no permitido'
        })
    
    try:
        # Obtener datos del FormData
        empresa = request.POST.get('empresa', '').strip()
        cod_tarifa = request.POST.get('cod_tarifa', '').strip()
        
        # Validar campos requeridos
        if not empresa or not cod_tarifa:
            return JsonResponse({
                'exito': False,
                'mensaje': 'Empresa y código de tarifa son requeridos'
            })
        
        from tributario.models import Tarifas
        
        # Buscar la tarifa en la base de datos
        try:
            tarifa = Tarifas.objects.get(
                empresa=empresa, 
                cod_tarifa=cod_tarifa
            )
            
            return JsonResponse({
                'exito': True,
                'concepto': {
                    'cod_tarifa': tarifa.cod_tarifa or '',
                    'descripcion': tarifa.descripcion or '',
                    'valor': str(tarifa.valor) if tarifa.valor else '0',
                    'frecuencia': tarifa.frecuencia or '',
                    'tipo': tarifa.tipo or '',
                    'ano': str(tarifa.ano) if tarifa.ano else '',
                    'rubro': tarifa.rubro or ''
                },
                'mensaje': 'Concepto encontrado exitosamente'
            })
            
        except Tarifas.DoesNotExist:
            return JsonResponse({
                'exito': False,
                'mensaje': 'No se encontró un concepto con ese código'
            })
            
    except Exception as e:
        logger.error(f"Error al buscar concepto misceláneos: {str(e)}", exc_info=True)
        return JsonResponse({
            'exito': False,
            'mensaje': f'Error interno: {str(e)}'
        }, status=500)


@csrf_exempt
@catastro_require_auth
def cargar_actividades_ajax(request):
    """Vista AJAX para cargar actividades por empresa - Catastro"""
    if request.method == 'GET':
        try:
            empresa = request.GET.get('empresa', '').strip()
            
            if not empresa:
                return JsonResponse({
                    'exito': False,
                    'actividades': [],
                    'mensaje': 'Empresa es obligatoria'
                })
            
            # Cargar actividades de la tabla actividad filtradas por empresa
            try:
                from tributario.models import Actividad
                actividades = Actividad.objects.filter(empresa=empresa).order_by('codigo')
                
                # Convertir a lista de diccionarios para JSON
                actividades_list = []
                for actividad in actividades:
                    actividades_list.append({
                        'codigo': actividad.codigo,
                        'descripcion': actividad.descripcion or ''
                    })
                
                logger.info(f"Actividades cargadas para empresa {empresa}: {len(actividades_list)} actividades")
                
                return JsonResponse({
                    'exito': True,
                    'actividades': actividades_list,
                    'mensaje': f'{len(actividades_list)} actividades cargadas'
                })
            except Exception as e:
                logger.error(f"Error al cargar actividades: {str(e)}", exc_info=True)
                return JsonResponse({
                    'exito': False,
                    'actividades': [],
                    'mensaje': f'Error al cargar actividades: {str(e)}'
                })
                
        except Exception as e:
            logger.error(f"Error en carga AJAX de actividades: {str(e)}", exc_info=True)
            return JsonResponse({
                'exito': False,
                'actividades': [],
                'mensaje': f'Error en el servidor: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'exito': False,
        'actividades': [],
        'mensaje': 'Método no permitido'
    })


@csrf_exempt
@catastro_require_auth
def obtener_tarifas_rubro_ajax(request):
    """Vista AJAX para obtener tarifas por empresa - Catastro"""
    if request.method == 'GET':
        try:
            empresa = request.GET.get('empresa', '').strip()
            
            if not empresa:
                return JsonResponse({
                    'exito': False,
                    'tarifas': [],
                    'mensaje': 'Empresa es obligatoria'
                })
            
            # Obtener tarifas filtradas por empresa
            try:
                from tributario.models import Tarifas
                tarifas = Tarifas.objects.filter(empresa=empresa).order_by('cod_tarifa')
                
                # Convertir a formato JSON
                tarifas_data = []
                for tarifa in tarifas:
                    tarifas_data.append({
                        'cod_tarifa': tarifa.cod_tarifa or '',
                        'codigo': tarifa.cod_tarifa or '',  # Alias para compatibilidad
                        'descripcion': tarifa.descripcion or '',
                        'valor': str(tarifa.valor) if tarifa.valor else '0',
                        'rubro': tarifa.rubro or '',
                        'empresa': tarifa.empresa or ''
                    })
                
                logger.info(f"Tarifas cargadas para empresa {empresa}: {len(tarifas_data)} tarifas")
                
                return JsonResponse({
                    'exito': True,
                    'tarifas': tarifas_data,
                    'mensaje': f'Se encontraron {len(tarifas_data)} tarifas'
                })
            except Exception as e:
                logger.error(f"Error al obtener tarifas: {str(e)}", exc_info=True)
                return JsonResponse({
                    'exito': False,
                    'tarifas': [],
                    'mensaje': f'Error al obtener tarifas: {str(e)}'
                })
                
        except Exception as e:
            logger.error(f"Error en obtener tarifas AJAX: {str(e)}", exc_info=True)
            return JsonResponse({
                'exito': False,
                'tarifas': [],
                'mensaje': f'Error en el servidor: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'exito': False,
        'tarifas': [],
        'mensaje': 'Método no permitido'
    })


@csrf_exempt
@catastro_require_auth
def buscar_actividad_ajax(request):
    """Vista AJAX para buscar actividad por empresa y código - Catastro"""
    if request.method == 'GET':
        try:
            empresa = request.GET.get('empresa', '').strip()
            codigo = request.GET.get('codigo', '').strip()
            
            if not empresa or not codigo:
                return JsonResponse({
                    'exito': False,
                    'existe': False,
                    'descripcion': '',
                    'mensaje': 'Empresa y código son obligatorios'
                })
            
            try:
                from tributario.models import Actividad
                
                # Buscar actividad
                actividad = Actividad.objects.filter(
                    empresa=empresa,
                    codigo=codigo
                ).first()
                
                if actividad:
                    return JsonResponse({
                        'exito': True,
                        'existe': True,
                        'descripcion': actividad.descripcion or '',
                        'mensaje': 'Actividad encontrada'
                    })
                else:
                    return JsonResponse({
                        'exito': False,
                        'existe': False,
                        'descripcion': '',
                        'mensaje': 'No se encontró una actividad con ese código'
                    })
                    
            except Exception as e:
                logger.error(f"Error al buscar actividad: {str(e)}", exc_info=True)
                return JsonResponse({
                    'exito': False,
                    'existe': False,
                    'descripcion': '',
                    'mensaje': f'Error al buscar actividad: {str(e)}'
                })
                
        except Exception as e:
            logger.error(f"Error en buscar actividad AJAX: {str(e)}", exc_info=True)
            return JsonResponse({
                'exito': False,
                'existe': False,
                'descripcion': '',
                'mensaje': f'Error en el servidor: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'exito': False,
        'existe': False,
        'descripcion': '',
        'mensaje': 'Método no permitido'
    })


@csrf_exempt
@catastro_require_auth
@csrf_exempt
def enviar_a_caja_ajax(request):
    """Vista AJAX para enviar transacción a caja e insertar en pagovariostemp - Catastro"""
    if request.method == 'POST':
        try:
            from datetime import datetime
            from decimal import Decimal
            from django.http import JsonResponse
            from django.db import connection, transaction
            from tributario.models import PagoVariosTemp, NoRecibos
            
            # Obtener empresa de la sesión de catastro
            empresa = request.session.get('catastro_empresa')
            if not empresa:
                # Si no hay en sesión, intentar del POST
                empresa = request.POST.get('empresa', '').strip()
            
            # Procesar datos del formulario
            fecha_str = request.POST.get('fecha', '').strip()
            dni = request.POST.get('dni', '').strip()
            nombre = request.POST.get('nombre', '').strip()
            direccion = request.POST.get('direccion', '').strip()
            comentario = request.POST.get('comentario', '').strip()
            oficina = request.POST.get('oficina', '').strip()
            
            # Validar campos obligatorios
            if not empresa or not fecha_str or not dni or not nombre:
                return JsonResponse({
                    'exito': False,
                    'mensaje': 'Empresa, fecha, DNI y nombre son obligatorios'
                })
            
            # Convertir fecha
            try:
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'exito': False,
                    'mensaje': 'Formato de fecha inválido. Use YYYY-MM-DD'
                })
            
            # Generar número de recibo usando la secuencia de norecibos
            numero_recibo_decimal = NoRecibos.obtener_siguiente_numero()
            numero_recibo = f"REC-{numero_recibo_decimal}"
            
            # Log para debug
            logger.info(f"Enviando a caja (Catastro) - Empresa: {empresa}, Oficina: {oficina}, DNI: {dni}, Nombre: {nombre}")
            
            # Procesar conceptos del formulario
            conceptos_procesados = []
            total_general = Decimal('0.00')
            
            # Buscar todos los conceptos en el formulario
            i = 0
            while True:
                codigo = request.POST.get(f'form-{i}-codigo', '').strip()
                if not codigo:  # Si no hay más conceptos
                    break
                
                descripcion = request.POST.get(f'form-{i}-descripcion', '').strip()
                cantidad = request.POST.get(f'form-{i}-cantidad', '1').strip()
                vl_unit = request.POST.get(f'form-{i}-vl_unit', '0').strip()
                valor = request.POST.get(f'form-{i}-valor', '0').strip()
                
                if codigo and valor and float(valor.replace(',', '.')) > 0:
                    try:
                        cantidad_decimal = Decimal(cantidad.replace(',', '.')) if cantidad else Decimal('1.00')
                        vl_unit_decimal = Decimal(vl_unit.replace(',', '.')) if vl_unit else Decimal('0.00')
                        valor_decimal = Decimal(valor.replace(',', '.')) if valor else Decimal('0.00')
                        
                        # Obtener usuario de la sesión de catastro
                        usuario = request.session.get('catastro_usuario_nombre', request.session.get('usuario', 'SISTEMA'))
                        
                        # Crear registro en pagovariostemp.
                        # En algunos despliegues de Supabase, la columna `id` puede quedar sin identity
                        # luego de una migración/importación. Si falla por `id NULL`, aplicamos fallback.
                        payload = dict(
                            empresa=empresa,
                            recibo=numero_recibo_decimal,
                            codigo=codigo,
                            fecha=fecha,
                            identidad=dni,
                            nombre=nombre,
                            descripcion=descripcion,
                            valor=valor_decimal,
                            comentario=comentario,
                            oficina=oficina,
                            facturadora='',
                            aplicado='0',
                            traslado='0',
                            solvencia=0,
                            fecha_solv=None,
                            cantidad=cantidad_decimal,
                            vl_unit=vl_unit_decimal,
                            deposito=0,
                            cajero=usuario,
                            usuario=usuario,
                            referencia='',
                            banco='',
                            Tipofa=' ',
                            Rtm=' ',
                            expe=0,
                            pagodia=0,
                            rcaja=valor_decimal,
                            Rfechapag=fecha,
                            permiso=0,
                            Fechavence=None,
                            direccion=direccion,
                            prima='',
                            sexo='',
                            rtn=dni
                        )
                        try:
                            pago_temp = PagoVariosTemp.objects.create(**payload)
                        except Exception as create_error:
                            error_text = str(create_error).lower()
                            if ('null value' in error_text and 'column id' in error_text) or (
                                'pagovariostemp_id_seq' in error_text
                            ):
                                with transaction.atomic():
                                    with connection.cursor() as cursor:
                                        cursor.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM pagovariostemp")
                                        next_id = cursor.fetchone()[0]
                                    payload['id'] = int(next_id)
                                    pago_temp = PagoVariosTemp.objects.create(**payload)
                                logger.warning(
                                    "Fallback aplicado en pagovariostemp: id asignado manualmente (%s).",
                                    payload['id'],
                                )
                            else:
                                raise
                        
                        conceptos_procesados.append({
                            'codigo': codigo,
                            'descripcion': descripcion,
                            'cantidad': str(cantidad_decimal),
                            'vl_unit': str(vl_unit_decimal),
                            'valor': str(valor_decimal)
                        })
                        
                        total_general += valor_decimal
                        
                        # Log para debug
                        logger.info(f"Concepto guardado (Catastro) - Recibo: {numero_recibo}, Código: {codigo}, Oficina: {oficina}")
                        
                    except Exception as e:
                        logger.error(f"Error al procesar concepto {codigo}: {str(e)}", exc_info=True)
                        return JsonResponse({
                            'exito': False,
                            'mensaje': f'Error al procesar concepto {codigo}: {str(e)}'
                        })
                
                i += 1
            
            if not conceptos_procesados:
                return JsonResponse({
                    'exito': False,
                    'mensaje': 'No se encontraron conceptos válidos para procesar'
                })
            
            return JsonResponse({
                'exito': True,
                'mensaje': f'Transacción enviada a caja exitosamente. Número de recibo: {numero_recibo}',
                'numero_recibo': numero_recibo,
                'total_general': str(total_general),
                'conceptos_procesados': len(conceptos_procesados),
                'conceptos': conceptos_procesados
            })
            
        except Exception as e:
            logger.error(f"Error al enviar a caja (Catastro): {str(e)}", exc_info=True)
            return JsonResponse({
                'exito': False,
                'mensaje': f'Error: {str(e)}'
            })
    
    return JsonResponse({
        'exito': False,
        'mensaje': 'Método no permitido'
    })

def ver_soporte(request, numero_recibo):
    """Vista para mostrar el soporte de transacción usando datos de pagovariostemp - Catastro"""
    try:
        from tributario.models import PagoVariosTemp
        from decimal import Decimal
        from django.utils import timezone
        
        # Convertir número de recibo a formato numérico para búsqueda
        try:
            numero_limpio = str(numero_recibo).strip()
            empresa_recibo = None
            if '-' in numero_limpio:
                partes = [p for p in numero_limpio.split('-') if p]
                if len(partes) >= 2:
                    empresa_recibo = partes[0]
                    numero_limpio = partes[-1]
            elif numero_limpio.upper().startswith('REC-'):
                numero_limpio = numero_limpio[4:]
            recibo_numero = Decimal(numero_limpio)
        except (ValueError, TypeError):
            return render(request, 'error.html', {
                'error': f'Formato de número de recibo inválido: {numero_recibo}',
                'modulo': 'Catastro'
            })
        
        # Buscar todos los registros de pagovariostemp para este recibo
        filtros = {'recibo': recibo_numero}
        if empresa_recibo:
            filtros['empresa'] = empresa_recibo
        else:
            # Si no hay empresa en el número, usar la de la sesión
            empresa_sesion = request.session.get('catastro_empresa')
            if empresa_sesion:
                filtros['empresa'] = empresa_sesion
        
        pagos = PagoVariosTemp.objects.filter(**filtros).order_by('codigo')
        
        if not pagos.exists():
            return render(request, 'error.html', {
                'error': f'No se encontraron registros para el recibo {numero_recibo}',
                'modulo': 'Catastro'
            })
        
        # Obtener datos del primer registro para información general
        primer_pago = pagos.first()
        
        # Procesar conceptos desde los registros de pagovariostemp
        conceptos = []
        total_general = Decimal('0.00')
        
        for pago in pagos:
            concepto = {
                'codigo': pago.codigo,
                'descripcion': pago.descripcion or '',
                'cantidad': str(pago.cantidad) if pago.cantidad else '1',
                'vl_unit': str(pago.vl_unit) if pago.vl_unit else '0.00',
                'valor': str(pago.valor.quantize(Decimal('0.01')) if pago.valor is not None else Decimal('0.00'))
            }
            conceptos.append(concepto)
            total_general += pago.valor if pago.valor else Decimal('0.00')
        
        # Generar QR code para el recibo
        qr_data = (
            f"Municipio: {primer_pago.empresa or ''}\n"
            f"Recibo: {numero_recibo}\n"
            f"Contribuyente: {primer_pago.nombre or ''}\n"
            f"Total: L. {total_general:.2f}\n"
            f"Fecha: {primer_pago.fecha}"
        )
        import qrcode
        import io
        import base64
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        qr_code = base64.b64encode(buffer.getvalue()).decode()
        
        return render(request, 'soporte_simple.html', {
            'numero_recibo': numero_recibo,
            'fecha': primer_pago.fecha,
            'hora_actual': timezone.now().strftime('%H:%M:%S'),
            'dni': primer_pago.identidad,
            'nombre': primer_pago.nombre,
            'direccion': primer_pago.direccion,
            'comentario': (primer_pago.comentario or '').strip(),
            'usuario': (primer_pago.usuario or '').strip() or 'SISTEMA',
            'cajero': (primer_pago.cajero or '').strip() or 'N/A',
            'oficina': (primer_pago.oficina or '').strip() or 'N/A',
            'conceptos': conceptos,
            'total_general': total_general.quantize(Decimal('0.01')) if total_general else Decimal('0.00'),
            'qr_code': qr_code,
            'modulo': 'Catastro',
            'descripcion': 'Soporte de Transacción Miscelánea'
        })
        
    except Exception as e:
        logger.error(f"Error al generar soporte (Catastro): {str(e)}", exc_info=True)
        return render(request, 'error.html', {
            'error': f'Error al generar soporte: {str(e)}',
            'modulo': 'Catastro'
        })

# Importar vistas de tasas municipales y aplicar decoradores
from .views_tasas_municipales import configurar_tasas_municipales as _configurar_tasas_municipales, obtener_tarifas_rubro_bienes as _obtener_tarifas_rubro_bienes

# Aplicar decoradores a las vistas importadas
configurar_tasas_municipales = catastro_require_auth(_configurar_tasas_municipales)
obtener_tarifas_rubro_bienes = catastro_require_auth(_obtener_tarifas_rubro_bienes)

# ============================================================================
# VISTAS AJAX PARA TASAS MUNICIPALES
# ============================================================================

@catastro_require_auth
def ajax_tasas_municipales(request):
    """
    Vista AJAX para obtener las tasas municipales filtradas por empresa y clave
    Incluye la descripción del rubro desde la tabla rubros
    Optimizado para máximo rendimiento usando SQL directo con JOIN
    """
    try:
        from django.db import connection
        from decimal import Decimal
        
        empresa = request.GET.get('empresa', '').strip()
        clave = request.GET.get('clave', '').strip()
        
        if not empresa or not clave:
            return JsonResponse({
                'success': False,
                'message': 'Empresa y clave son requeridos'
            })
        
        # Usar SQL directo con LEFT JOIN para mejor rendimiento
        # Esto evita múltiples consultas y es más rápido que el ORM
        # Una sola consulta con JOIN es mucho más eficiente que múltiples consultas
        with connection.cursor() as cursor:
            # Consulta SQL optimizada con JOIN
            # MySQL manejará automáticamente las diferencias de collation
            sql = """
                SELECT 
                    tm.id,
                    tm.empresa,
                    tm.clave,
                    tm.rubro,
                    COALESCE(r.descripcion, '') as rubro_descripcion,
                    tm.cod_tarifa,
                    tm.valor,
                    tm.cuenta,
                    tm.cuentarez
                FROM tasassmunicipales tm
                LEFT JOIN rubros r ON tm.empresa = r.empresa AND tm.rubro = r.codigo
                WHERE tm.empresa = %s AND tm.clave = %s
                ORDER BY tm.rubro, tm.cod_tarifa
            """
            
            cursor.execute(sql, [empresa, clave])
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
        
        # Convertir resultados a lista de diccionarios
        tasas_result = []
        for row in results:
            row_dict = dict(zip(columns, row))
            tasas_result.append({
                'id': row_dict.get('id'),
                'empresa': row_dict.get('empresa') or '',
                'clave': row_dict.get('clave') or '',
                'rubro': row_dict.get('rubro') or '',
                'rubro_descripcion': (row_dict.get('rubro_descripcion') or '').strip(),
                'cod_tarifa': row_dict.get('cod_tarifa') or '',
                'valor': str(row_dict.get('valor', Decimal('0.00'))),
                'cuenta': row_dict.get('cuenta') or '',
                'cuentarez': row_dict.get('cuentarez') or '',
            })
        
        return JsonResponse({
            'success': True,
            'tasas': tasas_result,
            'total': len(tasas_result)
        })
        
    except Exception as e:
        logger.error(f"Error en ajax_tasas_municipales: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener tasas municipales: {str(e)}'
        })

@catastro_require_auth
def ajax_calcular_tasas_municipales(request):
    """
    Vista AJAX para calcular las tasas municipales (rubros que empiezan con T)
    Procesa cada rubro según su tipo:
    - Tipo "F" (Fijo): Usa el valor directamente de la tabla tarifas
    - Tipo "V" (Variable): Busca en planarbitio según el Avalúo total del bien inmueble
    """
    try:
        from .models import BDCata1, TasasMunicipales
        from tributario.models import Tarifas, PlanArbitrio
        from datetime import datetime
        from django.db.models import Q
        from decimal import Decimal, InvalidOperation
        
        empresa = request.GET.get('empresa', '').strip() or request.session.get('catastro_empresa', '')
        clave = request.GET.get('clave', '').strip()
        
        def parse_decimal_param(name):
            raw = request.GET.get(name)
            if raw is None or raw == '':
                return None
            try:
                return Decimal(str(raw))
            except (InvalidOperation, ValueError):
                return None

        num_viviendas_param = parse_decimal_param('num_viviendas')
        num_cuartos_param = parse_decimal_param('num_cuartos')
        num_apartamentos_param = parse_decimal_param('num_apartamentos')
        avaluo_total_override = parse_decimal_param('avaluo_total')
        
        if not empresa or not clave:
            return JsonResponse({
                'success': False,
                'message': 'Empresa y clave catastral son requeridos'
            })
        
        # Obtener el bien inmueble
        try:
            bien_inmueble = BDCata1.objects.get(empresa=empresa, cocata1=clave)
        except BDCata1.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'No se encontró un bien inmueble con la clave {clave}'
            })
        
        # Calcular el Avalúo total
        valor_tierra = bien_inmueble.bvl2tie or Decimal('0.00')
        edificaciones = bien_inmueble.mejoras or Decimal('0.00')
        detalles_adicionales = bien_inmueble.detalle or Decimal('0.00')
        cultivo_permanente = bien_inmueble.cultivo or Decimal('0.00')
        valor_declarado = bien_inmueble.declarado or Decimal('0.00')
        
        avaluo_total = valor_tierra + edificaciones + detalles_adicionales + cultivo_permanente + valor_declarado
        if avaluo_total_override is not None:
            avaluo_total = avaluo_total_override
        
        num_viviendas = num_viviendas_param if num_viviendas_param is not None else (bien_inmueble.vivienda or Decimal('0'))
        num_cuartos = num_cuartos_param if num_cuartos_param is not None else (bien_inmueble.cuartos or Decimal('0'))
        num_apartamentos = num_apartamentos_param if num_apartamentos_param is not None else (bien_inmueble.apartamentos or Decimal('0'))
        num_viviendas = max(num_viviendas, Decimal('0'))
        num_cuartos = max(num_cuartos, Decimal('0'))
        num_apartamentos = max(num_apartamentos, Decimal('0'))
        
        # Año predeterminado
        ano_predeterminado = int(datetime.now().year)
        
        # Función helper para determinar el año de la tarifa
        def determinar_ano_tarifa(tarifa_obj):
            if tarifa_obj and tarifa_obj.ano is not None:
                try:
                    return int(tarifa_obj.ano)
                except (TypeError, ValueError, InvalidOperation):
                    return ano_predeterminado
            return ano_predeterminado
        
        # Función helper para obtener el valor del plan por tipocat
        def obtener_valor_plan_por_tipocat(tasa_municipal, tipocat_objetivo, ano_objetivo):
            planes = PlanArbitrio.objects.filter(
                empresa=empresa,
                rubro=tasa_municipal.rubro,
                cod_tarifa=tasa_municipal.cod_tarifa,
                ano=ano_objetivo,
                tipocat=str(tipocat_objetivo)
            ).order_by('codigo', 'id')

            if not planes.exists():
                return Decimal('0.00')

            valor_encontrado = Decimal('0.00')
            for plan in planes:
                minimo = plan.minimo if plan.minimo is not None else Decimal('0.00')
                maximo = plan.maximo if plan.maximo is not None else Decimal('9999999999.99')
                if minimo <= avaluo_total <= maximo:
                    valor_encontrado = plan.valor or Decimal('0.00')
                    break

            if valor_encontrado == Decimal('0.00'):
                for plan in planes:
                    minimo = plan.minimo if plan.minimo is not None else Decimal('0.00')
                    if minimo <= avaluo_total:
                        valor_encontrado = plan.valor or Decimal('0.00')
                        break

            if valor_encontrado == Decimal('0.00') and planes.exists():
                valor_encontrado = planes.first().valor or Decimal('0.00')

            return valor_encontrado
        
        # Obtener todos los rubros de tasasmunicipales que empiezan con "T"
        tasas_municipales = TasasMunicipales.objects.filter(
            empresa=empresa,
            clave=clave,
            rubro__startswith='T'
        )
        
        if not tasas_municipales.exists():
            return JsonResponse({
                'success': True,
                'message': 'No se encontraron rubros de tasas municipales (que empiecen con T) para calcular',
                'calculadas': 0,
                'errores': []
            })
        
        # Procesar cada tasa municipal
        tasas_calculadas = 0
        errores = []
        ano_vigente = ano_predeterminado
        
        for tasa_municipal in tasas_municipales:
            try:
                # Buscar la tarifa correspondiente en la tabla tarifas
                tarifas_query = Tarifas.objects.filter(
                    empresa=empresa,
                    rubro=tasa_municipal.rubro,
                    cod_tarifa=tasa_municipal.cod_tarifa
                ).order_by('-ano')
                tarifa = tarifas_query.first()
                if not tarifa:
                    errores.append(f"Tarifa no encontrada: Rubro {tasa_municipal.rubro}, Cod_Tarifa {tasa_municipal.cod_tarifa}")
                    continue
                
                # Determinar el tipo de tarifa
                tipo_tarifa = (tarifa.tipo or '').strip().upper()
                nuevo_valor = Decimal('0.00')
                
                if tipo_tarifa == 'F':
                    # Tipo Fijo: Usar el valor directamente de la tarifa
                    nuevo_valor = tarifa.valor or Decimal('0.00')
                    
                elif tipo_tarifa == 'V':
                    try:
                        plan_ano = determinar_ano_tarifa(tarifa)
                        ano_vigente = plan_ano
                        
                        # Detectar si es solar baldío:
                        # Valor terreno > 0, edificación = 0, viviendas = 0, cuartos = 0, apartamentos = 0
                        es_solar_baldio = (
                            valor_tierra > Decimal('0.00') and
                            edificaciones == Decimal('0.00') and
                            num_viviendas == Decimal('0') and
                            num_cuartos == Decimal('0') and
                            num_apartamentos == Decimal('0')
                        )
                        
                        if es_solar_baldio:
                            # Solar baldío: usar exclusivamente categoría 2 del plan de arbitrio
                            nuevo_valor_variable = obtener_valor_plan_por_tipocat(tasa_municipal, '2', plan_ano)
                            logger.info(f"Solar baldío detectado para Rubro {tasa_municipal.rubro}: usando categoría 2, valor={nuevo_valor_variable}, Avalúo: {avaluo_total}")
                        else:
                            # Buscar valor base con tipocat = 1 (siempre se busca)
                            valor_base = obtener_valor_plan_por_tipocat(tasa_municipal, '1', plan_ano)
                            logger.debug(f"Valor base (tipocat=1) para Rubro {tasa_municipal.rubro}: {valor_base}, Año: {plan_ano}, Avalúo: {avaluo_total}")
                            
                            # Inicializar el valor variable
                            nuevo_valor_variable = Decimal('0.00')
                            
                            # Si todos los campos son cero, solo usar el valor de tipocat = 1
                            if num_viviendas == 0 and num_cuartos == 0 and num_apartamentos == 0:
                                nuevo_valor_variable = valor_base
                                logger.debug(f"Todos los contadores son cero, usando valor base: {valor_base}")
                            else:
                                # Si num_viviendas = 0, usar el valor base directamente
                                # Si num_viviendas > 0, multiplicar el valor base por num_viviendas
                                if num_viviendas == 0:
                                    nuevo_valor_variable = valor_base
                                    logger.debug(f"Viviendas=0, usando valor base: {valor_base}")
                                else:
                                    nuevo_valor_variable = valor_base * num_viviendas
                                    logger.debug(f"Viviendas={num_viviendas}, multiplicando: {valor_base} * {num_viviendas} = {nuevo_valor_variable}")
                                
                                # Si num_cuartos > 0, buscar tipocat = 3, multiplicar y sumar
                                if num_cuartos > 0:
                                    valor_cuartos = obtener_valor_plan_por_tipocat(tasa_municipal, '3', plan_ano)
                                    nuevo_valor_variable += valor_cuartos * num_cuartos
                                    logger.debug(f"Cuartos={num_cuartos}, valor_cuartos={valor_cuartos}, sumando: {valor_cuartos * num_cuartos}, total acumulado: {nuevo_valor_variable}")
                                
                                # Si num_apartamentos > 0, buscar tipocat = 4, multiplicar y sumar
                                if num_apartamentos > 0:
                                    valor_apartamentos = obtener_valor_plan_por_tipocat(tasa_municipal, '4', plan_ano)
                                    nuevo_valor_variable += valor_apartamentos * num_apartamentos
                                    logger.debug(f"Apartamentos={num_apartamentos}, valor_apartamentos={valor_apartamentos}, sumando: {valor_apartamentos * num_apartamentos}, total acumulado: {nuevo_valor_variable}")

                        nuevo_valor = nuevo_valor_variable.quantize(Decimal('0.01'))
                        logger.debug(f"Valor final calculado para Rubro {tasa_municipal.rubro}: {nuevo_valor} {'(SOLAR BALDÍO - Cat.2)' if es_solar_baldio else ''}")
                    except Exception as e:
                        errores.append(f"Error al buscar plan de arbitrio para Rubro {tasa_municipal.rubro}: {str(e)}")
                        logger.error(f"Error detallado al buscar plan de arbitrio: {str(e)}", exc_info=True)
                        continue
                else:
                    errores.append(f"Tipo de tarifa desconocido '{tipo_tarifa}' para Rubro {tasa_municipal.rubro}")
                    continue
                
                # Actualizar el valor en tasasmunicipales
                tasa_municipal.valor = nuevo_valor
                tasa_municipal.save(update_fields=['valor'])
                tasas_calculadas += 1
                
                logger.info(f"Tasa municipal calculada: Empresa={empresa}, Clave={clave}, Rubro={tasa_municipal.rubro}, Tipo={tipo_tarifa}, Valor={nuevo_valor}, Viviendas={num_viviendas}, Cuartos={num_cuartos}, Apartamentos={num_apartamentos}")
                
            except Exception as e:
                error_msg = f"Error al procesar tasa Rubro {tasa_municipal.rubro}: {str(e)}"
                errores.append(error_msg)
                logger.error(error_msg, exc_info=True)
                continue
        
        mensaje = f'Se calcularon {tasas_calculadas} tasa(s) municipal(es) exitosamente.'
        if errores:
            mensaje += f' Errores: {len(errores)}'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'calculadas': tasas_calculadas,
            'errores': errores,
            'avaluo_total': str(avaluo_total),
            'ano_vigente': ano_vigente
        })
        
    except Exception as e:
        logger.error(f"Error en ajax_calcular_tasas_municipales: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error al calcular tasas municipales: {str(e)}'
        })

@catastro_require_auth
def ajax_obtener_tasa_impositiva(request):
    """
    Vista AJAX para obtener la tasa impositiva según el tipo de ficha (Rural/Urbana)
    Ficha 1 = Urbana -> usa tasau
    Ficha 2 = Rural -> usa tasar
    Intenta usar el modelo primero, si falla usa SQL directo
    """
    try:
        from django.db import connection
        from core.models import Municipio
        
        empresa = request.GET.get('empresa', '')
        ficha = request.GET.get('ficha', '')
        
        if not empresa:
            return JsonResponse({
                'success': False,
                'message': 'Empresa es requerida'
            })
        
        if not ficha:
            return JsonResponse({
                'success': False,
                'message': 'Ficha es requerida'
            })
        
        # Intentar usar el modelo primero
        tasau_val = None
        tasar_val = None
        
        try:
            municipio = Municipio.objects.get(codigo=empresa)
            # Intentar acceder a los campos del modelo
            tasau_val = getattr(municipio, 'tasau', None)
            tasar_val = getattr(municipio, 'tasar', None)
        except (Municipio.DoesNotExist, AttributeError) as e:
            logger.warning(f"No se pudo obtener municipio del modelo o campos no disponibles: {str(e)}")
            tasau_val = None
            tasar_val = None
        
        # Si no se pudieron obtener del modelo, usar SQL directo
        if tasau_val is None or tasar_val is None:
            try:
                with connection.cursor() as cursor:
                    # Consulta SQL directa para obtener tasau y tasar
                    cursor.execute(
                        "SELECT tasau, tasar FROM municipio WHERE codigo = %s",
                        [empresa]
                    )
                    row = cursor.fetchone()
                    
                    if not row:
                        return JsonResponse({
                            'success': False,
                            'message': f'No se encontró el municipio con código {empresa}'
                        })
                    
                    tasau_val = row[0] if row[0] is not None else 0.00
                    tasar_val = row[1] if row[1] is not None else 0.00
                    
            except Exception as e:
                logger.error(f"Error al ejecutar consulta SQL: {str(e)}", exc_info=True)
                return JsonResponse({
                    'success': False,
                    'message': f'Error al obtener datos del municipio: {str(e)}'
                })
        
        # Determinar qué tasa usar según la ficha
        # Ficha 1 = Urbana -> tasau
        # Ficha 2 = Rural -> tasar
        tasa_impositiva = Decimal('0.00')
        if ficha == '1' or ficha == 1:
            # Ficha Urbana
            tasa_impositiva = Decimal(str(tasau_val)) if tasau_val is not None else Decimal('0.00')
        elif ficha == '2' or ficha == 2:
            # Ficha Rural
            tasa_impositiva = Decimal(str(tasar_val)) if tasar_val is not None else Decimal('0.00')
        
        return JsonResponse({
            'success': True,
            'tasa_impositiva': str(tasa_impositiva),
            'ficha': str(ficha),
            'tipo': 'Urbana' if (ficha == '1' or ficha == 1) else 'Rural'
        })
        
    except Exception as e:
        logger.error(f"Error en ajax_obtener_tasa_impositiva: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener tasa impositiva: {str(e)}'
        })

@catastro_require_auth
def ajax_verificar_registro_guardado(request):
    """
    Vista AJAX para verificar si un registro de bien inmueble ya está guardado en la BD
    """
    try:
        empresa = request.GET.get('empresa', '')
        cocata1 = request.GET.get('cocata1', '')
        
        if not empresa or not cocata1:
            return JsonResponse({
                'success': False,
                'existe': False,
                'message': 'Empresa y código catastral son requeridos'
            })
        
        # Verificar si existe el registro en bdcata1
        from .models import BDCata1
        existe = BDCata1.objects.filter(empresa=empresa, cocata1=cocata1).exists()
        
        return JsonResponse({
            'success': True,
            'existe': existe,
            'message': 'Registro encontrado' if existe else 'Registro no encontrado'
        })
        
    except Exception as e:
        logger.error(f"Error en ajax_verificar_registro_guardado: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'existe': False,
            'message': f'Error al verificar registro: {str(e)}'
        })

@catastro_require_auth
def ajax_obtener_datos_municipio(request):
    """
    Vista AJAX para obtener datos del municipio (vl_exento, por_concer)
    """
    try:
        from django.db import connection
        
        empresa = request.GET.get('empresa', '')
        
        if not empresa:
            return JsonResponse({
                'success': False,
                'message': 'Empresa es requerida'
            })
        
        # Usar consulta SQL directa para obtener vl_exento y por_concer
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT vl_exento, por_concer FROM municipio WHERE codigo = %s",
                    [empresa]
                )
                row = cursor.fetchone()
                
                if not row:
                    return JsonResponse({
                        'success': False,
                        'message': f'No se encontró el municipio con código {empresa}'
                    })
                
                vl_exento = float(row[0]) if row[0] is not None else 0.00
                por_concer = float(row[1]) if row[1] is not None else 0.00
                
        except Exception as e:
            logger.error(f"Error al ejecutar consulta SQL: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error al obtener datos del municipio: {str(e)}'
            })
        
        return JsonResponse({
            'success': True,
            'vl_exento': vl_exento,
            'por_concer': por_concer
        })
        
    except Exception as e:
        logger.error(f"Error en ajax_obtener_datos_municipio: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener datos del municipio: {str(e)}'
        })

@catastro_require_auth
@csrf_exempt
def ajax_guardar_tasa_municipal(request):
    """
    Vista AJAX para guardar una nueva tasa municipal
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Método no permitido'
        })
    
    try:
        empresa = request.POST.get('empresa', '')
        clave = request.POST.get('clave', '').strip()
        rubro = request.POST.get('rubro', '').strip()
        cod_tarifa = request.POST.get('cod_tarifa', '').strip() or None
        valor = request.POST.get('valor', '0.00')
        cuenta = request.POST.get('cuenta', '').strip()
        cuentarez = request.POST.get('cuentarez', '').strip()
        
        # Validaciones
        if not empresa:
            return JsonResponse({
                'success': False,
                'message': 'Empresa es requerida'
            })
        
        if not clave:
            return JsonResponse({
                'success': False,
                'message': 'Clave es requerida'
            })
        
        if not rubro:
            return JsonResponse({
                'success': False,
                'message': 'Rubro es requerido'
            })
        
        try:
            valor_decimal = Decimal(valor)
            if valor_decimal < 0:
                return JsonResponse({
                    'success': False,
                    'message': 'El valor debe ser mayor o igual a cero'
                })
        except (ValueError, InvalidOperation):
            return JsonResponse({
                'success': False,
                'message': 'El valor debe ser un número válido'
            })
        
        # Verificar si ya existe una tasa con la misma empresa, clave y rubro
        tasa_existente = TasasMunicipales.objects.filter(
            empresa=empresa,
            clave=clave,
            rubro=rubro
        ).first()
        
        if tasa_existente:
            return JsonResponse({
                'success': False,
                'message': 'Ya existe una tasa municipal con esta empresa, clave y rubro'
            })
        
        # Crear nueva tasa municipal
        nueva_tasa = TasasMunicipales.objects.create(
            empresa=empresa,
            clave=clave,
            rubro=rubro,
            cod_tarifa=cod_tarifa,
            valor=valor_decimal,
            cuenta=cuenta,
            cuentarez=cuentarez
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Tasa municipal guardada exitosamente',
            'tasa_id': nueva_tasa.id
        })
        
    except Exception as e:
        logger.error(f"Error en ajax_guardar_tasa_municipal: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error al guardar tasa municipal: {str(e)}'
        })

@catastro_require_auth
def ajax_convertir_latlng_a_utm(request):
    """
    Endpoint AJAX para convertir coordenadas lat/lng a UTM
    """
    lat = request.GET.get('lat') or request.POST.get('lat')
    lng = request.GET.get('lng') or request.POST.get('lng')
    
    if not lat or not lng:
        return JsonResponse({'error': 'Se requieren lat y lng'}, status=400)
    
    try:
        lat_float = float(lat)
        lng_float = float(lng)
        
        # Intentar importar si no está disponible
        global latlng_to_utm
        if not COORDENADAS_UTILS_AVAILABLE or latlng_to_utm is None:
            logger.warning("COORDENADAS_UTILS_AVAILABLE es False o funciones no disponibles, intentando importar nuevamente...")
            if not _importar_utilidades_coordenadas():
                logger.error("No se pudieron importar las utilidades de coordenadas")
                return JsonResponse({
                    'error': 'La librería pyproj no está instalada. Instale con: pip install pyproj',
                    'success': False
                }, status=503)
        
        if latlng_to_utm is None:
            logger.error("Función latlng_to_utm es None después de intentar importar")
            return JsonResponse({
                'error': 'La librería pyproj no está instalada. Instale con: pip install pyproj',
                'success': False
            }, status=503)
        
        logger.info(f"Convirtiendo coordenadas: lat={lat_float}, lng={lng_float}")
        logger.info(f"Tipo de latlng_to_utm: {type(latlng_to_utm)}")
        logger.info(f"latlng_to_utm es callable: {callable(latlng_to_utm)}")
        
        try:
            # Llamar a la función de conversión
            result = latlng_to_utm(lat_float, lng_float)
            logger.info(f"Resultado de conversión (tipo: {type(result)}): {result}")
            
            if result is None:
                logger.error("La función retornó None directamente")
                return JsonResponse({'error': 'Error en la conversión: la función retornó None'}, status=500)
            
            if not isinstance(result, tuple) or len(result) != 2:
                logger.error(f"La función retornó un tipo inesperado: {type(result)}")
                return JsonResponse({'error': f'Error en la conversión: tipo de retorno inesperado: {type(result)}'}, status=500)
            
            easting, northing = result
            logger.info(f"Resultado parseado: easting={easting}, northing={northing}")
            
            if easting is not None and northing is not None:
                return JsonResponse({
                    'easting': float(easting),
                    'northing': float(northing),
                    'success': True
                })
            else:
                logger.error(f"Conversión retornó None: easting={easting}, northing={northing}")
                # Retornar error más descriptivo cuando pyproj no está disponible
                error_msg = 'Error en la conversión de coordenadas lat/lng a UTM. '
                if not COORDENADAS_UTILS_AVAILABLE:
                    error_msg += 'La librería pyproj no está instalada. Instale con: pip install pyproj'
                else:
                    error_msg += 'Verifique que las coordenadas sean válidas y que pyproj esté instalado correctamente.'
                return JsonResponse({'error': error_msg, 'success': False}, status=503)
        except Exception as conv_error:
            logger.error(f"Excepción al llamar latlng_to_utm: {str(conv_error)}", exc_info=True)
            import traceback
            logger.error(f"Traceback completo: {traceback.format_exc()}")
            error_msg = f'Error al ejecutar conversión: {str(conv_error)}'
            if 'pyproj' in str(conv_error).lower() or 'module' in str(conv_error).lower():
                error_msg = 'La librería pyproj no está instalada. Instale con: pip install pyproj'
            return JsonResponse({'error': error_msg, 'success': False}, status=503)
            
    except (ValueError, TypeError) as e:
        return JsonResponse({'error': f'Coordenadas inválidas: {str(e)}', 'success': False}, status=400)
    except NameError as e:
        logger.error(f"NameError al convertir lat/lng a UTM: {str(e)}")
        return JsonResponse({
            'error': 'La librería pyproj no está instalada. Instale con: pip install pyproj',
            'success': False
        }, status=503)
    except Exception as e:
        logger.error(f"Error al convertir lat/lng a UTM: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'Error interno: {str(e)}', 'success': False}, status=500)

@catastro_require_auth
def ajax_convertir_utm_a_latlng(request):
    """
    Endpoint AJAX para convertir coordenadas UTM a lat/lng
    """
    easting = request.GET.get('easting') or request.POST.get('easting')
    northing = request.GET.get('northing') or request.POST.get('northing')
    
    if not easting or not northing:
        return JsonResponse({'error': 'Se requieren easting y northing'}, status=400)
    
    try:
        easting_float = float(easting)
        northing_float = float(northing)
        
        # Intentar importar si no está disponible
        global utm_to_latlng
        if not COORDENADAS_UTILS_AVAILABLE or utm_to_latlng is None:
            logger.warning("COORDENADAS_UTILS_AVAILABLE es False o funciones no disponibles, intentando importar nuevamente...")
            if not _importar_utilidades_coordenadas():
                logger.error("No se pudieron importar las utilidades de coordenadas")
                return JsonResponse({
                    'error': 'La librería pyproj no está instalada. Instale con: pip install pyproj',
                    'success': False
                }, status=503)
        
        if utm_to_latlng is None:
            logger.error("Función utm_to_latlng es None después de intentar importar")
            return JsonResponse({
                'error': 'La librería pyproj no está instalada. Instale con: pip install pyproj',
                'success': False
            }, status=503)
        
        logger.info(f"Convirtiendo coordenadas UTM: easting={easting_float}, northing={northing_float}")
        try:
            lat, lng = utm_to_latlng(easting_float, northing_float)
            logger.info(f"Resultado de conversión: lat={lat}, lng={lng}")
            
            if lat is not None and lng is not None:
                return JsonResponse({
                    'lat': lat,
                    'lng': lng,
                    'success': True
                })
            else:
                logger.error(f"Conversión retornó None: lat={lat}, lng={lng}")
                # Retornar error más descriptivo cuando pyproj no está disponible
                error_msg = 'Error en la conversión de coordenadas UTM a lat/lng. '
                if not COORDENADAS_UTILS_AVAILABLE:
                    error_msg += 'La librería pyproj no está instalada. Instale con: pip install pyproj'
                else:
                    error_msg += 'Verifique que las coordenadas UTM sean válidas y que pyproj esté instalado correctamente.'
                return JsonResponse({'error': error_msg, 'success': False}, status=503)
        except Exception as conv_error:
            logger.error(f"Excepción al llamar utm_to_latlng: {str(conv_error)}", exc_info=True)
            error_msg = f'Error al ejecutar conversión: {str(conv_error)}'
            if 'pyproj' in str(conv_error).lower() or 'module' in str(conv_error).lower():
                error_msg = 'La librería pyproj no está instalada. Instale con: pip install pyproj'
            return JsonResponse({'error': error_msg, 'success': False}, status=503)
            
    except (ValueError, TypeError) as e:
        return JsonResponse({'error': f'Coordenadas inválidas: {str(e)}', 'success': False}, status=400)
    except NameError as e:
        logger.error(f"NameError al convertir UTM a lat/lng: {str(e)}")
        return JsonResponse({
            'error': 'La librería pyproj no está instalada. Instale con: pip install pyproj',
            'success': False
        }, status=503)
    except Exception as e:
        logger.error(f"Error al convertir UTM a lat/lng: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'Error interno: {str(e)}', 'success': False}, status=500)
