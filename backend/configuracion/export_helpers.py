# -*- coding: utf-8 -*-
"""Exportación tabular a Excel y PDF para catálogos de configuración."""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse


def cell_str(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Sí' if value else 'No'
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, 'f')
    return str(value).strip()


def queryset_to_matrix(queryset, field_names, headers=None):
    """Lista de filas [str, ...] y cabeceras."""
    if headers is None:
        headers = [fn.replace('_', ' ').title() for fn in field_names]
    rows = []
    for obj in queryset:
        rows.append([cell_str(getattr(obj, fn, None)) for fn in field_names])
    return headers, rows


def model_non_pk_fields(model):
    return [f.name for f in model._meta.fields if not f.primary_key]


def build_http_excel(filename_base, sheet_title, headers, rows):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None, 'openpyxl'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else 'Datos'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for c in col:
            max_len = max(max_len, len(str(c.value or '')))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    return response, None


def build_http_pdf(filename_base, title, headers, rows):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return None, 'reportlab'

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(title or 'Listado', styles['Title']))
    story.append(Spacer(1, 0.4 * cm))

    table_data = [headers] + rows
    col_count = len(headers)
    avail = landscape(A4)[0] - 2 * cm
    col_w = avail / max(col_count, 1)
    tbl = Table(table_data, colWidths=[col_w] * col_count, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    return response, None
